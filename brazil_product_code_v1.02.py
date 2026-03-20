# -*- coding: utf-8 -*-
import os
import re
import csv
import time
import shutil
import logging
import sqlite3
import itertools
import unicodedata
import difflib
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED

from brazil_tool.core.models import Invoice, Item
from brazil_tool.core.parser import parse_invoice_from_text
from brazil_tool.core.pdf import extract_text_from_pdf
from brazil_tool.core.llm import run_llm_assist, apply_llm_result
from brazil_tool.core.utils import calculate_similarity, br_to_float
from brazil_tool.core.report_parser import CollectionReportParser
from brazil_tool.core.statement_parser import BankStatementParser
from brazil_tool.db.payment_manager import PaymentManager


try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None
try:
    import fitz  # PyMuPDF
except ImportError:
    fitz = None
try:
    from PIL import Image
except ImportError:
    Image = None
try:
    from pdf2image import convert_from_path
except ImportError:
    convert_from_path = None
try:
    import pytesseract
except ImportError:
    pytesseract = None
try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    HAS_MATPLOTLIB = True
except Exception:
    plt = None
    FigureCanvas = None
    Figure = None
    HAS_MATPLOTLIB = False

import PySide6
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
                             QPushButton, QFileDialog, QTableWidget, QTableWidgetItem,
                             QProgressBar, QLabel, QLineEdit, QCheckBox, QHeaderView,
                             QDialog, QDialogButtonBox, QTextEdit, QSizePolicy,
                             QMessageBox, QAbstractItemView, QTabWidget, QComboBox,
                             QMenu, QWidgetAction, QFrame, QStyledItemDelegate, QStyleOptionProgressBar, QStyle, QInputDialog,
                             QListWidget, QListWidgetItem, QSpinBox, QRadioButton, QButtonGroup,
                             QDoubleSpinBox, QFormLayout, QGroupBox, QDateEdit, QSplitter,
                             QTreeWidget, QTreeWidgetItem, QStackedWidget)
from PySide6.QtCore import QThread, Signal, Qt, QSettings, QByteArray, QTimer, QRect, QUrl
from PySide6.QtGui import (QPalette, QColor, QIcon, QDesktopServices, QFont, QPixmap, 
                           QImage, QPainter, QPen, QBrush, QPolygon, QShortcut, QKeySequence)

# --- Helpers ---
def format_date_gui(date_str):
    """
    将各种日期格式统一转换为 YYYY/MM/DD 格式用于显示
    """
    if not date_str or not isinstance(date_str, str) or date_str.lower() in ['none', 'null', '']:
        return ""
    
    date_str = date_str.strip()
    try:
        # 如果已经是 YYYY/MM/DD，直接返回
        if re.match(r'^\d{4}/\d{2}/\d{2}$', date_str):
            return date_str

        # 尝试匹配 YYYY-MM-DD
        if '-' in date_str:
            d_part = date_str.split()[0]
            parts = d_part.split('-')
            if len(parts[0]) == 4: # YYYY-MM-DD
                dt = datetime.strptime(d_part, "%Y-%m-%d")
            else: # DD-MM-YYYY
                dt = datetime.strptime(d_part, "%d-%m-%Y")
            return dt.strftime("%Y/%m/%d")
            
        # 尝试匹配 DD/MM/YYYY
        if '/' in date_str:
            d_part = date_str.split()[0]
            parts = d_part.split('/')
            if len(parts[0]) == 4: # YYYY/MM/DD (covered above but just in case)
                dt = datetime.strptime(d_part, "%Y/%m/%d")
            else: # DD/MM/YYYY
                dt = datetime.strptime(d_part, "%d/%m/%Y")
            return dt.strftime("%Y/%m/%d")

        # 如果是 8 位数字 (YYYYMMDD)
        if len(date_str) == 8 and date_str.isdigit():
            return f"{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}"
            
        return date_str
    except:
        return date_str

def safe_parse_date(date_str):
    """
    尝试多种格式解析日期字符串，返回 datetime 对象或 None
    支持 DD/MM/YYYY, YYYY/MM/DD, YYYY-MM-DD, DD-MM-YYYY 等
    """
    if not date_str or not isinstance(date_str, str):
        return None
    date_str = date_str.strip()
    if not date_str or date_str.lower() in ['none', 'null', '']:
        return None
        
    # 提取日期部分 (有的带时间)
    d_part = date_str.split()[0]
    
    # 优先尝试常见的格式
    for fmt in ["%d/%m/%Y", "%Y/%m/%d", "%Y-%m-%d", "%d-%m-%Y"]:
        try:
            return datetime.strptime(d_part, fmt)
        except (ValueError, IndexError):
            continue
            
    # 如果带斜杠但不是上述格式，尝试手动拆分
    if '/' in d_part:
        parts = d_part.split('/')
        if len(parts) == 3:
            try:
                if len(parts[0]) == 4: # YYYY/MM/DD
                    return datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                elif len(parts[2]) == 4: # DD/MM/YYYY
                    return datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            except: pass

    return None

def safe_parse_date_to_date(date_str):
    """
    尝试多种格式解析日期字符串，返回 date 对象或 None
    """
    dt = safe_parse_date(date_str)
    return dt.date() if dt else None

SUMMARY_ROW_ROLE = "SUMMARY"
_FILTER_EXPR_ALLOWED_RE = re.compile(r'^[\d\.,\+\-\*\/\(\)\s]+$')
_FILTER_EXPR_NUMBER_RE = re.compile(r'\d[\d\.,]*')


def is_summary_item(item) -> bool:
    """判断单元格是否属于汇总行。"""
    if not item:
        return False
    if item.data(Qt.UserRole) == SUMMARY_ROW_ROLE:
        return True

    text = item.text().strip()
    return text.startswith("汇总 (") or text.startswith("汇 (")


def is_summary_row(table: QTableWidget, row: int) -> bool:
    """判断某一行是否为汇总行（兼容 role/text 双判定）。"""
    if table is None or row < 0 or row >= table.rowCount():
        return False

    first_item = table.item(row, 0)
    if is_summary_item(first_item):
        return True

    for c in range(table.columnCount()):
        item = table.item(row, c)
        if item and item.data(Qt.UserRole) == SUMMARY_ROW_ROLE:
            return True
    return False


def parse_filter_numeric_value(raw_value):
    """将筛选输入或单元格文本解析成数值，兼容巴西/美式金额与百分号。"""
    if raw_value is None:
        return None

    text = str(raw_value).strip()
    if not text:
        return None

    text = text.replace('％', '%')
    if text.endswith('%'):
        text = text[:-1].strip()

    value = br_to_float(text)
    if value is not None:
        return value

    fallback = text.replace(',', '').replace(' ', '')
    if not fallback:
        return None

    try:
        return float(fallback)
    except Exception:
        return None


def normalize_filter_expression(expr_str: str) -> str:
    """将带巴西金额格式的公式标准化为 Python 可安全解析的表达式。"""
    expr_str = str(expr_str).strip()
    if not _FILTER_EXPR_ALLOWED_RE.match(expr_str):
        raise ValueError("Unsafe characters in expression")

    def _replace_numeric_token(match):
        token = match.group(0)
        parsed = parse_filter_numeric_value(token)
        if parsed is None:
            raise ValueError(f"Invalid numeric literal: {token}")
        return repr(float(parsed))

    return _FILTER_EXPR_NUMBER_RE.sub(_replace_numeric_token, expr_str)


def check_filter_match(val_str, criterion):
    """
    检查值是否符合筛选条件 (支持 > < >= <= = != 运算符及简单计算式)
    """
    val_str = str(val_str).strip()
    
    # 1. 多选列表匹配
    if isinstance(criterion, list):
        match_val = val_str if val_str else "(空白)"
        return match_val in criterion

    # 2. 运算符匹配 (针对数字)
    query = str(criterion).strip().lower()
    
    # Regex for operators: >, <, >=, <=, =, !=
    # Followed by an expression
    op_match = re.match(r'^([><]=?|!=|=)\s*(.+)$', query)
    
    if op_match:
        op, expr_str = op_match.groups()
        try:
            cell_val = parse_filter_numeric_value(val_str)
            if cell_val is None:
                return False

            # Evaluate expression safely
            # 1. Normalize numeric literals so BR/US formats both work
            normalized_expr = normalize_filter_expression(expr_str)

            # 2. Safe arithmetic parsing (avoid eval)
            import ast

            def _safe_eval(node):
                if isinstance(node, ast.Expression):
                    return _safe_eval(node.body)
                if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
                    return float(node.value)
                if hasattr(ast, "Num") and isinstance(node, ast.Num):
                    return float(node.n)
                if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
                    val = _safe_eval(node.operand)
                    return val if isinstance(node.op, ast.UAdd) else -val
                if isinstance(node, ast.BinOp) and isinstance(node.op, (ast.Add, ast.Sub, ast.Mult, ast.Div)):
                    left = _safe_eval(node.left)
                    right = _safe_eval(node.right)
                    if isinstance(node.op, ast.Add):
                        return left + right
                    if isinstance(node.op, ast.Sub):
                        return left - right
                    if isinstance(node.op, ast.Mult):
                        return left * right
                    return left / right
                raise ValueError("Unsupported expression")

            parsed = ast.parse(normalized_expr, mode='eval')
            target_val = float(_safe_eval(parsed))
            
            if op == '>': return cell_val > target_val
            if op == '<': return cell_val < target_val
            if op == '>=': return cell_val >= target_val
            if op == '<=': return cell_val <= target_val
            if op == '=': return abs(cell_val - target_val) < 0.0001
            if op == '!=': return abs(cell_val - target_val) > 0.0001
        except Exception:
            pass # 不是数字或计算失败，回退到字符串匹配

    # 3. 字符串匹配
    if query.startswith("!"):
        # 反向筛选 (不包含)
        return query[1:] not in val_str.lower()
    else:
        # 正向筛选 (包含)
        return query in val_str.lower()

# --- Delegates ---
class ProgressBarDelegate(QStyledItemDelegate):
    def paint(self, painter, option, index):
        # Retrieve data: (percent, text)
        data = index.data(Qt.UserRole + 1000)
        if data:
            percent, text = data
            
            opts = QStyleOptionProgressBar()
            opts.rect = option.rect.adjusted(4, 2, -4, -2) # Add some padding
            opts.minimum = 0
            opts.maximum = 100
            opts.progress = int(percent)
            opts.text = text
            opts.textVisible = True
            opts.textAlignment = Qt.AlignCenter
            
            # Draw
            QApplication.style().drawControl(QStyle.CE_ProgressBar, opts, painter)
        else:
            QStyledItemDelegate.paint(self, painter, option, index)

# --- Filter Header Class ---
class FilterHeader(QHeaderView):
    filterChanged = Signal()

    def __init__(self, parent=None):
        super().__init__(Qt.Horizontal, parent)
        self._filters = {}  # col: filter_text
        self._padding = 4
        self._icon_size = 14
        self.setSectionsClickable(True)
        self.setSectionsMovable(True)  # 允许拖动表头
        self.setHighlightSections(True)
        self.setSortIndicatorShown(True)

    def paintSection(self, painter, rect, logicalIndex):
        painter.save()
        super().paintSection(painter, rect, logicalIndex)
        painter.restore()

        if logicalIndex in self._filters:
            # Draw active filter icon (Funnel filled)
            self._draw_indicator(painter, rect, active=True)
        else:
            # Draw inactive filter icon (Funnel outline) - only if enough space
            if rect.width() > 40:
                self._draw_indicator(painter, rect, active=False)

    def _draw_indicator(self, painter, rect, active):
        x = rect.right() - self._icon_size - self._padding - 2
        y = rect.center().y() - self._icon_size // 2
        
        # If sort indicator is visible and takes space, move left
        if self.isSortIndicatorShown() and self.sortIndicatorSection() == self.logicalIndexAt(rect.center().x()):
             x -= self._icon_size # simple shift

        painter.save()
        painter.setRenderHint(QPainter.Antialiasing)
        
        pen = QPen(Qt.black if not active else Qt.blue)
        pen.setWidth(1)
        painter.setPen(pen)
        
        if active:
            painter.setBrush(QBrush(Qt.blue))
        else:
            painter.setBrush(Qt.NoBrush)
            pen.setColor(QColor(150, 150, 150))
            painter.setPen(pen)

        # Draw Funnel
        #  \___/
        #   \ /
        #    v
        w = self._icon_size
        h = self._icon_size
        
        # Triangle part
        # p1(0,0) - p2(w,0) - p3(w/2, h)
        
        # Funnel shape: Top rect, bottom triangle
        # box_h = h // 3
        # p1(x, y), p2(x+w, y), p3(x+w, y+box_h), p4(x+w/2+1, y+h), p5(x+w/2-1, y+h), p6(x, y+box_h)
        
        poly = QPolygon()
        poly.append(QRect(x, y, w, h).topLeft())
        poly.append(QRect(x, y, w, h).topRight())
        poly.append(QRect(x, y, w, h).bottomRight() - PySide6.QtCore.QPoint(w//2 - 2, 0))
        poly.append(QRect(x, y, w, h).bottomLeft() + PySide6.QtCore.QPoint(w//2 - 2, 0))
        
        # Simplified triangle for now
        # painter.drawPolygon(poly)
        
        # Let's draw a simple "Y" or triangle
        poly2 = QPolygon()
        poly2.append(PySide6.QtCore.QPoint(x, y + 2))
        poly2.append(PySide6.QtCore.QPoint(x + w, y + 2))
        poly2.append(PySide6.QtCore.QPoint(x + w // 2, y + h - 2))
        painter.drawPolygon(poly2)

        painter.restore()

    def contextMenuEvent(self, event):
        """右键菜单：隐藏列"""
        logicalIndex = self.logicalIndexAt(event.pos())
        if logicalIndex == -1:
            return

        menu = QMenu(self)
        
        # Hide action
        hide_action = menu.addAction("👁 隐藏此列")
        menu.addSeparator()
        
        # Column Settings (Restore)
        config_action = menu.addAction("⚙️ 列设置(显示/排序)...")
        
        action = menu.exec(event.globalPos())
        
        if action == hide_action:
            # Hide the column
            # We need to find the parent table widget
            view = self.parent()
            if isinstance(view, QAbstractItemView):
                view.setColumnHidden(logicalIndex, True)
                
        elif action == config_action:
            # Try to find the main window or call the config dialog directly
            # The parent of Header is Table, parent of Table is usually the Tab/Window
            # But MainWindow has open_column_settings(table)
            
            # Since we don't have easy reference to MainWindow here, 
            # we can try to walk up or just instantiate the dialog if we have the class imported.
            # ColumnConfigDialog is available in this file.
            
            view = self.parent()
            if isinstance(view, QTableWidget):
                # We can construct the dialog here. 
                # Ideally, we should use the main window method to ensure consistency (saving state etc),
                # but direct usage is also fine if we trigger save state.
                
                # Let's try to find MainWindow to call save_column_state afterwards
                top_level = view.window()
                
                dialog = ColumnConfigDialog(view, top_level)
                if dialog.exec():
                    dialog.apply_settings()
                    if hasattr(top_level, "save_column_state"):
                        top_level.save_column_state()

    def mousePressEvent(self, event):
        # Check if click is on the filter icon area (right side of section)
        logicalIndex = self.logicalIndexAt(event.pos())
        if logicalIndex == -1:
            super().mousePressEvent(event)
            return

        rect = self.sectionViewportPosition(logicalIndex)
        # We approximate the rect width from visual index
        # Actually easier: get sectionSize
        sec_size = self.sectionSize(logicalIndex)
        sec_pos = self.sectionViewportPosition(logicalIndex)
        
        # Click area: Rightmost 20 pixels
        click_area_width = 24
        # Adjust if sorted?
        
        # Local pos x in section
        local_x = event.pos().x() - sec_pos
        
        if local_x > (sec_size - click_area_width):
            self.showFilterMenu(logicalIndex, event.globalPos())
        else:
            super().mousePressEvent(event)

    def showFilterMenu(self, logicalIndex, globalPos):
        menu = QMenu(self)
        menu.setMinimumWidth(250)
        
        # 1. 搜索框与公式框
        input_action = QWidgetAction(menu)
        input_frame = QFrame()
        input_layout = QVBoxLayout(input_frame)
        input_layout.setContentsMargins(4, 4, 4, 4)
        
        # [新增] 自定义公式输入
        formula_label = QLabel("自定义筛选公式 (如 > 100+20):")
        formula_edit = QLineEdit()
        formula_edit.setPlaceholderText("支持 >, <, =, != 及 +-*/")
        
        # Pre-fill if current filter is a string (formula)
        current_filter_val = self._filters.get(logicalIndex)
        if isinstance(current_filter_val, str):
            formula_edit.setText(current_filter_val)
            
        input_layout.addWidget(formula_label)
        input_layout.addWidget(formula_edit)
        
        # 列表搜索框
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("过滤列表项...")
        input_layout.addWidget(search_edit)
        
        # 2. 同类项列表
        list_widget = QListWidget()
        list_widget.setMaximumHeight(300)
        
        # 提取当前列所有唯一值
        unique_values = set()
        has_empty = False
        view = self.parent()
        if isinstance(view, QTableWidget):
            for r in range(view.rowCount()):
                if is_summary_row(view, r):
                    continue

                item = view.item(r, logicalIndex)
                txt = item.text().strip() if item else ""
                if not txt:
                    has_empty = True
                else:
                    unique_values.add(txt)
        
        sorted_vals = sorted(list(unique_values))
        if has_empty:
            sorted_vals.insert(0, "(空白)")
        
        # 添加 "全选" 项
        all_item = QListWidgetItem("(全选/显示全部)")
        all_item.setFlags(all_item.flags() | Qt.ItemIsUserCheckable)
        all_item.setCheckState(Qt.Checked)
        list_widget.addItem(all_item)
        
        # 获取当前已选状态 (if it's a list)
        current_list_filter = current_filter_val if isinstance(current_filter_val, list) else []
        
        items = []
        for val in sorted_vals:
            item = QListWidgetItem(val)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            # 如果没有激活列表筛选(且不是公式筛选)，默认全选；如果有列表筛选，按状态勾选
            if not current_list_filter or val in current_list_filter:
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)
            list_widget.addItem(item)
            items.append(item)
            
        input_layout.addWidget(list_widget)
        
        # 搜索框联动逻辑
        def filter_list(text):
            for i in range(1, list_widget.count()):
                it = list_widget.item(i)
                it.setHidden(text.lower() not in it.text().lower())
        search_edit.textChanged.connect(filter_list)
        
        # 全选联动逻辑
        def on_all_clicked(item):
            if item == all_item:
                state = all_item.checkState()
                for i in range(1, list_widget.count()):
                    list_widget.item(i).setCheckState(state)
        list_widget.itemChanged.connect(on_all_clicked)
        
        input_action.setDefaultWidget(input_frame)
        menu.addAction(input_action)
        
        # 3. 操作按钮
        apply_action = menu.addAction("✅ 应用筛选")
        clear_action = menu.addAction("🧹 清除此列筛选")
        menu.addSeparator()
        sort_asc = menu.addAction("🔼 升序排列")
        sort_desc = menu.addAction("🔽 降序排列")
        
        action = menu.exec(globalPos)
        
        if action == apply_action:
            # Check formula first
            formula_text = formula_edit.text().strip()
            if formula_text:
                self._filters[logicalIndex] = formula_text
            else:
                selected = []
                # 如果全选被勾选且没有隐藏任何项，视为清除筛选
                is_all_checked = (all_item.checkState() == Qt.Checked)
                
                if not is_all_checked:
                    for i in range(1, list_widget.count()):
                        it = list_widget.item(i)
                        if it.checkState() == Qt.Checked:
                            selected.append(it.text())
                    
                    if selected:
                        self._filters[logicalIndex] = selected
                    else:
                         # Nothing selected -> Filter everything (show nothing)
                         # Assuming unchecking all means hide all
                         if all_item.checkState() == Qt.Unchecked:
                             self._filters[logicalIndex] = []
                         else:
                             if logicalIndex in self._filters: del self._filters[logicalIndex]
                else:
                    if logicalIndex in self._filters: del self._filters[logicalIndex]
                
            self.filterChanged.emit()
            
        elif action == clear_action:
            if logicalIndex in self._filters:
                del self._filters[logicalIndex]
                self.filterChanged.emit()
        elif action == sort_asc:
            top_level = view.window() if isinstance(view, QTableWidget) else None
            if hasattr(top_level, "sort_with_summary"):
                top_level.sort_with_summary(view, logicalIndex, Qt.AscendingOrder)
            elif isinstance(view, QTableWidget):
                view.sortItems(logicalIndex, Qt.AscendingOrder)
        elif action == sort_desc:
            top_level = view.window() if isinstance(view, QTableWidget) else None
            if hasattr(top_level, "sort_with_summary"):
                top_level.sort_with_summary(view, logicalIndex, Qt.DescendingOrder)
            elif isinstance(view, QTableWidget):
                view.sortItems(logicalIndex, Qt.DescendingOrder)

    def get_filter_text(self, col):
        return self._filters.get(col, "")

    def clear_filters(self):
        """清除所有列的筛选"""
        self._filters.clear()
        self.filterChanged.emit()

import json
import base64
import io
import urllib.request
import urllib.error
import sqlite3
from datetime import datetime, timedelta

# --- 全局配置与日志---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
SETTINGS_FILE = "danfe_batch_gui_settings.ini"

# --- GUI 资源（Base64编码---
APP_ICON_B64 = "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAAn/SURBVHhe7Zt7bFxVFcd/M/fO3W5v12sD22gDxm0MExKkKZqgSCoK0aSoiBRpUaKq8AEf+CAgaiISiY8EESXyAw0gH1ESqGgS0SQhVoA2qY1pE5sYjG1s7N3udvf27s7M/fHcu/fO3t3b3W4n/Z/8d87MmfM7v3POzJmZgYCAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgIC-";

# --- 正则与文本处理工具---
# (已移除：使用 brazil_tool.core 替代)




def export_qtable(table: QTableWidget, parent: QWidget, filename_prefix: str = "export"):
    """通用表格导出函数 (支持颜色导出到Excel，同时也支持CSV)"""
    if table.rowCount() == 0:
        QMessageBox.warning(parent, "无数据", "表格为空，无法导出")
        return

    default_name = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path, _ = QFileDialog.getSaveFileName(parent, "导出表格", default_name, "Excel Files (*.xlsx);;CSV Files (*.csv)")
    if not path:
        return

    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        export_multiple_qtables([(table, "Sheet1")], parent, filename_prefix, provided_path=path)
    else:
        # CSV Export
        try:
            headers = []
            for c in range(table.columnCount()):
                item = table.horizontalHeaderItem(c)
                headers.append(item.text() if item else f"Col {c}")
            
            rows = []
            for r in range(table.rowCount()):
                if table.isRowHidden(r): continue
                if is_summary_row(table, r):
                    continue
                
                row_data = []
                for c in range(table.columnCount()):
                    widget = table.cellWidget(r, c)
                    if isinstance(widget, QProgressBar):
                        row_data.append(widget.text())
                    else:
                        item = table.item(r, c)
                        row_data.append(item.text().replace('\n', ' ') if item else "")
                rows.append(row_data)
                
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(rows)
            QMessageBox.information(parent, "导出成功", f"已保存到: {path}")
        except Exception as e:
            QMessageBox.critical(parent, "导出失败", f"CSV 导出错误: {e}")

def export_multiple_qtables(tables_info: list, parent: QWidget, filename_prefix: str = "export", provided_path: str = None):
    """
    导出多个表格到同一个 Excel 文件的不同 Sheet 中
    tables_info: list of (QTableWidget, sheet_name)
    """
    if not tables_info:
        return

    # 检查是否有数据
    valid_tables = [t for t, name in tables_info if t.rowCount() > 0]
    if not valid_tables:
        QMessageBox.warning(parent, "无数据", "所有表格均为空，无法导出")
        return

    if provided_path:
        path = provided_path
    else:
        default_name = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(parent, "导出对账结果", default_name, "Excel Files (*.xlsx)")
    
    if not path:
        return

    try:
        if openpyxl is None:
            raise ImportError("未安装 openpyxl")
        
        from openpyxl.styles import PatternFill
        wb = openpyxl.Workbook()
        # 移除默认 sheet
        wb.remove(wb.active)

        for table, sheet_name in tables_info:
            if table.rowCount() == 0: continue
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Collect Headers
            headers = []
            for c in range(table.columnCount()):
                item = table.horizontalHeaderItem(c)
                headers.append(item.text() if item else f"Col {c}")
            ws.append(headers)

            excel_row_idx = 2
            for r in range(table.rowCount()):
                if table.isRowHidden(r): continue
                if is_summary_row(table, r):
                    continue
                
                for c in range(table.columnCount()):
                    widget = table.cellWidget(r, c)
                    val = ""
                    bg_color = None
                    
                    if isinstance(widget, QProgressBar):
                        val = widget.text()
                    else:
                        item = table.item(r, c)
                        if item:
                            val = item.text().replace('\n', ' ')
                            brush = item.background()
                            if brush != Qt.NoBrush:
                                col = brush.color()
                                if col.isValid() and col not in [QColor(Qt.white), QColor(Qt.transparent)]:
                                    bg_color = col.name().replace('#', '')
                    
                    cell = ws.cell(row=excel_row_idx, column=c+1, value=val)
                    if bg_color:
                        try:
                            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                            cell.fill = fill
                        except: pass
                excel_row_idx += 1
        
        wb.save(path)
        QMessageBox.information(parent, "导出成功", f"数据已保存到: {path}")
        
        # 询问是否打开
        reply = QMessageBox.question(parent, '打开文件', '要现在打开导出的文件吗？',
                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            QDesktopServices.openUrl(PySide6.QtCore.QUrl.fromLocalFile(path))

    except Exception as e:
        QMessageBox.critical(parent, "导出失败", f"错误详情: {e}")


# --- Mapping System ---
class MappingManager:
    """管理产品和往来单位的标准化映射，并在入库前做基础清洗。"""

    PRODUCT_METADATA_FIELDS = (
        "status",
        "source",
        "confidence",
        "updated_at",
        "reviewed_by",
        "sample_text",
        "notes",
    )
    PRODUCT_SECTION_MARKERS = (
        "TRANSPORTADOR/VOLUMES TRANSPORTADOS",
        "DADOS ADICIONAIS",
        "INFORMACOES COMPLEMENTARES",
        "INFORMAÇÕES COMPLEMENTARES",
        "RESERVADO AO FISCO",
        "NOME/RAZAO SOCIAL",
        "NOME/RAZÃO SOCIAL",
    )
    PRODUCT_TAIL_MARKERS = (
        "DESENVOLVIDO POR",
        "HTTP://",
        "HTTPS://",
        "WWW.",
    )
    PRODUCT_MONEY_RE = re.compile(
        r'^\s*(?:R\$\s*)?(?:\d{1,3}(?:\.\d{3})+|\d+),\d{2}\s*$'
    )
    PRODUCT_PRICE_PREFIX_RE = re.compile(
        r'^\s*(?:R\$\s*)?(?:\d{1,3}(?:\.\d{3})+|\d+),\d{2}(?:\s*[-–—:]?\s*|\s+)'
    )
    PRODUCT_PHONE_RE = re.compile(r'\(?\d{2}\)?\s*\d{4,5}-?\d{4}')
    PRODUCT_ONLY_DIGITS_RE = re.compile(r'^\d+$')
    PRODUCT_CODE_TEXT_RE = re.compile(r'[^A-Z0-9._\-/]')

    def __init__(self, filepath="mapping_db.json"):
        self.filepath = filepath
        self.data = {"products": {}, "partners": {}}
        self.audit = {"dropped_products": [], "dropped_partners": []}
        self.load()

    @staticmethod
    def _clean_text(value) -> str:
        if value is None:
            return ""
        text = str(value).replace(" ", " ").replace("　", " ")
        text = text.replace("\r", " ").replace("\n", " ")
        return re.sub(r'\s+', ' ', text).strip()

    @classmethod
    def _looks_like_monetary_text(cls, value) -> bool:
        text = cls._clean_text(value)
        if not text:
            return False
        if cls.PRODUCT_MONEY_RE.match(text):
            return True
        return bool(re.match(r'^\s*(?:R\$\s*)?\d+[\.,]\d{2}\s*$', text))

    @classmethod
    def _normalize_product_key(cls, code) -> str:
        if code is None:
            return ""
        return re.sub(r'\D+', '', str(code))

    @classmethod
    def _normalize_product_code_value(cls, value) -> str:
        text = cls._clean_text(value)
        if not text:
            return ""
        if cls._looks_like_monetary_text(text):
            return ""
        upper = text.upper()
        if re.search(r'[A-Z]', upper):
            upper = cls.PRODUCT_CODE_TEXT_RE.sub('', upper)
            upper = re.sub(r'-{2,}', '-', upper).strip('-._/')
            return upper
        digits = re.sub(r'\D+', '', upper)
        if not digits:
            return ""
        if len(digits) > 32:
            return ""
        return digits

    @classmethod
    def _trim_product_tail(cls, text: str) -> str:
        cleaned = text
        upper = cleaned.upper()
        cut_positions = []
        for marker in cls.PRODUCT_SECTION_MARKERS + cls.PRODUCT_TAIL_MARKERS:
            pos = upper.find(marker)
            if pos > 0:
                cut_positions.append(pos)
        if cut_positions:
            cleaned = cleaned[:min(cut_positions)].strip(' -–—|;:,')
        return cleaned.strip()

    @classmethod
    def _looks_like_product_noise(cls, text: str) -> bool:
        if not text:
            return True
        upper = text.upper()
        if cls._looks_like_monetary_text(text):
            return True
        if cls.PRODUCT_PHONE_RE.search(text) and len(re.sub(r'\D+', '', text)) <= 13:
            return True
        if any(marker in upper for marker in cls.PRODUCT_SECTION_MARKERS):
            return True
        if 'CONTRATAÇÃO DO FRETE' in upper or 'CONTRATACAO DO FRETE' in upper:
            return True
        if cls.PRODUCT_ONLY_DIGITS_RE.match(re.sub(r'\s+', '', text)):
            return True
        if not re.search(r'[A-ZÀ-ÿ]', text):
            return True
        if len(text) > 220:
            return True
        return False

    @classmethod
    def _clean_product_name_text(cls, value) -> str:
        text = cls._clean_text(value)
        if not text:
            return ""

        prev = None
        while prev != text:
            prev = text
            text = cls.PRODUCT_PRICE_PREFIX_RE.sub('', text).strip()
            text = re.sub(r'^[\-–—|:;,./\\]+', '', text).strip()

        text = cls._trim_product_tail(text)
        text = re.sub(r'\s+', ' ', text).strip(' -–—|:;,')

        if cls._looks_like_product_noise(text):
            return ""
        return text

    @classmethod
    def _product_entry_score(cls, entry: dict) -> tuple:
        return (
            1 if entry.get('std_code') else 0,
            1 if entry.get('std_name') else 0,
            1 if entry.get('status') == 'reviewed' else 0,
            float(entry.get('confidence') or 0.0),
            len(entry.get('sample_text') or ''),
        )

    @classmethod
    def _merge_product_entries(cls, current: Optional[dict], incoming: Optional[dict]) -> Optional[dict]:
        if not current:
            return incoming
        if not incoming:
            return current
        preferred = incoming if cls._product_entry_score(incoming) > cls._product_entry_score(current) else current
        secondary = current if preferred is incoming else incoming
        merged = dict(preferred)
        for field in ('std_code', 'std_name', *cls.PRODUCT_METADATA_FIELDS):
            if not merged.get(field) and secondary.get(field):
                merged[field] = secondary.get(field)
        return merged

    @classmethod
    def _sanitize_product_entry(
        cls,
        key,
        value,
        *,
        keep_incomplete: bool = True,
        source: Optional[str] = None,
        status: Optional[str] = None,
        reviewed_by: Optional[str] = None,
        confidence: Optional[float] = None,
        sample_text: Optional[str] = None,
        notes: Optional[str] = None,
    ) -> Tuple[str, Optional[dict], List[str]]:
        issues: List[str] = []
        norm_key = cls._normalize_product_key(key)
        if not norm_key:
            issues.append('empty_key')
            return '', None, issues

        raw_entry = value if isinstance(value, dict) else {'std_code': '', 'std_name': value}
        std_code = cls._normalize_product_code_value(raw_entry.get('std_code', ''))
        raw_std_name = raw_entry.get('std_name', '')
        std_name = cls._clean_product_name_text(raw_std_name)
        if raw_std_name and not std_name:
            issues.append('std_name_rejected')

        entry = {'std_code': std_code, 'std_name': std_name}
        for field in cls.PRODUCT_METADATA_FIELDS:
            val = raw_entry.get(field)
            if val not in (None, ''):
                entry[field] = cls._clean_text(val) if isinstance(val, str) else val

        if source not in (None, ''):
            entry['source'] = cls._clean_text(source)
        if reviewed_by not in (None, ''):
            entry['reviewed_by'] = cls._clean_text(reviewed_by)
        if sample_text not in (None, ''):
            cleaned_sample = cls._clean_product_name_text(sample_text)
            if cleaned_sample:
                entry['sample_text'] = cleaned_sample
        if notes not in (None, ''):
            entry['notes'] = cls._clean_text(notes)
        if confidence not in (None, ''):
            try:
                entry['confidence'] = float(confidence)
            except (TypeError, ValueError):
                issues.append('invalid_confidence')

        if not entry.get('status'):
            entry['status'] = status or ('reviewed' if std_code and std_name else 'candidate')
        elif status not in (None, ''):
            entry['status'] = status

        entry['updated_at'] = cls._clean_text(entry.get('updated_at')) or datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if not std_code and not std_name:
            issues.append('empty_payload')
            if not keep_incomplete:
                return norm_key, None, issues
            return norm_key, None, issues
        return norm_key, entry, issues

    @classmethod
    def sanitize_legacy_product_mapping(cls, mapping: dict) -> Tuple[dict, dict]:
        cleaned = {}
        rejected = {}
        for raw_key, raw_value in (mapping or {}).items():
            norm_key = cls._normalize_product_key(raw_key)
            norm_value = cls._normalize_product_code_value(raw_value)
            if not norm_key or not norm_value:
                rejected[str(raw_key)] = raw_value
                continue
            if cls._looks_like_monetary_text(raw_key) or cls._looks_like_monetary_text(raw_value):
                rejected[str(raw_key)] = raw_value
                continue
            cleaned[norm_key] = norm_value
        return cleaned, rejected

    def load(self):
        self.audit = {"dropped_products": [], "dropped_partners": []}
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                self.data["products"] = loaded.get("products", {})
                self.data["partners"] = loaded.get("partners", {})
            except Exception as e:
                logging.error(f"Failed to load mappings: {e}")
                self.data = {"products": {}, "partners": {}}
        self.normalize_data()

    def normalize_data(self):
        clean_prods = {}
        dropped_products = []
        for raw_key, raw_value in (self.data.get('products') or {}).items():
            norm_key, entry, issues = self._sanitize_product_entry(raw_key, raw_value)
            if not norm_key or not entry:
                dropped_products.append({
                    'key': str(raw_key),
                    'value': raw_value,
                    'issues': issues,
                })
                continue
            clean_prods[norm_key] = self._merge_product_entries(clean_prods.get(norm_key), entry)
        self.data['products'] = clean_prods

        clean_parts = {}
        dropped_partners = []
        for raw_key, raw_value in (self.data.get('partners') or {}).items():
            clean_k = self._normalize_partner_key(raw_key)
            clean_v = self._clean_text(raw_value)
            if not clean_k or not clean_v or self._looks_like_monetary_text(clean_v):
                dropped_partners.append({
                    'key': str(raw_key),
                    'value': raw_value,
                })
                continue
            clean_parts[clean_k] = clean_v
        self.data['partners'] = clean_parts
        self.audit = {
            'dropped_products': dropped_products,
            'dropped_partners': dropped_partners,
        }

    @staticmethod
    def _normalize_partner_key(key: str) -> str:
        if key is None:
            return ""
        raw = str(key).strip()
        if not raw:
            return ""

        digits = re.sub(r'\D+', '', raw)
        if digits and (len(digits) >= 11 or re.search(r'[.\-/]', raw)):
            return digits

        return re.sub(r'\s+', ' ', raw).strip().upper()

    def save(self):
        self.normalize_data()
        try:
            with open(self.filepath, 'w', encoding='utf-8') as f:
                json.dump(self.data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            logging.error(f"Failed to save mappings: {e}")

    def get_product_std(self, code):
        if not code:
            return None
        res = self.data['products'].get(str(code))
        if res:
            return res
        norm_code = self._normalize_product_key(code)
        if not norm_code:
            return None
        return self.data['products'].get(norm_code)

    @staticmethod
    def _normalize_name_for_match(name: str) -> str:
        if not name:
            return ""
        s = unicodedata.normalize("NFKD", str(name))
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = s.upper()
        s = re.sub(r'[^A-Z0-9]+', ' ', s)
        return re.sub(r'\s+', ' ', s).strip()

    def find_best_product_name_match(self, description, threshold=0.8):
        if not description:
            return None

        cleaned_description = self._clean_product_name_text(description) or self._clean_text(description)
        if not cleaned_description:
            return None

        best = None
        highest_score = 0.0
        norm_desc = self._normalize_name_for_match(cleaned_description)
        for key, value in self.data.get('products', {}).items():
            std_name = self._clean_product_name_text(value.get('std_name', ''))
            if not std_name:
                continue
            score_raw = calculate_similarity(cleaned_description, std_name)
            if norm_desc:
                score_norm = calculate_similarity(norm_desc, self._normalize_name_for_match(std_name))
                score = max(score_raw, score_norm)
            else:
                score = score_raw
            if score > highest_score:
                highest_score = score
                best = {
                    'source_key': key,
                    'std_code': value.get('std_code', ''),
                    'std_name': std_name,
                    'score': score,
                    'status': value.get('status', ''),
                }

        if highest_score >= threshold:
            return best
        return None

    def evaluate_product_match(self, code, description, name_threshold=0.86):
        """
        Evaluate mapping with conflict-aware status.
        Status: AUTO_PASS / CODE_ONLY / NAME_ONLY / CONFLICT / UNMAPPED
        """
        code_std = self.get_product_std(code)
        name_std = self.find_best_product_name_match(description, threshold=name_threshold)

        code_candidate = None
        if code_std:
            code_candidate = {
                'std_code': code_std.get('std_code', ''),
                'std_name': code_std.get('std_name', ''),
                'score': 1.0,
            }

        def same_target(a, b):
            if not a or not b:
                return False
            a_code = str(a.get('std_code', '')).strip()
            b_code = str(b.get('std_code', '')).strip()
            if a_code and b_code and a_code == b_code:
                return True
            a_name = self._normalize_name_for_match(a.get('std_name', ''))
            b_name = self._normalize_name_for_match(b.get('std_name', ''))
            return bool(a_name and b_name and a_name == b_name)

        if code_candidate and name_std:
            if same_target(code_candidate, name_std):
                return {
                    'status': 'AUTO_PASS',
                    'code_candidate': code_candidate,
                    'name_candidate': name_std,
                }
            return {
                'status': 'CONFLICT',
                'code_candidate': code_candidate,
                'name_candidate': name_std,
            }
        if code_candidate:
            return {
                'status': 'CODE_ONLY',
                'code_candidate': code_candidate,
                'name_candidate': None,
            }
        if name_std:
            return {
                'status': 'NAME_ONLY',
                'code_candidate': None,
                'name_candidate': name_std,
            }
        return {'status': 'UNMAPPED', 'code_candidate': None, 'name_candidate': None}

    def get_product_smart(self, code, description, threshold=0.8):
        std = self.get_product_std(code)
        if std:
            return std
        best_match = self.find_best_product_name_match(description, threshold=threshold)
        if best_match:
            return {
                'std_code': best_match.get('std_code', ''),
                'std_name': best_match.get('std_name', ''),
            }
        return None

    def set_product_std(
        self,
        code,
        std_code,
        std_name,
        *,
        source: str = 'manual',
        status: Optional[str] = None,
        reviewed_by: str = '',
        confidence: Optional[float] = None,
        sample_text: str = '',
        notes: str = '',
    ):
        current = self.get_product_std(code) or {}
        _, entry, _ = self._sanitize_product_entry(
            code,
            {
                'std_code': std_code,
                'std_name': std_name,
                'status': current.get('status', ''),
                'source': current.get('source', ''),
                'confidence': current.get('confidence', ''),
                'updated_at': current.get('updated_at', ''),
                'reviewed_by': current.get('reviewed_by', ''),
                'sample_text': current.get('sample_text', ''),
                'notes': current.get('notes', ''),
            },
            source=source,
            status=status,
            reviewed_by=reviewed_by,
            confidence=confidence if confidence is not None else current.get('confidence'),
            sample_text=sample_text or current.get('sample_text', ''),
            notes=notes or current.get('notes', ''),
        )
        norm_code = self._normalize_product_key(code)
        if norm_code and entry:
            self.data['products'][norm_code] = self._merge_product_entries(current, entry)
            return True
        if norm_code:
            self.data['products'].pop(norm_code, None)
        return False

    def get_partner_std(self, key):
        if key is None:
            return None
        direct = self.data['partners'].get(key)
        if direct:
            return direct
        return self.data['partners'].get(self._normalize_partner_key(key))

    def set_partner_std(self, key, std_name):
        norm_key = self._normalize_partner_key(key)
        clean_name = self._clean_text(std_name)
        if norm_key and clean_name and not self._looks_like_monetary_text(clean_name):
            self.data['partners'][norm_key] = clean_name

# --- Tags Manager ---
class TagManager:
    def __init__(self, filepath="user_tags.json"):
        self.filepath = filepath
        self.tags = [] # list of {"name": str, "color": str}
        self.load()

    def load(self):
        if os.path.exists(self.filepath):
            try:
                with open(self.filepath, 'r', encoding='utf-8') as f:
                    self.tags = json.load(f)
            except: self.tags = []
        if not self.tags:
            self.tags = [{"name": "待确认", "color": "#FFCC00"}, {"name": "已核实", "color": "#99FF99"}]

    def save(self):
        try:
            with open(self.filepath, 'w', encoding='utf-8') as f:
                json.dump(self.tags, f, ensure_ascii=False)
        except: pass

    def get_tags(self):
        return self.tags

    def add_tag(self, name, color="#FFFFFF"):
        for t in self.tags:
            if t['name'] == name: return
        self.tags.append({"name": name, "color": color})
        self.save()

    def remove_tag(self, name):
        self.tags = [t for t in self.tags if t['name'] != name]
        self.save()

class TagEditDialog(QDialog):
    def __init__(self, tag_mgr: TagManager, current_tags: list, parent=None):
        super().__init__(parent)
        self.mgr = tag_mgr
        self.selected_tags = set(current_tags)
        self.setWindowTitle("标签管理 (User Tags)")
        self.resize(300, 400)
        
        layout = QVBoxLayout(self)
        
        self.list_widget = QListWidget()
        layout.addWidget(self.list_widget)
        
        self.refresh_list()
        
        # Add New Tag
        add_layout = QHBoxLayout()
        self.new_tag_input = QLineEdit()
        self.new_tag_input.setPlaceholderText("新标签名称?..")
        btn_add = QPushButton("添加")
        btn_add.clicked.connect(self.add_new_tag)
        add_layout.addWidget(self.new_tag_input)
        add_layout.addWidget(btn_add)
        layout.addLayout(add_layout)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def refresh_list(self):
        self.list_widget.clear()
        for t in self.mgr.get_tags():
            item = QListWidgetItem(t['name'])
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            if t['name'] in self.selected_tags:
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)
            # Set background color
            brush = QBrush(QColor(t['color']))
            item.setBackground(brush)
            self.list_widget.addItem(item)

    def add_new_tag(self):
        name = self.new_tag_input.text().strip()
        if name:
            import random
            color = f"#{random.randint(0xAAAAAA, 0xFFFFFF):06x}" # Lighter colors for background
            self.mgr.add_tag(name, color)
            self.refresh_list()
            self.new_tag_input.clear()

    def accept(self):
        self.selected_tags = set()
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item.checkState() == Qt.Checked:
                self.selected_tags.add(item.text())
        super().accept()

class SimilaritySearchConfigDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("相似性搜索设置")
        self.resize(300, 200)
        
        layout = QVBoxLayout(self)
        
        # Criteria Group
        grp = QFrame()
        grp.setFrameShape(QFrame.StyledPanel)
        grp_layout = QVBoxLayout(grp)
        grp_layout.addWidget(QLabel("<b>匹配依据:</b>"))
        
        self.rb_both = QRadioButton("同时匹配 编码(Key) 与 名称(Name)")
        self.rb_both.setChecked(True)
        self.rb_key = QRadioButton("仅匹配 编码(Key)")
        self.rb_name = QRadioButton("仅匹配 名称(Name)")
        
        self.bg = QButtonGroup()
        self.bg.addButton(self.rb_both, 0)
        self.bg.addButton(self.rb_key, 1)
        self.bg.addButton(self.rb_name, 2)
        
        grp_layout.addWidget(self.rb_both)
        grp_layout.addWidget(self.rb_key)
        grp_layout.addWidget(self.rb_name)
        layout.addWidget(grp)
        
        # Threshold
        thresh_layout = QHBoxLayout()
        thresh_layout.addWidget(QLabel("最小相似度 (%):"))
        self.spin_thresh = QSpinBox()
        self.spin_thresh.setRange(10, 100)
        self.spin_thresh.setValue(60)
        self.spin_thresh.setSingleStep(5)
        thresh_layout.addWidget(self.spin_thresh)
        layout.addLayout(thresh_layout)
        
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)
        
    def get_settings(self):
        mode = "both"
        if self.rb_key.isChecked(): mode = "key"
        elif self.rb_name.isChecked(): mode = "name"
        
        return {
            "mode": mode,
            "threshold": self.spin_thresh.value() / 100.0
        }
class SearchResultsDialog(QDialog):
    def __init__(self, results, parent_table, target_info="", parent=None):
        """
        results: list of (row_idx, score, key, name)
        parent_table: the QTableWidget we are operating on
        target_info: The original item info that was used for searching
        """
        super().__init__(parent)
        self.results = results
        self.parent_table = parent_table
        self.target_info = target_info
        self.setWindowTitle(f"搜索结果 (找到 {len(results)} 条相似项)")
        self.resize(850, 500)
        
        layout = QVBoxLayout(self)
        
        # Info
        layout.addWidget(QLabel(f"<b>搜索基准 (Target):</b> {target_info}"))
        layout.addWidget(QLabel("勾选项目以执行批量操作:"))
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["选择", "匹配度", "搜索基准 (Target)", "匹配Key (原始编码)", "匹配名称 (标准名称)"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        
        self.populate()
        layout.addWidget(self.table)
        
        # Actions
        btn_layout = QHBoxLayout()
        btn_merge_name = QPushButton("🔀 并入名称 (Merge Name)")
        btn_merge_key = QPushButton("🔑 并入Key (Merge Key)")
        btn_export = QPushButton("📤 导出结果")
        btn_close = QPushButton("关闭")
        
        btn_merge_name.clicked.connect(self.merge_name)
        btn_merge_key.clicked.connect(self.merge_key)
        btn_export.clicked.connect(lambda: export_qtable(self.table, self, "search_results"))
        btn_close.clicked.connect(self.accept)
        
        btn_layout.addWidget(btn_merge_name)
        btn_layout.addWidget(btn_merge_key)
        btn_layout.addWidget(btn_export)
        btn_layout.addStretch()
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)
        
    def populate(self):
        self.table.setRowCount(len(self.results))
        for i, (row_idx, score, key, name) in enumerate(self.results):
            # Checkbox
            chk_item = QTableWidgetItem()
            chk_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            chk_item.setCheckState(Qt.Unchecked)
            self.table.setItem(i, 0, chk_item)
            
            # Score
            score_item = QTableWidgetItem(f"{score*100:.1f}%")
            score_item.setTextAlignment(Qt.AlignCenter)
            if score > 0.85:
                score_item.setForeground(QBrush(QColor("green")))
            self.table.setItem(i, 1, score_item)

            # Target (New)
            self.table.setItem(i, 2, QTableWidgetItem(str(self.target_info)))
            
            # Key
            self.table.setItem(i, 3, QTableWidgetItem(str(key)))
            
            # Name
            self.table.setItem(i, 4, QTableWidgetItem(str(name)))
            
            # Store row index in first item
            chk_item.setData(Qt.UserRole, row_idx)

    def get_selected_rows(self):
        selected_indices = []
        for i in range(self.table.rowCount()):
            if self.table.item(i, 0).checkState() == Qt.Checked:
                # Retrieve original table row index
                row_idx = self.table.item(i, 0).data(Qt.UserRole)
                selected_indices.append(row_idx)
        return selected_indices

    def merge_name(self):
        rows = self.get_selected_rows()
        if not rows:
            QMessageBox.warning(self, "提示", "请先勾选要操作的行")
            return
            
        # Default name from first selected
        first_row_idx = rows[0]
        # Assuming parent table col 2 is Name (valid for Product table). 
        # For Partner table it is col 1. We need to know which table type.
        # Heuristic: Check column count. Prod=3, Partner=2.
        col_idx = 2 if self.parent_table.columnCount() == 3 else 1
        
        default_name = self.parent_table.item(first_row_idx, col_idx).text()
        
        name, ok = QInputDialog.getText(self, "并入名称", "请输入统一的标准名 ", text=default_name)
        if ok and name:
            for r in rows:
                self.parent_table.setItem(r, col_idx, QTableWidgetItem(name))
            QMessageBox.information(self, "完成", f"已更新 {len(rows)} 行")

    def merge_key(self):
        rows = self.get_selected_rows()
        if not rows:
            QMessageBox.warning(self, "提示", "请先勾选要操作的行")
            return
            
        # This is strictly "Merge to Key's Standard Value" logic
        # For Product Table (3 cols): Key | Std Code | Std Name
        # We pick one row's Key, find its Std values, and apply to others.
        
        # For Partner Table (2 cols): Key | Std Name
        # We pick one row's Key, find its Std Name, and apply.
        
        is_prod = (self.parent_table.columnCount() == 3)
        
        # Build options list from selected items in THIS dialog
        options = []
        row_map = {} # option_idx -> original_row_idx
        
        for i in range(self.table.rowCount()):
            if self.table.item(i, 0).checkState() == Qt.Checked:
                orig_idx = self.table.item(i, 0).data(Qt.UserRole)
                k = self.table.item(i, 2).text()
                n = self.table.item(i, 3).text()
                options.append(f"{k} | {n}")
                row_map[len(options)-1] = orig_idx
                
        item, ok = QInputDialog.getItem(self, "选择基准", "选择作为标准的项 (将其标准值应用到所有勾选行):", options, 0, False)
        if ok and item:
            # Find which original row was selected as base
            sel_idx = options.index(item)
            base_row = row_map[sel_idx]
            
            if is_prod:
                std_code = self.parent_table.item(base_row, 1).text()
                std_name = self.parent_table.item(base_row, 2).text()
                for r in rows:
                    self.parent_table.setItem(r, 1, QTableWidgetItem(std_code))
                    self.parent_table.setItem(r, 2, QTableWidgetItem(std_name))
            else:
                std_name = self.parent_table.item(base_row, 1).text()
                for r in rows:
                    self.parent_table.setItem(r, 1, QTableWidgetItem(std_name))
            
            QMessageBox.information(self, "完成", f"已更新 {len(rows)} 行")

class ProductConflictBatchDialog(QDialog):
    """产品映射冲突/待确认的批量处理窗口"""
    def __init__(self, conflicts: list, parent=None):
        super().__init__(parent)
        self.conflicts = conflicts or []
        self.decisions = []  # list of (conflict_dict, strategy)
        self.setWindowTitle(f"产品映射冲突处理 (共 {len(self.conflicts)} 条)")
        self.resize(1200, 620)

        layout = QVBoxLayout(self)
        self.summary_label = QLabel("")
        layout.addWidget(self.summary_label)

        toolbar = QHBoxLayout()
        btn_all_code = QPushButton("全部设为编码")
        btn_all_name = QPushButton("全部设为名称")
        btn_all_ignore = QPushButton("全部设为忽略")
        btn_all_code.clicked.connect(lambda: self.apply_all_strategy("code"))
        btn_all_name.clicked.connect(lambda: self.apply_all_strategy("name"))
        btn_all_ignore.clicked.connect(lambda: self.apply_all_strategy("ignored"))
        toolbar.addWidget(btn_all_code)
        toolbar.addWidget(btn_all_name)
        toolbar.addWidget(btn_all_ignore)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "处理策略", "状态", "发票", "行号", "原始编码", "原始品目名", "编码候选", "名称候选", "名称相似度"
        ])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(7, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeToContents)
        layout.addWidget(self.table)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.button(QDialogButtonBox.Ok).setText("应用选择")
        btn_box.button(QDialogButtonBox.Cancel).setText("取消")
        btn_box.accepted.connect(self.accept_with_validation)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

        self.populate_table()

    @staticmethod
    def _candidate_text(cand: dict) -> str:
        if not cand:
            return "-"
        return f"{cand.get('std_code', '-') or '-'} | {cand.get('std_name', '-') or '-'}"

    def _combo_for_conflict(self, conflict: dict) -> QComboBox:
        combo = QComboBox()
        combo.addItem("忽略本条", "ignored")

        has_code = bool(conflict.get("code_candidate"))
        has_name = bool(conflict.get("name_candidate"))
        status = str(conflict.get("status", "UNMAPPED"))

        if has_code:
            combo.addItem("以编码为准", "code")
        if has_name:
            combo.addItem("以名称为准", "name")

        default_strategy = "ignored"
        if status == "CONFLICT":
            default_strategy = "code" if has_code else ("name" if has_name else "ignored")
        elif status == "CODE_ONLY" and has_code:
            default_strategy = "code"
        elif status == "NAME_ONLY" and has_name:
            default_strategy = "name"

        for i in range(combo.count()):
            if combo.itemData(i) == default_strategy:
                combo.setCurrentIndex(i)
                break
        return combo

    def populate_table(self):
        self.table.setRowCount(len(self.conflicts))
        code_count = 0
        name_count = 0
        conflict_count = 0
        for i, c in enumerate(self.conflicts):
            status = str(c.get("status", "UNMAPPED"))
            if status == "CODE_ONLY":
                code_count += 1
            elif status == "NAME_ONLY":
                name_count += 1
            elif status == "CONFLICT":
                conflict_count += 1

            combo = self._combo_for_conflict(c)
            self.table.setCellWidget(i, 0, combo)
            self.table.setItem(i, 1, QTableWidgetItem(status))
            self.table.setItem(i, 2, QTableWidgetItem(str(c.get("invoice_number", ""))))
            self.table.setItem(i, 3, QTableWidgetItem(str(int(c.get("item_index", 0)) + 1)))
            self.table.setItem(i, 4, QTableWidgetItem(str(c.get("raw_code", ""))))
            self.table.setItem(i, 5, QTableWidgetItem(str(c.get("raw_name", ""))))
            self.table.setItem(i, 6, QTableWidgetItem(self._candidate_text(c.get("code_candidate"))))
            self.table.setItem(i, 7, QTableWidgetItem(self._candidate_text(c.get("name_candidate"))))
            self.table.setItem(i, 8, QTableWidgetItem(f"{float(c.get('name_score', 0.0)):.2f}"))
            self.table.item(i, 8).setTextAlignment(Qt.AlignCenter)

        self.summary_label.setText(
            f"待处理总数: {len(self.conflicts)} | CONFLICT: {conflict_count} | CODE_ONLY: {code_count} | NAME_ONLY: {name_count}"
        )

    def apply_all_strategy(self, strategy: str):
        for row in range(self.table.rowCount()):
            combo = self.table.cellWidget(row, 0)
            if not isinstance(combo, QComboBox):
                continue
            for i in range(combo.count()):
                if combo.itemData(i) == strategy:
                    combo.setCurrentIndex(i)
                    break

    def accept_with_validation(self):
        self.decisions = []
        for row in range(self.table.rowCount()):
            conflict = self.conflicts[row]
            combo = self.table.cellWidget(row, 0)
            if not isinstance(combo, QComboBox):
                continue
            strategy = combo.currentData()
            if strategy == "code" and not conflict.get("code_candidate"):
                QMessageBox.warning(self, "策略无效", f"第 {row+1} 行没有编码候选，无法选择“以编码为准”。")
                return
            if strategy == "name" and not conflict.get("name_candidate"):
                QMessageBox.warning(self, "策略无效", f"第 {row+1} 行没有名称候选，无法选择“以名称为准”。")
                return
            self.decisions.append((conflict, strategy))
        self.accept()

class MappingLibraryDialog(QDialog):
    """综合映射库管理窗口"""
    def __init__(self, mapping_mgr: MappingManager, invoices: list, parent=None):
        super().__init__(parent)
        self.mgr = mapping_mgr
        self.invoices = invoices
        self.setWindowTitle("标准映射 (Standardization Library)")
        self.resize(900, 600)
        
        layout = QVBoxLayout(self)
        
        self.tabs = QTabWidget()
        self.prod_tab = QWidget()
        self.part_tab = QWidget()
        
        self.setup_product_tab()
        self.setup_partner_tab()
        
        self.tabs.addTab(self.prod_tab, "📦 产品映射 (Products)")
        self.tabs.addTab(self.part_tab, "🏢 往来单位映 (Partners)")
        
        layout.addWidget(self.tabs)
        
        # Bottom Buttons
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def setup_product_tab(self):
        layout = QVBoxLayout(self.prod_tab)
        
        # Tools
        tb = QHBoxLayout()
        btn_add = QPushButton(" 添加")
        btn_del = QPushButton(" 删除")
        btn_scan = QPushButton("🔍 扫描现有产品")
        btn_export = QPushButton("📤 导出")
        
        # View Toggle
        self.btn_prod_view = QPushButton("👁 视图: 列表")
        self.btn_prod_view.setCheckable(True)
        self.btn_prod_view.clicked.connect(self.toggle_prod_view)
        
        btn_add.clicked.connect(self.ensure_prod_table_view)
        btn_add.clicked.connect(self.add_prod_row)
        
        btn_del.clicked.connect(self.ensure_prod_table_view)
        btn_del.clicked.connect(lambda: self.del_row(self.prod_table))
        
        btn_scan.clicked.connect(self.ensure_prod_table_view)
        btn_scan.clicked.connect(self.scan_products)
        
        btn_export.clicked.connect(lambda: export_qtable(self.prod_table, self, "products_mapping"))
        
        tb.addWidget(btn_add)
        tb.addWidget(btn_del)
        tb.addWidget(btn_scan)
        tb.addWidget(btn_export)
        
        # 搜索功能
        tb.addSpacing(20)
        tb.addWidget(QLabel("🔍 搜索:"))
        self.prod_search_input = QLineEdit()
        self.prod_search_input.setPlaceholderText("输入编码或名称...")
        self.prod_search_input.setFixedWidth(200)
        self.prod_search_input.textChanged.connect(lambda t: self.filter_mapping_table(self.prod_table, t))
        tb.addWidget(self.prod_search_input)
        
        tb.addStretch()
        tb.addWidget(self.btn_prod_view)
        layout.addLayout(tb)
        
        # Stack
        self.prod_stack = QStackedWidget()
        
        # Page 1: Table
        self.prod_table = QTableWidget()
        self.prod_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.prod_table.customContextMenuRequested.connect(self.show_prod_menu)
        self.prod_table.setColumnCount(3)
        self.prod_table.setHorizontalHeaderLabels(["原始编码 (Key)", "标准国内编码", "标准名称"])
        self.prod_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.prod_stack.addWidget(self.prod_table)
        
        # Page 2: Tree
        self.prod_tree = QTreeWidget()
        self.prod_tree.setHeaderLabels(["标准名称 / 原始编码", "标准编码", "类型"])
        self.prod_tree.setColumnWidth(0, 400)
        self.prod_stack.addWidget(self.prod_tree)
        
        layout.addWidget(self.prod_stack)
        
        # Load Data
        for code, info in self.mgr.data["products"].items():
            self.add_prod_row(code, info.get("std_code", ""), info.get("std_name", ""))

    def ensure_prod_table_view(self):
        if self.btn_prod_view.isChecked():
            self.btn_prod_view.setChecked(False)
            self.toggle_prod_view()

    def toggle_prod_view(self):
        if self.btn_prod_view.isChecked():
            self.btn_prod_view.setText("👁 视图: 分组 (只读)")
            self.populate_prod_tree()
            self.prod_stack.setCurrentWidget(self.prod_tree)
        else:
            self.btn_prod_view.setText("👁 视图: 列表")
            self.prod_stack.setCurrentWidget(self.prod_table)

    def populate_prod_tree(self):
        self.prod_tree.clear()
        groups = {} # (std_name, std_code) -> [list of orig_codes]
        
        # Read from Table
        for r in range(self.prod_table.rowCount()):
            orig = self.prod_table.item(r, 0).text()
            scode = self.prod_table.item(r, 1).text()
            sname = self.prod_table.item(r, 2).text()
            
            # Normalize key
            k_name = sname.strip() if sname else "(未命 "
            k_code = scode.strip()
            
            key = (k_name, k_code)
            if key not in groups: groups[key] = []
            groups[key].append(orig)
            
        # Sort groups by name
        sorted_keys = sorted(groups.keys(), key=lambda x: x[0])
        
        for (sname, scode), origs in groups.items():
            # Group Node
            grp_item = QTreeWidgetItem(self.prod_tree)
            grp_item.setText(0, sname)
            grp_item.setText(1, scode)
            grp_item.setText(2, f"包含 {len(origs)} 个原始项")
            # Gray background for group
            for c in range(3):
                grp_item.setBackground(c, QBrush(QColor(240, 240, 240)))
            grp_item.setFont(0, QFont("Arial", 9, QFont.Bold))
            
            for orig in origs:
                child = QTreeWidgetItem(grp_item)
                child.setText(0, orig)
                child.setText(2, "原始编码")
            
        self.prod_tree.expandAll()

    def setup_partner_tab(self):
        layout = QVBoxLayout(self.part_tab)
        
        # Tools
        tb = QHBoxLayout()
        btn_add = QPushButton(" 添加")
        btn_del = QPushButton(" 删除")
        btn_scan = QPushButton("🔍 扫描现有单位")
        btn_export = QPushButton("📤 导出")
        
        # View Toggle
        self.btn_part_view = QPushButton("👁 视图: 列表")
        self.btn_part_view.setCheckable(True)
        self.btn_part_view.clicked.connect(self.toggle_part_view)
        
        btn_add.clicked.connect(self.ensure_part_table_view)
        btn_add.clicked.connect(self.add_part_row)
        
        btn_del.clicked.connect(self.ensure_part_table_view)
        btn_del.clicked.connect(lambda: self.del_row(self.part_table))
        
        btn_scan.clicked.connect(self.ensure_part_table_view)
        btn_scan.clicked.connect(self.scan_partners)
        
        btn_export.clicked.connect(lambda: export_qtable(self.part_table, self, "partners_mapping"))
        
        tb.addWidget(btn_add)
        tb.addWidget(btn_del)
        tb.addWidget(btn_scan)
        tb.addWidget(btn_export)
        
        # 搜索功能
        tb.addSpacing(20)
        tb.addWidget(QLabel("🔍 搜索:"))
        self.part_search_input = QLineEdit()
        self.part_search_input.setPlaceholderText("输入税号或名称...")
        self.part_search_input.setFixedWidth(200)
        self.part_search_input.textChanged.connect(lambda t: self.filter_mapping_table(self.part_table, t))
        tb.addWidget(self.part_search_input)
        
        tb.addStretch()
        tb.addWidget(self.btn_part_view)
        layout.addLayout(tb)

        # Stack
        self.part_stack = QStackedWidget()
        
        # Page 1: Table
        self.part_table = QTableWidget()
        self.part_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.part_table.customContextMenuRequested.connect(self.show_part_menu)
        self.part_table.setColumnCount(2)
        self.part_table.setHorizontalHeaderLabels(["识别Key (CNPJ/名称)", "标准显示名称"])
        self.part_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.part_stack.addWidget(self.part_table)
        
        # Page 2: Tree
        self.part_tree = QTreeWidget()
        self.part_tree.setHeaderLabels(["标准名称 / 原始Key", "类型"])
        self.part_tree.setColumnWidth(0, 400)
        self.part_stack.addWidget(self.part_tree)
        
        layout.addWidget(self.part_stack)
        
        # Load Data
        for key, name in self.mgr.data["partners"].items():
            self.add_part_row(key, name)

    def filter_mapping_table(self, table, keyword):
        """通用映射表过滤逻辑"""
        keyword = keyword.strip().lower()
        for r in range(table.rowCount()):
            match = False
            for c in range(table.columnCount()):
                item = table.item(r, c)
                if item and keyword in item.text().lower():
                    match = True
                    break
            table.setRowHidden(r, not match)

    def ensure_part_table_view(self):
        if self.btn_part_view.isChecked():
            self.btn_part_view.setChecked(False)
            self.toggle_part_view()

    def toggle_part_view(self):
        if self.btn_part_view.isChecked():
            self.btn_part_view.setText("👁 视图: 分组 (只读)")
            self.populate_part_tree()
            self.part_stack.setCurrentWidget(self.part_tree)
        else:
            self.btn_part_view.setText("👁 视图: 列表")
            self.part_stack.setCurrentWidget(self.part_table)

    def populate_part_tree(self):
        self.part_tree.clear()
        groups = {} # std_name -> [list of orig_keys]
        
        for r in range(self.part_table.rowCount()):
            key = self.part_table.item(r, 0).text()
            name = self.part_table.item(r, 1).text()
            
            grp = name.strip() if name else "(未命 "
            if grp not in groups: groups[grp] = []
            groups[grp].append(key)
            
        # Sort keys
        sorted_keys = sorted(groups.keys())
        
        for grp_name in sorted_keys:
            keys = groups[grp_name]
            grp_item = QTreeWidgetItem(self.part_tree)
            grp_item.setText(0, grp_name)
            grp_item.setText(1, f"包含 {len(keys)} 个原始项")
            for c in range(2):
                grp_item.setBackground(c, QBrush(QColor(240, 240, 240)))
            grp_item.setFont(0, QFont("Arial", 9, QFont.Bold))
            
            for k in keys:
                child = QTreeWidgetItem(grp_item)
                child.setText(0, k)
                child.setText(1, "原始Key")
                
        self.part_tree.expandAll()

    # --- Menu Handlers ---
    def show_prod_menu(self, pos):
        item = self.prod_table.itemAt(pos)
        if not item: return
        
        menu = QMenu(self)
        menu.addAction("🔍 寻找相似 (Search Similar)", self.search_similar_prod)
        menu.addSeparator()
        menu.addAction("🔀 并入名称 (Merge to Name)", self.merge_prod_name)
        menu.addAction("🔑 并入Key (Merge to Key)", self.merge_prod_key)
        menu.exec(self.prod_table.viewport().mapToGlobal(pos))

    def show_part_menu(self, pos):
        item = self.part_table.itemAt(pos)
        if not item: return
        
        menu = QMenu(self)
        menu.addAction("🔍 寻找相似 (Search Similar)", self.search_similar_part)
        menu.addSeparator()
        menu.addAction("🔀 并入名称 (Merge to Name)", self.merge_part_name)
        menu.addAction("🔑 并入Key (Merge to Key)", self.merge_part_key)
        menu.exec(self.part_table.viewport().mapToGlobal(pos))

    # --- Search Similar Logic ---
    def calculate_similarity(self, s1, s2):
        if not s1 or not s2: return 0.0
        s1, s2 = str(s1).strip().lower(), str(s2).strip().lower()
        if not s1 or not s2: return 0.0
        
        if s1 == s2: return 1.0
        
        # Normalize: remove non-alphanumeric
        n1 = re.sub(r'[^a-z0-9]', '', s1)
        n2 = re.sub(r'[^a-z0-9]', '', s2)
        if n1 and n2 and n1 == n2: return 0.95
        
        # Substring
        if n1 and n2 and (n1 in n2 or n2 in n1): 
            # Penalize very short substrings matches
            if len(n1) < 4 or len(n2) < 4: return 0.7
            return 0.85
        
        # Fuzzy
        return difflib.SequenceMatcher(None, s1, s2).ratio()

    def search_similar_prod(self):
        current_row = self.prod_table.currentRow()
        if current_row < 0: 
            QMessageBox.warning(self, "提示", "请先选择一行作为查找基准")
            return
        
        target_key = self.prod_table.item(current_row, 0).text()
        target_name = self.prod_table.item(current_row, 2).text()
        
        # Get Config
        dlg = SimilaritySearchConfigDialog(self)
        if not dlg.exec(): return
        settings = dlg.get_settings()
        
        mode = settings["mode"]
        threshold = settings["threshold"]
        
        matches = []
        for r in range(self.prod_table.rowCount()):
            if r == current_row: continue
            
            key = self.prod_table.item(r, 0).text()
            name = self.prod_table.item(r, 2).text()
            
            score = 0.0
            s_key = self.calculate_similarity(target_key, key)
            s_name = self.calculate_similarity(target_name, name)
            
            if mode == "key":
                score = s_key
            elif mode == "name":
                score = s_name
            else:
                score = max(s_key, s_name)
            
            if score >= threshold: 
                matches.append((r, score, key, name))
        
        matches.sort(key=lambda x: x[1], reverse=True)
        
        if not matches:
            reply = QMessageBox.question(self, "未找到匹配", 
                                       f"未找到相似度高于 {threshold*100:.0f}% 的项。\n是否手动从列表中选择？",
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                # Prepare all items
                all_items = []
                for k, v in self.data.get('products', {}).items():
                    nm = v.get('std_name', '')
                    all_items.append(f"{k} | {nm}")
                all_items.sort()
                
                item, ok = QInputDialog.getItem(self, "手动选择", "请选择标准项 (Key | Name):", all_items, 0, False)
                if ok and item:
                    try:
                        sel_key = item.split(" | ")[0].strip()
                        sel_name = item.split(" | ")[1].strip()
                        self.prod_table.setItem(current_row, 1, QTableWidgetItem(sel_key))
                        self.prod_table.setItem(current_row, 2, QTableWidgetItem(sel_name))
                        QMessageBox.information(self, "成功", "已手动应用选择。")
                    except Exception as e:
                        QMessageBox.warning(self, "错误", f"应用选择时出错: {str(e)}")
            return
            
        # Show Results with target context
        target_info = f"{target_key} | {target_name}"
        res_dlg = SearchResultsDialog(matches, self.prod_table, target_info, self)
        res_dlg.exec()

    def search_similar_part(self):
        current_row = self.part_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "提示", "请先选择一行作为查找基准 ")
            return
        
        target_key = self.part_table.item(current_row, 0).text()
        target_name = self.part_table.item(current_row, 1).text()
        
        # Get Config
        dlg = SimilaritySearchConfigDialog(self)
        if not dlg.exec(): return
        settings = dlg.get_settings()
        
        mode = settings["mode"]
        threshold = settings["threshold"]
        
        matches = []
        for r in range(self.part_table.rowCount()):
            if r == current_row: continue
            
            key = self.part_table.item(r, 0).text()
            name = self.part_table.item(r, 1).text()
            
            score = 0.0
            s_key = self.calculate_similarity(target_key, key)
            s_name = self.calculate_similarity(target_name, name)
            
            if mode == "key":
                score = s_key
            elif mode == "name":
                score = s_name
            else:
                score = max(s_key, s_name)
            
            if score >= threshold: 
                matches.append((r, score, key, name))
        
        matches.sort(key=lambda x: x[1], reverse=True)
        
        if not matches:
            reply = QMessageBox.question(self, "未找到匹配", 
                                       f"未找到相似度高于 {threshold*100:.0f}% 的项。\n是否手动从列表中选择？",
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                # Prepare all items
                all_items = []
                # Partner data structure is simpler: key -> name
                for k, v in self.data.get('partners', {}).items():
                    all_items.append(f"{k} | {v}")
                all_items.sort()
                
                item, ok = QInputDialog.getItem(self, "手动选择", "请选择标准项 (Key | Name):", all_items, 0, False)
                if ok and item:
                    try:
                        sel_key = item.split(" | ")[0].strip()
                        sel_name = item.split(" | ")[1].strip()
                        # self.part_table is current
                        self.part_table.setItem(current_row, 0, QTableWidgetItem(sel_key))
                        self.part_table.setItem(current_row, 1, QTableWidgetItem(sel_name))
                        QMessageBox.information(self, "成功", "已手动应用选择。")
                    except Exception as e:
                        QMessageBox.warning(self, "错误", f"应用选择时出错: {str(e)}")
            return
            
        # Show Results with target context
        target_info = f"{target_key} | {target_name}"
        res_dlg = SearchResultsDialog(matches, self.part_table, target_info, self)
        res_dlg.exec()

    # --- Merge Logic ---
    def merge_prod_name(self):
        """Product: Merge selected rows to a single Standard Name"""
        rows = sorted(set(i.row() for i in self.prod_table.selectedItems()))
        if not rows: return
        
        # Default text from first selected row's std name
        default_name = self.prod_table.item(rows[0], 2).text()
        
        name, ok = QInputDialog.getText(self, "并入名称", "请输入统一的标准名 ", text=default_name)
        if ok and name:
            for r in rows:
                self.prod_table.setItem(r, 2, QTableWidgetItem(name))

    def merge_prod_key(self):
        """Product: Pick a key from selection, apply its Std Name/Code to all"""
        rows = sorted(set(i.row() for i in self.prod_table.selectedItems()))
        if not rows: return
        
        # Create list of "Key - Name" for user to pick
        options = []
        for r in rows:
            key = self.prod_table.item(r, 0).text()
            name = self.prod_table.item(r, 2).text()
            options.append(f"{key} | {name}")
            
        item, ok = QInputDialog.getItem(self, "并入Key", "选择作为标准的Key (将应用其标准值到所有选中 :", options, 0, False)
        if ok and item:
            # Extract index from selection
            idx = options.index(item)
            target_row = rows[idx]
            target_std_code = self.prod_table.item(target_row, 1).text()
            target_std_name = self.prod_table.item(target_row, 2).text()
            
            for r in rows:
                self.prod_table.setItem(r, 1, QTableWidgetItem(target_std_code))
                self.prod_table.setItem(r, 2, QTableWidgetItem(target_std_name))

    def merge_part_name(self):
        """Partner: Merge selected rows to a single Standard Name"""
        rows = sorted(set(i.row() for i in self.part_table.selectedItems()))
        if not rows: return
        
        default_name = self.part_table.item(rows[0], 1).text()
        
        name, ok = QInputDialog.getText(self, "并入名称", "请输入统一的标准名 ", text=default_name)
        if ok and name:
            for r in rows:
                self.part_table.setItem(r, 1, QTableWidgetItem(name))

    def merge_part_key(self):
        """Partner: Pick a key, apply its name to all"""
        rows = sorted(set(i.row() for i in self.part_table.selectedItems()))
        if not rows: return
        
        options = []
        for r in rows:
            key = self.part_table.item(r, 0).text()
            name = self.part_table.item(r, 1).text()
            options.append(f"{key} | {name}")
            
        item, ok = QInputDialog.getItem(self, "并入Key", "选择作为标准的Key (将应用其名称到所有选中 :", options, 0, False)
        if ok and item:
            idx = options.index(item)
            target_row = rows[idx]
            target_name = self.part_table.item(target_row, 1).text()
            # If target name is empty, maybe use Key itself?
            if not target_name:
                target_name = self.part_table.item(target_row, 0).text()
            
            for r in rows:
                self.part_table.setItem(r, 1, QTableWidgetItem(target_name))

    def add_prod_row(self, c="", sc="", sn=""):
        r = self.prod_table.rowCount()
        self.prod_table.insertRow(r)
        self.prod_table.setItem(r, 0, QTableWidgetItem(str(c)))
        self.prod_table.setItem(r, 1, QTableWidgetItem(str(sc)))
        self.prod_table.setItem(r, 2, QTableWidgetItem(str(sn)))

    def add_part_row(self, k="", n=""):
        r = self.part_table.rowCount()
        self.part_table.insertRow(r)
        self.part_table.setItem(r, 0, QTableWidgetItem(str(k)))
        self.part_table.setItem(r, 1, QTableWidgetItem(str(n)))

    def del_row(self, table):
        rows = set(i.row() for i in table.selectedItems())
        for r in sorted(rows, reverse=True):
            table.removeRow(r)

    def scan_products(self):
        """Scan invoices for unique product codes"""
        seen = set()
        for r in range(self.prod_table.rowCount()):
            seen.add(self.prod_table.item(r, 0).text())
            
        count = 0
        for inv in self.invoices:
            for item in inv.itens:
                c = item.codigo_produto
                if c and c not in seen:
                    self.add_prod_row(c, item.codigo_domestico or "", item.descricao or "")
                    seen.add(c)
                    count += 1
        QMessageBox.information(self, "扫描结果", f"新增 {count} 个产品编")

    def scan_partners(self):
        """Scan invoices for issuers/recipients"""
        seen = set()
        for r in range(self.part_table.rowCount()):
            seen.add(self.part_table.item(r, 0).text())
            
        count = 0
        def try_add(key, name):
            nonlocal count
            if key and key not in seen:
                self.add_part_row(key, name)
                seen.add(key)
                count += 1

        for inv in self.invoices:
            # Issuer
            if inv.emitente_cnpj:
                clean = re.sub(r'\D', '', inv.emitente_cnpj)
                try_add(clean, inv.emitente_nome)
            elif inv.emitente_nome:
                try_add(inv.emitente_nome, inv.emitente_nome)
                
            # Recipient
            if inv.destinatario_cnpj:
                clean = re.sub(r'\D', '', inv.destinatario_cnpj)
                try_add(clean, inv.destinatario_nome)
            elif inv.destinatario_nome:
                try_add(inv.destinatario_nome, inv.destinatario_nome)
                
        QMessageBox.information(self, "扫描结果", f"新增 {count} 个单位Key")

    def accept(self):
        """Save all data back to manager"""
        new_prod = {}
        dropped_prod = 0
        for r in range(self.prod_table.rowCount()):
            raw_key = self.prod_table.item(r, 0).text().strip()
            if not raw_key:
                continue
            existing = self.mgr.get_product_std(raw_key) or {}
            _, sanitized, _ = self.mgr._sanitize_product_entry(
                raw_key,
                {
                    "std_code": self.prod_table.item(r, 1).text().strip(),
                    "std_name": self.prod_table.item(r, 2).text().strip(),
                    "status": existing.get("status", ""),
                    "source": existing.get("source", "mapping_library_dialog") or "mapping_library_dialog",
                    "confidence": existing.get("confidence", ""),
                    "updated_at": existing.get("updated_at", ""),
                    "reviewed_by": existing.get("reviewed_by", ""),
                    "sample_text": existing.get("sample_text", ""),
                    "notes": existing.get("notes", ""),
                },
                source=existing.get("source", "mapping_library_dialog") or "mapping_library_dialog",
                status=existing.get("status") or None,
            )
            if sanitized:
                norm_key = self.mgr._normalize_product_key(raw_key)
                new_prod[norm_key] = self.mgr._merge_product_entries(new_prod.get(norm_key), sanitized)
            else:
                dropped_prod += 1
        self.mgr.data["products"] = new_prod

        new_part = {}
        for r in range(self.part_table.rowCount()):
            raw_key = self.part_table.item(r, 0).text().strip()
            raw_value = self.part_table.item(r, 1).text().strip()
            norm_key = self.mgr._normalize_partner_key(raw_key)
            clean_value = self.mgr._clean_text(raw_value)
            if norm_key and clean_value:
                new_part[norm_key] = clean_value
        self.mgr.data["partners"] = new_part

        self.mgr.save()
        if dropped_prod:
            QMessageBox.information(self, "映射清洗", f"已自动跳过 {dropped_prod} 条无效或污染的产品映射。")
        super().accept()

# --- GUI 组件 ---
class ProductMappingDialog(QDialog):
    """产品编码映射管理对话框"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.setWindowTitle("产品编码映射管理")
        self.setMinimumSize(700, 500)

        layout = QVBoxLayout(self)

        # 说明
        info_label = QLabel("管理巴西产品编码到国内编码的映射关系。可以手动添加或从表格中自动提取。")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        # 映射表格
        self.mapping_table = QTableWidget()
        self.mapping_table.setColumnCount(2)
        self.mapping_table.setHorizontalHeaderLabels(["巴西产品编码", "国内编码"])
        self.mapping_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.mapping_table)

        # 按钮
        button_layout = QHBoxLayout()
        self.add_btn = QPushButton("添加映射")
        self.remove_btn = QPushButton("删除选中")
        self.auto_extract_btn = QPushButton("🔍 从当前数据提取")
        self.import_btn = QPushButton("📥 导入映射")
        self.export_btn = QPushButton("📤 导出映射")

        self.add_btn.clicked.connect(self.add_mapping)
        self.remove_btn.clicked.connect(self.remove_mapping)
        self.auto_extract_btn.clicked.connect(self.auto_extract_mappings)
        self.import_btn.clicked.connect(self.import_mappings)
        self.export_btn.clicked.connect(self.export_mappings)

        button_layout.addWidget(self.add_btn)
        button_layout.addWidget(self.remove_btn)
        button_layout.addWidget(self.auto_extract_btn)
        button_layout.addWidget(self.import_btn)
        button_layout.addWidget(self.export_btn)
        layout.addLayout(button_layout)

        # 对话框按钮
        dialog_buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        dialog_buttons.accepted.connect(self.accept)
        dialog_buttons.rejected.connect(self.reject)
        layout.addWidget(dialog_buttons)

        # 加载现有映射
        self.load_mappings()

    def load_mappings(self):
        """从父窗口加载映射到表"""
        self.mapping_table.setRowCount(0)
        for br_code, cn_code in self.parent_window.product_code_mapping.items():
            row = self.mapping_table.rowCount()
            self.mapping_table.insertRow(row)
            self.mapping_table.setItem(row, 0, QTableWidgetItem(str(br_code)))
            self.mapping_table.setItem(row, 1, QTableWidgetItem(str(cn_code)))

    def add_mapping(self):
        """添加新的映射"""
        row = self.mapping_table.rowCount()
        self.mapping_table.insertRow(row)
        self.mapping_table.setItem(row, 0, QTableWidgetItem(""))
        self.mapping_table.setItem(row, 1, QTableWidgetItem(""))

    def remove_mapping(self):
        """删除选中的映"""
        selected_rows = set(item.row() for item in self.mapping_table.selectedItems())
        for row in sorted(selected_rows, reverse=True):
            self.mapping_table.removeRow(row)

    def auto_extract_mappings(self):
        """从当前数据中提取已有的映"""
        count = 0
        for invoice in self.parent_window.invoices:
            for item in invoice.itens:
                if item.codigo_produto and item.codigo_domestico:
                    # 检查是否已存在
                    if item.codigo_produto not in self.parent_window.product_code_mapping:
                        self.parent_window.product_code_mapping[item.codigo_produto] = item.codigo_domestico
                        count += 1

        self.load_mappings()
        QMessageBox.information(self, "提取完成", f"已提 {count} 个新映射")

    def import_mappings(self):
        """从Excel或JSON文件导入映射"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "导入映射文件",
            "",
            "Excel Files (*.xlsx *.xls);;JSON Files (*.json);;All Files (*.*)"
        )
        if not file_path:
            return

        try:
            imported = {}
            file_ext = os.path.splitext(file_path)[1].lower()

            if file_ext in ['.xlsx', '.xls']:
                # 读取Excel文件
                if openpyxl is None:
                    QMessageBox.critical(self, "导入失败", "未安装openpyxl库，无法读取Excel文件。\n请安  pip install openpyxl")
                    return

                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active

                # 读取数据 (假设第一列是巴西编码，第二列是国内编码，第一行可能是表头)
                row_count = 0
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
                    if len(row) < 2:
                        continue

                    br_code = str(row[0]).strip() if row[0] else ""
                    cn_code = str(row[1]).strip() if row[1] else ""

                    # 跳过表头和空
                    if not br_code or not cn_code:
                        continue
                    if row_idx == 1 and (br_code.lower() in ['codigo', 'code', '编码', '巴西编码'] or
                                         cn_code.lower() in ['codigo', 'code', '编码', '国内编码']):
                        continue

                    imported[br_code] = cn_code
                    row_count += 1

                wb.close()

            elif file_ext == '.json':
                # 读取JSON文件
                with open(file_path, 'r', encoding='utf-8') as f:
                    imported = json.load(f)
            else:
                QMessageBox.warning(self, "不支持的格式", f"不支持的文件格式: {file_ext}\n请选择 .xlsx, .xls  .json 文件")
                return

            # 合并映射
            self.parent_window.product_code_mapping.update(imported)

            # 自动保存映射文件
            self.parent_window.save_product_mapping()

            # 刷新显示
            self.load_mappings()

            QMessageBox.information(self, "导入成功", f"已导 {len(imported)} 个映射\n映射文件已自动保")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"无法读取文件: {e}")

    def export_mappings(self):
        """导出映射到JSON文件"""
        file_path, _ = QFileDialog.getSaveFileName(self, "导出映射文件", "", "JSON Files (*.json)")
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(self.parent_window.product_code_mapping, f, ensure_ascii=False, indent=2)
                QMessageBox.information(self, "导出成功", f"已导 {len(self.parent_window.product_code_mapping)} 个映")
            except Exception as e:
                QMessageBox.critical(self, "导出失败", f"无法写入文件: {e}")

    def accept(self):
        """保存映射并关闭"""
        # 从表格收集映射
        new_mapping = {}
        for row in range(self.mapping_table.rowCount()):
            br_code_item = self.mapping_table.item(row, 0)
            cn_code_item = self.mapping_table.item(row, 1)
            if br_code_item and cn_code_item:
                br_code = br_code_item.text().strip()
                cn_code = cn_code_item.text().strip()
                if br_code and cn_code:
                    new_mapping[br_code] = cn_code

        # 更新父窗口的映射
        self.parent_window.product_code_mapping = new_mapping
        self.parent_window.save_product_mapping()

        super().accept()

class ThemeManager:

    """主题管理 """

    @staticmethod

    def apply_theme(app, theme_name):

        app.setStyle("Fusion")

        

        if theme_name == "Dark":

            dark_palette = QPalette()

            dark_palette.setColor(QPalette.Window, QColor(53, 53, 53))

            dark_palette.setColor(QPalette.WindowText, Qt.white)

            dark_palette.setColor(QPalette.Base, QColor(25, 25, 25))

            dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))

            dark_palette.setColor(QPalette.ToolTipBase, Qt.white)

            dark_palette.setColor(QPalette.ToolTipText, Qt.white)

            dark_palette.setColor(QPalette.Text, Qt.white)

            dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))

            dark_palette.setColor(QPalette.ButtonText, Qt.white)

            dark_palette.setColor(QPalette.BrightText, Qt.red)

            dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))

            dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))

            dark_palette.setColor(QPalette.HighlightedText, Qt.black)

            app.setPalette(dark_palette)

        elif theme_name == "Light":

            # Standard light palette

            app.setPalette(QApplication.style().standardPalette())

        else: # System or default

            # Reset to system default

            app.setPalette(QApplication.style().standardPalette())



class WorkerThread(QThread):

    """用于在后台线程处理发票提取和解析的任务"""

    progress = Signal(int)

    finished = Signal(list)

    log = Signal(str)



    def __init__(self, file_paths, config):

        super().__init__()

        self.file_paths = file_paths

        self.config = config

        self._is_running = True



    def stop(self):

        self._is_running = False



    def run(self):

        results = []

        total_files = len(self.file_paths)

        if total_files == 0:

            self.finished.emit([])

            return



        prefer_pymupdf = self.config.get("prefer_pymupdf", True)

        enable_ocr = self.config.get("enable_ocr", False)

        ocr_lang = self.config.get("ocr_lang", "por")

        poppler_path = self.config.get("poppler_path")

        max_workers = self.config.get("max_workers", 4)

        enable_llm = self.config.get("enable_llm", False)



        processed_count = 0

        

        # 使用有界并发：仅保持 max_workers 个任务在跑，停止时不再补充新任务
        executor = ThreadPoolExecutor(max_workers=max_workers)
        try:
            file_iter = iter(self.file_paths)
            future_to_file = {}

            initial_workers = min(max_workers, total_files)
            for _ in range(initial_workers):
                fp = next(file_iter, None)
                if fp is None:
                    break
                fut = executor.submit(
                    self.process_single_file,
                    fp,
                    prefer_pymupdf,
                    enable_ocr,
                    ocr_lang,
                    poppler_path,
                    enable_llm,
                )
                future_to_file[fut] = fp

            while future_to_file:
                if not self._is_running:
                    for fut in list(future_to_file.keys()):
                        fut.cancel()
                    break

                done, _ = wait(list(future_to_file.keys()), timeout=0.2, return_when=FIRST_COMPLETED)
                if not done:
                    continue

                for future in done:
                    file_name = future_to_file.pop(future, "")
                    try:
                        invoice = future.result()
                        if invoice:
                            results.append(invoice)
                    except Exception as e:
                        self.log.emit(f"处理文件 {file_name} 时发生错误: {e}")

                    processed_count += 1
                    progress_val = int((processed_count / total_files) * 100)
                    self.progress.emit(progress_val)

                    if self._is_running:
                        next_file = next(file_iter, None)
                        if next_file:
                            new_future = executor.submit(
                                self.process_single_file,
                                next_file,
                                prefer_pymupdf,
                                enable_ocr,
                                ocr_lang,
                                poppler_path,
                                enable_llm,
                            )
                            future_to_file[new_future] = next_file
        finally:
            executor.shutdown(wait=False, cancel_futures=True)

        self.finished.emit(results)



    def process_single_file(self, file_path, prefer_pymupdf, enable_ocr, ocr_lang, poppler_path, enable_llm):

        """处理单个文件的内部函数"""

        file_name = os.path.basename(file_path)

        self.log.emit(f"开始处理: {file_name}")

        

        try:
            # 1. 提取文本
            text, meta = extract_text_from_pdf(
                file_path, 
                prefer_pymupdf=prefer_pymupdf,
                enable_ocr=enable_ocr,
                ocr_lang=ocr_lang,
                poppler_path=poppler_path
            )
            
            # 2. 解析发票
            invoice = parse_invoice_from_text(text, file_name)
            invoice.file_path = file_path # 记录绝对路径
            invoice.extract_meta = meta
            
            # 3. LLM 辅助 (如果启用)

            if enable_llm and self.config.get("llm_endpoint"):
                try:
                    # 注意：这里需要确保 run_llm_assist 是线程安全的，或者在这里加锁
                    # 简单起见，假设 API 支持并发
                    llm_res = run_llm_assist(file_path, text, self.config)
                    if llm_res:
                        apply_llm_result(invoice, llm_res)
                except Exception as llm_err:
                    self.log.emit(f"LLM 辅助失败，已跳过: {file_name} | {llm_err}")

            

            return invoice

        except Exception as e:

            self.log.emit(f"文件 {file_name} 解析失败: {e}")

            return None


class DataLoadWorker(QThread):
    """后台线程加载 JSON 数据，避免阻塞主界面"""
    progress = Signal(int, int)  # current, total
    finished = Signal(list, dict)  # invoices, recon_data
    error = Signal(str)
    log = Signal(str)

    def __init__(self, filepath: str):
        super().__init__()
        self.filepath = filepath
        self._is_running = True

    def stop(self):
        self._is_running = False

    def run(self):
        try:
            self.log.emit(f"开始加载数据文件: {self.filepath}")
            with open(self.filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)

            invoices = []
            inv_list = data.get("invoices", [])
            total = len(inv_list)
            self.log.emit(f"共 {total} 条发票数据，开始解析...")

            for idx, inv_dict in enumerate(inv_list):
                if not self._is_running:
                    self.log.emit("加载被中断")
                    return

                # 每 500 条发送一次进度
                if idx % 500 == 0:
                    self.progress.emit(idx, total)

                # 提取 payment_history（不影响后续处理）
                inv_dict.pop("payment_history", [])

                # 提取 items
                items_data = inv_dict.pop("itens", [])
                items = [Item(**item_dict) for item_dict in items_data]

                # 创建 Invoice 对象
                inv_dict["itens"] = items
                invoice = Invoice(**inv_dict)
                invoices.append(invoice)

            self.progress.emit(total, total)
            recon_data = data.get("recon_data", {})
            self.log.emit(f"数据解析完成，共 {len(invoices)} 条发票")
            self.finished.emit(invoices, recon_data)

        except Exception as e:
            self.error.emit(str(e))


class InvoiceRowsLoadWorker(QThread):
    """后台加载收款发票列表，避免主线程阻塞。"""
    finished = Signal(int, list)  # request_id, rows
    error = Signal(int, str)

    def __init__(self, db, request_id: int, parent=None):
        super().__init__(parent)
        self.db = db
        self.request_id = request_id
        self._is_running = True

    def stop(self):
        self._is_running = False

    @staticmethod
    def _to_plain_row(row):
        if isinstance(row, dict):
            return dict(row)
        try:
            return dict(row)
        except Exception:
            return {}

    def run(self):
        try:
            raw_rows = self.db.get_invoices()
            if not self._is_running:
                return

            rows = []
            for row in raw_rows:
                if not self._is_running:
                    return
                rows.append(self._to_plain_row(row))

            self.finished.emit(self.request_id, rows)
        except Exception as e:
            self.error.emit(self.request_id, str(e))


class InstallmentRowsLoadWorker(QThread):
    """后台加载分期明细，避免主线程阻塞。"""
    finished = Signal(int, int, list)  # request_id, invoice_id, rows
    error = Signal(int, int, str)  # request_id, invoice_id, message

    def __init__(self, db, request_id: int, invoice_id: int, parent=None):
        super().__init__(parent)
        self.db = db
        self.request_id = request_id
        self.invoice_id = invoice_id
        self._is_running = True

    def stop(self):
        self._is_running = False

    @staticmethod
    def _to_plain_row(row):
        if isinstance(row, dict):
            return dict(row)
        try:
            return dict(row)
        except Exception:
            return {}

    def run(self):
        try:
            raw_rows = self.db.get_installments(self.invoice_id)
            if not self._is_running:
                return

            rows = []
            for row in raw_rows:
                if not self._is_running:
                    return
                rows.append(self._to_plain_row(row))

            self.finished.emit(self.request_id, self.invoice_id, rows)
        except Exception as e:
            self.error.emit(self.request_id, self.invoice_id, str(e))


class TransactionRowsLoadWorker(QThread):
    """后台加载账户流水及余额，避免主线程阻塞。"""
    finished = Signal(int, int, float, list)  # request_id, account_id, balance, rows
    error = Signal(int, int, str)  # request_id, account_id, message

    def __init__(self, db, request_id: int, account_id: int, parent=None):
        super().__init__(parent)
        self.db = db
        self.request_id = request_id
        self.account_id = account_id
        self._is_running = True

    def stop(self):
        self._is_running = False

    @staticmethod
    def _to_plain_row(row):
        if isinstance(row, dict):
            return dict(row)
        try:
            return dict(row)
        except Exception:
            return {}

    def run(self):
        try:
            rows_raw = self.db.get_transactions(self.account_id)
            if not self._is_running:
                return

            current_balance = self.db.get_account_balance(self.account_id)
            if not self._is_running:
                return

            rows = []
            for row in rows_raw:
                if not self._is_running:
                    return
                rows.append(self._to_plain_row(row))

            self.finished.emit(self.request_id, self.account_id, float(current_balance or 0.0), rows)
        except Exception as e:
            self.error.emit(self.request_id, self.account_id, str(e))


class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("设置")
        self.setMinimumWidth(450)
        
        self.settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)
        
        layout = QVBoxLayout(self)
        
        # OCR 设置
        self.ocr_checkbox = QCheckBox("启用 OCR (当标准文本提取失败时)")
        self.ocr_checkbox.setChecked(self.settings.value("enable_ocr", False, type=bool))
        layout.addWidget(self.ocr_checkbox)
        
        ocr_lang_layout = QHBoxLayout()
        ocr_lang_label = QLabel("OCR 语言:")
        self.ocr_lang_input = QLineEdit(self.settings.value("ocr_lang", "por"))
        ocr_lang_layout.addWidget(ocr_lang_label)
        ocr_lang_layout.addWidget(self.ocr_lang_input)
        layout.addLayout(ocr_lang_layout)

        # Poppler 路径
        poppler_layout = QHBoxLayout()
        poppler_label = QLabel("Poppler 路径 (可 :")
        self.poppler_path_input = QLineEdit(self.settings.value("poppler_path", ""))
        self.poppler_browse_btn = QPushButton("浏览")
        self.poppler_browse_btn.clicked.connect(self.browse_poppler)
        poppler_layout.addWidget(poppler_label)
        poppler_layout.addWidget(self.poppler_path_input)
        poppler_layout.addWidget(self.poppler_browse_btn)
        layout.addLayout(poppler_layout)

        # Tesseract 路径
        tesseract_layout = QHBoxLayout()
        tesseract_label = QLabel("Tesseract-OCR 路径:")
        self.tesseract_path_input = QLineEdit(self.settings.value("tesseract_cmd", ""))
        self.tesseract_browse_btn = QPushButton("浏览")
        self.tesseract_browse_btn.clicked.connect(self.browse_tesseract)
        tesseract_layout.addWidget(tesseract_label)
        tesseract_layout.addWidget(self.tesseract_path_input)
        tesseract_layout.addWidget(self.tesseract_browse_btn)
        layout.addLayout(tesseract_layout)

        # PDF 提取引擎
        pdf_engine_layout = QHBoxLayout()
        pdf_engine_label = QLabel("PDF 文本提取引擎:")
        self.pymupdf_checkbox = QCheckBox("优先使用 PyMuPDF (推荐)")
        self.pymupdf_checkbox.setChecked(self.settings.value("prefer_pymupdf", True, type=bool))
        pdf_engine_layout.addWidget(pdf_engine_label)
        pdf_engine_layout.addWidget(self.pymupdf_checkbox)
        layout.addLayout(pdf_engine_layout)

        # 多模 LLM 设置
        llm_title = QLabel("多模 LLM 设置 (LM Studio / 其他 OpenAI 兼容接口)")
        llm_title.setWordWrap(True)
        layout.addWidget(llm_title)

        self.llm_checkbox = QCheckBox("启用 LLM 辅助解析 (不覆盖已识别字段)")
        self.llm_checkbox.setChecked(self.settings.value("enable_llm", False, type=bool))
        layout.addWidget(self.llm_checkbox)

        llm_mode_layout = QHBoxLayout()
        llm_mode_label = QLabel("接口模式:")
        self.llm_mode_input = QComboBox()
        self.llm_mode_input.addItems(["local", "remote"])
        current_mode = self.settings.value("llm_mode", "local")
        if current_mode in ["local", "remote"]:
            self.llm_mode_input.setCurrentText(current_mode)
        llm_mode_layout.addWidget(llm_mode_label)
        llm_mode_layout.addWidget(self.llm_mode_input)
        layout.addLayout(llm_mode_layout)

        llm_endpoint_layout = QHBoxLayout()
        llm_endpoint_label = QLabel("API Endpoint:")
        default_endpoint = self.settings.value("llm_endpoint", "http://localhost:1234/v1")
        self.llm_endpoint_input = QLineEdit(default_endpoint)
        llm_endpoint_layout.addWidget(llm_endpoint_label)
        llm_endpoint_layout.addWidget(self.llm_endpoint_input)
        layout.addLayout(llm_endpoint_layout)

        llm_model_layout = QHBoxLayout()
        llm_model_label = QLabel("模型名称:")
        self.llm_model_input = QLineEdit(self.settings.value("llm_model", ""))
        llm_model_layout.addWidget(llm_model_label)
        llm_model_layout.addWidget(self.llm_model_input)
        layout.addLayout(llm_model_layout)

        llm_key_layout = QHBoxLayout()
        llm_key_label = QLabel("API Key:")
        self.llm_api_key_input = QLineEdit(self.settings.value("llm_api_key", ""))
        self.llm_api_key_input.setEchoMode(QLineEdit.Password)
        llm_key_layout.addWidget(llm_key_label)
        llm_key_layout.addWidget(self.llm_api_key_input)
        layout.addLayout(llm_key_layout)

        self.llm_multimodal_checkbox = QCheckBox("启用多模 (图片 + 文本)")
        self.llm_multimodal_checkbox.setChecked(self.settings.value("llm_use_multimodal", True, type=bool))
        layout.addWidget(self.llm_multimodal_checkbox)

        llm_timeout_layout = QHBoxLayout()
        llm_timeout_label = QLabel("超时 (秒):")
        self.llm_timeout_input = QLineEdit(self.settings.value("llm_timeout_sec", "30"))
        llm_timeout_layout.addWidget(llm_timeout_label)
        llm_timeout_layout.addWidget(self.llm_timeout_input)
        layout.addLayout(llm_timeout_layout)

        llm_chars_layout = QHBoxLayout()
        llm_chars_label = QLabel("最大文本长度:")
        self.llm_max_chars_input = QLineEdit(self.settings.value("llm_max_chars", "8000"))
        llm_chars_layout.addWidget(llm_chars_label)
        llm_chars_layout.addWidget(self.llm_max_chars_input)
        layout.addLayout(llm_chars_layout)

        # 特殊本地规则 (Special Local Rules)
        self.local_rules_checkbox = QCheckBox("启用特殊的本地规则 (Enable Special Local Rules)")
        self.local_rules_checkbox.setToolTip("例如：参考号识别规则 (2060A1710 -> Date) 等")
        self.local_rules_checkbox.setChecked(self.settings.value("enable_local_rules", True, type=bool))
        layout.addWidget(self.local_rules_checkbox)

        # --- 自动核对设置 ---
        layout.addWidget(QLabel("<b>自动核对设置 (Auto Reconciliation)</b>"))
        recon_grid = QGridLayout()
        
        self.recon_fuzzy_cb = QCheckBox("启用模糊匹配 (Fuzzy Match)")
        self.recon_fuzzy_cb.setChecked(self.settings.value("recon_fuzzy_match", True, type=bool))
        
        self.recon_merged_cb = QCheckBox("启用合并匹配 (Merged Match)")
        self.recon_merged_cb.setChecked(self.settings.value("recon_merged_match", True, type=bool))
        
        recon_grid.addWidget(self.recon_fuzzy_cb, 0, 0)
        recon_grid.addWidget(self.recon_merged_cb, 0, 1)
        
        recon_grid.addWidget(QLabel("相似度阈值 (0.1-1.0):"), 1, 0)
        self.recon_sim_input = QDoubleSpinBox()
        self.recon_sim_input.setRange(0.1, 1.0)
        self.recon_sim_input.setSingleStep(0.05)
        self.recon_sim_input.setValue(float(self.settings.value("recon_similarity_threshold", 0.6)))
        recon_grid.addWidget(self.recon_sim_input, 1, 1)
        
        recon_grid.addWidget(QLabel("金额容差 (Tolerance):"), 2, 0)
        self.recon_tol_input = QDoubleSpinBox()
        self.recon_tol_input.setRange(0.0, 10.0)
        self.recon_tol_input.setSingleStep(0.01)
        self.recon_tol_input.setValue(float(self.settings.value("recon_amount_tolerance", 0.05)))
        recon_grid.addWidget(self.recon_tol_input, 2, 1)
        
        layout.addLayout(recon_grid)

        # Network Settings
        network_layout = QVBoxLayout()
        network_layout.addWidget(QLabel("<b>Network Settings (Multi-user)</b>"))
        
        self.network_enabled_cb = QCheckBox("Enable Network Mode")
        self.network_enabled_cb.setChecked(self.settings.value("network_enabled", False, type=bool))
        network_layout.addWidget(self.network_enabled_cb)
        
        server_url_layout = QHBoxLayout()
        server_url_layout.addWidget(QLabel("Server URL:"))
        self.server_url_input = QLineEdit(self.settings.value("server_url", "http://localhost:8000"))
        server_url_layout.addWidget(self.server_url_input)
        network_layout.addLayout(server_url_layout)
        
        layout.addLayout(network_layout)

        # Color Settings
        layout.addWidget(QLabel("<b>Status Colors</b>"))
        color_layout = QGridLayout()
        self.colors = {}
        color_keys = [
            ("color_paid", "Paid", "#285028"),
            ("color_partial", "Partial", "#646428"),
            ("color_pending", "Pending", "#642828"),
            ("color_overdue", "Overdue", "#ffaa00")
        ]
        
        for i, (key, label, default) in enumerate(color_keys):
            val = self.settings.value(key, default)
            self.colors[key] = val
            
            btn = QPushButton(label)
            btn.setStyleSheet(f"background-color: {val}; color: white;")
            btn.clicked.connect(lambda checked=False, k=key, b=btn: self.pick_color(k, b))
            color_layout.addWidget(btn, 0, i)
            
        layout.addLayout(color_layout)

        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def pick_color(self, key, button):
        from PySide6.QtWidgets import QColorDialog
        initial_color = QColor(self.colors[key])
        color = QColorDialog.getColor(initial_color, self, "选择标记颜色")
        if color.isValid():
            self.colors[key] = color.name()
            button.setStyleSheet(f"background-color: {color.name()}; color: white; border: 1px solid gray;")

    def browse_poppler(self):
        dir_path = QFileDialog.getExistingDirectory(self, "选择 Poppler/bin 目录")
        if dir_path:
            self.poppler_path_input.setText(dir_path)

    def browse_tesseract(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择 Tesseract-OCR 可执行文件", filter="tesseract.exe (*.exe)")
        if file_path:
            self.tesseract_path_input.setText(file_path)

    def accept(self):
        self.settings.setValue("enable_ocr", self.ocr_checkbox.isChecked())
        self.settings.setValue("ocr_lang", self.ocr_lang_input.text())
        self.settings.setValue("poppler_path", self.poppler_path_input.text())
        self.settings.setValue("tesseract_cmd", self.tesseract_path_input.text())
        self.settings.setValue("prefer_pymupdf", self.pymupdf_checkbox.isChecked())
        self.settings.setValue("enable_llm", self.llm_checkbox.isChecked())
        self.settings.setValue("llm_mode", self.llm_mode_input.currentText())
        self.settings.setValue("llm_endpoint", self.llm_endpoint_input.text())
        self.settings.setValue("llm_model", self.llm_model_input.text())
        self.settings.setValue("llm_api_key", self.llm_api_key_input.text())
        self.settings.setValue("llm_use_multimodal", self.llm_multimodal_checkbox.isChecked())
        self.settings.setValue("llm_timeout_sec", self.llm_timeout_input.text())
        self.settings.setValue("llm_max_chars", self.llm_max_chars_input.text())
        self.settings.setValue("enable_local_rules", self.local_rules_checkbox.isChecked())
        
        # Save Recon Settings
        self.settings.setValue("recon_fuzzy_match", self.recon_fuzzy_cb.isChecked())
        self.settings.setValue("recon_merged_match", self.recon_merged_cb.isChecked())
        self.settings.setValue("recon_similarity_threshold", self.recon_sim_input.value())
        self.settings.setValue("recon_amount_tolerance", self.recon_tol_input.value())

        # Save Network Settings
        self.settings.setValue("network_enabled", self.network_enabled_cb.isChecked())
        self.settings.setValue("server_url", self.server_url_input.text())
        
        # 保存颜色设置
        for key, val in self.colors.items():
            self.settings.setValue(key, val)
        
        # 应用 Tesseract 路径
        tesseract_cmd = self.tesseract_path_input.text()
        if tesseract_cmd and os.path.exists(tesseract_cmd):
            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

        super().accept()

class _LegacyPaymentManager:
    """收款系统数据库管 (Legacy - Use brazil_tool.db.payment_manager)"""
    def __init__(self, db_path="invoice_payment.db"):
        self.db_path = db_path
        self.init_db()

    def init_db(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        # 发票
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS payment_invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_number TEXT UNIQUE,
                issuer_name TEXT,
                issue_date TEXT,
                total_amount REAL,
                terms_count INTEGER DEFAULT 1,
                status TEXT DEFAULT 'Unpaid',
                file_name TEXT,
                created_at TEXT,
                natureza_operacao TEXT
            )
        ''')
        
        # 字段迁移 (payment_invoices)
        cursor.execute("PRAGMA table_info(payment_invoices)")
        columns = [info[1] for info in cursor.fetchall()]
        if "natureza_operacao" not in columns:
            cursor.execute("ALTER TABLE payment_invoices ADD COLUMN natureza_operacao TEXT")
        if "destinatario_name" not in columns:
            cursor.execute("ALTER TABLE payment_invoices ADD COLUMN destinatario_name TEXT")
        if "destinatario_cnpj" not in columns:
            cursor.execute("ALTER TABLE payment_invoices ADD COLUMN destinatario_cnpj TEXT")
        if "issuer_cnpj" not in columns:
            cursor.execute("ALTER TABLE payment_invoices ADD COLUMN issuer_cnpj TEXT")
        if "created_at" not in columns:
            cursor.execute("ALTER TABLE payment_invoices ADD COLUMN created_at TEXT")

        # 分期
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS payment_installments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_id INTEGER,
                term_number INTEGER,
                due_date TEXT,
                amount REAL,
                paid_amount REAL DEFAULT 0,
                penalty REAL DEFAULT 0,
                status TEXT DEFAULT 'Pending',
                paid_date TEXT,
                note TEXT,
                account_id INTEGER,
                transaction_id INTEGER,
                created_at TEXT,
                FOREIGN KEY(invoice_id) REFERENCES payment_invoices(id)
            )
        ''')
        
        # 字段迁移 (payment_installments)
        cursor.execute("PRAGMA table_info(payment_installments)")
        columns = [info[1] for info in cursor.fetchall()]
        if "paid_amount" not in columns:
            cursor.execute("ALTER TABLE payment_installments ADD COLUMN paid_amount REAL DEFAULT 0")
        if "note" not in columns:
            cursor.execute("ALTER TABLE payment_installments ADD COLUMN note TEXT")
        if "penalty" not in columns:
            cursor.execute("ALTER TABLE payment_installments ADD COLUMN penalty REAL DEFAULT 0")
        if "account_id" not in columns:
            cursor.execute("ALTER TABLE payment_installments ADD COLUMN account_id INTEGER")
        if "transaction_id" not in columns:
            cursor.execute("ALTER TABLE payment_installments ADD COLUMN transaction_id INTEGER")
        if "created_at" not in columns:
            cursor.execute("ALTER TABLE payment_installments ADD COLUMN created_at TEXT")

        # 账户
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE,
                bank_info TEXT,
                currency TEXT DEFAULT 'BRL',
                initial_balance REAL DEFAULT 0,
                current_balance REAL DEFAULT 0,
                note TEXT,
                is_active INTEGER DEFAULT 1
            )
        ''')

        # 账户流水
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS account_transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_id INTEGER,
                date TEXT,
                type TEXT, -- 'Income', 'Expense'
                amount REAL,
                description TEXT,
                related_invoice_id INTEGER,
                related_installment_id INTEGER,
                created_at TEXT,
                FOREIGN KEY(account_id) REFERENCES accounts(id)
            )
        ''')

        # 预收账款 (New)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS customer_advances (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_name TEXT,
                customer_cnpj TEXT,
                amount REAL,
                remaining_amount REAL,
                date TEXT,
                description TEXT,
                account_id INTEGER,
                transaction_id INTEGER,
                created_at TEXT,
                status TEXT DEFAULT 'Open',
                FOREIGN KEY(account_id) REFERENCES accounts(id)
            )
        ''')

        conn.commit()
        conn.close()

    def add_advance(self, customer_name, customer_cnpj, amount, date, description, account_id, transaction_id):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO customer_advances (customer_name, customer_cnpj, amount, remaining_amount, date, description, account_id, transaction_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (customer_name, customer_cnpj, amount, amount, date, description, account_id, transaction_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
            return cursor.lastrowid
        finally:
            conn.close()

    def get_advances_by_customer(self, customer_cnpj=None, customer_name=None):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        if customer_cnpj:
            clean_cnpj = re.sub(r'\D', '', customer_cnpj)
            cursor.execute("SELECT * FROM customer_advances WHERE REPLACE(REPLACE(REPLACE(customer_cnpj, '.', ''), '/', ''), '-', '') = ? AND remaining_amount > 0 AND status = 'Open'", (clean_cnpj,))
        elif customer_name:
            cursor.execute("SELECT * FROM customer_advances WHERE customer_name = ? AND remaining_amount > 0 AND status = 'Open'", (customer_name,))
        else:
            cursor.execute("SELECT * FROM customer_advances WHERE remaining_amount > 0 AND status = 'Open'")
        rows = cursor.fetchall()
        conn.close()
        return rows

    def apply_advance_to_installment(self, advance_id, installment_id, amount_to_apply):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            # 1. Update advance remaining amount
            cursor.execute("UPDATE customer_advances SET remaining_amount = remaining_amount - ? WHERE id = ?", (amount_to_apply, advance_id))
            cursor.execute("UPDATE customer_advances SET status = 'Used' WHERE id = ? AND remaining_amount < 0.01", (advance_id,))
            
            # 2. Update installment paid amount
            cursor.execute("SELECT paid_amount, amount, invoice_id FROM payment_installments WHERE id = ?", (installment_id,))
            inst = cursor.fetchone()
            if inst:
                curr_paid = inst[0] or 0.0
                total_amt = inst[1]
                inv_id = inst[2]
                
                new_paid = curr_paid + amount_to_apply
                new_status = 'Paid' if new_paid >= total_amt - 0.01 else 'Partial'
                
                cursor.execute('''
                    UPDATE payment_installments 
                    SET paid_amount = ?, status = ?, paid_date = ?, note = ?
                    WHERE id = ?
                ''', (new_paid, new_status, datetime.now().strftime("%Y-%m-%d"), f"Written off from advance ID {advance_id}", installment_id))
                
                self._refresh_invoice_status(cursor, inv_id)
            
            conn.commit()
            return True
        except Exception as e:
            print(f"Error applying advance: {e}")
            conn.rollback()
            return False
        finally:
            conn.close()

    # --- Account Management ---
    def add_account(self, name, bank_info="", currency="BRL", initial_balance=0.0, note=""):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO accounts (name, bank_info, currency, initial_balance, current_balance, note)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (name, bank_info, currency, initial_balance, initial_balance, note))
            conn.commit()
            return cursor.lastrowid
        except sqlite3.IntegrityError:
            return None # Duplicate name
        finally:
            conn.close()

    def update_account(self, account_id, name, bank_info, currency, note, is_active=1):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('''
                UPDATE accounts SET name=?, bank_info=?, currency=?, note=?, is_active=?
                WHERE id=?
            ''', (name, bank_info, currency, note, is_active, account_id))
            conn.commit()
        finally:
            conn.close()

    def delete_account(self, account_id):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            # Check for transactions
            cursor.execute("SELECT count(*) FROM account_transactions WHERE account_id=?", (account_id,))
            if cursor.fetchone()[0] > 0:
                # Soft delete
                cursor.execute("UPDATE accounts SET is_active=0 WHERE id=?", (account_id,))
            else:
                cursor.execute("DELETE FROM accounts WHERE id=?", (account_id,))
            conn.commit()
        finally:
            conn.close()

    def get_accounts(self, active_only=True):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        if active_only:
            cursor.execute("SELECT * FROM accounts WHERE is_active=1 ORDER BY name")
        else:
            cursor.execute("SELECT * FROM accounts ORDER BY name")
        rows = cursor.fetchall()
        conn.close()
        return rows

    def get_account_balance(self, account_id):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT current_balance FROM accounts WHERE id=?", (account_id,))
        row = cursor.fetchone()
        conn.close()
        return row[0] if row else 0.0

    # --- Transaction Management ---
    def add_transaction(self, account_id, date, trans_type, amount, description, related_invoice_id=None, related_installment_id=None):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            # Add transaction
            cursor.execute('''
                INSERT INTO account_transactions (account_id, date, type, amount, description, related_invoice_id, related_installment_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (account_id, date, trans_type, amount, description, related_invoice_id, related_installment_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            trans_id = cursor.lastrowid

            # Update account balance
            if trans_type == 'Income':
                cursor.execute("UPDATE accounts SET current_balance = current_balance + ? WHERE id=?", (amount, account_id))
            else:
                cursor.execute("UPDATE accounts SET current_balance = current_balance - ? WHERE id=?", (amount, account_id))
            
            conn.commit()
            return trans_id
        finally:
            conn.close()

    def delete_transaction(self, trans_id):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT * FROM account_transactions WHERE id=?", (trans_id,))
            trans = cursor.fetchone()
            if not trans: return

            # Revert balance
            if trans['type'] == 'Income':
                cursor.execute("UPDATE accounts SET current_balance = current_balance - ? WHERE id=?", (trans['amount'], trans['account_id']))
            else:
                cursor.execute("UPDATE accounts SET current_balance = current_balance + ? WHERE id=?", (trans['amount'], trans['account_id']))

            # Delete transaction
            cursor.execute("DELETE FROM account_transactions WHERE id=?", (trans_id,))
            
            # If linked to installment, revert installment status
            if trans['related_installment_id']:
                inst_id = trans['related_installment_id']
                
                # Fetch current installment data
                cursor.execute("SELECT paid_amount, invoice_id FROM payment_installments WHERE id=?", (inst_id,))
                inst_row = cursor.fetchone()
                
                if inst_row:
                    curr_paid = inst_row['paid_amount'] or 0.0
                    # Subtract the deleted transaction amount
                    new_paid = max(0.0, curr_paid - trans['amount'])
                    
                    new_status = 'Pending'
                    new_paid_date = None
                    
                    if new_paid > 0.01:
                        new_status = 'Partial'
                        # If partial, keep the date (or we could fetch the latest trans date, but keeping simple for now)
                        # We won't clear paid_date if it's still partial, unless we want to be very precise.
                        # For now, let's just clear it only if Pending.
                    
                    # Only clear transaction_id if it matched the deleted one (though typically it holds the latest)
                    cursor.execute('''
                        UPDATE payment_installments 
                        SET transaction_id = CASE WHEN transaction_id=? THEN NULL ELSE transaction_id END,
                            paid_amount = ?,
                            status = ?,
                            paid_date = CASE WHEN ? <= 0.01 THEN NULL ELSE paid_date END
                        WHERE id=?
                    ''', (trans_id, new_paid, new_status, new_paid, inst_id))

                    # Refresh Invoice Status
                    self._refresh_invoice_status(cursor, inst_row['invoice_id'])

            conn.commit()
        finally:
            conn.close()

    def get_transactions(self, account_id, limit=100):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('''
            SELECT t.*, i.invoice_number, i.destinatario_name
            FROM account_transactions t
            LEFT JOIN payment_invoices i ON t.related_invoice_id = i.id
            WHERE t.account_id=? 
            ORDER BY t.date DESC, t.created_at DESC LIMIT ?
        ''', (account_id, limit))
        rows = cursor.fetchall()
        conn.close()
        return rows

    def upsert_invoice(self, invoice_data: dict) -> int:
        """插入或更新发票，返回发票ID"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            cursor.execute("SELECT id FROM payment_invoices WHERE invoice_number = ?", (invoice_data['invoice_number'],))
            row = cursor.fetchone()

            if row:
                inv_id = row[0]
                cursor.execute('''
                    UPDATE payment_invoices
                    SET issuer_name=?, issue_date=?, total_amount=?, file_name=?, natureza_operacao=?, destinatario_name=?, destinatario_cnpj=?, issuer_cnpj=?
                    WHERE id=?
                ''', (
                    invoice_data['issuer_name'],
                    invoice_data['issue_date'],
                    invoice_data['total_amount'],
                    invoice_data['file_name'],
                    invoice_data.get('natureza_operacao', ''),
                    invoice_data.get('destinatario_name', ''),
                    invoice_data.get('destinatario_cnpj', ''),
                    invoice_data.get('issuer_cnpj', ''),
                    inv_id
                ))
            else:
                cursor.execute('''
                    INSERT INTO payment_invoices (invoice_number, issuer_name, issue_date, total_amount, file_name, created_at, status, natureza_operacao, destinatario_name, destinatario_cnpj, issuer_cnpj)
                    VALUES (?, ?, ?, ?, ?, ?, 'Unpaid', ?, ?, ?, ?)
                ''', (
                    invoice_data['invoice_number'],
                    invoice_data['issuer_name'],
                    invoice_data['issue_date'],
                    invoice_data['total_amount'],
                    invoice_data['file_name'],
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    invoice_data.get('natureza_operacao', ''),
                    invoice_data.get('destinatario_name', ''),
                    invoice_data.get('destinatario_cnpj', ''),
                    invoice_data.get('issuer_cnpj', '')
                ))
                inv_id = cursor.lastrowid

            conn.commit()
            return inv_id
        finally:
            conn.close()

    def batch_upsert_invoices(self, invoices_data: list) -> dict:
        """批量插入或更新发票（使用单个事务，大幅提升性能）

        Returns:
            dict: {invoice_number: invoice_id}
        """
        if not invoices_data:
            return {}

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        result = {}
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            # 1. 批量获取已存在的发票
            numbers = [d['invoice_number'] for d in invoices_data]
            placeholders = ','.join('?' * len(numbers))
            cursor.execute(f"SELECT id, invoice_number FROM payment_invoices WHERE invoice_number IN ({placeholders})", numbers)
            existing = {row[1]: row[0] for row in cursor.fetchall()}

            # 2. 分类处理
            to_insert = []
            to_update = []
            for data in invoices_data:
                inv_num = data['invoice_number']
                if inv_num in existing:
                    to_update.append((
                        data['issuer_name'],
                        data['issue_date'],
                        data['total_amount'],
                        data['file_name'],
                        data.get('natureza_operacao', ''),
                        data.get('destinatario_name', ''),
                        data.get('destinatario_cnpj', ''),
                        data.get('issuer_cnpj', ''),
                        existing[inv_num]
                    ))
                    result[inv_num] = existing[inv_num]
                else:
                    to_insert.append((
                        inv_num,
                        data['issuer_name'],
                        data['issue_date'],
                        data['total_amount'],
                        data['file_name'],
                        now,
                        data.get('natureza_operacao', ''),
                        data.get('destinatario_name', ''),
                        data.get('destinatario_cnpj', ''),
                        data.get('issuer_cnpj', '')
                    ))

            # 3. 批量更新
            if to_update:
                cursor.executemany('''
                    UPDATE payment_invoices
                    SET issuer_name=?, issue_date=?, total_amount=?, file_name=?,
                        natureza_operacao=?, destinatario_name=?, destinatario_cnpj=?, issuer_cnpj=?
                    WHERE id=?
                ''', to_update)

            # 4. 批量插入
            if to_insert:
                cursor.executemany('''
                    INSERT INTO payment_invoices
                    (invoice_number, issuer_name, issue_date, total_amount, file_name, created_at, status,
                     natureza_operacao, destinatario_name, destinatario_cnpj, issuer_cnpj)
                    VALUES (?, ?, ?, ?, ?, ?, 'Unpaid', ?, ?, ?, ?)
                ''', to_insert)

                # 获取新插入的 ID
                cursor.execute(f"SELECT id, invoice_number FROM payment_invoices WHERE invoice_number IN ({placeholders})", numbers)
                for row in cursor.fetchall():
                    result[row[1]] = row[0]

            conn.commit()
            return result
        finally:
            conn.close()

    def generate_payment_plan(self, invoice_id: int, terms: int, start_date: datetime = None, interval_days: int = 30):
        """生成/重置分期计划 (非破坏 """
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            # 获取发票信息
            cursor.execute("SELECT total_amount, issue_date FROM payment_invoices WHERE id=?", (invoice_id,))
            row = cursor.fetchone()
            if not row: return
            
            total_amount, issue_date_str = row
            
            # 解析日期作为基准 (如果没有提供 start_date)
            base_date = datetime.now()
            if not start_date:
                base_date = safe_parse_date(issue_date_str) or datetime.now()
            else:
                base_date = start_date

            # 获取所有已有支付额
            cursor.execute("SELECT SUM(paid_amount) FROM payment_installments WHERE invoice_id=?", (invoice_id,))
            total_paid_so_far = cursor.fetchone()[0] or 0.0
            
            # 还需要付多少本金
            remaining_principal = round(total_amount - total_paid_so_far, 2)
            
            if remaining_principal <= 0.01:
                cursor.execute("UPDATE payment_invoices SET terms_count=? WHERE id=?", (terms, invoice_id))
                conn.commit()
                return

            # 删除未支付分
            cursor.execute("""
                DELETE FROM payment_installments 
                WHERE invoice_id=? AND status = 'Pending' AND paid_amount = 0 AND (note IS NULL OR note = '')
            """, (invoice_id,))
            
            # 计算保留下来的期
            cursor.execute("SELECT count(*) FROM payment_installments WHERE invoice_id=?", (invoice_id,))
            kept_count = cursor.fetchone()[0]
            
            # 还需要补充生成的期数
            new_terms_to_gen = terms - kept_count
            if new_terms_to_gen <= 0:
                new_terms_to_gen = 1 

            term_amount = round(remaining_principal / new_terms_to_gen, 2)
            
            for i in range(1, new_terms_to_gen + 1):
                actual_term_num = kept_count + i
                # 计算日期: 如果指定 start_date，第一期就 start_date ( start_date + interval")
                # 通常 start_date  第一期到期日"
                if start_date and i == 1 and kept_count == 0:
                    due_date = start_date
                elif start_date:
                    # 如果 start_date，后续期数基 start_date
                    # 逻辑: Term 1 = start_date
                    # Term 2 = start_date + interval
                    due_date = start_date + timedelta(days=interval_days * (i - 1)) if kept_count == 0 else start_date + timedelta(days=interval_days * i) # 简化逻辑
                    
                    # 更稳妥的逻辑                     # 如果是全新的计划 (kept_count=0):
                    #   i=1 -> start_date
                    #   i=2 -> start_date + interval
                    # 如果已有保留期数 (kept=1):
                    #   i=1 (实际上是   -> start_date (用户指定 接下来第一 的日志
                    #   i=2 -> start_date + interval
                    if i == 1:
                        due_date = start_date
                    else:
                        due_date = start_date + timedelta(days=interval_days * (i - 1))
                else:
                    # 默认首期为发票日期 (actual_term_num=1 时偏移为0)
                    due_date = base_date + timedelta(days=interval_days * (actual_term_num - 1))

                current_amount = term_amount if i < new_terms_to_gen else round(remaining_principal - term_amount * (i-1), 2)
                
                cursor.execute('''
                    INSERT INTO payment_installments (invoice_id, term_number, due_date, amount, paid_amount, penalty, status)
                    VALUES (?, ?, ?, ?, 0, 0, 'Pending')
                ''', (invoice_id, actual_term_num, due_date.strftime("%Y-%m-%d"), current_amount))
            
            cursor.execute("UPDATE payment_invoices SET terms_count=? WHERE id=?", (terms, invoice_id))
            conn.commit()
        finally:
            conn.close()

    def get_all_installments_for_export(self, invoice_number: str) -> List[dict]:
        """为导 JSON 获取所有分期数"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT pi.* FROM payment_installments pi
                JOIN payment_invoices i ON pi.invoice_id = i.id
                WHERE i.invoice_number = ?
            """, (invoice_number,))
            return [dict(row) for row in cursor.fetchall()]
        finally:
            conn.close()

    def restore_installments_from_import(self, invoice_number: str, installments: List[dict]):
        """从导入的 JSON 恢复分期数据"""
        if not installments: return
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            # 找到发票ID
            cursor.execute("SELECT id FROM payment_invoices WHERE invoice_number = ?", (invoice_number,))
            row = cursor.fetchone()
            if not row: return
            inv_id = row[0]
            
            # 先清除旧的分
            cursor.execute("DELETE FROM payment_installments WHERE invoice_id = ?", (inv_id,))
            
            # 插入新分
            for inst in installments:
                cursor.execute("""
                    INSERT INTO payment_installments 
                    (invoice_id, term_number, due_date, amount, paid_amount, penalty, status, paid_date, note)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    inv_id, inst.get('term_number'), inst.get('due_date'),
                    inst.get('amount'), inst.get('paid_amount'), inst.get('penalty'),
                    inst.get('status'), inst.get('paid_date'), inst.get('note')
                ))
            conn.commit()
        finally:
            conn.close()

    def get_invoices(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('''
            SELECT 
                i.*,
                (SELECT count(*) FROM payment_installments pi WHERE pi.invoice_id = i.id) as total_terms_count,
                (SELECT count(*) FROM payment_installments pi WHERE pi.invoice_id = i.id AND pi.status = 'Paid') as paid_terms_count,
                (SELECT SUM(paid_amount) FROM payment_installments pi WHERE pi.invoice_id = i.id) as total_paid_amount
            FROM payment_invoices i 
            ORDER BY created_at DESC
        ''')
        rows = cursor.fetchall()
        conn.close()
        return rows

    def get_installments(self, invoice_id):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM payment_installments WHERE invoice_id=? ORDER BY term_number", (invoice_id,))
        rows = cursor.fetchall()
        conn.close()
        return rows

    def update_installment_field(self, installment_id: int, field: str, value: any):
        """手动更新分期字段"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute(f"UPDATE payment_installments SET {field}=? WHERE id=?", (value, installment_id))
            conn.commit()
            # 重新计算发票状
            cursor.execute("SELECT invoice_id FROM payment_installments WHERE id=?", (installment_id,))
            inv_id = cursor.fetchone()[0]
            self._refresh_invoice_status(cursor, inv_id)
            conn.commit()
        finally:
            conn.close()

    def update_installment_status(self, installment_id, status):
        """仅更新状态字符串，核心收款逻辑已迁移至 register_payment_batch"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("UPDATE payment_installments SET status=? WHERE id=?", (status, installment_id))
            
            cursor.execute("SELECT invoice_id FROM payment_installments WHERE id=?", (installment_id,))
            inv_id = cursor.fetchone()[0]
            self._refresh_invoice_status(cursor, inv_id)
            conn.commit()
        finally:
            conn.close()

    def _refresh_invoice_status(self, cursor, invoice_id):
        """重新计算并更新发票的总体状"""
        cursor.execute("SELECT count(*) FROM payment_installments WHERE invoice_id=? AND status='Pending'", (invoice_id,))
        pending_count = cursor.fetchone()[0]
        cursor.execute("SELECT count(*) FROM payment_installments WHERE invoice_id=? AND status='Paid'", (invoice_id,))
        paid_count = cursor.fetchone()[0]
        cursor.execute("SELECT count(*) FROM payment_installments WHERE invoice_id=? AND status='Partial'", (invoice_id,))
        partial_count = cursor.fetchone()[0]
        
        if pending_count == 0 and partial_count == 0 and paid_count > 0:
            inv_status = 'Paid'
        elif paid_count > 0 or partial_count > 0:
            inv_status = 'Partial'
        else:
            inv_status = 'Unpaid'
        
        cursor.execute("UPDATE payment_invoices SET status=? WHERE id=?", (inv_status, invoice_id))

    def find_invoice_id_by_number(self, number: str):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT id FROM payment_invoices WHERE invoice_number = ?", (number,))
            row = cursor.fetchone()
            return row[0] if row else None
        finally:
            conn.close()

    def get_all_existing_invoice_numbers(self) -> set:
        """批量获取所有已存在的发票号（用于快速判断是否需要同步）"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT invoice_number FROM payment_invoices")
            return {row[0] for row in cursor.fetchall()}
        finally:
            conn.close()

    def get_account_id_for_invoice(self, invoice_id: int):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT account_id FROM account_transactions WHERE related_invoice_id=?", (invoice_id,))
            row = cursor.fetchone()
            return row[0] if row else None
        finally:
            conn.close()

    def get_all_installments_extended(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT 
                    i.issuer_name, i.destinatario_name, i.destinatario_cnpj, i.issuer_cnpj,
                    p.due_date, p.amount, p.paid_amount, p.status, i.invoice_number
                FROM payment_installments p
                JOIN payment_invoices i ON p.invoice_id = i.id
            """)
            return cursor.fetchall()
        finally:
            conn.close()

    def search_pending_installments(self, patterns: List[str]):
        if not patterns: return []
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        try:
            query = """
                SELECT i.id as invoice_id, p.id as installment_id, p.amount, 
                       i.destinatario_name, i.destinatario_cnpj, p.due_date, i.invoice_number 
                FROM payment_installments p
                JOIN payment_invoices i ON p.invoice_id = i.id
                WHERE p.status != 'Paid' AND (
            """
            conditions = []
            params = []
            for pat in patterns:
                conditions.append("i.invoice_number LIKE ?")
                params.append(pat)
            
            query += " OR ".join(conditions) + ")"
            
            cursor.execute(query, params)
            return cursor.fetchall()
        finally:
            conn.close()

class PaymentPlanDialog(QDialog):
    """分期计划设置对话"""
    def __init__(self, parent=None, initial_date=None):
        super().__init__(parent)
        self.setWindowTitle("设置分期计划")
        self.resize(400, 250)
        
        layout = QVBoxLayout(self)
        
        from PySide6.QtWidgets import QFormLayout, QSpinBox, QDateEdit, QDialogButtonBox, QCheckBox
        
        form_layout = QFormLayout()
        
        self.terms_spin = QSpinBox()
        self.terms_spin.setRange(1, 60)
        self.terms_spin.setValue(1)
        form_layout.addRow("分期 ", self.terms_spin)
        
        # Checkbox to use invoice date
        self.use_invoice_date_cb = QCheckBox("使用每张发票的开票日期作为首期基准")
        self.use_invoice_date_cb.setChecked(False)
        self.use_invoice_date_cb.toggled.connect(self.on_use_invoice_date_toggled)
        layout.addWidget(self.use_invoice_date_cb)
        
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        
        # 使用传入的初始日期，如果没有则使用当前日期
        if initial_date:
            if isinstance(initial_date, str):
                d = safe_parse_date_to_date(initial_date)
                if d:
                    self.start_date_edit.setDate(d)
                else:
                    self.start_date_edit.setDate(datetime.now().date())
            elif hasattr(initial_date, 'date'):
                self.start_date_edit.setDate(initial_date.date())
            else:
                self.start_date_edit.setDate(initial_date)
        else:
            self.start_date_edit.setDate(datetime.now().date())
            
        self.start_date_edit.setDisplayFormat("yyyy/MM/dd")
        form_layout.addRow("首期到期 ", self.start_date_edit)
        
        self.interval_spin = QSpinBox()
        self.interval_spin.setRange(1, 365)
        self.interval_spin.setValue(30)
        self.interval_spin.setSuffix(" ")
        form_layout.addRow("间隔时间:", self.interval_spin)
        
        layout.addLayout(form_layout)
        
        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
    def on_use_invoice_date_toggled(self, checked):
        self.start_date_edit.setEnabled(not checked)
        
    def get_data(self):
        use_inv_date = self.use_invoice_date_cb.isChecked()
        start_date = None
        if not use_inv_date:
            start_date = datetime(
                self.start_date_edit.date().year(),
                self.start_date_edit.date().month(),
                self.start_date_edit.date().day()
            )
            
        return {
            "terms": self.terms_spin.value(),
            "start_date": start_date,
            "use_invoice_date": use_inv_date,
            "interval": self.interval_spin.value()
        }
class AdvanceWriteoffDialog(QDialog):
    """预收账款核销对话框"""
    def __init__(self, db, customer_cnpj, customer_name, installment_id, amount_due, parent=None):
        super().__init__(parent)
        self.db = db
        self.customer_cnpj = customer_cnpj
        self.customer_name = customer_name
        self.installment_id = installment_id
        self.amount_due = amount_due
        self.selected_advance_id = None
        
        self.setWindowTitle(f"预收核销 - {customer_name}")
        self.resize(600, 400)
        self.setup_ui()
        self.load_advances()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        info_label = QLabel(f"正在为客户 <b>{self.customer_name}</b> 核销分期款项\n待收金额: <b>R$ {self.amount_due:,.2f}</b>")
        info_label.setStyleSheet("font-size: 11pt;")
        layout.addWidget(info_label)
        
        layout.addWidget(QLabel("选择可用的预收款记录:"))
        
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(["ID", "日期", "描述", "原始金额", "可用余额"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        layout.addWidget(self.table)
        
        btn_layout = QHBoxLayout()
        self.btn_apply = QPushButton("✅ 确认核销")
        self.btn_apply.setEnabled(False)
        self.btn_apply.clicked.connect(self.accept)
        
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(self.btn_apply)
        layout.addLayout(btn_layout)
        
        self.table.itemSelectionChanged.connect(self.on_selection_changed)

    def load_advances(self):
        advances = self.db.get_advances_by_customer(self.customer_cnpj, self.customer_name)
        self.table.setRowCount(len(advances))
        for i, adv in enumerate(advances):
            self.table.setItem(i, 0, QTableWidgetItem(str(adv['id'])))
            self.table.setItem(i, 1, QTableWidgetItem(format_date_gui(str(adv['date']))))
            self.table.setItem(i, 2, QTableWidgetItem(str(adv['description'])))
            self.table.setItem(i, 3, QTableWidgetItem(f"{adv['amount']:,.2f}"))
            self.table.setItem(i, 4, QTableWidgetItem(f"{adv['remaining_amount']:,.2f}"))
            self.table.item(i, 0).setData(Qt.UserRole, adv)
            
        if not advances:
            self.table.setRowCount(1)
            self.table.setSpan(0, 0, 1, 5)
            self.table.setItem(0, 0, QTableWidgetItem("该客户暂无可用预收款记录"))
            self.table.item(0, 0).setTextAlignment(Qt.AlignCenter)

    def on_selection_changed(self):
        selected_rows = self.table.selectedItems()
        self.btn_apply.setEnabled(len(selected_rows) > 0)

    def get_selected_advance(self):
        row = self.table.currentRow()
        if row >= 0:
            return self.table.item(row, 0).data(Qt.UserRole)
        return None

class PaymentSystemTab(QWidget):
    def __init__(self, main_window):
        super().__init__(main_window)
        self.main_window = main_window
        self._invoice_load_seq = 0
        self._installment_load_seq = 0
        self._invoice_view_state = {}
        self._installment_view_state = {}
        self._invoice_worker = None
        self._installment_worker = None
        self._summary_refresh_timer = QTimer(self)
        self._summary_refresh_timer.setSingleShot(True)
        self._summary_refresh_timer.timeout.connect(self._refresh_main_summary_table)
        
        # Load settings
        settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)
        network_enabled = settings.value("network_enabled", False, type=bool)
        server_url = settings.value("server_url", "http://localhost:8000")
        
        if network_enabled:
            try:
                from brazil_tool.db.network_payment_manager import NetworkPaymentManager
                self.db = NetworkPaymentManager(server_url)
                logging.info("Using Network DB: %s", server_url)
            except ImportError as e:
                logging.warning("Failed to import NetworkPaymentManager: %s", e)
                self.db = PaymentManager()
        else:
            self.db = PaymentManager()
            
        self.setup_ui()
        self.load_invoices()

    def _open_local_db(self):
        db_path = getattr(self.db, "db_path", None)
        if not db_path:
            QMessageBox.warning(self, "模式限制", "当前操作仅支持本地数据库模式。请关闭 Network Mode 后重试。")
            return None
        return sqlite3.connect(db_path)

    @staticmethod
    def _to_dict(row):
        if isinstance(row, dict):
            return row
        try:
            return dict(row)
        except Exception:
            return {}

    @staticmethod
    def _to_float(val, default=0.0):
        try:
            if val is None:
                return default
            return float(val)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _to_iso_date(date_str):
        d_obj = safe_parse_date(date_str)
        if d_obj:
            return d_obj.strftime("%Y-%m-%d")
        return datetime.now().strftime("%Y-%m-%d")

    @staticmethod
    def _build_invoice_search_patterns(clean_num):
        if not clean_num:
            return []

        search_patterns = [f"%{clean_num}%"]
        if len(clean_num) == 4 and clean_num.isdigit():
            serie = clean_num[0]
            number = clean_num[1:]
            search_patterns.append(f"%.{int(serie):03d}.{int(number):03d}")
        elif len(clean_num) > 4 and clean_num.isdigit():
            serie = clean_num[:-3]
            number = clean_num[-3:]
            search_patterns.append(f"%.{int(serie):03d}.{int(number):03d}")
        return search_patterns

    def _get_invoice_map(self):
        invoice_map = {}
        for row in self.db.get_invoices():
            inv = self._to_dict(row)
            inv_id = inv.get("id")
            if inv_id is None:
                continue
            try:
                invoice_map[int(inv_id)] = inv
            except (TypeError, ValueError):
                continue
        return invoice_map

    def _get_installment_map_for_invoice(self, invoice_id):
        inst_map = {}
        for row in self.db.get_installments(invoice_id):
            inst = self._to_dict(row)
            inst_id = inst.get("id")
            if inst_id is None:
                continue
            try:
                inst_map[int(inst_id)] = inst
            except (TypeError, ValueError):
                continue
        return inst_map

    def _collect_pending_installments(self):
        pending = []
        invoices = self._get_invoice_map()
        for inv_id, inv in invoices.items():
            for inst_row in self.db.get_installments(inv_id):
                inst = self._to_dict(inst_row)
                status = str(inst.get("status", "")).strip().title()
                if status == "Paid":
                    continue

                amount = self._to_float(inst.get("amount"))
                penalty = self._to_float(inst.get("penalty"))
                paid_amount = self._to_float(inst.get("paid_amount"))
                due_remaining = max(0.0, amount + penalty - paid_amount)
                if due_remaining <= 0.01:
                    continue

                pending.append({
                    "inv_id": inv_id,
                    "inst_id": int(inst.get("id")),
                    "invoice_number": str(inv.get("invoice_number", "")),
                    "amount": amount,
                    "penalty": penalty,
                    "paid_amount": paid_amount,
                    "due_remaining": due_remaining,
                    "due_date": inst.get("due_date"),
                    "destinatario_name": inv.get("destinatario_name", ""),
                    "destinatario_cnpj": inv.get("destinatario_cnpj", ""),
                    "term_number": inst.get("term_number"),
                })
        return pending

    def setup_ui(self):
        layout = QHBoxLayout(self)

        # --- 左侧：发票列 ---
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        left_layout.addWidget(QLabel("<b>发票列表</b> (右键可批量操 "))
        
        self.invoice_table = QTableWidget()
        self.invoice_table.setColumnCount(10)
        self.invoice_table.setHorizontalHeaderLabels(["ID", "发票", "Natureza (类型)", "开票人", "日期", "总金", "实收总额", "待收余额", "状", "进度"])
        self.invoice_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.invoice_table.horizontalHeader().setStretchLastSection(True)
        self.invoice_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.invoice_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.invoice_table.itemDoubleClicked.connect(self.on_invoice_selected)
        self.invoice_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.invoice_table.customContextMenuRequested.connect(self.show_invoice_context_menu)
        # Render progress via delegate instead of per-row widgets to reduce UI lag.
        self.invoice_table.setItemDelegateForColumn(9, ProgressBarDelegate(self.invoice_table))
        self.invoice_table.setColumnWidth(9, 110)
        
        # Use Filter Header
        inv_header = FilterHeader(self.invoice_table)
        inv_header.filterChanged.connect(lambda: self.main_window.apply_header_filters(self.invoice_table))
        # Custom sorting
        self.invoice_table.setSortingEnabled(False)
        inv_header.sectionClicked.connect(self.main_window.on_header_clicked)
        self.invoice_table.setHorizontalHeader(inv_header)
        
        left_layout.addWidget(self.invoice_table)
        
        btn_layout = QHBoxLayout()
        refresh_btn = QPushButton("刷新列表")
        refresh_btn.clicked.connect(self.load_invoices)
        
        export_inv_btn = QPushButton("📤 导出列表")
        export_inv_btn.clicked.connect(lambda: export_qtable(self.invoice_table, self, "payment_invoices"))
        
        btn_layout.addWidget(refresh_btn)
        btn_layout.addWidget(export_inv_btn)
        left_layout.addLayout(btn_layout)

        # --- 右侧：详情与操作 ---
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        
        # 1. 概览面板 (Dashboard)
        from PySide6.QtWidgets import QGroupBox, QFormLayout
        self.summary_group = QGroupBox("当前发票概览")
        summary_layout = QVBoxLayout(self.summary_group)
        
        # 统计数据
        stats_layout = QHBoxLayout()
        self.lbl_total = QLabel("应收(含滞纳): 0.00")
        self.lbl_paid = QLabel("已收: 0.00")
        self.lbl_balance = QLabel("待收余额: 0.00")
        
        # 美化字体
        font_bold = QFont()
        font_bold.setBold(True)
        font_bold.setPointSize(10)
        self.lbl_balance.setFont(font_bold)
        self.lbl_balance.setStyleSheet("color: #ff5555;") # 红色余额
        
        stats_layout.addWidget(self.lbl_total)
        stats_layout.addWidget(self.lbl_paid)
        stats_layout.addWidget(self.lbl_balance)
        summary_layout.addLayout(stats_layout)
        
        # 进度
        self.summary_progress = QProgressBar()
        self.summary_progress.setTextVisible(True)
        self.summary_progress.setFormat("支付进度: %p%")
        summary_layout.addWidget(self.summary_progress)

        # 关联流水跳转按钮
        self.btn_view_trans = QPushButton("🏦 查看关联账户流水 (View Transactions)")
        self.btn_view_trans.setToolTip("跳转到账户管理并显示该发票的所有收款记录")
        self.btn_view_trans.clicked.connect(self.goto_account_details)
        summary_layout.addWidget(self.btn_view_trans)
        
        right_layout.addWidget(self.summary_group)

        # 2. 分期详情列表
        header_layout = QHBoxLayout()
        header_layout.addWidget(QLabel("<b>分期详情</b> (双击单元格可编辑)"))
        export_inst_btn = QPushButton("📤 导出明细")
        export_inst_btn.setFixedSize(80, 24)
        export_inst_btn.clicked.connect(lambda: export_qtable(self.installment_table, self, "payment_installments"))
        header_layout.addWidget(export_inst_btn)
        header_layout.addStretch()
        right_layout.addLayout(header_layout)
        
        self.installment_table = QTableWidget()
        self.installment_table.setColumnCount(9)
        self.installment_table.setHorizontalHeaderLabels(["ID", "期数", "到期", "本金", "滞纳", "实收", "收款日期", "状", "备注"])
        self.installment_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.installment_table.horizontalHeader().setStretchLastSection(True)
        self.installment_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.installment_table.itemChanged.connect(self.on_installment_cell_changed)
        
        # Use Filter Header
        inst_header = FilterHeader(self.installment_table)
        inst_header.filterChanged.connect(lambda: self.main_window.apply_header_filters(self.installment_table))
        # Custom sorting
        self.installment_table.setSortingEnabled(False)
        inst_header.sectionClicked.connect(self.main_window.on_header_clicked)
        self.installment_table.setHorizontalHeader(inst_header)

        right_layout.addWidget(self.installment_table)
        
        # 3. 操作
        action_group = QGroupBox("操作")
        action_layout_v = QVBoxLayout(action_group)
        
        # 快捷状态操作
        status_btns_layout = QHBoxLayout()
        self.pay_btn = QPushButton("💰 登记选中收款")
        self.unpay_btn = QPushButton("🔄 重置为未付")
        self.advance_writeoff_btn = QPushButton("⚖️ 预收核销")
        self.advance_writeoff_btn.setToolTip("使用客户之前的预收款来核销当前选中的分期")
        
        self.pay_btn.clicked.connect(lambda: self.set_payment_status('Paid'))
        self.unpay_btn.clicked.connect(lambda: self.set_payment_status('Pending'))
        self.advance_writeoff_btn.clicked.connect(self.open_advance_writeoff_dialog)
        
        status_btns_layout.addWidget(self.pay_btn)
        status_btns_layout.addWidget(self.unpay_btn)
        status_btns_layout.addWidget(self.advance_writeoff_btn)
        action_layout_v.addLayout(status_btns_layout)

        # 重置计划
        plan_layout = QHBoxLayout()
        from PySide6.QtWidgets import QComboBox
        self.terms_combo = QComboBox()
        self.terms_combo.addItems(["1期", "2期", "3期", "4期", "5期", "6期", "12期"])
        self.gen_plan_btn = QPushButton("🔄 重新生成计划")
        self.gen_plan_btn.setToolTip("保留已支付记录，重新分配剩余金额")
        self.gen_plan_btn.clicked.connect(self.generate_plan)
        
        plan_layout.addWidget(QLabel("目标期数:"))
        plan_layout.addWidget(self.terms_combo)
        plan_layout.addWidget(self.gen_plan_btn)
        action_layout_v.addLayout(plan_layout)
        
        right_layout.addWidget(action_group)

        # 分割
        from PySide6.QtWidgets import QSplitter
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 3)
        layout.addWidget(splitter)
        
        # 注册表格统计
        if hasattr(self.main_window, 'register_table_for_stats'):
            self.main_window.register_table_for_stats(self.invoice_table)
            self.main_window.register_table_for_stats(self.installment_table)
        
        self.current_invoice_id = None
        self._loading_installments = False

    def show_invoice_context_menu(self, pos):
        """显示发票列表右键菜单 (支持多选批量操 """
        from PySide6.QtWidgets import QMenu
        item = self.invoice_table.itemAt(pos)
        if not item: return
        
        # 获取所有选中的发票ID
        selected_items = self.invoice_table.selectedItems()
        selected_rows = sorted(list(set(i.row() for i in selected_items)))
        
        if not selected_rows:
            selected_rows = [item.row()]
            
        inv_ids = []
        for r in selected_rows:
            try:
                inv_ids.append(int(self.invoice_table.item(r, 0).text()))
            except: pass
            
        if not inv_ids: return
        
        menu = QMenu(self)
        
        # 如果只选了一行，显示导航菜单
        if len(inv_ids) == 1:
            row = selected_rows[0]
            inv_num = self.invoice_table.item(row, 1).text()
            inv_index = -1
            if hasattr(self, 'main_window') and self.main_window.invoices:
                for i, inv in enumerate(self.main_window.invoices):
                    curr_num = inv.numero or f"FILE-{inv.file_name}"
                    if curr_num == inv_num:
                        inv_index = i
                        break
            
            if inv_index >= 0:
                nav_menu = self.main_window.create_navigation_menu(inv_index, "payment")
                menu.addActions(nav_menu.actions())
                menu.addSeparator()

        count_txt = f" ({len(inv_ids)}张发 " if len(inv_ids) > 1 else ""
        
        action_plan = menu.addAction(f"📅 设置分期计划{count_txt}")
        action_upload_pdf = menu.addAction(f"📄 补充/载入发票 PDF{count_txt}")
        menu.addSeparator()
        
        # 新增查看 PDF
        if len(inv_ids) == 1:
            row = selected_rows[0]
            # 找到主列表中的索引
            inv_index = -1
            inv_num = self.invoice_table.item(row, 1).text()
            if self.main_window.invoices:
                for idx, inv in enumerate(self.main_window.invoices):
                    if (inv.numero or f"FILE-{inv.file_name}") == inv_num:
                        inv_index = idx
                        break
            if inv_index >= 0:
                action_pdf = menu.addAction("📄 查看源文件 PDF (Open)")
                action_pdf.triggered.connect(lambda: self.main_window.view_source_pdf(inv_index))
                menu.addSeparator()

        action_pay_all = menu.addAction(f"💰 登记整单收款{count_txt}")
        action_unpay_all = menu.addAction(f"🔄 重置整单未付{count_txt}")
        menu.addSeparator()
        action_goto_account = menu.addAction("🏦 转到账户明细")
        
        action = menu.exec(self.invoice_table.viewport().mapToGlobal(pos))
        
        if action == action_plan:
            self.open_payment_plan_dialog(inv_ids)
        elif action == action_upload_pdf:
            self.upload_invoice_pdf(inv_ids)
        elif action == action_pay_all:
            self.batch_update_invoice_status(inv_ids, 'Paid')
        elif action == action_unpay_all:
            self.batch_update_invoice_status(inv_ids, 'Pending')
        elif action == action_goto_account:
            if self.main_window and hasattr(self.main_window, 'account_tab'):
                # Try to find and filter transactions for the first selected invoice
                target_inv_id = inv_ids[0]
                found = self.main_window.account_tab.filter_transactions_by_invoice(target_inv_id)
                
                if found:
                    self.main_window.tab_widget.setCurrentWidget(self.main_window.account_tab)
                else:
                    QMessageBox.information(self, "未收", "该发票尚未关联任何收款记录")

    def upload_invoice_pdf(self, invoice_ids):
        """为选中的发票补充或更换源 PDF 文件"""
        if not invoice_ids: return
        
        # 选择 PDF 文件
        file_path, _ = QFileDialog.getOpenFileName(self, "选择发票源文件 PDF", "", "PDF Files (*.pdf)")
        if not file_path: return
        
        invoice_dir = "发票"
        if not os.path.exists(invoice_dir):
            os.makedirs(invoice_dir)

        invoice_map = self._get_invoice_map()
        success_count = 0
        for inv_id in invoice_ids:
            inv = invoice_map.get(int(inv_id))
            if not inv:
                continue

            inv_num = str(inv.get("invoice_number") or inv_id)
            
            # 清理文件名中的非法字符
            safe_num = "".join([c for c in inv_num if c.isalnum() or c in (' ', '_', '-')]).strip()
            if not safe_num:
                safe_num = str(inv_id)
            new_filename = f"{safe_num}.pdf"
            dest_path = os.path.join(invoice_dir, new_filename)
            
            try:
                # 如果目标文件已存在，询问是否覆盖 (只在处理第一个发票时询问)
                if os.path.exists(dest_path) and success_count == 0:
                    reply = QMessageBox.question(self, "文件已存在", 
                                               f"文件 {new_filename} 已存在，是否覆盖？",
                                               QMessageBox.Yes | QMessageBox.No)
                    if reply == QMessageBox.No:
                        # 尝试使用带时间戳的文件名
                        new_filename = f"{safe_num}_{int(time.time())}.pdf"
                        dest_path = os.path.join(invoice_dir, new_filename)

                shutil.copy2(file_path, dest_path)

                current_desc = str(inv.get("description") or "")
                new_desc = current_desc.replace("⚠️ NEED PDF (Temp Created)", "").replace("⚠️ NEED PDF", "").strip()

                invoice_data = {
                    "invoice_number": inv_num,
                    "issuer_name": str(inv.get("issuer_name") or ""),
                    "issue_date": str(inv.get("issue_date") or ""),
                    "total_amount": self._to_float(inv.get("total_amount")),
                    "file_name": new_filename,
                    "natureza_operacao": str(inv.get("natureza_operacao") or ""),
                    "destinatario_name": str(inv.get("destinatario_name") or ""),
                    "destinatario_cnpj": str(inv.get("destinatario_cnpj") or ""),
                    "issuer_cnpj": str(inv.get("issuer_cnpj") or ""),
                    "description": new_desc,
                }
                self.db.upsert_invoice(invoice_data)
                success_count += 1
            except Exception as e:
                self.main_window.log_message(f"❌ 复制文件失败 ({inv_id}): {str(e)}")

        if success_count > 0:
            QMessageBox.information(self, "成功", f"已成功为 {success_count} 张发票补充 PDF 源文件。")
            self.load_invoices()
            # 同步更新 MainWindow 的数据列表 (如果有)
            if hasattr(self.main_window, 'load_all_data'):
                self.main_window.load_all_data()

    def open_payment_plan_dialog(self, invoice_ids):
        """打开分期计划设置对话 (支持批量)"""
        if isinstance(invoice_ids, int):
            invoice_ids = [invoice_ids]
            
        # 尝试获取首个选中发票的日期作为默认日期
        initial_date = None
        if invoice_ids:
            first_id = str(invoice_ids[0])
            for r in range(self.invoice_table.rowCount()):
                id_item = self.invoice_table.item(r, 0)
                if id_item and id_item.text() == first_id:
                    date_item = self.invoice_table.item(r, 4) # 日期列
                    if date_item:
                        initial_date = date_item.text()
                    break

        dialog = PaymentPlanDialog(self, initial_date=initial_date)
        if dialog.exec():
            data = dialog.get_data()
            
            count_str = f" {len(invoice_ids)} 张发票" if len(invoice_ids) > 1 else "当前发票"
            
            start_date_str = "每张发票的开票日期" if data.get('use_invoice_date') else data['start_date'].strftime('%Y-%m-%d')

            reply = QMessageBox.question(self, "确认", 
                f"确定要为{count_str}生成 {data['terms']} 期计划吗？\n"
                f"首期: {start_date_str}\n"
                f"间隔: {data['interval']} 天\n"
                "注意: 已支付的记录将被保留",
                QMessageBox.Yes | QMessageBox.No)
                
            if reply == QMessageBox.Yes:
                # Progress Bar
                progress = QProgressBar()
                progress.setRange(0, len(invoice_ids))
                progress.setWindowTitle("正在生成计划...")
                # Center the progress bar window roughly
                if self.main_window:
                    geo = self.main_window.geometry()
                    progress.move(geo.center().x() - 150, geo.center().y() - 50)
                progress.resize(300, 50)
                progress.show()
                failed_plan_ids = []
                for i, inv_id in enumerate(invoice_ids):
                    ok = self.db.generate_payment_plan(
                        inv_id, 
                        data['terms'], 
                        start_date=data['start_date'], # If use_invoice_date is True, this is None, which triggers the backend logic to use issue_date
                        interval_days=data['interval']
                    )
                    if ok is False:
                        failed_plan_ids.append(inv_id)
                    progress.setValue(i+1)
                    QApplication.processEvents()
                
                progress.close()
                if failed_plan_ids:
                    QMessageBox.warning(
                        self,
                        "部分失败",
                        f"有 {len(failed_plan_ids)} 张发票生成计划失败，请检查网络/数据库状态。",
                    )
                
                if self.current_invoice_id in invoice_ids:
                    self.load_installments(self.current_invoice_id)
                self.load_invoices()

    def open_advance_writeoff_dialog(self):
        """打开预收核销对话框"""
        selected_rows = self.installment_table.selectedItems()
        if not selected_rows:
            QMessageBox.warning(self, "未选中", "请先在右侧列表中选中需要核销的分期。")
            return
            
        row = selected_rows[0].row()
        inst_id = int(self.installment_table.item(row, 0).text())

        if not self.current_invoice_id:
            QMessageBox.warning(self, "未选中", "请先在左侧选择对应发票。")
            return

        invoice_map = self._get_invoice_map()
        invoice = invoice_map.get(int(self.current_invoice_id))
        inst_map = self._get_installment_map_for_invoice(self.current_invoice_id)
        inst = inst_map.get(inst_id)

        if not invoice or not inst:
            QMessageBox.critical(self, "错误", "无法获取分期关联的客户信息")
            return
            
        cust_name = str(invoice.get("destinatario_name") or "")
        cust_cnpj = str(invoice.get("destinatario_cnpj") or "")
        balance = max(0.0, self._to_float(inst.get("amount")) + self._to_float(inst.get("penalty")) - self._to_float(inst.get("paid_amount")))
        
        if balance <= 0.01:
            QMessageBox.information(self, "无需核销", "该分期已结清。")
            return
            
        dlg = AdvanceWriteoffDialog(self.db, cust_cnpj, cust_name, inst_id, balance, self)
        if dlg.exec() == QDialog.Accepted:
            adv = dlg.get_selected_advance()
            if adv:
                # 确定核销金额 (取可用余额和待收余额的最小值)
                apply_amt = min(adv['remaining_amount'], balance)
                
                # 弹出确认框确认金额
                amt_str, ok = QInputDialog.getText(self, "确认核销金额", 
                                                f"可用预收: R$ {adv['remaining_amount']:,.2f}\n待收余额: R$ {balance:,.2f}\n请输入核销金额:", 
                                                text=f"{apply_amt:.2f}")
                if not ok: return
                try:
                    final_amt = float(amt_str.replace(',', ''))
                except:
                    QMessageBox.warning(self, "输入错误", "请输入有效的数字。")
                    return
                
                if final_amt > adv['remaining_amount'] + 0.01:
                    QMessageBox.warning(self, "余额不足", "核销金额不能超过预收款余额。")
                    return
                
                success = self.db.apply_advance_to_installment(adv['id'], inst_id, final_amt)
                if success:
                    QMessageBox.information(self, "成功", f"已成功从预收款中核销 R$ {final_amt:,.2f}")
                    self.load_invoices()
                    if self.current_invoice_id:
                        self.load_installments(self.current_invoice_id)
                else:
                    QMessageBox.critical(self, "失败", "核销过程中发生错误。")

    def batch_update_invoice_status(self, invoice_ids, status):
        """批量更新发票所有分期状 (支持单ID或ID列表)"""
        if isinstance(invoice_ids, int):
            invoice_ids = [invoice_ids]
            
        all_installment_ids = []
        for inv_id in invoice_ids:
            rows = self.db.get_installments(inv_id)
            if rows:
                all_installment_ids.extend([r['id'] for r in rows])
        
        if not all_installment_ids: return
        
        if status == 'Paid':
            self.register_payment_batch(all_installment_ids)
        else:
            # Revert to Pending
            reply = QMessageBox.question(self, "确认", f"确定要重置选中 {len(invoice_ids)} 张发票为未付吗？\n关联的账户交易记录将被删除，余额将回滚", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                for inst_id in all_installment_ids:
                    self.revert_payment(inst_id)
                self.load_invoices()
                if self.current_invoice_id in invoice_ids:
                    self.load_installments(self.current_invoice_id)

    def load_invoices(self):
        # --- 保存当前选中的发票ID和滚动条位置 ---
        selected_id = None
        table = self.invoice_table
        curr_row = table.currentRow()
        if curr_row >= 0:
            id_item = table.item(curr_row, 0)
            if id_item:
                try:
                    selected_id = int(id_item.text())
                except Exception:
                    pass

        self._invoice_load_seq += 1
        req_id = self._invoice_load_seq
        self._invoice_view_state[req_id] = {
            "selected_id": selected_id,
            "v_scroll": table.verticalScrollBar().value(),
            "h_scroll": table.horizontalScrollBar().value(),
        }
        # 清理过期缓存，避免状态字典持续增长
        for key in list(self._invoice_view_state.keys()):
            if key < req_id - 5:
                self._invoice_view_state.pop(key, None)

        if self._invoice_worker and self._invoice_worker.isRunning():
            self._invoice_worker.stop()

        worker = InvoiceRowsLoadWorker(self.db, req_id, self)
        worker.finished.connect(self._on_invoice_rows_loaded)
        worker.error.connect(self._on_invoice_rows_error)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        self._invoice_worker = worker
        worker.start()

    def _on_invoice_rows_loaded(self, req_id, rows):
        if req_id != self._invoice_load_seq:
            return
        self._invoice_worker = None
        view_state = self._invoice_view_state.pop(req_id, {})
        self._apply_invoice_rows(rows, view_state)

    def _on_invoice_rows_error(self, req_id, error_msg):
        if req_id != self._invoice_load_seq:
            return
        self._invoice_worker = None
        if hasattr(self.main_window, "log_message"):
            self.main_window.log_message(f"加载发票列表失败: {error_msg}")
        else:
            logging.error("Failed to load invoice rows: %s", error_msg)

    def _apply_invoice_rows(self, rows, view_state):
        table = self.invoice_table
        table.setSortingEnabled(False)
        table.blockSignals(True)
        table.setUpdatesEnabled(False)

        try:
            table.clearContents()
            table.setRowCount(len(rows))

            # 获取自己的 CNPJ 用于识别对手方
            my_cnpj = self.main_window.identify_self_cnpj()
            settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)
            color_paid = QColor(settings.value("color_paid", "#285028"))
            color_partial = QColor(settings.value("color_partial", "#646428"))
            white = QColor("white")

            # 缓存标准化查询，避免重复计算
            mapping_mgr = self.main_window.mapping_mgr
            std_by_cnpj = {}
            std_by_name = {}

            for i, row in enumerate(rows):
                inv_id_val = row.get("id")
                inv_id = str(inv_id_val if inv_id_val is not None else "")
                inv_num = str(row.get("invoice_number", ""))
                nature_raw = row.get("natureza_operacao")
                nature = str(nature_raw) if nature_raw is not None else ""

                # --- 智能往来单位显示与校准 ---
                raw_issuer = str(row.get("issuer_name") or "")
                raw_dest = str(row.get("destinatario_name") or "")
                emit_cnpj = re.sub(r'\D', '', str(row.get("issuer_cnpj") or ""))
                dest_cnpj = re.sub(r'\D', '', str(row.get("destinatario_cnpj") or ""))

                if my_cnpj and emit_cnpj == my_cnpj:
                    partner_name = raw_dest
                    target_cnpj = dest_cnpj
                else:
                    partner_name = raw_issuer
                    target_cnpj = emit_cnpj

                std_name = ""
                if target_cnpj:
                    std_name = std_by_cnpj.get(target_cnpj)
                    if std_name is None:
                        std_name = mapping_mgr.get_partner_std(target_cnpj) or ""
                        std_by_cnpj[target_cnpj] = std_name

                if not std_name:
                    std_name = std_by_name.get(partner_name)
                    if std_name is None:
                        std_name = mapping_mgr.get_partner_std(partner_name) or ""
                        std_by_name[partner_name] = std_name

                display_partner = std_name if std_name else partner_name
                # --------------------------------

                date = format_date_gui(str(row.get("issue_date", "")))
                amount_val = self._to_float(row.get("total_amount"))
                paid_amount_val = self._to_float(row.get("total_paid_amount"))
                balance_val = amount_val - paid_amount_val

                status = str(row.get("status", ""))
                bg_color = None
                if status == 'Paid':
                    bg_color = color_paid
                elif status == 'Partial':
                    bg_color = color_partial

                items = [
                    QTableWidgetItem(inv_id),
                    QTableWidgetItem(inv_num),
                    QTableWidgetItem(nature),
                    QTableWidgetItem(display_partner),
                    QTableWidgetItem(date),
                    QTableWidgetItem(f"{amount_val:.2f}"),
                    QTableWidgetItem(f"{paid_amount_val:.2f}"),
                    QTableWidgetItem(f"{balance_val:.2f}"),
                    QTableWidgetItem(status)
                ]

                for col, item in enumerate(items):
                    if bg_color:
                        item.setBackground(bg_color)
                        item.setForeground(white)
                    table.setItem(i, col, item)

                paid_cnt = int(row.get("paid_terms_count") or 0)
                total_cnt = int(row.get("total_terms_count") or 0)
                pct = int((paid_cnt / total_cnt) * 100) if total_cnt > 0 else 0
                if pct > 100:
                    pct = 100

                progress_item = QTableWidgetItem()
                progress_item.setData(Qt.DisplayRole, pct)  # for sorting
                progress_item.setData(Qt.UserRole + 1000, (pct, f"{paid_cnt}/{total_cnt}"))
                table.setItem(i, 9, progress_item)

                id_cell = table.item(i, 0)
                if id_cell:
                    id_cell.setData(Qt.UserRole, inv_id_val)
        finally:
            table.setUpdatesEnabled(True)
            table.blockSignals(False)

        table.setSortingEnabled(False) # Ensure false for custom sort

        # --- 恢复选中的发票ID和滚动条位置 ---
        selected_id = view_state.get("selected_id")
        if selected_id is not None:
            table.blockSignals(True)
            for r in range(table.rowCount()):
                item = table.item(r, 0)
                if item and item.text() == str(selected_id):
                    table.selectRow(r)
                    table.setCurrentCell(r, 0)
                    break
            table.blockSignals(False)

        table.verticalScrollBar().setValue(view_state.get("v_scroll", 0))
        table.horizontalScrollBar().setValue(view_state.get("h_scroll", 0))
        # ----------------------------------------

        # 延迟刷新主汇总表，合并短时间内多次请求
        if self.main_window and hasattr(self.main_window, "payment_tab") and self.main_window.payment_tab is self:
            self._summary_refresh_timer.start(120)

        # Update Summary Row
        self.main_window.update_summary_row(self.invoice_table)

    def _refresh_main_summary_table(self):
        if not self.main_window:
            return
        try:
            self.main_window.populate_summary_table()
        except Exception as e:
            logging.error(f"Failed to refresh main summary table: {e}")

    def on_invoice_selected(self, item):
        row = item.row()
        inv_id = int(self.invoice_table.item(row, 0).text())
        self.current_invoice_id = inv_id
        self.load_installments(inv_id)
        
    def update_summary_display(self, rows):
        """更新概览面板数据"""
        total_due = 0.0
        total_paid = 0.0
        
        for row in rows:
            row_data = row if isinstance(row, dict) else self._to_dict(row)
            # 应收 = 本金 + 滞纳金
            total_due += self._to_float(row_data.get('amount')) + self._to_float(row_data.get('penalty'))
            total_paid += self._to_float(row_data.get('paid_amount'))
            
        balance = total_due - total_paid
        #
        if balance < 0: balance = 0 # 允许负数（多付）
        
        self.lbl_total.setText(f"应收(含滞纳): {total_due:,.2f}")
        self.lbl_paid.setText(f"已收: {total_paid:,.2f}")
        self.lbl_balance.setText(f"待收余额: {balance:,.2f}")
        
        if total_due > 0:
            pct = int((total_paid / total_due) * 100)
            if pct > 100: pct = 100
            self.summary_progress.setValue(pct)
        else:
            self.summary_progress.setValue(0)

    def load_installments(self, invoice_id):
        if invoice_id is None:
            return

        # --- 保存当前选中的分期ID和滚动条位置 ---
        selected_inst_id = None
        table = self.installment_table
        curr_row = table.currentRow()
        if curr_row >= 0:
            inst_id_item = table.item(curr_row, 0)
            if inst_id_item:
                try:
                    selected_inst_id = int(inst_id_item.text())
                except Exception:
                    pass

        self._loading_installments = True
        self._installment_load_seq += 1
        req_id = self._installment_load_seq
        self._installment_view_state[req_id] = {
            "invoice_id": int(invoice_id),
            "selected_inst_id": selected_inst_id,
            "v_scroll": table.verticalScrollBar().value(),
            "h_scroll": table.horizontalScrollBar().value(),
        }
        # 清理过期缓存，避免状态字典持续增长
        for key in list(self._installment_view_state.keys()):
            if key < req_id - 5:
                self._installment_view_state.pop(key, None)

        if self._installment_worker and self._installment_worker.isRunning():
            self._installment_worker.stop()

        worker = InstallmentRowsLoadWorker(self.db, req_id, int(invoice_id), self)
        worker.finished.connect(self._on_installment_rows_loaded)
        worker.error.connect(self._on_installment_rows_error)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        self._installment_worker = worker
        worker.start()

    def _on_installment_rows_loaded(self, req_id, invoice_id, rows):
        if req_id != self._installment_load_seq:
            return
        self._installment_worker = None
        view_state = self._installment_view_state.pop(req_id, {})

        # 更新概览
        self.update_summary_display(rows)

        table = self.installment_table
        settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)
        color_paid = QColor(settings.value("color_paid", "#285028"))
        color_partial = QColor(settings.value("color_partial", "#646428"))
        color_pending = QColor(settings.value("color_pending", "#642828"))
        white = QColor("white")

        table.blockSignals(True)
        table.setUpdatesEnabled(False)
        try:
            table.clearContents()
            table.setRowCount(len(rows))

            for i, row in enumerate(rows):
                id_item = QTableWidgetItem(str(row.get('id', "")))
                id_item.setFlags(id_item.flags() ^ Qt.ItemIsEditable) # ID不可编辑

                term_item = QTableWidgetItem(f"{row.get('term_number', '')} 期")
                term_item.setFlags(term_item.flags() ^ Qt.ItemIsEditable)

                due_item = QTableWidgetItem(format_date_gui(str(row.get('due_date', ""))))
                amount_item = QTableWidgetItem(f"{self._to_float(row.get('amount')):.2f}")
                penalty_item = QTableWidgetItem(f"{self._to_float(row.get('penalty')):.2f}")

                paid_item = QTableWidgetItem(f"{self._to_float(row.get('paid_amount')):.2f}")
                paid_item.setFlags(paid_item.flags() ^ Qt.ItemIsEditable) # 实收不可直接编辑，强制走登记

                paid_date_item = QTableWidgetItem(format_date_gui(str(row.get('paid_date') or "")))
                paid_date_item.setFlags(paid_date_item.flags() ^ Qt.ItemIsEditable) # 收款日期不可直接编辑

                status = str(row.get('status', ""))
                status_item = QTableWidgetItem(status)
                status_item.setFlags(status_item.flags() ^ Qt.ItemIsEditable)

                if status == 'Paid':
                    status_item.setBackground(color_paid)
                    status_item.setForeground(white)
                elif status == 'Partial':
                    status_item.setBackground(color_partial)
                    status_item.setForeground(white)
                elif status == 'Pending':
                    status_item.setBackground(color_pending)
                    status_item.setForeground(white)

                note_item = QTableWidgetItem(str(row.get('note') or ""))

                table.setItem(i, 0, id_item)
                table.setItem(i, 1, term_item)
                table.setItem(i, 2, due_item)
                table.setItem(i, 3, amount_item)
                table.setItem(i, 4, penalty_item)
                table.setItem(i, 5, paid_item)
                table.setItem(i, 6, paid_date_item)
                table.setItem(i, 7, status_item)
                table.setItem(i, 8, note_item)
        finally:
            table.setUpdatesEnabled(True)
            table.blockSignals(False)

        # --- 恢复选中的分期ID和滚动条位置 ---
        selected_inst_id = view_state.get("selected_inst_id")
        if selected_inst_id is not None:
            table.blockSignals(True)
            for r in range(table.rowCount()):
                item = table.item(r, 0)
                if item and item.text() == str(selected_inst_id):
                    table.selectRow(r)
                    table.setCurrentCell(r, 0)
                    break
            table.blockSignals(False)

        table.verticalScrollBar().setValue(view_state.get("v_scroll", 0))
        table.horizontalScrollBar().setValue(view_state.get("h_scroll", 0))
        # ----------------------------------------

        self._loading_installments = False

        # Update Summary Row
        self.main_window.update_summary_row(self.installment_table)

    def _on_installment_rows_error(self, req_id, invoice_id, error_msg):
        if req_id != self._installment_load_seq:
            return
        self._installment_worker = None
        self._loading_installments = False
        if hasattr(self.main_window, "log_message"):
            self.main_window.log_message(f"加载分期失败 (invoice_id={invoice_id}): {error_msg}")
        else:
            logging.error("Failed to load installment rows: %s", error_msg)

    def shutdown_workers(self):
        if self._invoice_worker and self._invoice_worker.isRunning():
            self._invoice_worker.stop()
            self._invoice_worker.wait(300)
        if self._installment_worker and self._installment_worker.isRunning():
            self._installment_worker.stop()
            self._installment_worker.wait(300)

    def on_installment_cell_changed(self, item):
        if self._loading_installments or not self.current_invoice_id:
            return
            
        row = item.row()
        col = item.column()
        val = item.text().strip()
        
        inst_id_item = self.installment_table.item(row, 0)
        if not inst_id_item: return
        
        try:
            inst_id = int(inst_id_item.text())
        except ValueError:
            return # Ignore summary rows or non-int IDs
        
        field_map = {2: "due_date", 3: "amount", 4: "penalty", 5: "paid_amount", 6: "paid_date", 8: "note"}
        if col in field_map:
            db_val = val
            if col in [3, 4, 5]: # 数字转换
                db_val = br_to_float(val) or 0.0
            
            ok = self.db.update_installment_field(inst_id, field_map[col], db_val)
            if ok is False:
                QMessageBox.warning(self, "保存失败", "分期字段更新失败，请检查网络/数据库连接。")
                self.load_installments(self.current_invoice_id)
                self.load_invoices()
                return
            
            # 移除自动变更状态逻辑，强制要求点击按钮进行“登记收款”以选择账户
            
            # 刷新界面状
            self.load_installments(self.current_invoice_id)
            self.load_invoices()

    def generate_plan(self):
        if not self.current_invoice_id:
            QMessageBox.warning(self, "提示", "请先选择一张发票")
            return
            
        # 从文本中提取数字，而不是依赖索引
        terms_text = self.terms_combo.currentText()
        import re
        m = re.search(r'\d+', terms_text)
        terms = int(m.group(0)) if m else 1
        
        reply = QMessageBox.question(self, "确认", f"确定要为当前发票生成 {terms} 期还款计划吗？\n这将覆盖旧的计划", QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            # 尝试获取发票日期作为起始日期
            initial_date = None
            for r in range(self.invoice_table.rowCount()):
                id_item = self.invoice_table.item(r, 0)
                if id_item and id_item.text() == str(self.current_invoice_id):
                    date_item = self.invoice_table.item(r, 4)
                    if date_item:
                        initial_date_str = date_item.text()
                        initial_date = safe_parse_date(initial_date_str) or initial_date
                    break

            plan_ok = self.db.generate_payment_plan(self.current_invoice_id, terms, start_date=initial_date)
            if plan_ok is False:
                QMessageBox.warning(self, "生成失败", "分期计划生成失败，请检查网络/数据库连接。")
                return
            self.load_installments(self.current_invoice_id)
            self.load_invoices() # 刷新状 
    def set_payment_status(self, status):
        # 获取所有选中的行
        selected_rows = set()
        for item in self.installment_table.selectedItems():
            selected_rows.add(item.row())
            
        if not selected_rows:
            # 如果没有选中项，尝试使用当前行
            curr = self.installment_table.currentRow()
            if curr >= 0:
                selected_rows.add(curr)
            else:
                return

        ids = []
        for row in selected_rows:
            inst_id_item = self.installment_table.item(row, 0)
            if inst_id_item:
                try:
                    ids.append(int(inst_id_item.text()))
                except ValueError:
                    continue # Ignore summary rows
        
        if not ids: return

        if status == 'Paid':
            self.register_payment_batch(ids)
        else:
            # Revert
            reply = QMessageBox.question(self, "确认", "确定要重置为未付吗？\n关联的账户交易记录将被删除，余额将回滚", QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                for inst_id in ids:
                    self.revert_payment(inst_id)
                self.load_installments(self.current_invoice_id)
                self.load_invoices()

    def register_payment_batch(self, installment_ids):
        """批量登记收款 (带账户选择)"""
        if not installment_ids:
            return
        if not self.current_invoice_id:
            QMessageBox.warning(self, "错误", "请先选择发票后再进行登记。")
            return

        inst_map = self._get_installment_map_for_invoice(self.current_invoice_id)
        total_due = 0.0
        for inst_id in installment_ids:
            inst = inst_map.get(inst_id)
            if not inst:
                continue
            full_due = self._to_float(inst.get("amount")) + self._to_float(inst.get("penalty")) - self._to_float(inst.get("paid_amount"))
            total_due += max(0.0, full_due)
            
        dlg = PaymentRegistrationDialog(self.db, default_amount=total_due, parent=self)
        if dlg.exec():
            data = dlg.get_data()
            account_id = data['account_id']
            pay_date = self._to_iso_date(data['date'])
            note = str(data['note']).strip()
            penalty_input = self._to_float(data.get('penalty', 0.0))
            
            # Use entered amount only if single item, otherwise distribute fully?
            is_single = (len(installment_ids) == 1)
            entered_amount = max(0.0, self._to_float(data['amount']))
            
            if not account_id:
                QMessageBox.warning(self, "错误", "必须选择收款账户")
                return

            if penalty_input > 0 and installment_ids:
                first_inst_id = installment_ids[0]
                first_inst = inst_map.get(first_inst_id)
                if first_inst:
                    new_penalty = self._to_float(first_inst.get("penalty")) + penalty_input
                    penalty_ok = self.db.update_installment_field(first_inst_id, "penalty", round(new_penalty, 2))
                    if penalty_ok is False:
                        QMessageBox.warning(self, "更新失败", "手续费更新失败，请检查网络/数据库连接。")
                        return
                    inst_map = self._get_installment_map_for_invoice(self.current_invoice_id)

            success_count = 0
            failed_count = 0
            for inst_id in installment_ids:
                inst = inst_map.get(inst_id)
                if not inst:
                    failed_count += 1
                    continue

                full_due = max(
                    0.0,
                    self._to_float(inst.get("amount")) + self._to_float(inst.get("penalty")) - self._to_float(inst.get("paid_amount")),
                )
                if is_single:
                    pay_amount = min(entered_amount, full_due)
                else:
                    pay_amount = full_due

                if pay_amount <= 0.01:
                    continue

                trans_desc = f"收款-分期#{inst_id}"
                if note:
                    trans_desc += f" ({note})"

                ok = self.db.register_payment(inst_id, pay_amount, pay_date, account_id, trans_desc)
                if ok:
                    success_count += 1
                else:
                    failed_count += 1

            if success_count > 0 and failed_count == 0:
                QMessageBox.information(self, "成功", "收款登记成功")
            elif success_count > 0:
                QMessageBox.warning(self, "部分成功", f"成功 {success_count} 笔，失败 {failed_count} 笔。")
            else:
                QMessageBox.critical(self, "错误", "收款登记失败，未能处理任何分期。")

            if self.current_invoice_id:
                self.load_installments(self.current_invoice_id)
            self.load_invoices()

            # 自动刷新账户管理模块的数据
            if self.main_window and hasattr(self.main_window, 'account_tab'):
                self.main_window.account_tab.load_accounts()

    def revert_payment(self, inst_id):
        """回滚支付 (删除关联交易)"""
        try:
            target_inv_id = self.current_invoice_id
            inst = None

            if target_inv_id:
                inst = self._get_installment_map_for_invoice(target_inv_id).get(inst_id)

            if not inst:
                for inv_id in self._get_invoice_map().keys():
                    cand = self._get_installment_map_for_invoice(inv_id).get(inst_id)
                    if cand:
                        inst = cand
                        target_inv_id = inv_id
                        break

            if not inst:
                QMessageBox.warning(self, "错误", f"未找到分期记录: {inst_id}")
                return

            trans_id = inst.get("transaction_id")
            if trans_id:
                delete_ok = self.db.delete_transaction(trans_id)
                if delete_ok is False:
                    QMessageBox.warning(self, "撤销失败", "删除流水记录失败，请检查网络/数据库连接。")
                    return

            reset_steps = [
                ("paid_amount", 0),
                ("paid_date", None),
                ("transaction_id", None),
                ("account_id", None),
                ("note", ""),
                ("status", "Pending"),
            ]
            for field_name, field_value in reset_steps:
                ok = self.db.update_installment_field(inst_id, field_name, field_value)
                if ok is False:
                    QMessageBox.warning(self, "撤销失败", f"更新字段 {field_name} 失败，请检查网络/数据库连接。")
                    return

            invoice_id = inst.get("invoice_id") or target_inv_id
            if invoice_id:
                refresh_ok = self.db.refresh_invoice_status(int(invoice_id))
                if refresh_ok is False:
                    QMessageBox.warning(self, "状态刷新失败", "发票状态刷新失败，请稍后重试。")
                    return

            # 自动刷新账户管理模块的数据
            if self.main_window and hasattr(self.main_window, 'account_tab'):
                self.main_window.account_tab.load_accounts()
        except Exception as e:
            QMessageBox.critical(self, "错误", f"回滚失败: {e}")

    def goto_account_details(self):
        """跳转到账户管理页并过滤出当前发票的所有交易记录"""
        if not self.current_invoice_id:
            QMessageBox.warning(self, "提示", "请先选择一张发票")
            return
            
        if self.main_window and hasattr(self.main_window, 'account_tab'):
            found = self.main_window.account_tab.filter_transactions_by_invoice(self.current_invoice_id)
            if found:
                self.main_window.tab_widget.setCurrentWidget(self.main_window.account_tab)
            else:
                QMessageBox.information(self, "未发现记录", "该发票及其分期目前尚未关联任何收款流水。")

    def import_collection_report(self):
        """导入银行收款/逾期报告 PDF 并自动对账"""
        paths, _ = QFileDialog.getOpenFileNames(self, "选择银行报告 PDF", "", "PDF Files (*.pdf)")
        if not paths: return

        # 选择入账账户
        accounts = self.db.get_accounts()
        if not accounts:
            QMessageBox.warning(self, "错误", "请先在'账户管理'中创建一个收款账户")
            return
            
        acc_names = [f"{a['name']} (余额: {a['current_balance']:.2f})" for a in accounts]
        acc_choice, ok = QInputDialog.getItem(self, "选择入账账户", "报告中的收款将记入以下账户:", acc_names, 0, False)
        if not ok: return
        account_id = accounts[acc_names.index(acc_choice)]['id']

        total_matched = 0
        total_records = 0
        pending_cache = None
        
        for path in paths:
            self.main_window.log_message(f"正在读取报告: {os.path.basename(path)}")
            # 1. 提取与解析
            text, _ = extract_text_from_pdf(path)
            records = CollectionReportParser.parse_report(text)
            total_records += len(records)

            # 2. 数据库匹配并登记
            for rec in records:
                raw_ref = rec.get('invoice_ref', '')
                enable_local = QSettings(SETTINGS_FILE, QSettings.IniFormat).value("enable_local_rules", True, type=bool)
                ref_info = CollectionReportParser.parse_invoice_reference(raw_ref, enable_local)
                clean_num = ref_info.get('invoice') if ref_info else CollectionReportParser.clean_invoice_number(raw_ref, enable_local)

                term_num = ref_info.get('term_number')
                matches = []
                if clean_num:
                    patterns = self._build_invoice_search_patterns(clean_num)
                    matches = self.db.search_pending_installments(patterns, term_num)

                rec_amount = self._to_float(rec.get('amount'))

                # fallback: 按待收金额匹配
                if not matches and rec_amount > 0:
                    if pending_cache is None:
                        pending_cache = self._collect_pending_installments()

                    amt_matches = [m for m in pending_cache if abs(self._to_float(m.get("due_remaining")) - rec_amount) < 0.5]
                    if amt_matches:
                        if rec.get('due_date'):
                            rec_due = safe_parse_date(rec.get('due_date'))
                            if rec_due:
                                amt_matches.sort(
                                    key=lambda x: abs(
                                        (
                                            (safe_parse_date(str(x.get("due_date"))) or rec_due) - rec_due
                                        ).days
                                    )
                                )
                        matches = amt_matches

                best_match = None
                due_remaining = 0.0
                for m in matches:
                    row = self._to_dict(m)
                    inst_id = row.get("installment_id", row.get("inst_id"))
                    if not inst_id:
                        continue

                    db_due = row.get("due_date")
                    if ref_info and ref_info.get('day') and db_due:
                        try:
                            dt_obj = datetime.strptime(str(db_due), "%Y-%m-%d")
                            if dt_obj.day != ref_info['day'] or dt_obj.month != ref_info['month']:
                                continue
                        except Exception:
                            pass

                    amount = self._to_float(row.get("amount"))
                    penalty = self._to_float(row.get("penalty"))
                    paid_amount = self._to_float(row.get("paid_amount"))
                    due_amt = max(0.0, amount + penalty - paid_amount)
                    if abs(due_amt - rec_amount) < 1.0:
                        best_match = row
                        due_remaining = due_amt
                        break

                if best_match and str(rec.get('status', '')).strip().lower() == "paid":
                    inst_id = best_match.get("installment_id", best_match.get("inst_id"))
                    pay_amount = min(rec_amount, due_remaining if due_remaining > 0 else rec_amount)
                    if inst_id and pay_amount > 0.01:
                        pay_date = self._to_iso_date(rec.get('pay_date'))
                        ok = self.db.register_payment(inst_id, pay_amount, pay_date, account_id, f"自动导入回单: {raw_ref}")
                        if ok:
                            total_matched += 1
                            if pending_cache is not None:
                                for p in pending_cache:
                                    if int(p.get("inst_id")) == int(inst_id):
                                        p["paid_amount"] = self._to_float(p.get("paid_amount")) + pay_amount
                                        p["due_remaining"] = max(0.0, self._to_float(p.get("due_remaining")) - pay_amount)
                                        break

        self.load_invoices()
        if self.main_window.account_tab: self.main_window.account_tab.load_accounts()
        
        msg = f"报告处理完成！\n共解析到 {total_records} 条记录\n成功自动匹配并登记: {total_matched} 笔收款"
        QMessageBox.information(self, "导入结果", msg)
        self.main_window.log_message(msg)

    def audit_bank_statement(self):
        """导入银行流水(Extrato)并进行二次对账验证"""
        paths, _ = QFileDialog.getOpenFileNames(self, "选择银行流水 PDF", "", "PDF Files (*.pdf)")
        if not paths: return

        # 1. 选择核对的入账账户
        accounts = self.db.get_accounts()
        if not accounts:
            QMessageBox.warning(self, "错误", "请先创建账户")
            return
        acc_names = [f"{a['name']}" for a in accounts]
        acc_choice, ok = QInputDialog.getItem(self, "选择流水所属账户", "系统将流水对比后记入此账户:", acc_names, 0, False)
        if not ok: return
        account_id = accounts[acc_names.index(acc_choice)]['id']

        # 2. 解析所有流水
        all_trans = []
        for path in paths:
            text, _ = extract_text_from_pdf(path)
            all_trans.extend(BankStatementParser.parse_statement(text))
        
        if not all_trans:
            QMessageBox.information(self, "无数据", "未在文件中识别到有效的收入款项。")
            return

        # 3. 智能对账 (匹配数据库)
        matched_results = []
        unmatched_list = []
        pending_entries = self._collect_pending_installments()
        
        for tr in all_trans:
            match = None
            tr_amount = self._to_float(tr.get('amount'))
            
            # 策略 A: 优先使用 CNPJ 匹配
            if tr.get('cnpj'):
                clean_cnpj = re.sub(r'\D', '', tr.get('cnpj', ''))
                cnpj_candidates = [
                    p for p in pending_entries
                    if re.sub(r'\D', '', str(p.get('destinatario_cnpj', ''))) == clean_cnpj
                    and abs(self._to_float(p.get('due_remaining')) - tr_amount) < 1.0
                ]
                cnpj_candidates.sort(key=lambda x: str(x.get("due_date") or "9999-12-31"))
                if cnpj_candidates:
                    match = cnpj_candidates[0]
            
            # 策略 B: 如果没匹配上，尝试纯金额匹配 (金额一致)
            if not match:
                exact_candidates = [p for p in pending_entries if abs(self._to_float(p.get('due_remaining')) - tr_amount) < 0.01]
                exact_candidates.sort(key=lambda x: str(x.get("due_date") or "9999-12-31"))
                if exact_candidates:
                    match = exact_candidates[0]
            
            if match:
                matched_results.append((tr, match))
                pending_entries = [p for p in pending_entries if int(p.get("inst_id")) != int(match.get("inst_id"))]
            else:
                unmatched_list.append(tr)

        # 4. 显示核对结果并确认
        unmatched_count = len(unmatched_list)
        if not matched_results:
            QMessageBox.warning(self, "核对失败", f"未能自动匹配到任何数据库记录。\n解析到收入: {len(all_trans)} 笔\n其中未匹配: {unmatched_count} 笔")
            return

        # 构造确认列表
        confirm_msg = f"智能核对完成！发现 {len(matched_results)} 笔匹配记录：\n\n"
        for tr, db in matched_results[:10]: # 仅显示前10笔
            confirm_msg += f"- 流水: {tr['date']} | R$ {tr['amount']:.2f} -> 匹配发票: {db.get('invoice_number', '')} ({db.get('destinatario_name', '')})\n"
        
        if len(matched_results) > 10:
            confirm_msg += f"...及其他 {len(matched_results)-10} 笔。\n"
            
        confirm_msg += f"\n共有 {unmatched_count} 笔流水未能识别。是否立即执行自动入账登记？"
        
        reply = QMessageBox.question(self, "确认自动入账", confirm_msg, QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            # 5. 执行入账更新
            success_count = 0
            for tr, db in matched_results:
                try:
                    d_fmt = self._to_iso_date(tr.get('date'))
                    tr_amount = self._to_float(tr.get('amount'))
                    due_remaining = self._to_float(db.get('due_remaining'))
                    pay_amount = min(tr_amount, due_remaining if due_remaining > 0 else tr_amount)
                    if pay_amount <= 0.01:
                        continue

                    ok = self.db.register_payment(
                        db.get('inst_id'),
                        pay_amount,
                        d_fmt,
                        account_id,
                        f"银行流水自动核销: {tr.get('desc', '')}",
                    )
                    if ok:
                        success_count += 1
                except Exception:
                    continue
            
            if unmatched_list:
                # 2. 弹窗确认
                headers = ["日期", "摘要/描述", "金额", "税号 (CNPJ)", "凭证号 (Ref)"]
                keys = ["date", "desc", "amount", "cnpj", "doc_ref"]
                
                dlg = DataPreviewDialog(
                    "确认预收账款 (Unmatched Incomes)", 
                    unmatched_list, 
                    headers, 
                    keys, 
                    parent=self,
                    instruction="⚠️ 以下流水未匹配到发票。\n请保留要登记为【预收账款】的记录，不需要入账的请选中并按 Delete 删除。"
                )
                
                if dlg.exec():
                    # 3. 处理用户确认后的列表
                    for tr in dlg.result_data:
                        try:
                            d_fmt = self._to_iso_date(tr.get('date'))
                            tr_amount = self._to_float(tr.get('amount'))
                            if tr_amount <= 0.01:
                                continue
                            
                            # 获取或推测客户名称
                            cust_name = tr.get('desc', '')
                            st_cnpj = tr.get('cnpj', '')
                            if st_cnpj:
                                clean_cnpj = re.sub(r'\D', '', st_cnpj)
                                std_name = self.main_window.mapping_mgr.get_partner_std(clean_cnpj)
                                if std_name: cust_name = std_name
                            
                            trans_id = self.db.add_transaction(
                                account_id,
                                d_fmt,
                                'Income',
                                tr_amount,
                                f"预收账款(流水审计未匹配): {tr.get('desc', '')}",
                            )
                            if trans_id:
                                self.db.add_advance(
                                    cust_name,
                                    st_cnpj,
                                    tr_amount,
                                    d_fmt,
                                    f"银行流水审计入账(预收): {tr.get('desc', '')}",
                                    account_id,
                                    trans_id,
                                )
                                success_count += 1
                        except Exception:
                            continue
            
            self.load_invoices()
            if self.main_window.account_tab: self.main_window.account_tab.load_accounts()
            QMessageBox.information(self, "完成", f"已成功完成 {success_count} 笔账目的自动审计与登记。")

class DataPreviewWidget(QWidget):
    """可复用的数据预览表格组件"""
    jump_to_date_requested = Signal(str)
    stats_updated = Signal(str)

    def __init__(self, data, headers, key_map, instruction=None, highlight_index=None, row_colors=None, parent=None):
        super().__init__(parent)
        self.data = data
        self.headers = headers
        self.key_map = key_map
        self.row_colors = row_colors or {}
        
        layout = QVBoxLayout(self)
        if instruction:
            layout.addWidget(QLabel(instruction))
            
        self.table = QTableWidget()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        
        # Use FilterHeader
        self.header = FilterHeader(self.table)
        self.header.filterChanged.connect(self.apply_filters)
        self.table.setHorizontalHeader(self.header)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.itemSelectionChanged.connect(self.calculate_stats)
        self.table.setSortingEnabled(True)
        
        # Context Menu
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        self.populate_table()
        layout.addWidget(self.table)
        
        # Auto Sort by Date (Ascending)
        self.auto_sort_date()
        
        if highlight_index is not None and 0 <= highlight_index < self.table.rowCount():
            self.table.selectRow(highlight_index)
            self.table.scrollToItem(self.table.item(highlight_index, 0))
            
        self.lbl_count = QLabel(f"共 {len(data)} 条记录")
        layout.addWidget(self.lbl_count)
        
        QShortcut(QKeySequence.Delete, self.table, self.delete_selected_rows)

    def apply_filters(self):
        """应用表头筛选"""
        filters = self.header._filters
        rows = self.table.rowCount()
        
        for r in range(rows):
            visible = True
            for c, criterion in filters.items():
                item = self.table.item(r, c)
                val = item.text().strip() if item else ""
                
                if not check_filter_match(val, criterion):
                    visible = False
                    break
            
            self.table.setRowHidden(r, not visible)
            
    def auto_sort_date(self):
        """Find first date column and sort ascending"""
        date_col = -1
        # Check keys first
        for i, key in enumerate(self.key_map):
            if 'date' in key.lower():
                date_col = i
                break
        
        # Fallback to headers
        if date_col == -1:
            for i, h in enumerate(self.headers):
                if "日期" in h or "Date" in h:
                    date_col = i
                    break
                    
        if date_col != -1:
            self.table.sortItems(date_col, Qt.AscendingOrder)

    def calculate_stats(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            self.stats_updated.emit("")
            return
            
        count = len(selected_items)
        values = []
        for item in selected_items:
            # Skip hidden rows (filtered out)
            if item.row() >= 0 and self.table.isRowHidden(item.row()):
                continue
                
            try:
                raw_txt = item.text().strip()
                if not raw_txt or raw_txt == "-": continue
                # 使用标准的 br_to_float
                clean_txt = re.sub(r'\(.*?\)', '', raw_txt).strip()
                val = br_to_float(clean_txt)
                if val is not None:
                    values.append(val)
            except:
                pass
        
        txt_parts = [f"计数: {count}"]
        if values:
            total = sum(values)
            avg = total / len(values)
            txt_parts.append(f"求和: {total:,.2f}")
            txt_parts.append(f"平均: {avg:,.4f}")
            
        self.stats_updated.emit("  |  ".join(txt_parts))

    def populate_table(self):
        self.table.setSortingEnabled(False) # Disable sorting while populating
        self.table.setRowCount(len(self.data))
        self.table.blockSignals(True)
        for i, row_dict in enumerate(self.data):
            # 优先从 row_colors 获取，其次尝试读取数据项中自带的 _ui_color (用于同步核对结果)
            bg_color = self.row_colors.get(i) or row_dict.get('_ui_color')
            
            for j, key in enumerate(self.key_map):
                val = row_dict.get(key, "")
                if isinstance(val, (float, int)) and not isinstance(val, bool):
                    val_str = f"{val:,.2f}" if isinstance(val, float) else str(val)
                else:
                    # Apply global date formatting if it looks like a date
                    val_str = format_date_gui(str(val)) if val is not None else ""
                
                item = QTableWidgetItem(val_str)
                if bg_color:
                    item.setBackground(bg_color)
                
                # Store original data in the first column item
                if j == 0:
                    item.setData(Qt.UserRole, row_dict)
                
                self.table.setItem(i, j, item)
        self.table.blockSignals(False)
        self.table.setSortingEnabled(True) # Re-enable sorting

    def show_context_menu(self, pos):
        item = self.table.itemAt(pos)
        if not item: return
        row = item.row()
        
        menu = QMenu(self)
        
        # Find potential dates in this row
        date_candidates = []
        # Use data from UserRole if available, otherwise fallback to self.data index (unreliable if sorted)
        # Better: rely on UserRole data we just added
        first_item = self.table.item(row, 0)
        record = first_item.data(Qt.UserRole) if first_item else None
        
        if record:
            # Prioritize specific keys
            for k in ["pay_date", "date", "due_date"]:
                val = record.get(k)
                if val and isinstance(val, str) and len(val) >= 8: # Simple check
                    date_candidates.append(val)
                    break # Take the first valid one
        
        if date_candidates:
            date_str = date_candidates[0]
            action_jump = menu.addAction(f"📅 跳转到对应日期 ({date_str})")
            action_jump.triggered.connect(lambda: self.jump_to_date_requested.emit(date_str))
            
        action_del = menu.addAction("🗑️ 删除此行")
        action_del.triggered.connect(self.delete_selected_rows)
        
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def delete_selected_rows(self):
        rows = sorted(set(index.row() for index in self.table.selectedIndexes()), reverse=True)
        if not rows: return
        if QMessageBox.question(self, "确认删除", f"确定删除选中的 {len(rows)} 行吗？") == QMessageBox.Yes:
            for r in rows:
                self.table.removeRow(r)
            self.lbl_count.setText(f"共 {self.table.rowCount()} 条记录")

    def get_data(self):
        new_data = []
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        for r in range(rows):
            # Retrieve original dict to preserve hidden fields (like _ui_color, _source_path)
            item_0 = self.table.item(r, 0)
            original_dict = item_0.data(Qt.UserRole) if item_0 else {}
            
            row_dict = original_dict.copy() if original_dict else {}

            for c in range(cols):
                if c < len(self.key_map):
                    key = self.key_map[c]
                    text = self.table.item(r, c).text().strip()
                    if key in ['amount', 'paid_amount']:
                        try: val = float(text.replace(',', ''))
                        except: val = 0.0
                        row_dict[key] = val
                    else:
                        row_dict[key] = text
            new_data.append(row_dict)
        return new_data

class DataPreviewDialog(QDialog):
    """通用数据预览与编辑弹窗"""
    def __init__(self, title, data, headers, key_map, parent=None, instruction=None, highlight_index=None, row_colors=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(1000, 600)
        
        layout = QVBoxLayout(self)
        self.preview_widget = DataPreviewWidget(data, headers, key_map, instruction, highlight_index, row_colors)
        layout.addWidget(self.preview_widget)
        
        btn_layout = QHBoxLayout()
        
        # Stats Label
        self.stats_label = QLabel("")
        self.stats_label.setStyleSheet("color: #007bff; font-weight: bold; margin-right: 15px;")
        self.stats_label.setTextInteractionFlags(Qt.TextSelectableByMouse) # Enable selection
        btn_layout.addWidget(self.stats_label)
        self.preview_widget.stats_updated.connect(self.stats_label.setText)
        
        btn_layout.addStretch()
        
        btn_export = QPushButton("📤 导出")
        btn_export.clicked.connect(lambda: export_qtable(self.preview_widget.table, self, "preview_export"))
        btn_layout.addWidget(btn_export)
        
        btn_cancel = QPushButton("取消导入")
        btn_cancel.clicked.connect(self.reject)
        btn_confirm = QPushButton("✅ 确认并加载")
        btn_confirm.clicked.connect(self.save_and_accept)
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_confirm)
        layout.addLayout(btn_layout)
        
        self.result_data = []

    def save_and_accept(self):
        try:
            self.result_data = self.preview_widget.get_data()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "数据错误", f"读取表格数据失败: {str(e)}")

class UnifiedPreviewDialog(QDialog):
    """统一数据预览弹窗 (包含多个标签页)"""
    def __init__(self, parent=None, report_data=None, stmt_data=None, 
                 highlight_report_idx=None, highlight_stmt_idx=None,
                 report_colors=None, stmt_colors=None):
        super().__init__(parent)
        self.setWindowTitle("源数据统一预览")
        self.resize(1100, 700)
        
        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)
        
        # Reports Tab
        self.report_widget = None
        if report_data is not None:
            headers = ["客户名称", "发票/参考号", "金额", "到期日", "支付日", "状态"]
            keys = ["name", "invoice_ref", "amount", "due_date", "pay_date", "status"]
            self.report_widget = DataPreviewWidget(report_data, headers, keys, 
                                                 "🔍 收款报告数据 (双击修改)", 
                                                 highlight_report_idx, report_colors)
            self.report_widget.jump_to_date_requested.connect(lambda d: self.handle_jump(d, target='stmt'))
            self.tabs.addTab(self.report_widget, "收款报告 (Baixados)")

        # Statements Tab
        self.stmt_widget = None
        if stmt_data is not None:
            headers = ["交易日期", "描述/备注", "金额", "CNPJ", "银行"]
            keys = ["date", "desc", "amount", "cnpj", "bank"]
            self.stmt_widget = DataPreviewWidget(stmt_data, headers, keys, 
                                               "🔍 银行流水数据 (双击修改)", 
                                               highlight_stmt_idx, stmt_colors)
            self.stmt_widget.jump_to_date_requested.connect(lambda d: self.handle_jump(d, target='report'))
            self.tabs.addTab(self.stmt_widget, "银行流水 (Extrato)")
            
        # Select tab based on highlights
        if highlight_stmt_idx is not None and self.stmt_widget:
            self.tabs.setCurrentWidget(self.stmt_widget)
        elif highlight_report_idx is not None and self.report_widget:
            self.tabs.setCurrentWidget(self.report_widget)
            
        # Buttons
        btn_layout = QHBoxLayout()
        
        # Stats Label
        self.stats_label = QLabel("")
        self.stats_label.setStyleSheet("color: #007bff; font-weight: bold; margin-right: 15px;")
        self.stats_label.setTextInteractionFlags(Qt.TextSelectableByMouse) # Enable selection
        btn_layout.addWidget(self.stats_label)
        
        if self.report_widget:
            self.report_widget.stats_updated.connect(self.stats_label.setText)
        if self.stmt_widget:
            self.stmt_widget.stats_updated.connect(self.stats_label.setText)
        
        btn_layout.addStretch()
        
        btn_export = QPushButton("📤 导出当前表")
        btn_export.clicked.connect(self.export_current_tab)
        
        btn_close = QPushButton("关闭") # Just viewing usually
        btn_close.clicked.connect(self.reject)
        
        btn_save = QPushButton("💾 保存修改")
        btn_save.clicked.connect(self.save_and_accept)
        
        btn_layout.addWidget(btn_export)
        btn_layout.addWidget(btn_close)
        btn_layout.addWidget(btn_save)
        layout.addLayout(btn_layout)
        
        self.result_reports = []
        self.result_stmts = []

    def export_current_tab(self):
        current_widget = self.tabs.currentWidget()
        if isinstance(current_widget, DataPreviewWidget):
            idx = self.tabs.currentIndex()
            tab_text = self.tabs.tabText(idx)
            # Sanitize tab text for filename
            safe_name = tab_text.split('(')[0].strip().replace('/', '_')
            export_qtable(current_widget.table, self, f"preview_{safe_name}")

    def handle_jump(self, date_str, target):
        """跳转到另一侧的对应日期 (Priority: Pay Date > Due Date)"""
        target_widget = self.stmt_widget if target == 'stmt' else self.report_widget
        if not target_widget:
            QMessageBox.information(self, "提示", "目标数据为空，无法跳转。")
            return
            
        try:
            # Parse source date
            src_date = safe_parse_date_to_date(date_str)
            if not src_date: raise ValueError("Parse failed")
        except:
            QMessageBox.warning(self, "格式错误", f"无法识别日期格式: {date_str}")
            return
            
        # Search in target data with scoring
        # Score 0 = Exact Primary Match (Best)
        # Score 10+diff = Close Primary Match
        # Score 100 = Exact Secondary Match
        # Score 110+diff = Close Secondary Match
        
        best_idx = -1
        best_score = 99999
        
        primary_key = 'pay_date' if target == 'report' else 'date'
        secondary_key = 'due_date' if target == 'report' else None
        
        data = target_widget.data
        for i, row in enumerate(data):
            # 1. Check Primary Key
            p_val = row.get(primary_key)
            if p_val:
                d_obj = safe_parse_date_to_date(p_val)
                if d_obj:
                    diff = abs((d_obj - src_date).days)
                    
                    score = 9999
                    if diff == 0: score = 0
                    elif diff <= 5: score = 10 + diff
                    
                    if score < best_score:
                        best_score = score
                        best_idx = i
            
            if best_score == 0: break # Found exact match
            
            # 2. Check Secondary Key (only for reports)
            if secondary_key:
                s_val = row.get(secondary_key)
                if s_val:
                    d_obj = safe_parse_date_to_date(s_val)
                    if d_obj:
                        diff = abs((d_obj - src_date).days)
                        
                        score = 9999
                        if diff == 0: score = 100
                        elif diff <= 5: score = 110 + diff
                        
                        if score < best_score:
                            best_score = score
                            best_idx = i
            
        if best_idx != -1:
            self.tabs.setCurrentWidget(target_widget)
            target_widget.table.selectRow(best_idx)
            target_widget.table.scrollToItem(target_widget.table.item(best_idx, 0))
        else:
            QMessageBox.information(self, "未找到", f"在目标列表中未找到 {date_str} 附近 (5天内) 的记录。")

    def save_and_accept(self):
        try:
            if self.report_widget:
                self.result_reports = self.report_widget.get_data()
            if self.stmt_widget:
                self.result_stmts = self.stmt_widget.get_data()
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e))

class ReconciliationWorker(QThread):
    """异步核对工作线程，防止 UI 卡死"""
    progress = Signal(str)
    finished = Signal(list)
    error = Signal(str)

    def __init__(self, report_records, statement_records, mapping_mgr, enable_local_rules=True, tolerance=0.05):
        super().__init__()
        self.report_records = report_records
        self.statement_records = statement_records
        self.mapping_mgr = mapping_mgr
        self.enable_local_rules = enable_local_rules
        self.tolerance = tolerance

    def run(self):
        try:
            # 不使用 deepcopy 以保持对象引用一致，但需小心处理可能发生的 UI 线程冲突
            # 我们在主线程已经准备好了数据，这里主要进行只读计算
            results = self.do_reconciliation()
            self.finished.emit(results)
        except Exception as e:
            import traceback
            self.error.emit(f"核对出错: {str(e)}\n{traceback.format_exc()}")

    def _find_subset_match(self, items, target_amt, max_size=12, tolerance=0.05, max_combos=200000):
        """支持近似匹配的子集和搜索 (升级版：增加贪婪预搜索)"""
        if not items: return None, 0.0
        import time
        from itertools import combinations
        
        # 1. 贪婪预搜索 (针对超多项分拆场景)
        if len(items) > 12:
            search_items = sorted(items, key=lambda x: x['amount'], reverse=True)
            current_sum = 0
            greedy_subset = []
            for item in search_items:
                if current_sum + item['amount'] <= target_amt + tolerance:
                    current_sum += item['amount']
                    greedy_subset.append(item)
                if abs(current_sum - target_amt) <= tolerance:
                    return greedy_subset, current_sum - target_amt
        
        # 2. 标准组合搜索
        search_items = sorted(items, key=lambda x: x['amount'], reverse=True)
        search_limit = 80 if target_amt > 10000 else 45
        search_items = search_items[:search_limit] 
        n = len(search_items)
        best_combo = None
        min_diff = float('inf')
        
        checked = 0
        start_time = time.time()
        
        if sum(x['amount'] for x in search_items) < target_amt - tolerance:
            return None, 0.0

        for size in range(1, min(n, max_size) + 1):
            if checked > max_combos or time.time() - start_time > 30: break

            for combo in combinations(search_items, size):
                checked += 1
                if checked % 20000 == 0:
                    if checked > max_combos or time.time() - start_time > 35: break

                sum_amt = sum(x['amount'] for x in combo)
                diff = abs(sum_amt - target_amt)
                
                if diff < 1e-5: return combo, 0.0 
                
                if diff <= tolerance and diff < min_diff:
                    min_diff = diff
                    best_combo = combo
                    # 如果找到了非常接近的解，提前结束
                    if diff < 0.01: return combo, (sum_amt - target_amt)
            
            if checked > max_combos: break
        
        if best_combo:
            actual_sum = sum(x['amount'] for x in best_combo)
            return best_combo, (actual_sum - target_amt)
        return None, 0.0

    def do_reconciliation(self):
        recon_results = []
        used_statement_indices = set()
        used_report_indices = set()
        
        # --- [优化] 只匹配收入：过滤掉所有负额或零额流水 ---
        original_stmts = self.statement_records
        # 强化噪音过滤：彻底屏蔽余额汇总和系统利息
        # 使用无音标版本以增强鲁棒性
        noise_keywords = [
            "SDO CTA/APL", "REND PAGO APLIC", "RES APLIC AUT", "SDO CTA ANT",
            "SALDO TOTAL", "DISPONIVEL", "SALDO ANTERIOR", "SDO CTA/APL AUTOM",
            "SALDO FINAL", "SALDO INICIAL", "SALDO DO DIA", "SISPAG"
        ]
        
        valid_stmt_indices = []
        for i, s in enumerate(original_stmts):
            amt = s.get('amount', 0)
            raw_desc = s.get('desc', '').upper()
            
            # 简单去音标处理，处理如 DISPONÍVEL -> DISPONIVEL
            import unicodedata
            norm_desc = unicodedata.normalize('NFKD', raw_desc).encode('ascii', 'ignore').decode('ascii').upper()
            
            if amt > 0:
                if any(k in norm_desc for k in noise_keywords): continue
                valid_stmt_indices.append(i)
        
        self.progress.emit(f"预热中... 有效收入流水: {len(valid_stmt_indices)} 条")
        
        for idx in valid_stmt_indices:
            st = original_stmts[idx]
            st.pop('_ui_color', None)
            st['_dt'] = safe_parse_date_to_date(st['date'])
            # 修正异常历史日期 (如 2020 修正为 2025)
            if st['_dt'] and st['_dt'].year < 2024:
                try: st['_dt'] = st['_dt'].replace(year=2025)
                except: pass

            if not st.get('cnpj'):
                # 匹配带格式的 CNPJ/CPF
                found = re.findall(r'(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})|(\d{3}\.\d{3}\.\d{3}-\d{2})', st.get('desc', ''))
                if found: 
                    st['cnpj'] = found[0][0] or found[0][1]
                else:
                    # 匹配纯数字的 CNPJ (14位) 或 CPF (11位)
                    found_plain = re.findall(r'\b\d{14}\b|\b\d{11}\b', st.get('desc', ''))
                    if found_plain: st['cnpj'] = found_plain[0]
            
            st_cnpj_clean = re.sub(r'\D', '', st.get('cnpj', '')) if st.get('cnpj') else ""
            # 剥离银行前缀以提取核心单位名
            raw_desc = st.get('desc', '').upper()
            # 增强正则：支持更多变体如 PIX RECEBIDO, PIX TRANSF, PIX QRS, etc.
            clean_desc = re.sub(r'^(PIX|TED|DOC|BOLETO|RECEBIMENTO|TRANSFERENCIA|RECEBIMENTOS)S?\s+(RECEBIDO|ENVIADO|RECEBIDA|TRANSF|QRS|PAGTO|TIT)?\s*', '', raw_desc)
            st['_std_partner'] = self.mapping_mgr.get_partner_std(st_cnpj_clean) or \
                                 self.mapping_mgr.get_partner_std(clean_desc) or clean_desc

        valid_report_indices = [i for i, r in enumerate(self.report_records) if r.get('amount', 0) > 0]

        for idx in valid_report_indices:
            rep = self.report_records[idx]
            rep.pop('_ui_color', None)
            rep['_dt_pay'] = safe_parse_date_to_date(rep.get('pay_date'))
            rep['_dt_due'] = safe_parse_date_to_date(rep.get('due_date'))
            rep['_dt_best'] = rep['_dt_pay'] or rep['_dt_due']
            # 修正报表年份异常
            if rep['_dt_best'] and rep['_dt_best'].year < 2024:
                try: rep['_dt_best'] = rep['_dt_best'].replace(year=2025)
                except: pass

            std_p = self.mapping_mgr.get_partner_std(rep.get('name', '')) or rep.get('name', '')
            if "SHPP" in rep.get('name', '').upper() and "NORTE" not in std_p.upper():
                std_p = "NORTETOOLS (SHPP)"
            rep['_std_partner'] = std_p
            # 提取来源上下文 (文件名中的关键账户信息)
            source_file = str(rep.get('_source_file', '')).upper()
            rep['_source_context'] = ""
            for key in ["ARGOTECH", "MARTELUX", "DEYUN", "CAIYA", "NORTETOOLS", "COMPBRASIL", "ZHEMAX", "GEOLOC"]:
                if key in source_file:
                    rep['_source_context'] = key
                    break

        # --- Phase 0: Quick FIFO Exact Match (High Speed) ---
        self.progress.emit("阶段 0: 快速匹配完全一致金额...")
        for r_idx in valid_report_indices:
            if r_idx in used_report_indices: continue
            rep = self.report_records[r_idx]
            dt_rep = rep['_dt_best']
            if not dt_rep: continue
            
            # 搜索金额完全一致且日期在 3 天内的
            for s_idx in valid_stmt_indices:
                if s_idx in used_statement_indices: continue
                st = original_stmts[s_idx]
                if abs(st['amount'] - rep['amount']) < 0.001:
                    date_diff = abs((st['_dt'] - dt_rep).days) if st['_dt'] else 999
                    if date_diff <= 3:
                        used_report_indices.add(r_idx)
                        used_statement_indices.add(s_idx)
                        
                        ref_info = CollectionReportParser.parse_invoice_reference(rep['invoice_ref'], self.enable_local_rules)
                        clean_num = ref_info.get('invoice') if ref_info else CollectionReportParser.clean_invoice_number(rep['invoice_ref'], self.enable_local_rules)

                        recon_results.append({
                            "report": rep, "statement": st, "type": "STRONG",
                            "note": "⚡ 快速精确匹配", "s_idx": s_idx,
                            "ref_info": ref_info, "clean_num": clean_num
                        })
                        break

        # --- Phase 1: 1-to-1 Strong Match (Enhanced with Interest Bias) ---
        self.progress.emit("阶段 1: 正在进行 1对1 智能核对...")
        for r_idx in valid_report_indices:
            if r_idx in used_report_indices: continue
            rep = self.report_records[r_idx]
            rep_amt = rep['amount']
            dt_rep = rep['_dt_best']
            if not dt_rep: continue
            
            ref_info = CollectionReportParser.parse_invoice_reference(rep['invoice_ref'], self.enable_local_rules)
            clean_num = ref_info.get('invoice') if ref_info else CollectionReportParser.clean_invoice_number(rep['invoice_ref'], self.enable_local_rules)

            best_s_idx = -1
            best_score = -1
            
            for s_idx in valid_stmt_indices:
                if s_idx in used_statement_indices: continue
                st = original_stmts[s_idx]
                st_amt = st['amount']
                st_desc_upper = st.get('desc', '').upper()
                
                # --- [升级] 利息自适应容差逻辑 (Late Fee Logic) ---
                is_settlement = any(k in st_desc_upper for k in ["REDE VISA", "REDE MAST", "CIELO", "GETNET", "STONE", "PAGSEGURO"])
                
                is_amt_match = False
                current_tol = self.tolerance
                
                # 基础容差
                if abs(st_amt - rep_amt) <= current_tol:
                    is_amt_match = True
                
                # 滞纳金逻辑 (Late Fee/Interest):
                # 如果流水金额 > 报告金额，且支付日期晚于到期日，放宽正向容差
                days_late = (st['_dt'] - dt_rep).days if st['_dt'] else 0
                if st_amt > rep_amt:
                    # 默认允许 5% 的正向溢价
                    max_surplus_pct = 0.05
                    # 如果确认迟付，且迟付时间较长，允许更高的滞纳金 (例如 10% 或更多)
                    if days_late > 5:
                        # 估算：罚款 2% + 利息 1%/月。3个月约 5%，6个月约 8%。
                        # 放宽至 12% 以覆盖大多数情况
                        max_surplus_pct = 0.12
                    
                    if (st_amt - rep_amt) / rep_amt <= max_surplus_pct:
                        if not is_settlement: is_amt_match = True
                
                # 结算扣费逻辑
                elif is_settlement and rep_amt > st_amt and (rep_amt - st_amt) / rep_amt <= 0.045:
                    is_amt_match = True # 结算类允许 4.5% 负向偏差
                
                if not is_amt_match: continue
                
                sim = calculate_similarity(rep['_std_partner'], st['_std_partner'])
                is_pix = "PIX" in st_desc_upper
                
                if is_pix and sim < 0.6:
                    rep_keywords = [k for k in re.split(r'[\s\-]+', rep['_std_partner']) if len(k) > 3]
                    if any(k.upper() in st_desc_upper for k in rep_keywords): sim = 0.8
                
                if ("SHPP" in st['desc'].upper() and "NORTE" in rep['_std_partner'].upper()): sim = max(sim, 0.9)
                
                date_diff = abs(days_late)
                
                # --- [升级] 动态时间窗口逻辑 ---
                is_exact_amt = abs(st_amt - rep_amt) < 0.01
                # 针对已知延迟严重的单位或金额完全一致的项放宽限制
                max_days = 65 
                if any(k in st['desc'].upper() or k in rep['_std_partner'].upper() for k in ["NORTE", "SHPP", "DANPLER", "ARGOTECH", "GENIOBOX", "AMERICA"]):
                    max_days = 90
                
                # 如果金额完全一致且相似度极高，允许极长时间跨度 (180天)
                if is_exact_amt and sim >= 0.85:
                    max_days = 180
                
                score = 0
                if date_diff > max_days:
                    # 如果金额完全一致但日期超限，记录为 SUSPECT 候选
                    if is_exact_amt and sim >= 0.7 and date_diff <= 220:
                        score = max(score, 30) # 给予基础分以进入后续判断
                    else:
                        continue
                
                ref_match = clean_num and clean_num in st.get('desc', '') if clean_num else False
                if ref_match: score += 85
                if sim >= 0.85: score += 40
                elif sim >= 0.6: score += 20
                elif sim >= 0.3: score += 10 # 增加中等相似度的基础分
                
                # 金额完全一致且日期接近，给予极大加分
                if is_exact_amt and date_diff <= 7: score += 50
                # 金额一致但日期较远，给予适度分数
                elif is_exact_amt: score += 35
                
                # 滞纳金匹配加分: 仅在已有一定匹配基础(名称或参考号)的情况下作为辅助信号
                # 防止纯金额接近但单位完全无关的项目误匹配
                if st_amt > rep_amt and days_late > 0:
                    if score >= 10: # 必须至少有名称相似度或参考号匹配
                        score += 25 # 降低加分幅度 (45 -> 25)

                if is_settlement and score < 30: score += 30
                score -= date_diff * 0.5 # 减小日期惩罚权重
                
                if score > best_score and score >= 35:
                    best_score = score
                    best_s_idx = s_idx
            
            if best_s_idx != -1:
                used_report_indices.add(r_idx)
                used_statement_indices.add(best_s_idx)
                
                final_st = original_stmts[best_s_idx]
                note = ""
                has_amt_diff = abs(final_st['amount'] - rep_amt) > 0.01
                date_diff = abs((dt_rep - final_st['_dt']).days) if final_st['_dt'] else 999
                days_late = (final_st['_dt'] - dt_rep).days if final_st['_dt'] else 0
                
                # 重新计算结算标志以确保备注正确
                st_desc_final = final_st.get('desc', '').upper()
                is_settlement_final = any(k in st_desc_final for k in ["REDE VISA", "REDE MAST", "CIELO", "GETNET", "STONE", "PAGSEGURO"])

                if final_st['amount'] > rep_amt + 0.01:
                    diff_val = final_st['amount'] - rep_amt
                    if days_late > 0:
                        note = f"⚠️ 包含滞纳金/利息 (延迟 {days_late} 天, 差额: {diff_val:.2f})"
                    else:
                        note = f"⚠️ 疑似含利息 (差额: {diff_val:.2f})"
                elif is_settlement_final and rep_amt > final_st['amount'] + 0.01:
                    note = f"💳 刷卡结算 (手续费约: {rep_amt - final_st['amount']:.2f})"
                
                # 只有金额完全一致(或明确是滞纳金)、日期在合理范围且分数达标才算 STRONG
                # 如果包含滞纳金 (score加分后)，也视为 STRONG
                is_strong = (best_score >= 80 and date_diff <= 65)
                m_type = "STRONG" if is_strong else "MEDIUM"
                
                # 判定为 SUSPECT (疑似) 的条件
                if not is_strong and not has_amt_diff and date_diff > 90:
                    m_type = "SUSPECT"
                    note = f"❓ 滞后到账 (延迟 {date_diff} 天)"
                
                recon_results.append({
                    "report": rep, "statement": final_st,
                    "type": m_type,
                    "ref_info": ref_info, "clean_num": clean_num,
                    "s_idx": best_s_idx, "note": note or final_st['desc']
                })

        # --- Phase 1.5: Affinity Pre-Lock (Keyword Enhanced) ---
        self.progress.emit("阶段 1.5: 正在预锁定强相关流水...")
        for s_idx in valid_stmt_indices:
            if s_idx in used_statement_indices: continue
            st = original_stmts[s_idx]
            st_desc = st['desc'].upper()
            st_std = st['_std_partner'].upper()
            
            target_unit = None
            if "SHPP" in st_desc: target_unit = "NORTE"
            elif "DANPLER" in st_desc: target_unit = "DANPLER"
            elif "ARGOTECH" in st_desc: target_unit = "ARGOTECH"
            elif len(st_std) > 5 and not any(k in st_std for k in ["PIX RECEBIDO", "BOLETOS", "EXTRATO"]):
                target_unit = st_std
            
            if not target_unit: continue
            
            for r_idx in valid_report_indices:
                if r_idx in used_report_indices: continue
                rep = self.report_records[r_idx]
                rep_std = rep['_std_partner'].upper()
                if target_unit in rep_std or rep_std in target_unit:
                    date_diff = abs((st['_dt'] - rep['_dt_best']).days) if st['_dt'] and rep['_dt_best'] else 999
                    if date_diff > 65: continue 
                    
                    is_settlement = any(k in st_desc for k in ["REDE VISA", "REDE MAST", "CIELO", "GETNET", "STONE"])
                    diff = st['amount'] - rep['amount']
                    tol_up = rep['amount'] * 0.05
                    tol_down = rep['amount'] * 0.045 if is_settlement else self.tolerance
                    
                    if -tol_down <= diff <= tol_up:
                        used_report_indices.add(r_idx)
                        used_statement_indices.add(s_idx)
                        note = "🎯 强相关单位预锁定"
                        if diff > self.tolerance: note = f"⚠️ 疑似含利息 (差额: {diff:,.2f})"
                        elif diff < -self.tolerance: note = f"💳 结算扣费 (差额: {diff:,.2f})"
                        
                        recon_results.append({
                            "report": rep, "statement": st, "type": "STRONG",
                            "note": note,
                            "s_idx": s_idx
                        })
                        break

        # --- Phase 2: N-to-1 Aggregate Match (Source Context Aware) ---
        self.progress.emit("阶段 2: 正在搜索聚合入账组合...")
        remaining_stmts = [i for i in valid_stmt_indices if i not in used_statement_indices]
        remaining_stmts.sort(key=lambda i: original_stmts[i]['_dt'] or datetime.min)

        for s_idx in remaining_stmts:
            st = original_stmts[s_idx]
            dt_st = st['_dt']
            if not dt_st: continue
            
            st_amt = st['amount']
            desc_upper = st.get('desc', '').upper()
            st_std = st.get('_std_partner', '').upper()
            is_settlement = any(k in desc_upper for k in ["REDE VISA", "REDE MAST", "CIELO", "GETNET", "STONE"])
            
            # 识别流水是否指定了 "来源账户" (e.g., RECEBIMENTOS ARGOTECH)
            target_source_context = None
            for key in ["ARGOTECH", "MARTELUX", "DEYUN", "CAIYA", "NORTETOOLS", "COMPBRASIL", "ZHEMAX", "GEOLOC"]:
                if key in desc_upper or key in st_std:
                    target_source_context = key
                    break

            # --- [升级] 大额聚合特殊配置 ---
            is_large = st_amt > 20000
            search_days_back = 95 if (is_large or any(k in desc_upper for k in ["NORTE", "PALACIO", "DEYUN", "CAIYA", "ARGOTECH", "GEOLOC", "SHPP", "MARTELUX", "DANPLER"])) else 45
            max_subset = 25 if is_large else 15
            
            candidates = []
            for r_idx in valid_report_indices:
                if r_idx in used_report_indices: continue
                rep = self.report_records[r_idx]
                rep_std = rep.get('_std_partner', '').upper()
                
                # 如果流水指定了账户 (Source Context)，则优先匹配该账户来源的报告，忽略单位名称
                is_source_match = False
                if target_source_context and rep.get('_source_context') == target_source_context:
                    is_source_match = True
                
                if "SHPP" in desc_upper or "NORTE" in desc_upper:
                    if not ("NORTE" in rep_std or "SHOPEE" in rep_std or "SHPP" in rep_std): continue
                
                is_shopee_norte = ("SHPP" in desc_upper and "NORTE" in rep_std)
                keyword_match = any(word in rep_std for word in re.split(r'[\s\-]+', st_std) if len(word) > 4) if len(st_std) > 4 else False

                # [优化] 检查原始描述是否为银行聚合流水（无具体单位名）
                is_generic_batch = any(k in desc_upper for k in ["BOLETOS RECEBIDOS", "BOLETO RECEBIDO", "PIX RECEBIDO", "RECEBIMENTOS", "TRANSFERENCIA RECEBIDA"])

                if not is_source_match and not is_generic_batch:
                    if st_std and not any(k in st_std for k in ["BOLETOS RECEBIDOS", "PIX RECEBIDO", "EXTRATO", "TRANSFERENCIA", "REDE VISA", "REDE MAST"]):
                        if st_std not in rep_std and rep_std not in st_std and not is_shopee_norte and not keyword_match: continue
                
                dt_rep = rep['_dt_best']
                # 聚合入账通常发生在支付后一定窗口内，允许少量跨日误差
                if dt_rep and -15 <= (dt_st - dt_rep).days <= search_days_back:
                    candidates.append(rep)
            
            if not candidates: continue
            candidates.sort(key=lambda x: abs((dt_st - x['_dt_best']).days) if x.get('_dt_best') else 999)
            
            # 动态容差：大额聚合允许 0.1% 的金额误差 (原 0.02%)
            current_batch_tol = self.tolerance
            if is_settlement: current_batch_tol = max(self.tolerance, st['amount'] * 0.045)
            if is_large: current_batch_tol = max(current_batch_tol, st_amt * 0.001) 
            
            best_subset, batch_diff = self._find_subset_match(candidates, st['amount'], max_size=max_subset, tolerance=current_batch_tol)
            
            if best_subset:
                used_statement_indices.add(s_idx)
                is_perfect = abs(batch_diff) <= self.tolerance
                for rep in best_subset:
                    for ridx, robj in enumerate(self.report_records):
                        if robj is rep:
                            used_report_indices.add(ridx)
                            break
                    
                    # 细化备注
                    note = f"聚合入账 (误差 {batch_diff:,.2f})"
                    if not is_perfect:
                        if batch_diff > self.tolerance:
                            # Reports > Statement => 收少了 => 结算扣费或折扣
                            tag = "💳 结算扣费" if is_settlement else "🧾 折扣/差异"
                            note = f"{tag} (差额: {-batch_diff:,.2f})"
                        elif batch_diff < -self.tolerance:
                            # Reports < Statement => 收多了 => 利息
                            note = f"⚠️ 疑似含利息 (差额: {-batch_diff:,.2f})"
                    
                    if target_source_context:
                        note += f" [账户匹配: {target_source_context}]"

                    recon_results.append({
                        "report": rep, "statement": st,
                        "type": "BATCH" if is_perfect else "PARTIAL",
                        "note": note,
                        "batch_diff": batch_diff,
                        "s_idx": s_idx
                    })

        # --- Phase 3: 1-to-N Split Payment Match ---
        self.progress.emit("阶段 3: 正在分析大规模分拆支付项...")
        remaining_reps_indices = [i for i in valid_report_indices if i not in used_report_indices]
        remaining_reps_indices.sort(key=lambda i: self.report_records[i]['amount'], reverse=True)

        for r_idx in remaining_reps_indices:
            rep = self.report_records[r_idx]
            if rep['amount'] < 100: continue
            remaining_stmts_objs = [original_stmts[i] for i in valid_stmt_indices if i not in used_statement_indices]
            rep_amt = rep['amount']
            is_ultra = rep_amt > 30000
            max_w = 90 if (is_ultra or "NORTE" in rep['_std_partner'].upper()) else 45
            
            candidates = []
            rep_std = rep['_std_partner'].upper()
            rep_keywords = [k for k in re.split(r'[\s\-]+', rep_std) if len(k) > 4]

            for s in remaining_stmts_objs:
                # 分拆支付可能跨度很大
                if not s['_dt'] or abs((s['_dt'] - rep['_dt_best']).days) > max_w: continue
                s_std = s.get('_std_partner', '').upper()
                s_desc = s['desc'].upper()
                
                is_match = (s_std in rep_std or rep_std in s_std)
                if not is_match and any(k in s_desc for k in rep_keywords): is_match = True
                if "SHPP" in s_desc and "NORTE" in rep_std: is_match = True
                
                has_clear_unit = not any(k in s_std for k in ["BOLETOS RECEBIDOS", "PIX RECEBIDO", "EXTRATO", "TRANSFERENCIA"]) and len(s_std) > 3
                
                if is_match or not has_clear_unit:
                    candidates.append(s)
            
            if not candidates: continue
            # 升级动态容差：大额分拆允许 0.05% 的总额误差
            search_tol = max(self.tolerance, rep_amt * 0.0005 if is_ultra else 0.5)
            best_stmt_subset, stmt_diff = self._find_subset_match(candidates, rep_amt, max_size=40, tolerance=search_tol, max_combos=300000)
            
            if best_stmt_subset:
                used_report_indices.add(r_idx)
                is_perfect = abs(stmt_diff) <= self.tolerance
                for st_obj in best_stmt_subset:
                    found_s_idx = -1
                    for s_idx, s_orig in enumerate(original_stmts):
                        if s_orig is st_obj:
                            found_s_idx = s_idx
                            used_statement_indices.add(s_idx)
                            break
                    recon_results.append({
                        "report": rep, "statement": st_obj,
                        "type": "SPLIT" if is_perfect else "SPLIT_PARTIAL",
                        "note": f"分拆匹配 (误差 {stmt_diff:,.2f})",
                        "batch_diff": stmt_diff,
                        "s_idx": found_s_idx
                    })

        # --- Phase 3.5: Generic Batch Statement Match (BOLETOS RECEBIDOS 无单位名) ---
        self.progress.emit("阶段 3.5: 匹配无单位名的银行聚合流水...")
        generic_batch_keywords = ["BOLETOS RECEBIDOS", "BOLETO RECEBIDO"]

        remaining_generic_stmts = []
        for s_idx in valid_stmt_indices:
            if s_idx in used_statement_indices: continue
            st = original_stmts[s_idx]
            desc_upper = st.get('desc', '').upper()
            # 识别无单位名的聚合流水
            if any(k in desc_upper for k in generic_batch_keywords):
                remaining_generic_stmts.append(s_idx)

        for s_idx in remaining_generic_stmts:
            st = original_stmts[s_idx]
            dt_st = st['_dt']
            st_amt = st['amount']
            if not dt_st or st_amt < 100: continue

            # [优化] 大额聚合使用更宽的时间窗口
            is_large = st_amt > 50000
            search_window_back = 10 if is_large else 5   # 流水日期之前
            search_window_forward = 15 if is_large else 10  # 流水日期之后

            # 寻找支付日期在时间窗口内的报告
            candidates = []
            for r_idx in valid_report_indices:
                if r_idx in used_report_indices: continue
                rep = self.report_records[r_idx]
                dt_pay = rep.get('_dt_pay') or rep.get('_dt_best')
                if not dt_pay: continue

                # 支付日期在流水日期前后的窗口内
                day_diff = (dt_st - dt_pay).days
                if -search_window_back <= day_diff <= search_window_forward:
                    candidates.append(rep)

            if not candidates: continue

            # [优化] 大额聚合使用更大的容差 (允许包含利息或扣费)
            current_batch_tol = max(self.tolerance, st_amt * 0.02) if is_large else max(self.tolerance, st_amt * 0.01)
            best_subset, batch_diff = self._find_subset_match(candidates, st_amt, max_size=35, tolerance=current_batch_tol)

            if best_subset:
                used_statement_indices.add(s_idx)
                is_perfect = abs(batch_diff) <= self.tolerance

                for rep in best_subset:
                    for ridx, robj in enumerate(self.report_records):
                        if robj is rep:
                            used_report_indices.add(ridx)
                            break

                    note = f"🏦 银行聚合入账 (误差 {batch_diff:,.2f})"
                    if not is_perfect:
                        if batch_diff > self.tolerance:
                            note = f"💳 银行聚合 (扣费: {-batch_diff:,.2f})"
                        elif batch_diff < -self.tolerance:
                            note = f"⚠️ 银行聚合 (含利息: {-batch_diff:,.2f})"

                    recon_results.append({
                        "report": rep, "statement": st,
                        "type": "BATCH" if is_perfect else "PARTIAL",
                        "note": note,
                        "batch_diff": batch_diff,
                        "s_idx": s_idx
                    })

        # --- Phase 4: M-to-N Cluster Match (Total Sum per Partner) ---
        self.progress.emit("阶段 4: 正在进行多对多(M:N)汇总匹配...")
        
        # 1. Group remaining items by Standard Partner
        rem_reps_by_partner = {}
        for i, r in enumerate(self.report_records):
            if i not in used_report_indices:
                p = r.get('_std_partner', '').upper().strip()
                if not p or len(p) < 3: continue
                if p not in rem_reps_by_partner: rem_reps_by_partner[p] = []
                rem_reps_by_partner[p].append(i)
        
        rem_stmts_by_partner = {}
        for i, s in enumerate(original_stmts):
            if i not in used_statement_indices and s['amount'] > 0:
                p = s.get('_std_partner', '').upper().strip()
                if not p or len(p) < 3: continue
                # Skip generic names for cluster matching safety
                if any(k in p for k in ["BOLETOS", "PIX RECEBIDO", "EXTRATO", "TRANSFERENCIA", "CLIENTE"]): continue
                
                if p not in rem_stmts_by_partner: rem_stmts_by_partner[p] = []
                rem_stmts_by_partner[p].append(i)
        
        # [优化] 扩展单位名匹配：对 NORTE/NORTETOOLS 等进行关键词聚合
        # 将流水中含有特定关键词的归入同一组
        keyword_groups = {
            "NORTETOOLS": ["NORTE", "NORTETOOLS", "SHPP"],
            "PALACIO": ["PALACIO"],
            "DANPLER": ["DANPLER"],
        }

        for kw_group, keywords in keyword_groups.items():
            # 找到报告中包含这些关键词的
            matching_rep_indices = []
            for partner, indices in rem_reps_by_partner.items():
                if any(kw in partner for kw in keywords):
                    matching_rep_indices.extend(indices)

            # 找到流水中包含这些关键词的（检查原始描述）
            matching_stmt_indices = []
            for s_idx in valid_stmt_indices:
                if s_idx in used_statement_indices: continue
                st = original_stmts[s_idx]
                if st['amount'] <= 0: continue
                desc_upper = st.get('desc', '').upper()
                std_upper = st.get('_std_partner', '').upper()
                if any(kw in desc_upper or kw in std_upper for kw in keywords):
                    matching_stmt_indices.append(s_idx)

            if not matching_rep_indices or not matching_stmt_indices: continue

            total_rep_amt = sum(self.report_records[i]['amount'] for i in matching_rep_indices)
            total_stmt_amt = sum(original_stmts[i]['amount'] for i in matching_stmt_indices)
            diff = total_stmt_amt - total_rep_amt

            # 对于大客户，允许较大的差异（可能包含预付款或其他调整）
            is_match = abs(diff) / (total_rep_amt or 1) < 0.15  # 15% 容差

            if is_match and len(matching_rep_indices) > 0:
                self.progress.emit(f"关键词组 [{kw_group}]: 报告 {len(matching_rep_indices)} 笔 {total_rep_amt:,.2f} vs 流水 {len(matching_stmt_indices)} 笔 {total_stmt_amt:,.2f}")

                for ri in matching_rep_indices: used_report_indices.add(ri)
                for si in matching_stmt_indices: used_statement_indices.add(si)

                target_s_idx = sorted(matching_stmt_indices, key=lambda i: original_stmts[i]['amount'], reverse=True)[0]
                target_stmt = original_stmts[target_s_idx]

                note = f"🔗 大客户汇总匹配 [{kw_group}] (差额 {diff:,.2f})"

                for ri in matching_rep_indices:
                    rep = self.report_records[ri]
                    recon_results.append({
                        "report": rep, "statement": target_stmt,
                        "type": "BATCH",
                        "note": note,
                        "batch_diff": diff,
                        "s_idx": target_s_idx
                    })

        # 2. Try to match totals (精确单位名匹配)
        common_partners = set(rem_reps_by_partner.keys()) & set(rem_stmts_by_partner.keys())

        for partner in common_partners:
            r_indices = rem_reps_by_partner[partner]
            s_indices = rem_stmts_by_partner[partner]
            
            # Sum totals
            total_rep_amt = sum(self.report_records[i]['amount'] for i in r_indices)
            total_stmt_amt = sum(original_stmts[i]['amount'] for i in s_indices)
            
            diff = total_stmt_amt - total_rep_amt
            
            # Rule: Match if difference is small (exact or small fee/interest)
            # Tolerance: 0.05 absolute, or up to 2% if looks like fees/interest
            is_match = False
            note = ""
            
            if abs(diff) <= self.tolerance:
                is_match = True
                note = f"M:N 汇总完美匹配 (误差 {diff:.2f})"
            elif abs(diff) < 5.0 or abs(diff) / (total_rep_amt or 1) < 0.02:
                is_match = True
                note = f"M:N 汇总近似匹配 (差额 {diff:.2f})"
            
            if is_match:
                # Mark all as used
                for ri in r_indices: used_report_indices.add(ri)
                for si in s_indices: used_statement_indices.add(si)
                
                # Create result entries
                # We link ALL reports to the FIRST statement visually, but note implies Group match
                # Ideally we should link N reports to M statements. 
                # The current data structure is 1 Report -> 1 Statement (or None).
                # So we map each Report to the *first* statement of the group (or distribute them?)
                # To keep it simple: Map each Report to the Statement with closest amount? 
                # Or just map all to the first/largest statement to show they are accounted for.
                
                # Let's map all Reports to the Largest Statement in the group for visibility
                target_s_idx = sorted(s_indices, key=lambda i: original_stmts[i]['amount'], reverse=True)[0]
                target_stmt = original_stmts[target_s_idx]
                
                # 1) 所有报告挂到目标流水，保持结果可视化与组匹配语义一致
                for ri in r_indices:
                    rep = self.report_records[ri]
                    recon_results.append({
                        "report": rep, "statement": target_stmt,
                        "type": "BATCH", # Reuse BATCH for M:N
                        "note": note + f" [含 {len(s_indices)} 笔流水]",
                        "batch_diff": diff, # Global diff
                        "s_idx": target_s_idx
                    })
                
                # 2) 其余流水复用最后一条报告，确保流水侧能标记为已匹配
                for si in s_indices:
                    if si == target_s_idx: continue
                    last_rep = self.report_records[r_indices[-1]]
                    recon_results.append({
                        "report": last_rep, "statement": original_stmts[si],
                        "type": "BATCH",
                        "note": "M:N 关联流水",
                        "s_idx": si
                    })

        # --- Final cleanup: Ensure all reports are listed and have reference info ---
        existing_rep_ids = set(id(res['report']) for res in recon_results)
        for r_idx in valid_report_indices:
            rep = self.report_records[r_idx]
            if id(rep) not in existing_rep_ids:
                recon_results.append({
                    "report": rep, 
                    "statement": None, 
                    "type": "NONE"
                })
        
        # [核心修复] 统一补全所有条目的解析信息
        for res in recon_results:
            rep = res['report']
            if 'ref_info' not in res or 'clean_num' not in res:
                ref_info = CollectionReportParser.parse_invoice_reference(rep['invoice_ref'], self.enable_local_rules)
                clean_num = ref_info.get('invoice') if ref_info else CollectionReportParser.clean_invoice_number(rep['invoice_ref'], self.enable_local_rules)
                res['ref_info'] = ref_info
                res['clean_num'] = clean_num

        return recon_results

class ManualMatchSelectionDialog(QDialog):
    """人工手动匹配选择对话框"""
    def __init__(self, title, items_with_idx, headers, keys, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(900, 600)
        self.items_with_idx = items_with_idx # list of (orig_idx, data_dict)
        self.keys = keys
        
        layout = QVBoxLayout(self)
        
        # 搜索栏
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("🔍 快速搜索:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入关键字过滤列表...")
        self.search_input.textChanged.connect(self.filter_table)
        search_layout.addWidget(self.search_input)
        layout.addLayout(search_layout)
        
        self.table = QTableWidget()
        self.table.setColumnCount(len(headers) + 1)
        self.table.setHorizontalHeaderLabels(["选择"] + headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        self.table.setRowCount(len(items_with_idx))
        for i, (orig_idx, data) in enumerate(items_with_idx):
            chk_item = QTableWidgetItem()
            chk_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            chk_item.setCheckState(Qt.Unchecked)
            chk_item.setData(Qt.UserRole, orig_idx)
            self.table.setItem(i, 0, chk_item)
            
            for j, key in enumerate(keys):
                val = data.get(key, "")
                if key == 'amount': 
                    display_val = f"{float(val):,.2f}" if val is not None else "0.00"
                elif 'date' in key.lower():
                    display_val = format_date_gui(str(val))
                else:
                    display_val = str(val)
                
                item = QTableWidgetItem(display_val)
                if key == 'amount': item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(i, j + 1, item)
        
        layout.addWidget(self.table)
        
        # 底部按钮
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(QLabel("💡 支持勾选多个项目进行批量匹配 (M:N)"))
        btn_layout.addStretch()
        
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        btn_layout.addWidget(btn_box)
        layout.addLayout(btn_layout)

    def filter_table(self, text):
        text = text.lower().strip()
        for i in range(self.table.rowCount()):
            match = False
            if not text:
                match = True
            else:
                for j in range(self.table.columnCount()):
                    item = self.table.item(i, j)
                    if item and text in item.text().lower():
                        match = True
                        break
            self.table.setRowHidden(i, not match)

    def set_selected_indices(self, indices):
        for i in range(self.table.rowCount()):
            orig_idx = self.table.item(i, 0).data(Qt.UserRole)
            if orig_idx in indices:
                self.table.item(i, 0).setCheckState(Qt.Checked)

    def get_selected_indices(self):
        selected = []
        for i in range(self.table.rowCount()):
            if self.table.item(i, 0).checkState() == Qt.Checked:
                selected.append(self.table.item(i, 0).data(Qt.UserRole))
        return selected

class AutoReconciliationTab(QWidget):
    """自动化对账模块：自动核对银行报告与银行流水"""
    def __init__(self, main_window):
        super().__init__(main_window)
        self.main_window = main_window
        self.db = self.main_window.payment_tab.db
        
        self.report_records = []
        self.statement_records = []
        self.recon_results = [] # 存储对账结果详细信息，用于反查
        
        self.setup_ui()

    def _open_local_db(self):
        db_path = getattr(self.db, "db_path", None)
        if not db_path:
            QMessageBox.warning(self, "模式限制", "该操作仅支持本地数据库模式。请关闭 Network Mode 后重试。")
            return None
        return sqlite3.connect(db_path)

    @staticmethod
    def _to_float(val, default=0.0):
        try:
            if val is None:
                return default
            return float(val)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _to_iso_date(date_str):
        d_obj = safe_parse_date(date_str)
        if d_obj:
            return d_obj.strftime("%Y-%m-%d")
        return datetime.now().strftime("%Y-%m-%d")

    @staticmethod
    def _row_value(item, key, idx):
        if isinstance(item, (dict, sqlite3.Row)):
            try:
                return item[key]
            except (KeyError, IndexError, TypeError):
                pass
        if isinstance(item, (list, tuple)) and idx < len(item):
            return item[idx]
        return None

    @staticmethod
    def _build_invoice_search_patterns(clean_num):
        if not clean_num:
            return []
        patterns = [f"%{clean_num}%"]
        if len(clean_num) == 4 and clean_num.isdigit():
            serie = clean_num[0]
            number = clean_num[1:]
            patterns.append(f"%.{int(serie):03d}.{int(number):03d}")
        elif len(clean_num) > 4 and clean_num.isdigit():
            serie = clean_num[:-3]
            number = clean_num[-3:]
            patterns.append(f"%.{int(serie):03d}.{int(number):03d}")
        return patterns

    def _get_installment_by_id(self, invoice_id, installment_id):
        if not invoice_id or not installment_id:
            return None
        for row in self.db.get_installments(invoice_id):
            try:
                if int(row['id']) == int(installment_id):
                    return row
            except Exception:
                continue
        return None

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # 1. 顶部控制区
        top_ctrl = QGroupBox("对账源数据导入")
        ctrl_layout = QHBoxLayout(top_ctrl)
        
        self.btn_load_reports = QPushButton("📥 1. 导入收款报告 (Baixados)")
        self.btn_load_reports.setStyleSheet("background-color: #f0f0f0; padding: 8px;")
        self.btn_load_reports.clicked.connect(self.load_reports)

        self.btn_load_statements = QPushButton("🏦 2. 导入银行流水 (Extrato)")
        self.btn_load_statements.setStyleSheet("background-color: #f0f0f0; padding: 8px;")
        self.btn_load_statements.clicked.connect(self.load_statements)

        self.btn_preview_data = QPushButton("👁️ 预览源数据")
        self.btn_preview_data.setToolTip("查看已导入的收款报告和银行流水")
        self.btn_preview_data.clicked.connect(lambda: self.open_unified_preview())
        
        self.btn_run_recon = QPushButton("🚀 开始自动智能核对")
        self.btn_run_recon.setStyleSheet("background-color: #5bc0de; color: white; font-weight: bold; padding: 8px;")
        self.btn_run_recon.clicked.connect(self.run_reconciliation)
        
        self.btn_clear_recon = QPushButton("🧹 清空数据")
        self.btn_clear_recon.setToolTip("重置所有已加载的报告和流水数据")
        self.btn_clear_recon.clicked.connect(self.clear_data)
        
        ctrl_layout.addWidget(self.btn_load_reports)
        ctrl_layout.addWidget(self.btn_load_statements)
        ctrl_layout.addWidget(self.btn_preview_data)
        ctrl_layout.addSpacing(20)
        ctrl_layout.addWidget(self.btn_run_recon)
        ctrl_layout.addWidget(self.btn_clear_recon)
        ctrl_layout.addStretch()
        
        layout.addWidget(top_ctrl)
        
        # 2. 状态摘要
        self.lbl_summary = QLabel("等待导入数据... (请先导入回单报告和银行流水)")
        self.lbl_summary.setStyleSheet("color: #666; font-style: italic;")
        layout.addWidget(self.lbl_summary)
        
        # 3. 对账结果显示区域 (使用标签页)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        # Tab 1: 核心对账结果 (基于报告)
        self.tab_recon = QWidget()
        recon_layout = QVBoxLayout(self.tab_recon)
        recon_layout.setContentsMargins(0, 0, 0, 0)
        
        self.table = QTableWidget()
        headers = [
            "选择", "对账状态", "往来单位", "关联发票/参考号", "到期日", "支付日", "报告金额", "流水日期", "流水金额", "差异", "来源文件", "摘要/流水说明"
        ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        
        # 修复：统一使用 FilterHeader 提升功能，并解决调整长度的反向跳动问题
        header = FilterHeader(self.table)
        header.filterChanged.connect(lambda: self.main_window.apply_header_filters(self.table))
        self.table.setHorizontalHeader(header)
        
        # 修复：所有列设为 Interactive，允许自由调整而不产生相互挤压
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setColumnWidth(0, 50) # 选择
        self.table.setColumnWidth(1, 120) # 状态
        self.table.setColumnWidth(2, 200) # 单位
        self.table.setColumnWidth(4, 90) # 到期日
        self.table.setColumnWidth(5, 90) # 支付日
        self.table.setColumnWidth(10, 150) # 来源文件
        
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # 启用右键菜单
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
        recon_layout.addWidget(self.table)
        self.tabs.addTab(self.tab_recon, "核对结果 (基于报告)")
        
        # Tab 2: 银行流水视图 (显示所有流水及匹配状态)
        self.tab_stmt = QWidget()
        stmt_layout = QVBoxLayout(self.tab_stmt)
        stmt_layout.setContentsMargins(0, 0, 0, 0)
        
        self.stmt_table = QTableWidget()
        stmt_headers = ["状态", "交易日期", "流水描述/备注", "金额", "对方CNPJ", "来源文件", "匹配信息 (关联报告)"]
        self.stmt_table.setColumnCount(len(stmt_headers))
        self.stmt_table.setHorizontalHeaderLabels(stmt_headers)
        
        self.stmt_table.setColumnWidth(5, 150) # 来源文件
        
        stmt_header = FilterHeader(self.stmt_table)
        stmt_header.filterChanged.connect(lambda: self.main_window.apply_header_filters(self.stmt_table))
        self.stmt_table.setHorizontalHeader(stmt_header)
        
        self.stmt_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.stmt_table.horizontalHeader().setStretchLastSection(True)
        self.stmt_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.stmt_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # 启用右键菜单 (stmt_table)
        self.stmt_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.stmt_table.customContextMenuRequested.connect(self.show_stmt_context_menu)
        
        stmt_layout.addWidget(self.stmt_table)
        self.tabs.addTab(self.tab_stmt, "银行流水视图 (全量)")

        # 4. 底部操作
        bottom_layout = QHBoxLayout()
        self.btn_batch_confirm = QPushButton("✅ 批量确认入账 (Match & Pay)")
        self.btn_batch_confirm.setEnabled(False)
        self.btn_batch_confirm.clicked.connect(self.batch_confirm_payments)
        
        self.btn_export_recon = QPushButton("📤 导出对账差异表")
        self.btn_export_recon.clicked.connect(lambda: export_multiple_qtables([
            (self.table, "核对差异"),
            (self.stmt_table, "流水详情")
        ], self, "reconciliation_results"))
        
        bottom_layout.addWidget(self.btn_batch_confirm)
        bottom_layout.addStretch()
        bottom_layout.addWidget(self.btn_export_recon)
        layout.addLayout(bottom_layout)
        
        # 注册表格统计
        if hasattr(self.main_window, 'register_table_for_stats'):
            self.main_window.register_table_for_stats(self.table)
            self.main_window.register_table_for_stats(self.stmt_table)

    def load_reports(self):
        paths, _ = QFileDialog.getOpenFileNames(self, "选择收款报告 PDF (Baixados)", "", "PDF Files (*.pdf)")
        if not paths: return
        
        raw_records = []
        for p in paths:
            abs_p = os.path.abspath(p)
            file_name = os.path.basename(p)
            text, _ = extract_text_from_pdf(p)
            records = CollectionReportParser.parse_report(text)
            for r in records:
                r['_source_path'] = abs_p
                r['_source_file'] = file_name
            raw_records.extend(records)
        
        if not raw_records:
            QMessageBox.information(self, "无数据", "未从文件中解析出任何记录。")
            return
            
        # 预览
        headers = ["客户名称", "发票/参考号", "金额", "到期日", "支付日", "状态", "来源文件"]
        keys = ["name", "invoice_ref", "amount", "due_date", "pay_date", "status", "_source_file"]
        
        dlg = DataPreviewDialog("预览收款报告数据", raw_records, headers, keys, self)
        if dlg.exec() == QDialog.Accepted:
            self.report_records = dlg.result_data
            self.btn_load_reports.setText(f"✅ 已加载报告 ({len(self.report_records)} 条)")
            self.update_summary()
            # 自动保存
            if hasattr(self.main_window, 'auto_save_file'):
                self.main_window.save_data_to_file(self.main_window.auto_save_file, include_payment_history=False, pretty=False)

    def load_statements(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "选择银行流水文件 (Extrato)", "",
            "All Supported (*.pdf *.xlsx *.xls);;PDF Files (*.pdf);;Excel Files (*.xlsx *.xls)"
        )
        if not paths: return

        raw_records = []
        for p in paths:
            abs_p = os.path.abspath(p)
            file_name = os.path.basename(p)
            file_ext = os.path.splitext(p)[1].lower()

            if file_ext in ['.xlsx', '.xls']:
                # 解析 Excel 文件
                records = self._parse_excel_statement(abs_p, file_name)
            else:
                # 解析 PDF 文件
                text, _ = extract_text_from_pdf(p)
                records = BankStatementParser.parse_statement(text)
                for r in records:
                    r['_source_path'] = abs_p
                    r['_source_file'] = file_name

            raw_records.extend(records)

        if not raw_records:
            QMessageBox.information(self, "无数据", "未从文件中解析出任何流水记录。")
            return

        # 预览
        headers = ["交易日期", "描述/备注", "金额", "CNPJ", "来源文件"]
        keys = ["date", "desc", "amount", "cnpj", "_source_file"]

        dlg = DataPreviewDialog("预览银行流水数据", raw_records, headers, keys, self)
        if dlg.exec() == QDialog.Accepted:
            self.statement_records = dlg.result_data
            self.btn_load_statements.setText(f"✅ 已加载流水 ({len(self.statement_records)} 条)")
            self.update_summary()
            # 自动保存
            if hasattr(self.main_window, 'auto_save_file'):
                self.main_window.save_data_to_file(self.main_window.auto_save_file, include_payment_history=False, pretty=False)

    def _parse_excel_statement(self, file_path: str, file_name: str) -> list:
        """解析 Excel 格式的银行流水文件

        支持的格式：
        - 标准5列格式：交易日期, 描述/备注, 金额(收入), CNPJ/支出, 来源文件
        - 其他格式通过表头自动识别
        """
        records = []
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = wb.active

            # 读取表头
            headers = []
            for cell in sheet[1]:
                headers.append(str(cell.value or "").strip())

            self.main_window.log_message(f"Excel 表头: {headers}")

            # 判断是否是标准5列格式（交易日期, 描述/备注, 金额, CNPJ, 来源文件）
            is_standard_format = (
                len(headers) >= 4 and
                '日期' in headers[0] and
                '描述' in headers[1]
            )

            # 读取数据行
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(c is None or str(c).strip() == '' for c in row):
                    continue

                record = {
                    'date': '',
                    'desc': '',
                    'amount': 0.0,
                    'cnpj': '',
                    '_source_path': file_path,
                    '_source_file': file_name
                }

                if is_standard_format:
                    # 标准格式: A=日期, B=描述, C=收入, D=支出, E=来源文件
                    # 日期
                    if len(row) > 0 and row[0]:
                        date_val = row[0]
                        if isinstance(date_val, (int, float)):
                            record['date'] = str(int(date_val))
                        else:
                            record['date'] = str(date_val).strip()

                    # 描述
                    if len(row) > 1 and row[1]:
                        record['desc'] = str(row[1]).strip()

                    # 金额计算: 收入(C列) - 支出(D列)
                    credit = 0.0  # 收入
                    debit = 0.0   # 支出

                    if len(row) > 2 and row[2]:
                        credit = self._parse_amount(row[2])

                    if len(row) > 3 and row[3]:
                        # D列在表头写的是CNPJ，但实际数据是支出金额
                        val = row[3]
                        parsed = self._parse_amount(val)
                        if parsed > 0:
                            debit = parsed
                        elif isinstance(val, str) and re.search(r'\d{14}', re.sub(r'\D', '', val)):
                            # 确实是CNPJ
                            record['cnpj'] = str(val).strip()

                    # 最终金额：收入为正，支出为负
                    if credit > 0:
                        record['amount'] = credit
                    elif debit > 0:
                        record['amount'] = -debit
                else:
                    # 非标准格式，尝试智能解析
                    for idx, val in enumerate(row):
                        if val is None:
                            continue
                        val_str = str(val).strip()

                        # 日期识别（YYYYMMDD 或 DD/MM/YYYY）
                        if not record['date'] and re.match(r'^\d{8}$', val_str):
                            record['date'] = val_str
                        elif not record['date'] and re.match(r'^\d{2}/\d{2}/\d{4}$', val_str):
                            record['date'] = val_str

                        # 金额识别
                        elif idx > 0 and record['amount'] == 0.0:
                            amt = self._parse_amount(val)
                            if amt != 0.0:
                                record['amount'] = amt

                        # 描述识别（最长的字符串）
                        elif len(val_str) > len(record['desc']) and not val_str.replace('.', '').replace(',', '').isdigit():
                            record['desc'] = val_str

                # 跳过余额行和空记录
                desc_upper = record['desc'].upper()
                if record['desc'] and 'SALDO' not in desc_upper and record['amount'] != 0.0:
                    records.append(record)

            wb.close()
            self.main_window.log_message(f"从 Excel 解析出 {len(records)} 条流水记录")

        except Exception as e:
            self.main_window.log_message(f"解析 Excel 流水文件失败: {e}")
            import traceback
            self.main_window.log_message(traceback.format_exc())

        return records

    def _parse_amount(self, value) -> float:
        """解析金额值"""
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        # 处理巴西格式 1.234,56
        s = str(value).strip()
        s = re.sub(r'[R$\s]', '', s)
        if ',' in s and '.' in s:
            s = s.replace('.', '').replace(',', '.')
        elif ',' in s:
            s = s.replace(',', '.')
        try:
            return float(s)
        except:
            return 0.0

    def clear_data(self):
        """清空所有对账数据并重置 UI"""
        self.report_records = []
        self.statement_records = []
        self.table.setRowCount(0)
        self.btn_load_reports.setText("📥 1. 导入收款报告 (Baixados)")
        self.btn_load_statements.setText("🏦 2. 导入银行流水 (Extrato)")
        self.btn_batch_confirm.setEnabled(False)
        self.update_summary()
        self.main_window.log_message("已清空自动化对账模块的数据")
        # 自动保存
        if hasattr(self.main_window, 'auto_save_file'):
            self.main_window.save_data_to_file(self.main_window.auto_save_file, include_payment_history=False, pretty=False)

    def get_state(self):
        """获取当前对账模块的状态数据用于保存"""
        return {
            "report_records": self.report_records,
            "statement_records": self.statement_records
        }

    def set_state(self, data):
        """恢复对账模块的状态数据"""
        if not data: return
        self.report_records = data.get("report_records", [])
        self.statement_records = data.get("statement_records", [])
        
        if self.report_records:
            self.btn_load_reports.setText(f"✅ 已加载报告 ({len(self.report_records)} 条)")
        if self.statement_records:
            self.btn_load_statements.setText(f"✅ 已加载流水 ({len(self.statement_records)} 条)")
        self.update_summary()

    def update_summary(self):
        self.lbl_summary.setText(f"已就绪：报告记录 {len(self.report_records)} 条 | 银行流水记录 {len(self.statement_records)} 条")

    def run_reconciliation(self):
        """核心自动核对算法：开启异步工作线程 (V2.0 线程版)"""
        if not self.report_records:
            QMessageBox.warning(self, "缺失数据", "请先导入回单报告记录")
            return
            
        # 预先分配临时索引，确保在整个核对和 UI 渲染过程中身份稳定 (不受内存地址改变影响)
        for idx, r in enumerate(self.report_records): r['_temp_idx'] = idx
        for idx, s in enumerate(self.statement_records): s['_temp_idx'] = idx

        # 准备进度提示
        self.btn_run_recon.setEnabled(False)
        self.btn_run_recon.setText("⌛ 正在智能核对中...")
        self.main_window.log_message("🚀 启动智能核对工作线程...")
        
        # 准备配置
        settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)
        enable_local = settings.value("enable_local_rules", True, type=bool)
        tolerance = float(settings.value("recon_amount_tolerance", 0.05))
        
        # 创建工作线程 (不再使用 deepcopy 以便直接引用对象)
        self.worker = ReconciliationWorker(
            self.report_records,
            self.statement_records,
            self.main_window.mapping_mgr,
            enable_local,
            tolerance
        )
        
        self.worker.progress.connect(lambda msg: self.main_window.log_message(f"  > {msg}"))
        self.worker.finished.connect(self.on_reconciliation_finished)
        self.worker.error.connect(self.on_reconciliation_error)
        self.worker.start()

    def on_reconciliation_error(self, err_msg):
        self.btn_run_recon.setEnabled(True)
        self.btn_run_recon.setText("🚀 开始自动智能核对")
        QMessageBox.critical(self, "核对失败", err_msg)

    def on_reconciliation_finished(self, recon_results):
        """核对完成后的 UI 渲染逻辑 (索引增强版)"""
        self.btn_run_recon.setEnabled(True)
        self.btn_run_recon.setText("🚀 开始自动智能核对")
        
        self.table.setRowCount(0)
        self.table.setSortingEnabled(False)
        
        # 核心修复 1：预先清除所有颜色，并统一分配稳定索引
        for r in self.report_records: r.pop('_ui_color', None)
        for s in self.statement_records: s.pop('_ui_color', None)
        
        # 获取所有发票的支付状态 (用于标记已入账)
        payment_status_map = {}
        try:
            db_invoices = self.db.get_invoices()
            for row in db_invoices:
                paid_count = row.get('paid_terms_count', 0)
                total_count = row.get('total_terms_count', 0)
                if total_count > 0 and paid_count >= total_count:
                    status = "PAID"
                elif paid_count > 0:
                    status = "PARTIAL"
                else:
                    status = "UNPAID"
                payment_status_map[row['invoice_number']] = status
        except: pass

        for res in recon_results:
            rep = res['report']
            st = res.get('statement')
            res['r_idx'] = rep.get('_temp_idx', -1)
            res['s_idx'] = st.get('_temp_idx', -1) if st else -1

        # 1. 统一定义噪音过滤逻辑 (需与 do_reconciliation 保持一致)
        noise_keywords = [
            "SDO CTA/APL", "REND PAGO APLIC", "RES APLIC AUT", "SDO CTA ANT",
            "SALDO TOTAL", "DISPONIVEL", "SALDO ANTERIOR", "SDO CTA/APL AUTOM",
            "SALDO FINAL", "SALDO INICIAL", "SALDO DO DIA", "SISPAG"
        ]
        
        def is_noise(desc):
            import unicodedata
            norm = unicodedata.normalize('NFKD', str(desc).upper()).encode('ascii', 'ignore').decode('ascii').upper()
            return any(k in norm for k in noise_keywords)

        # 2. 基础数据统计 (汇总和合计)
        total_reports_amt = sum(r.get('amount', 0) for r in self.report_records if r.get('amount', 0) != 0)
        total_stmts_amt = sum(s.get('amount', 0) for s in self.statement_records if s.get('amount', 0) != 0 and not is_noise(s.get('desc', '')))
        
        # 3. 统计分析
        unmatched_reps_objs = [r['report'] for r in recon_results if r['statement'] is None]
        used_stmt_indices = set(r['s_idx'] for r in recon_results if r.get('s_idx', -1) != -1)
        unmatched_stmts_objs = [s for i, s in enumerate(self.statement_records) 
                                if i not in used_stmt_indices and s.get('amount', 0) != 0 and not is_noise(s.get('desc', ''))]
        
        total_rem_rep = sum(r['amount'] for r in unmatched_reps_objs)
        total_rem_stmt = sum(s['amount'] for s in unmatched_stmts_objs)
        
        # 4. 计算已匹配项目的内部差异 (及平均数)
        matched_r_indices = set()
        matched_s_indices = set()
        matched_rep_sum = 0.0
        matched_stmt_sum = 0.0
        
        for res in recon_results:
            if res['type'] == 'NONE': continue
            if res['r_idx'] != -1 and res['r_idx'] not in matched_r_indices:
                matched_rep_sum += res['report']['amount']
                matched_r_indices.add(res['r_idx'])
            if res['s_idx'] != -1 and res['s_idx'] not in matched_s_indices:
                matched_stmt_sum += res['statement']['amount']
                matched_s_indices.add(res['s_idx'])
                
        match_discrepancy = matched_rep_sum - matched_stmt_sum
        avg_matched_rep = (matched_rep_sum / len(matched_r_indices)) if matched_r_indices else 0.0
        
        # 总净差异
        total_gap = total_reports_amt - total_stmts_amt
        
        summary_msg = (
            f"📊 [对账概览 - 汇总与合计]\n"
            f"   • 报告总额: {total_reports_amt:,.2f} | 有效流水总额: {total_stmts_amt:,.2f}\n"
            f"   • 全局净差异: {total_gap:,.2f} (账面总资金缺口)\n"
            f"   --------------------------------------------------\n"
            f"   • 成功匹配: {len(matched_r_indices)} 笔报告 <-> {len(matched_s_indices)} 笔流水\n"
            f"   • 匹配总额: {matched_rep_sum:,.2f} | 平均金额: {avg_matched_rep:,.4f}\n"
            f"   • 匹配内差异: {match_discrepancy:,.2f} (含手续费/利息/聚合误差)\n"
            f"   • 完全未匹配: 报告 {total_rem_rep:,.2f} | 流水 {total_rem_stmt:,.2f}\n"
            f"   --------------------------------------------------"
        )
        
        self.main_window.log_message(summary_msg)
        
        # UI 颜色
        COLOR_STRONG = QColor("#dff0d8") 
        COLOR_MEDIUM = QColor("#e8f5e9")
        COLOR_BATCH = QColor("#d9edf7")
        COLOR_PARTIAL = QColor("#fff3e0") 
        COLOR_SUSPECT = QColor("#fcf8e3")
        COLOR_NONE = QColor("#f2dede")

        self.table.setRowCount(len(recon_results))
        for i, res in enumerate(recon_results):
            rep, st, m_type = res['report'], res['statement'], res['type']
            
            chk_item = QTableWidgetItem()
            chk_item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            chk_item.setCheckState(Qt.Checked if m_type in ["STRONG", "MEDIUM", "BATCH", "PARTIAL", "SPLIT", "SPLIT_PARTIAL"] else Qt.Unchecked)
            self.table.setItem(i, 0, chk_item)

            bg_color = COLOR_NONE
            status_txt = "❌ 缺失流水"
            diff_val = (rep['amount'] - st['amount']) if st else rep['amount']
            stmt_amt_display = f"{st['amount']:,.2f}" if st else "-"

            if m_type == "STRONG": status_txt, bg_color = "✅ 完美匹配", COLOR_STRONG
            elif m_type == "MEDIUM": status_txt, bg_color = "🟢 智能匹配", COLOR_MEDIUM
            elif m_type == "BATCH": 
                status_txt, bg_color = "📦 组合匹配", COLOR_BATCH
                stmt_amt_display += " (组合)"
                diff_val = res.get('batch_diff', 0.0)
            elif m_type == "SPLIT":
                status_txt, bg_color = "✂️ 分拆匹配", COLOR_BATCH
                stmt_amt_display += " (分拆)"
                diff_val = res.get('batch_diff', 0.0)
            elif m_type == "PARTIAL":
                status_txt, bg_color = "🧩 部分匹配 (组合)", COLOR_PARTIAL
                stmt_amt_display += " (组合)"
                diff_val = res.get('batch_diff', 0.0)
            elif m_type == "SPLIT_PARTIAL":
                status_txt, bg_color = "🧩 部分匹配 (分拆)", COLOR_PARTIAL
                stmt_amt_display += " (分拆)"
                diff_val = res.get('batch_diff', 0.0)
            elif m_type == "SUSPECT": status_txt, bg_color = "❓ 疑似匹配", COLOR_SUSPECT

            # 标记关联发票是否存在于系统中 (Check if invoice exists)
            invoice_col_idx = 3 # "关联发票/参考号" is at index 3 in row_data, which is col 3+1 in table
            inv_num = res.get('clean_num') or (res['ref_info'].get('invoice') if res.get('ref_info') else None)
            exists_in_db = False
            p_status = "UNPAID"
            if inv_num:
                exists_in_db = bool(self.db.find_invoice_id_by_number(inv_num))
                p_status = payment_status_map.get(inv_num, "UNPAID")

            status_prefix = ""
            if p_status == "PAID":
                status_prefix = "💰 [已入账] "
            elif p_status == "PARTIAL":
                status_prefix = "⏳ [部分入账] "

            row_data = [
                status_prefix + status_txt, rep.get('_std_partner', 'Unknown'),
                f"{res.get('clean_num', '')} (Raw: {rep['invoice_ref']})" if res.get('clean_num') else rep['invoice_ref'],
                format_date_gui(rep.get('due_date', '')), format_date_gui(rep.get('pay_date', '')),
                f"{rep['amount']:,.2f}", format_date_gui(st['date']) if st else "-",
                stmt_amt_display, f"{diff_val:,.2f}", rep.get('_source_file', '-'),
                (res.get('note') or (st['desc'] if st else "(未匹配)"))
            ]
            for col_idx, val in enumerate(row_data):
                item = QTableWidgetItem(str(val))
                item.setBackground(bg_color)
                
                # 特别标注已存在的发票
                if col_idx == 2 and exists_in_db: # "关联发票" 列
                    font = item.font()
                    font.setBold(True)
                    font.setUnderline(True)
                    item.setFont(font)
                    item.setForeground(QColor("#0056b3")) # 深蓝色
                    item.setToolTip(f"✨ 系统中已存在此发票 ({inv_num})，右键可跳转。")

                if col_idx in [5, 7, 8]: item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table.setItem(i, col_idx + 1, item)
            chk_item.setData(Qt.UserRole, res)

        # 填充流水视图 (核心修复：使用 s_idx 映射)
        stmt_info_map = {}
        for res in recon_results:
            s_idx = res.get('s_idx', -1)
            if s_idx != -1:
                if s_idx not in stmt_info_map: 
                    stmt_info_map[s_idx] = {'type': res['type'], 'count': 0, 'reps': []}
                stmt_info_map[s_idx]['count'] += 1
                stmt_info_map[s_idx]['reps'].append(res['report'].get('name', 'Unknown'))
        
        # 建立类型到颜色的映射，确保两边视图颜色一致
        TYPE_TO_COLOR = {
            "STRONG": COLOR_STRONG,
            "MEDIUM": COLOR_MEDIUM,
            "BATCH": COLOR_BATCH,
            "SPLIT": COLOR_BATCH,
            "PARTIAL": COLOR_PARTIAL,
            "SPLIT_PARTIAL": COLOR_PARTIAL,
            "SUSPECT": COLOR_SUSPECT,
            "NONE": COLOR_NONE
        }

        self.stmt_table.setRowCount(len(self.statement_records))
        for i, st in enumerate(self.statement_records):
            info = stmt_info_map.get(i)
            if info:
                bg_color = TYPE_TO_COLOR.get(info['type'], COLOR_STRONG)
                status_str = f"✅ 已匹配 ({info['type']})"
            else:
                bg_color = QColor("#ffffff")
                status_str = "❌ 未匹配"
            
            st['_ui_color'] = bg_color if info else None
            match_info = (", ".join(info['reps'][:2]) + ("..." if len(info['reps'])>2 else "")) if info else "-"
            row_data = [
                status_str, format_date_gui(st['date']), st['desc'], 
                f"{st['amount']:,.2f}", st.get('cnpj', ''), st.get('_source_file', '-'), match_info
            ]
            for c_idx, val in enumerate(row_data):
                item = QTableWidgetItem(str(val))
                item.setBackground(bg_color)
                self.stmt_table.setItem(i, c_idx, item)
                if c_idx == 0: 
                    item.setData(Qt.UserRole, st)
                    item.setData(Qt.UserRole + 1, st.get('_temp_idx', i)) # 关键：存储原始流水索引，用于聚焦功能

        self.recon_results = recon_results
        self.btn_batch_confirm.setEnabled(len(recon_results) > 0)
        self.auto_save_results(recon_results)
        QMessageBox.information(self, "核对完成", f"对账结束，流水视图已同步更新匹配状态。")

    def manual_match_report(self, recon_item):
        """人工手动匹配报告到流水"""
        if not recon_item: return
        report = recon_item['report']
        
        # 获取所有流水，并标记已使用的
        used_s_indices = set(res.get('s_idx', -1) for res in self.recon_results if res.get('s_idx', -1) != -1)
        # 当前已匹配到此报告的流水索引
        current_s_indices = [res.get('s_idx') for res in self.recon_results if res.get('report') is report and res.get('s_idx', -1) != -1]
        
        # 构造待选列表：未使用的 + 当前已匹配的
        available_stmts_with_idx = []
        for i, st in enumerate(self.statement_records):
            if i not in used_s_indices or i in current_s_indices:
                available_stmts_with_idx.append((i, st))
        
        if not available_stmts_with_idx:
            QMessageBox.information(self, "提示", "没有可用的银行流水（所有流水均已匹配）。")
            return

        dlg = ManualMatchSelectionDialog(f"人工匹配：{report.get('name')} ({report.get('amount', 0):,.2f})", 
                                        available_stmts_with_idx, 
                                        ["日期", "描述/备注", "金额", "来源文件"], 
                                        ["date", "desc", "amount", "_source_file"], 
                                        self)
        dlg.set_selected_indices(current_s_indices)
        
        if dlg.exec() == QDialog.Accepted:
            selected_s_indices = dlg.get_selected_indices()
            
            # 更新数据：移除旧关联，添加新关联
            # 1. 移除此 Report 的所有匹配
            self.recon_results = [res for res in self.recon_results if res.get('report') is not report]
            
            # 2. 如果没有任何选择，视为取消匹配
            if not selected_s_indices:
                self.recon_results.append({"report": report, "statement": None, "type": "NONE"})
            else:
                for s_idx in selected_s_indices:
                    st = self.statement_records[s_idx]
                    # 如果该流水之前匹配了别的 Report，移除那些关联 (确保 1个流水只能对应手动选择的这组)
                    self.recon_results = [res for res in self.recon_results if res.get('s_idx') != s_idx]
                    
                    self.recon_results.append({
                        "report": report,
                        "statement": st,
                        "type": "STRONG", 
                        "note": "👤 人工手动匹配",
                        "s_idx": s_idx,
                        "r_idx": report.get('_temp_idx', -1)
                    })
            
            # 3. 补全可能丢失的 Report 条目 (如果有的话)
            existing_reps = set(id(res['report']) for res in self.recon_results)
            for rep in self.report_records:
                if id(rep) not in existing_reps:
                    self.recon_results.append({"report": rep, "statement": None, "type": "NONE"})

            self.on_reconciliation_finished(self.recon_results)
            self.main_window.log_message(f"✅ 已人工更新报告 '{report.get('name')}' 的匹配关联")

    def manual_match_statement(self, stmt_data):
        """人工手动匹配流水到报告"""
        if not stmt_data: return
        s_idx = stmt_data.get('_temp_idx', -1)
        if s_idx == -1: return

        # 获取所有报告，并标记已使用的
        # 注意：这里我们允许一个报告匹配多个流水，所以 used_r_indices 只是参考
        used_r_indices = set(res.get('r_idx', -1) for res in self.recon_results if res.get('s_idx', -1) != -1)
        # 当前已匹配到此流水的报告索引
        current_r_indices = [res.get('r_idx') for res in self.recon_results if res.get('s_idx') == s_idx]
        
        # 构造待选列表
        available_reps_with_idx = []
        for i, rep in enumerate(self.report_records):
            if i not in used_r_indices or i in current_r_indices:
                available_reps_with_idx.append((i, rep))
                
        dlg = ManualMatchSelectionDialog(f"人工匹配流水：{stmt_data.get('desc')} ({stmt_data.get('amount', 0):,.2f})", 
                                        available_reps_with_idx, 
                                        ["单位名称", "参考号/发票", "金额", "日期"], 
                                        ["name", "invoice_ref", "amount", "due_date"], 
                                        self)
        dlg.set_selected_indices(current_r_indices)
        
        if dlg.exec() == QDialog.Accepted:
            selected_r_indices = dlg.get_selected_indices()
            
            # 1. 移除此 Statement 的所有旧匹配
            self.recon_results = [res for res in self.recon_results if res.get('s_idx') != s_idx]
            
            # 2. 添加新匹配
            for r_idx in selected_r_indices:
                rep = self.report_records[r_idx]
                # 如果该 Report 之前有别的匹配，先移除 (确保手动模式下关系清晰)
                self.recon_results = [res for res in self.recon_results if res.get('report') is not rep]
                
                self.recon_results.append({
                    "report": rep,
                    "statement": stmt_data,
                    "type": "STRONG", 
                    "note": "👤 人工手动匹配",
                    "s_idx": s_idx,
                    "r_idx": r_idx
                })
            
            # 3. 补全 NONE 条目
            existing_reps = set(id(res['report']) for res in self.recon_results)
            enable_local = QSettings(SETTINGS_FILE, QSettings.IniFormat).value("enable_local_rules", True, type=bool)
            for rep in self.report_records:
                if id(rep) not in existing_reps:
                    ref_info = CollectionReportParser.parse_invoice_reference(rep['invoice_ref'], enable_local)
                    clean_num = ref_info.get('invoice') if ref_info else CollectionReportParser.clean_invoice_number(rep['invoice_ref'], enable_local)
                    self.recon_results.append({
                        "report": rep, "statement": None, "type": "NONE",
                        "ref_info": ref_info, "clean_num": clean_num
                    })

            self.on_reconciliation_finished(self.recon_results)
            self.main_window.log_message(f"✅ 已人工更新流水 '{stmt_data.get('desc')}' 的匹配关联")

    def unmatch_recon_item(self, recon_item):
        """取消关联"""
        if not recon_item or recon_item.get('type') == 'NONE': return
        
        rep = recon_item['report']
        st = recon_item.get('statement')
        
        if QMessageBox.question(self, "取消匹配", f"确定要取消此项关联吗？\n\n报告：{rep.get('name')}\n流水：{st.get('desc') if st else '-'}") != QMessageBox.Yes:
            return
            
        # 移除该条目
        if recon_item in self.recon_results:
            self.recon_results.remove(recon_item)
            
            # 检查 Report 是否变为空，如果是，补一个 NONE 项
            has_other = any(res.get('report') is rep for res in self.recon_results)
            if not has_other:
                enable_local = QSettings(SETTINGS_FILE, QSettings.IniFormat).value("enable_local_rules", True, type=bool)
                ref_info = CollectionReportParser.parse_invoice_reference(rep['invoice_ref'], enable_local)
                clean_num = ref_info.get('invoice') if ref_info else CollectionReportParser.clean_invoice_number(rep['invoice_ref'], enable_local)
                self.recon_results.append({
                    "report": rep, "statement": None, "type": "NONE",
                    "ref_info": ref_info, "clean_num": clean_num
                })
        
        self.on_reconciliation_finished(self.recon_results)

    def auto_save_results(self, recon_results):
        """自动保存核对结果到 Excel"""
        if not recon_results: return
        
        # 创建输出目录
        out_dir = "reconciliation_results"
        if not os.path.exists(out_dir):
            try:
                os.makedirs(out_dir)
            except Exception:
                pass
            
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reconciliation_auto_save_{timestamp}.xlsx"
        filepath = os.path.join(out_dir, filename)
        
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reconciliation Results"
            
            # Headers
            headers = ["Type", "Partner", "Ref Info", "Clean Ref", "Report Amount", "Statement Date", "Statement Amount", "Diff", "Desc"]
            ws.append(headers)
            
            for res in recon_results:
                rep = res['report']
                st = res['statement']
                m_type = res['type']
                
                row = [
                    m_type,
                    rep['name'],
                    str(res.get('ref_info', '')),
                    str(res.get('clean_num', '')),
                    rep['amount'],
                    st['date'] if st else "",
                    st['amount'] if st else "",
                    (rep['amount'] - st['amount']) if st else rep['amount'],
                    st['desc'] if st else ""
                ]
                ws.append(row)
                
            wb.save(filepath)
            self.main_window.log_message(f"✅ 核对结果已自动保存至: {filepath}")
            
        except Exception as e:
            self.main_window.log_message(f"❌ 自动保存失败: {str(e)}")

    def show_context_menu(self, pos):
        """对账结果右键菜单 (Tab 1)"""
        item = self.table.itemAt(pos)
        if not item: return
        row = item.row()
        data = self.table.item(row, 0).data(Qt.UserRole)
        
        menu = QMenu(self)
        
        # Jump to Statement View
        if data and data.get('statement'):
            action_jump = menu.addAction("➡️ 跳转到银行流水视图对应项")
            action_jump.triggered.connect(lambda: self.jump_to_stmt_tab(data.get('statement')))
            menu.addSeparator()
            
        # Check for Invoice Match and Add Jump Action
        inv_num = None
        if data:
            if data.get('ref_info') and data['ref_info'].get('invoice'):
                inv_num = data['ref_info']['invoice']
            elif data.get('clean_num'):
                inv_num = data['clean_num']
        
        if inv_num:
            # Verify if invoice exists in DB
            # Refactored for Network Support
            inv_id = self.db.find_invoice_id_by_number(inv_num)
            
            if inv_id:
                action_inv = menu.addAction(f"📄 跳转到发票汇总 (Inv: {inv_num})")
                action_inv.triggered.connect(lambda: self.jump_to_invoice_summary(inv_num))
                
                # 新增：直接跳转到发票明细的选项
                action_detail = menu.addAction("🔍 跳转到发票明细 (Detail)")
                action_detail.triggered.connect(lambda: self.jump_to_invoice_detail(inv_num))
                
                menu.addSeparator()
        
        # New Actions: Find in Preview
        if data:
            action_open_file = menu.addAction("📂 打开来源文件 (PDF)")
            path = data['report'].get('_source_path')
            if path and os.path.exists(path):
                action_open_file.triggered.connect(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)))
            else:
                action_open_file.setEnabled(False)
            
            action_find_rep = menu.addAction("🔎 跳转到收款报告源日期 (Preview)")
            # 找到 report 在原始列表中的索引
            rep_idx = -1
            for idx, r_obj in enumerate(self.report_records):
                if r_obj is data.get('report'):
                    rep_idx = idx
                    break
            action_find_rep.triggered.connect(lambda: self.open_unified_preview(highlight_report_idx=rep_idx))
            
            if data.get('statement'):
                action_find_stmt = menu.addAction("🔎 跳转到银行流水源日期 (Preview)")
                st_idx = -1
                for idx, s_obj in enumerate(self.statement_records):
                    if s_obj is data.get('statement'):
                        st_idx = idx
                        break
                action_find_stmt.triggered.connect(lambda: self.open_unified_preview(highlight_stmt_idx=st_idx))
            
            menu.addSeparator()

            # --- [新增] 人工匹配与取消匹配 ---
            action_manual = menu.addAction("🤝 人工匹配/修改关联 (Manual Match)...")
            action_manual.triggered.connect(lambda: self.manual_match_report(data))
            
            if data.get('statement'):
                action_unmatch = menu.addAction("🔓 取消匹配关联 (Unmatch)")
                action_unmatch.triggered.connect(lambda: self.unmatch_recon_item(data))
            
            menu.addSeparator()

        if data and data.get('statement'):
            action_confirm = menu.addAction("💰 确认此笔单项入账")
            action_confirm.triggered.connect(lambda: self.batch_confirm_payments([data]))
        
        menu.addSeparator()
        if data and data.get('type') in ["BATCH", "PARTIAL", "SPLIT", "SPLIT_PARTIAL"]:
            action_focus = menu.addAction("🔍 聚焦显示此组关联项 (Focus)")
            action_focus.triggered.connect(lambda: self.focus_related_items(data))
            
        action_clear = menu.addAction("🧹 清除所有聚焦与筛选")
        action_clear.triggered.connect(self.clear_all_filters)
        
        menu.addSeparator()
        action_copy = menu.addAction("📋 复制此行")
        action_copy.triggered.connect(lambda: self._copy_row(row))
        
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def show_stmt_context_menu(self, pos):
        """银行流水视图右键菜单 (Tab 2)"""
        item = self.stmt_table.itemAt(pos)
        if not item: return
        row = item.row()
        stmt = self.stmt_table.item(row, 0).data(Qt.UserRole)
        
        menu = QMenu(self)
        
        if stmt:
            action_open_file = menu.addAction("📂 打开来源文件 (PDF)")
            path = stmt.get('_source_path')
            if path and os.path.exists(path):
                action_open_file.triggered.connect(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)))
            else:
                action_open_file.setEnabled(False)

            action_jump = menu.addAction("⬅️ 跳转到核对结果对应项")
            action_jump.triggered.connect(lambda: self.jump_to_recon_tab(stmt))
            
            action_find = menu.addAction("🔎 跳转到源流水日期 (Preview)")
            st_idx = -1
            for idx, s_obj in enumerate(self.statement_records):
                if s_obj is stmt:
                    st_idx = idx
                    break
            action_find.triggered.connect(lambda: self.open_unified_preview(highlight_stmt_idx=st_idx))
            
            menu.addSeparator()

            # --- [新增] 人工匹配与取消匹配 ---
            action_manual = menu.addAction("🤝 人工匹配到收款报告 (Manual Match)...")
            action_manual.triggered.connect(lambda: self.manual_match_statement(stmt))
            
            # 找到该流水关联的对账结果
            related_recon = None
            for res in self.recon_results:
                if res.get('statement') is stmt:
                    related_recon = res
                    break
            
            if related_recon:
                action_unmatch = menu.addAction("🔓 取消匹配关联 (Unmatch)")
                action_unmatch.triggered.connect(lambda: self.unmatch_recon_item(related_recon))
            
            menu.addSeparator()

            if related_recon and related_recon.get('type') in ["BATCH", "PARTIAL", "SPLIT", "SPLIT_PARTIAL"]:
                action_focus = menu.addAction("🔍 聚焦显示此组关联项 (Focus)")
                action_focus.triggered.connect(lambda: self.focus_related_items(related_recon))
            
            action_clear = menu.addAction("🧹 清除所有聚焦与筛选")
            action_clear.triggered.connect(self.clear_all_filters)
            menu.addSeparator()

        action_copy = menu.addAction("📋 复制此行")
        action_copy.triggered.connect(lambda: self._copy_stmt_row(row))
        
        menu.exec(self.stmt_table.viewport().mapToGlobal(pos))

    def focus_related_items(self, target_data):
        """聚焦显示与当前条目关联的所有对账项和流水项"""
        if not target_data: return
        
        # 记录当前焦点位置，以便在过滤后恢复
        active_table = self.table if self.tabs.currentIndex() == 0 else self.stmt_table
        curr_item = active_table.currentItem()
        
        m_type = target_data.get('type', '')
        s_idx = target_data.get('s_idx', -1)
        r_idx = target_data.get('r_idx', -1)
        
        # 1. 过滤 Tab 1 (对账结果表)
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            if not item: continue
            data = item.data(Qt.UserRole)
            if data == "SUMMARY": continue
            
            show = False
            if m_type in ["BATCH", "PARTIAL"]:
                # 组合匹配：展示所有共享同一个流水索引的项目
                if data and isinstance(data, dict) and data.get('s_idx') == s_idx: 
                    show = True
            elif m_type in ["SPLIT", "SPLIT_PARTIAL"]:
                # 分拆匹配：展示所有共享同一个报告索引的项目 (核心修复：改用 r_idx)
                if data and isinstance(data, dict) and data.get('r_idx') == r_idx: 
                    show = True
            else:
                # 单一匹配或未匹配：展示自身
                if data and isinstance(data, dict):
                    if data.get('r_idx') == r_idx and data.get('s_idx') == s_idx:
                        show = True
            
            self.table.setRowHidden(r, not show)
            
        # 2. 过滤 Tab 2 (银行流水视图)
        involved_s_indices = set()
        if m_type in ["BATCH", "PARTIAL"]:
            involved_s_indices.add(s_idx)
        elif m_type in ["SPLIT", "SPLIT_PARTIAL"]:
            # 寻找所有链接到此报告索引的流水
            for res in self.recon_results:
                if res.get('r_idx') == r_idx and res.get('s_idx', -1) != -1:
                    involved_s_indices.add(res['s_idx'])
        else:
            if s_idx != -1: involved_s_indices.add(s_idx)
            
        for r in range(self.stmt_table.rowCount()):
            item = self.stmt_table.item(r, 0)
            if not item: continue
            if item.data(Qt.UserRole) == "SUMMARY" or item.text().startswith("汇 ("): continue
            
            # 核心修复：使用存储在 UserRole+1 的原始流水索引
            orig_s_idx = item.data(Qt.UserRole + 1)
            self.stmt_table.setRowHidden(r, orig_s_idx not in involved_s_indices)

        self.main_window.update_summary_row(self.table)
        self.main_window.update_summary_row(self.stmt_table)
        
        # 恢复焦点并确保可见
        if curr_item:
            active_table.setCurrentItem(curr_item)
            active_table.scrollToItem(curr_item)
            active_table.setFocus()
            
        self.main_window.log_message(f"🔍 已进入聚焦模式：仅显示关联项 ({m_type})")

    def clear_all_filters(self):
        """清除所有聚焦和表头筛选器"""
        # 记录当前焦点位置
        active_table = self.table if self.tabs.currentIndex() == 0 else self.stmt_table
        curr_item = active_table.currentItem()
        
        # 清除 Table 1
        for r in range(self.table.rowCount()):
            self.table.setRowHidden(r, False)
        header = self.table.horizontalHeader()
        if isinstance(header, FilterHeader):
            header._filters.clear()
            header.viewport().update()
            
        # 清除 Table 2
        for r in range(self.stmt_table.rowCount()):
            self.stmt_table.setRowHidden(r, False)
        stmt_header = self.stmt_table.horizontalHeader()
        if isinstance(stmt_header, FilterHeader):
            stmt_header._filters.clear()
            stmt_header.viewport().update()
            
        self.main_window.update_summary_row(self.table)
        self.main_window.update_summary_row(self.stmt_table)
        
        # 恢复焦点
        if curr_item:
            active_table.setCurrentItem(curr_item)
            active_table.scrollToItem(curr_item)
            active_table.setFocus()
            
        self.main_window.log_message("🧹 已清除所有聚焦和筛选器")

    def _is_same_statement(self, s1, s2):
        """Helper to robustly compare two statement objects"""
        if s1 is s2: return True
        if s1 == s2: return True
        # Fallback: compare core fields
        keys = ['date', 'desc', 'amount', 'cnpj']
        try:
            # Use a small tolerance for float comparison just in case
            if abs(s1.get('amount', 0) - s2.get('amount', 0)) > 0.001: return False
            return all(s1.get(k) == s2.get(k) for k in keys if k != 'amount')
        except: return False

    def jump_to_stmt_tab(self, statement):
        if not statement: return
        self.tabs.setCurrentIndex(1) # Switch to Statement Tab
        
        # Find row
        for r in range(self.stmt_table.rowCount()):
            item = self.stmt_table.item(r, 0)
            st_data = item.data(Qt.UserRole)
            if self._is_same_statement(st_data, statement):
                self.stmt_table.selectRow(r)
                self.stmt_table.scrollToItem(item)
                self.stmt_table.setFocus()
                return
        
        QMessageBox.information(self, "提示", "在流水视图中未找到该记录。")

    def jump_to_recon_tab(self, statement):
        if not statement: return
        self.tabs.setCurrentIndex(0) # Switch to Recon Tab
        
        # Find row
        # Note: A statement might match multiple reports (Batch). We jump to the first one.
        for r in range(self.table.rowCount()):
            item = self.table.item(r, 0)
            data = item.data(Qt.UserRole)
            if data and data.get('statement'):
                if self._is_same_statement(data.get('statement'), statement):
                    self.table.selectRow(r)
                    self.table.scrollToItem(item)
                    self.table.setFocus()
                    return
        
        QMessageBox.information(self, "提示", "在核对结果中未找到该流水的匹配记录（可能是未匹配项）。")

    def jump_to_invoice_summary(self, invoice_number):
        """跳转到收款系统的发票列表并定位"""
        if not invoice_number: return
        
        # 1. Switch to Payment System Tab
        self.main_window.tab_widget.setCurrentWidget(self.main_window.payment_tab)
        
        # 2. Find row in invoice_table
        table = self.main_window.payment_tab.invoice_table
        
        found = False
        for r in range(table.rowCount()):
            item = table.item(r, 1) # Column 1 is Invoice Number
            if item and item.text().strip() == invoice_number.strip():
                table.selectRow(r)
                table.scrollToItem(item)
                table.setFocus()
                found = True
                break
        
        if not found:
            # Try approximate search if exact match failed (e.g. leading zeros)
            for r in range(table.rowCount()):
                item = table.item(r, 1)
                if item and invoice_number in item.text():
                    table.selectRow(r)
                    table.scrollToItem(item)
                    table.setFocus()
                    found = True
                    break

        if not found:
            QMessageBox.information(self, "未找到", f"在发票列表中未找到发票号: {invoice_number}")

    def jump_to_invoice_detail(self, invoice_number):
        """跳转到发票明细 Tab 并定位"""
        if not invoice_number: return
        
        # 1. 查找索引
        invoice_index = -1
        for idx, inv in enumerate(self.main_window.invoices):
            if inv.numero == invoice_number:
                invoice_index = idx
                break
        
        if invoice_index != -1:
            self.main_window.navigate_to_invoice(invoice_index, self.main_window.table)
        else:
            # 备选：通过表中的发票号搜索 (不依赖 self.invoices)
            table = self.main_window.table
            found = False
            for r in range(table.rowCount()):
                # 发票号在第1列 (Serie Acess Key 之后?) 需要确认 index
                # 默认 populate_table 使用的 key_map 是 combined_headers 映射
                # 遍历所有列查找
                for c in range(table.columnCount()):
                    item = table.item(r, c)
                    if item and item.text().strip() == invoice_number:
                        self.main_window.tab_widget.setCurrentWidget(self.main_window.tab1_container)
                        table.selectRow(r)
                        table.scrollToItem(item)
                        found = True
                        break
                if found: break
            
            if not found:
                QMessageBox.information(self, "未找到", f"在发票明细中未找到发票号: {invoice_number}")

    def _copy_stmt_row(self, row):
        txt = "\t".join([self.stmt_table.item(row, c).text() for c in range(self.stmt_table.columnCount())])
        QApplication.clipboard().setText(txt)

    def find_in_reports(self, record):
        if not record: return
        try:
            idx = self.report_records.index(record)
            self.open_unified_preview(highlight_report_idx=idx)
        except ValueError:
            QMessageBox.warning(self, "查找失败", "在原始报告列表中未找到该记录(可能已被清除或修改)。")

    def find_in_statements(self, record):
        if not record: return
        try:
            idx = self.statement_records.index(record)
            self.open_unified_preview(highlight_stmt_idx=idx)
        except ValueError:
            QMessageBox.warning(self, "查找失败", "在原始流水列表中未找到该记录(可能已被清除或修改)。")

    def open_unified_preview(self, highlight_report_idx=None, highlight_stmt_idx=None):
        if not self.report_records and not self.statement_records:
            QMessageBox.information(self, "无数据", "暂无任何导入的数据。")
            return

        # Prepare colors
        report_colors = {}
        for i, rec in enumerate(self.report_records):
            if rec.get('_ui_color'): report_colors[i] = rec['_ui_color']
            
        stmt_colors = {}
        for i, rec in enumerate(self.statement_records):
            if rec.get('_ui_color'): stmt_colors[i] = rec['_ui_color']

        dlg = UnifiedPreviewDialog(self, 
                                 report_data=self.report_records,
                                 stmt_data=self.statement_records,
                                 highlight_report_idx=highlight_report_idx,
                                 highlight_stmt_idx=highlight_stmt_idx,
                                 report_colors=report_colors,
                                 stmt_colors=stmt_colors)
                                 
        if dlg.exec() == QDialog.Accepted:
            # If user modified data
            if dlg.result_reports:
                self.report_records = dlg.result_reports
                self.btn_load_reports.setText(f"✅ 已加载报告 ({len(self.report_records)} 条)")
            
            if dlg.result_stmts:
                self.statement_records = dlg.result_stmts
                self.btn_load_statements.setText(f"✅ 已加载流水 ({len(self.statement_records)} 条)")
                
            self.update_summary()

    def _copy_row(self, row):
        txt = "\t".join([self.table.item(row, c).text() for c in range(1, self.table.columnCount())])
        QApplication.clipboard().setText(txt)

    def batch_confirm_payments(self, specific_rows=None):
        """批量或单笔确认入账"""
        rows_to_pay = []
        
        if specific_rows:
            rows_to_pay = specific_rows
        else:
            # 扫描勾选项
            for r in range(self.table.rowCount()):
                chk = self.table.item(r, 0)
                # 排除汇总行 (Skip summary rows)
                if chk and chk.data(Qt.UserRole) == "SUMMARY":
                    continue
                    
                if chk and chk.checkState() == Qt.Checked:
                    data = chk.data(Qt.UserRole)
                    if data and data['statement']:
                        rows_to_pay.append(data)
        
        if not rows_to_pay:
            QMessageBox.warning(self, "无可入账项", "请先勾选带有流水记录的项目。")
            return
            
        count = len(rows_to_pay)
        reply = QMessageBox.question(self, "确认入账", f"确定要将这 {count} 笔对账记录正式登记入账吗？\n余额将自动同步更新。", 
                                   QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes: return

        # 选账户
        accounts = self.db.get_accounts()
        if not accounts:
            QMessageBox.warning(self, "错误", "请先创建收款账户。")
            return
        acc_names = [a['name'] for a in accounts]
        acc_choice, ok = QInputDialog.getItem(self, "确认入账账户", "款项将统一记入：", acc_names, 0, False)
        if not ok: return
        account_id = accounts[acc_names.index(acc_choice)]['id']

        success_count = 0
        failed_count = 0
        
        for res in rows_to_pay:
            rep = res.get('report') or {}
            st = res.get('statement') or {}
            
            # 1. 寻找发票匹配 - 增强版逻辑
            enable_local = QSettings(SETTINGS_FILE, QSettings.IniFormat).value("enable_local_rules", True, type=bool)
            invoice_ref = rep.get('invoice_ref', '')
            ref_info = CollectionReportParser.parse_invoice_reference(invoice_ref, enable_local)
            clean_num = ref_info.get('invoice') if ref_info else CollectionReportParser.clean_invoice_number(invoice_ref, enable_local)

            rep_name = str(rep.get('name') or "")
            rep_std_name = self.main_window.mapping_mgr.get_partner_std(rep_name) or rep_name
            
            # 组合查询条件
            term_num = ref_info.get('term_number') if ref_info else None
            matches = self.db.search_pending_installments(self._build_invoice_search_patterns(clean_num), term_num) if clean_num else []
            
            best_match = None
            
            if len(matches) == 1:
                best_match = matches[0]
            elif len(matches) > 1:
                for m in matches:
                    db_dest_name = self._row_value(m, 'destinatario_name', 3)
                    db_dest_cnpj = self._row_value(m, 'destinatario_cnpj', 4)
                    db_due_date = self._row_value(m, 'due_date', 5) # YYYY-MM-DD
                    
                    # 增强匹配：如果解析出了日期，优先匹配日期
                    if ref_info and ref_info.get('day'):
                        try:
                            dt_obj = datetime.strptime(db_due_date, "%Y-%m-%d")
                            if dt_obj.day == ref_info['day'] and dt_obj.month == ref_info['month']:
                                best_match = m
                                break
                        except Exception:
                            pass

                    clean_cnpj = re.sub(r'\D', '', db_dest_cnpj) if db_dest_cnpj else ""
                    db_std_name = self.main_window.mapping_mgr.get_partner_std(clean_cnpj) or \
                                  self.main_window.mapping_mgr.get_partner_std(db_dest_name) or db_dest_name
                    db_std_name = str(db_std_name or "")
                    if rep_std_name.lower() in db_std_name.lower() or db_std_name.lower() in rep_std_name.lower():
                        best_match = m
                        break
            
            if best_match:
                inv_id = self._row_value(best_match, 'invoice_id', 0)
                inst_id = self._row_value(best_match, 'installment_id', 1)
                amount = self._to_float(self._row_value(best_match, 'amount', 2))
                paid_amount = self._to_float(self._row_value(best_match, 'paid_amount', 7))
                penalty = self._to_float(self._row_value(best_match, 'penalty', 8))

                if (paid_amount == 0.0 and penalty == 0.0) and inv_id and inst_id:
                    inst_row = self._get_installment_by_id(inv_id, inst_id)
                    if inst_row:
                        amount = self._to_float(inst_row['amount'])
                        paid_amount = self._to_float(inst_row['paid_amount'])
                        penalty = self._to_float(inst_row['penalty'])

                due_remaining = max(0.0, amount + penalty - paid_amount)
                pay_amount = self._to_float(st.get('amount', rep.get('amount', 0.0)))
                if due_remaining > 0.0:
                    pay_amount = min(pay_amount, due_remaining)

                if inst_id and pay_amount > 0.01:
                    pay_date = self._to_iso_date(st.get('date'))
                    ok = self.db.register_payment(
                        inst_id,
                        pay_amount,
                        pay_date,
                        account_id,
                        f"Auto-Reconcile: {invoice_ref} - {st.get('desc', '')}",
                    )
                    if ok:
                        success_count += 1
                    else:
                        failed_count += 1
                else:
                    failed_count += 1
            else:
                # 未匹配到现有发票
                # check if valid invoice structure is detected
                if ref_info and ref_info.get('invoice') and ref_info.get('term_number'):
                    inv_num = ref_info['invoice']
                    term_num = ref_info['term_number']
                    
                    reply_temp = QMessageBox.question(self, "创建临时发票", 
                        f"未找到发票 {inv_num} (第 {term_num} 期)。\n\n"
                        f"是否创建【临时发票】并补全前 {term_num} 期分期？\n"
                        f"(发票将被标记为 '⚠️ NEED PDF')",
                        QMessageBox.Yes | QMessageBox.No)
                        
                    if reply_temp == QMessageBox.Yes:
                        if self._create_temp_invoice_and_pay(account_id, rep, st, ref_info):
                            success_count += 1
                        else:
                            failed_count += 1
                        continue # Skip to next iteration

                # 预收账款逻辑: 未匹配到发票号，但已知客户名称
                prepay_amount = self._to_float(rep.get('amount', st.get('amount', 0.0)))
                if prepay_amount <= 0.01:
                    failed_count += 1
                    continue
                
                # --- 新增: 用户确认 ---
                reply = QMessageBox.question(self, "确认预收账款", 
                    f"未匹配到发票记录 (且不符合发票规则)。\n\n客户: {rep_name}\n金额: {prepay_amount:,.2f}\n流水: {st.get('desc', '')}\n\n是否作为【预收账款】存入？",
                    QMessageBox.Yes | QMessageBox.No)
                if reply != QMessageBox.Yes:
                    continue
                # ---------------------
                
                d_fmt = self._to_iso_date(st.get('date'))
                trans_id = self.db.add_transaction(
                    account_id,
                    d_fmt,
                    'Income',
                    prepay_amount,
                    f"预收账款(未匹配发票): {st.get('desc', '')} | 客户: {rep_name}",
                )

                if not trans_id:
                    failed_count += 1
                    continue

                st_cnpj = st.get('cnpj', '')
                adv_id = self.db.add_advance(
                    rep_name,
                    st_cnpj,
                    prepay_amount,
                    d_fmt,
                    f"银行流水自动入账(预收): {st.get('desc', '')}",
                    account_id,
                    trans_id,
                )
                if adv_id:
                    success_count += 1
                else:
                    failed_count += 1
        
        self.main_window.payment_tab.load_invoices()
        if self.main_window.account_tab: self.main_window.account_tab.load_accounts()
        if failed_count > 0:
            QMessageBox.warning(self, "完成", f"已成功处理 {success_count} 笔，失败 {failed_count} 笔。")
        else:
            QMessageBox.information(self, "完成", f"已成功处理 {success_count} 笔入账。")
        self.run_reconciliation() # 刷新当前核对状态列表

    def _create_temp_invoice_and_pay(self, account_id, rep, st, ref_info):
        """Create a temporary invoice/installments and register payment via manager APIs."""
        try:
            inv_num = str(ref_info['invoice'])
            target_term = int(ref_info['term_number'])
        except Exception:
            return False

        amount = self._to_float(rep.get('amount', st.get('amount', 0.0)))
        if amount <= 0.01:
            return False

        pay_date_fmt = self._to_iso_date(st.get('date'))
        due_date_str = rep.get('due_date') or st.get('date')
        d_due = safe_parse_date(due_date_str)
        due_date_obj = d_due if d_due else safe_parse_date(pay_date_fmt)
        due_date_fmt = due_date_obj.strftime("%Y-%m-%d") if due_date_obj else pay_date_fmt

        inv_id = self.db.find_invoice_id_by_number(inv_num)
        if not inv_id:
            total_estimated = round(amount * target_term, 2)
            invoice_data = {
                "invoice_number": inv_num,
                "issuer_name": "",
                "issue_date": due_date_fmt,
                "total_amount": total_estimated,
                "file_name": "",
                "natureza_operacao": "⚠️ NEED PDF (Temp Created)",
                "destinatario_name": rep.get('name', ''),
                "destinatario_cnpj": "",
                "issuer_cnpj": "",
                "description": "⚠️ NEED PDF (Temp Created)",
            }
            inv_id = self.db.upsert_invoice(invoice_data)
            if not inv_id or inv_id == -1:
                return False

        installments = self.db.get_installments(inv_id)
        term_map = {}
        for row in installments:
            try:
                row_data = dict(row) if not isinstance(row, dict) else row
                term_map[int(row_data['term_number'])] = row_data
            except Exception:
                continue

        if target_term not in term_map:
            base_due = due_date_obj if due_date_obj else datetime.now()
            first_due = base_due - timedelta(days=30 * (target_term - 1))
            plan_ok = self.db.generate_payment_plan(inv_id, target_term, start_date=first_due, interval_days=30)
            if plan_ok is False:
                return False

            installments = self.db.get_installments(inv_id)
            term_map = {}
            for row in installments:
                try:
                    row_data = dict(row) if not isinstance(row, dict) else row
                    term_map[int(row_data['term_number'])] = row_data
                except Exception:
                    continue

        inst_row = term_map.get(target_term)
        if not inst_row:
            return False

        due_remaining = max(
            0.0,
            self._to_float(inst_row.get('amount')) + self._to_float(inst_row.get('penalty')) - self._to_float(inst_row.get('paid_amount')),
        )
        pay_amount = min(amount, due_remaining if due_remaining > 0.0 else amount)
        if pay_amount <= 0.01:
            return False

        return bool(
            self.db.register_payment(
                int(inst_row['id']),
                pay_amount,
                pay_date_fmt,
                account_id,
                f"自动核对入账(Temp): {st.get('desc', '')}",
            )
        )

class ChartConfigDialog(QDialog):
    """图表配置与展示对话框"""
    def __init__(self, main_window, rows, parent=None):
        super().__init__(parent)
        self.main_window = main_window
        self.rows = rows
        self.setWindowTitle("图表分析 (Chart Analysis)")
        self.resize(1000, 700)
        self.setup_ui()
        self.plot()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Controls
        ctrl_frame = QFrame()
        ctrl_frame.setFrameShape(QFrame.StyledPanel)
        ctrl_layout = QHBoxLayout(ctrl_frame)
        
        ctrl_layout.addWidget(QLabel("分析维度 (X轴/分组):"))
        self.dim_combo = QComboBox()
        
        # Base options
        dims = ["按月份 (Month)", "按标准产品 (Product)", "按标准往来单位 (Partner)", "按发票 (Invoice)"]
        
        # Add dynamic headers from main window
        if hasattr(self.main_window, 'combined_headers'):
            for h in self.main_window.combined_headers:
                if h not in dims: # Avoid duplicates
                    dims.append(h)
                    
        self.dim_combo.addItems(dims)
        ctrl_layout.addWidget(self.dim_combo)
        
        ctrl_layout.addWidget(QLabel("统计指标 (Y轴/数值):"))
        self.metric_combo = QComboBox()
        self.metric_combo.addItems(["总金额 (Total Amount)", "实收金额 (Paid Amount)", "税额 (Tax)", "数量 (Quantity)", "计数 (Count)"])
        ctrl_layout.addWidget(self.metric_combo)
        
        ctrl_layout.addWidget(QLabel("图表类型:"))
        self.type_combo = QComboBox()
        self.type_combo.addItems(["折线图 (Line - 趋势)", "柱状图 (Bar - 对比)", "饼图 (Pie - 占比)"])
        ctrl_layout.addWidget(self.type_combo)
        
        btn_refresh = QPushButton("🔄 更新图表")
        btn_refresh.clicked.connect(self.plot)
        ctrl_layout.addWidget(btn_refresh)
        
        btn_export = QPushButton("💾 导出图表")
        btn_export.clicked.connect(self.export_chart)
        ctrl_layout.addWidget(btn_export)
        
        layout.addWidget(ctrl_frame)
        
        # Chart Area
        self.chart_container = QWidget()
        self.chart_layout = QVBoxLayout(self.chart_container)
        layout.addWidget(self.chart_container, 1)
        
        # Init matplotlib
        if HAS_MATPLOTLIB:
            try:
                import matplotlib.pyplot as plt
                plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS', 'sans-serif']
                plt.rcParams['axes.unicode_minus'] = False
            except:
                pass
            
            self.figure = Figure(figsize=(8, 6))
            self.canvas = FigureCanvas(self.figure)
            self.chart_layout.addWidget(self.canvas)
            self.has_matplotlib = True
        else:
            self.chart_layout.addWidget(QLabel("未安装 Matplotlib，无法显示图表"))
            self.has_matplotlib = False

    def plot(self):
        if not self.has_matplotlib: return
        try:
            import matplotlib.pyplot as plt
        except:
            return
            
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        dim_text = self.dim_combo.currentText()
        metric_idx = self.metric_combo.currentIndex()
        type_idx = self.type_combo.currentIndex()
        
        # 0: Total, 1: Paid, 2: Tax, 3: Qty, 4: Count
        # 0: Line, 1: Bar, 2: Pie
        
        # Data Aggregation
        data = {} # key -> value
        
        # Determine column index if dynamic
        col_index = -1
        if "按月份" not in dim_text and "按标准产品" not in dim_text and \
           "按标准往来单位" not in dim_text and "按发票" not in dim_text:
            if hasattr(self.main_window, 'combined_headers'):
                try:
                    col_index = self.main_window.combined_headers.index(dim_text)
                except: pass

        # Prepare helper to get dimension key
        def get_key(inv, item):
            if "按月份" in dim_text: # Month
                date = inv.data_emissao
                if date:
                    # Robust parsing
                    try:
                        if '/' in date:
                            parts = date.split('/')
                            return f"{parts[2]}-{parts[1]}" if len(parts) >= 3 else date[:7]
                        elif '-' in date:
                             return date[:7]
                    except: pass
                return "无日期"
            elif "按标准产品" in dim_text: # Product
                raw = item.codigo_produto or ""
                std = self.main_window.mapping_mgr.get_product_std(raw)
                return std.get("std_name") if std and std.get("std_name") else (item.descricao or raw or "Unknown")
            elif "按标准往来单位" in dim_text: # Partner
                raw_name = inv.destinatario_nome or inv.emitente_nome or "Unknown"
                raw_cnpj = inv.destinatario_cnpj or inv.emitente_cnpj
                if raw_cnpj:
                    clean = re.sub(r'\D', '', raw_cnpj)
                    std = self.main_window.mapping_mgr.get_partner_std(clean)
                    if std: return std
                return raw_name
            elif "按发票" in dim_text: # Invoice
                return inv.numero or f"FILE-{inv.file_name}"
            elif col_index >= 0: # Dynamic Column
                # Get row data
                try:
                    row_data = self.main_window.get_row_from_invoice_and_item(inv, item)
                    if col_index < len(row_data):
                        val = row_data[col_index]
                        return str(val) if val is not None else "Empty"
                except: pass
                return "Unknown"
            else:
                return "Unknown"

        def get_val(inv, item, p_data):
            if metric_idx == 0: return item.valor_total or 0.0 # Total
            elif metric_idx == 1: # Paid (Estimate pro-rata?)
                inv_total = inv.total_nota or 1.0
                ratio = (item.valor_total or 0) / inv_total if inv_total else 0
                return (p_data.get('paid_amount', 0.0) * ratio)
            elif metric_idx == 2: return (item.valor_icms or 0) + (item.valor_ipi or 0) # Tax
            elif metric_idx == 3: return item.quantidade or 0.0 # Qty
            elif metric_idx == 4: return 1 # Count
            return 0

        # Status map for payment info
        status_map = self.main_window.query_tab._status_map if hasattr(self.main_window, 'query_tab') else {}

        for inv, item in self.rows:
            k = get_key(inv, item)
            inv_key = inv.numero or f"FILE-{inv.file_name}"
            p_data = status_map.get(inv_key, {})
            v = get_val(inv, item, p_data)
            
            data[k] = data.get(k, 0.0) + v
            
        # Plotting
        sorted_keys = []
        if "按月份" in dim_text: # Sort Months
            sorted_keys = sorted(data.keys())
        else: # Sort by Value Desc
            sorted_keys = sorted(data.keys(), key=lambda x: data[x], reverse=True)[:20] # Top 20
            
        x_labels = sorted_keys
        y_values = [data[k] for k in sorted_keys]
        
        # Clean labels length
        clean_labels = [str(s)[:15]+".." if len(str(s))>15 else str(s) for s in x_labels]
        
        if not y_values:
            ax.text(0.5, 0.5, "无数据 (No Data)", ha='center', va='center')
            self.canvas.draw()
            return
            
        if type_idx == 2: # Pie
            # Filter negative or zero for pie
            pie_vals = []
            pie_labs = []
            for v, l in zip(y_values, clean_labels):
                if v > 0:
                    pie_vals.append(v)
                    pie_labs.append(l)
            if pie_vals:
                ax.pie(pie_vals, labels=pie_labs, autopct='%1.1f%%', startangle=90)
                ax.axis('equal')
            else:
                ax.text(0.5, 0.5, "数值过小无法显示饼图", ha='center')
                
        elif type_idx == 0: # Line
            ax.plot(clean_labels, y_values, marker='o', linestyle='-', color='#5bc0de')
            ax.grid(True, linestyle='--', alpha=0.7)
            plt.setp(ax.get_xticklabels(), rotation=45, ha="right", rotation_mode="anchor")
            
        else: # Bar
            ax.bar(clean_labels, y_values, color='#5cb85c')
            plt.setp(ax.get_xticklabels(), rotation=45, ha="right", rotation_mode="anchor")
            
        title_str = f"{self.dim_combo.currentText()} - {self.metric_combo.currentText()}"
        ax.set_title(title_str)
        self.figure.tight_layout()
        self.canvas.draw()

    def export_chart(self):
        """将当前图表导出为图片或PDF"""
        if not self.has_matplotlib:
            QMessageBox.warning(self, "导出提示", "未安装 Matplotlib，无法导出图表")
            return
            
        try:
            # 建议文件名
            suggested_name = f"chart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            
            path, selected_filter = QFileDialog.getSaveFileName(
                self, 
                "导出图表", 
                suggested_name, 
                "PNG Images (*.png);;JPEG Images (*.jpg);;PDF Files (*.pdf);;All Files (*)"
            )
            
            if path:
                # 检查是否已有后缀，如果没有则根据选中的过滤器添加
                if '.' not in os.path.basename(path):
                    if "PNG" in selected_filter: path += ".png"
                    elif "JPEG" in selected_filter: path += ".jpg"
                    elif "PDF" in selected_filter: path += ".pdf"
                
                self.figure.savefig(path, dpi=300, bbox_inches='tight')
                QMessageBox.information(self, "导出成功", f"图表已成功保存至:\n{path}")
        except Exception as e:
            logging.exception("Failed to export chart")
            QMessageBox.critical(self, "导出错误", f"导出图表时出错: {str(e)}")

class DataQueryTab(QWidget):
    def __init__(self, main_window: "MainWindow"):
        super().__init__()
        self.main_window = main_window
        self._status_map = {}
        layout = QVBoxLayout(self)

        # 搜索和聚合控件
        control_layout = QHBoxLayout()
        
        # 关键字
        control_layout.addWidget(QLabel("关键字"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("发票/公司/产品 ..")
        
        # 日期范围
        from PySide6.QtWidgets import QDateEdit
        control_layout.addWidget(QLabel("开始日期"))
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDisplayFormat("yyyy/MM/dd")
        # 默认值设为一年前以覆盖大部分情况
        self.start_date.setDate(datetime.now().date() - timedelta(days=365))
        
        control_layout.addWidget(self.start_date)
        
        control_layout.addWidget(QLabel("结束日期"))
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDisplayFormat("yyyy/MM/dd")
        self.end_date.setDate(datetime.now().date())
        control_layout.addWidget(self.end_date)
        
        # 启用日期过滤复选框
        self.use_date_check = QCheckBox("启用日期筛选")
        self.use_date_check.setChecked(False)
        control_layout.addWidget(self.use_date_check)

        self.search_btn = QPushButton("🔍 查询")
        self.clear_btn = QPushButton(" 清除")
        self.export_query_btn = QPushButton("📤 导出结果") # New export button
        self.chart_btn = QPushButton("📊 图表分析") # New Chart Button
        self.chart_btn.clicked.connect(self.open_chart_dialog)
        
        control_layout.addWidget(self.search_input)
        control_layout.addWidget(self.search_btn)
        control_layout.addWidget(self.clear_btn)
        control_layout.addWidget(self.export_query_btn)
        control_layout.addWidget(self.chart_btn)
        
        # 支付状态筛选
        control_layout.addWidget(QLabel("支付状态:"))
        self.status_filter_combo = QComboBox()
        self.status_filter_combo.addItems(["全部 (All)", "仅看已结清 (Paid)", "仅看有欠款 (Pending)", "仅看已逾期 (Overdue)"])
        self.status_filter_combo.currentIndexChanged.connect(self.refresh)
        control_layout.addWidget(self.status_filter_combo)
        
        # 第二行控件
        row2_layout = QHBoxLayout()
        
        # 聚合选择
        self.group_combo = QComboBox()
        self.group_combo.addItem("📋 不汇总 (Detailed List - All Items)") 
        self.group_combo.addItem("🏢 按标准往来单位 (By Standard Partner)")
        self.group_combo.addItem("📦 按标准产品 (By Standard Product)")
        self.group_combo.addItem("📅 按月份 (Month - 签发日期)")
        
        # Add all headers dynamically (Optional, but kept for flexibility)
        headers = self.main_window.combined_headers
        for h in headers:
            self.group_combo.addItem(h)
            
        self.group_combo.currentIndexChanged.connect(self.refresh)
        
        self.refresh_btn = QPushButton("🔄 刷新")

        row2_layout.addWidget(QLabel("汇总标准:"))
        row2_layout.addWidget(self.group_combo)
        row2_layout.addStretch()
        self.result_label = QLabel("匹配: 0")
        row2_layout.addWidget(self.result_label)
        row2_layout.addWidget(self.refresh_btn)
        
        layout.addLayout(control_layout)
        layout.addLayout(row2_layout) # Move row2_layout up

        # 财务看板面板 (Moved to be less intrusive or removed if Chart covers it, but kept for text summary)
        self.kpi_group = QGroupBox("当前筛选财务统计 (Dashboard)")
        kpi_layout = QHBoxLayout(self.kpi_group)
        self.lbl_kpi_total = QLabel("总额: 0.00")
        self.lbl_kpi_paid = QLabel("实收: 0.00")
        self.lbl_kpi_balance = QLabel("余额: 0.00")
        self.lbl_kpi_tax = QLabel("税额: 0.00")
        
        for lbl in [self.lbl_kpi_total, self.lbl_kpi_paid, self.lbl_kpi_balance, self.lbl_kpi_tax]:
            lbl.setStyleSheet("font-weight: bold; font-size: 10pt; color: #444;")
        self.lbl_kpi_balance.setStyleSheet("font-weight: bold; font-size: 10pt; color: #d9534f;") # 余额醒目红
        
        kpi_layout.addWidget(self.lbl_kpi_total)
        kpi_layout.addWidget(self.lbl_kpi_paid)
        kpi_layout.addWidget(self.lbl_kpi_balance)
        kpi_layout.addWidget(self.lbl_kpi_tax)
        kpi_layout.addStretch()
        
        layout.addWidget(self.kpi_group)

        # --- Table View ---
        self.table_container = QWidget()
        table_layout = QVBoxLayout(self.table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.setSpacing(0)

        self.table = QTableWidget()
        self.table.setSortingEnabled(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        
        # Use FilterHeader
        header = FilterHeader(self.table)
        header.filterChanged.connect(lambda: self.main_window.apply_header_filters(self.table))
        self.table.setHorizontalHeader(header)
        
        # 启用右键菜单
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
        table_layout.addWidget(self.table)
        layout.addWidget(self.table_container)
        
        # 注册表格统计
        if hasattr(self.main_window, 'register_table_for_stats'):
            self.main_window.register_table_for_stats(self.table)
        
        self.filtered_rows = [] # Init

        self.search_btn.clicked.connect(self.apply_filter)
        self.clear_btn.clicked.connect(self.clear_filter)
        self.refresh_btn.clicked.connect(self.refresh)
        self.export_query_btn.clicked.connect(self.export_query_result)

        QTimer.singleShot(100, self.init_table_config)

    def init_table_config(self):
        # Disable native sorting for summary row handling
        self.table.setSortingEnabled(False)
        self.table.horizontalHeader().sectionClicked.connect(self.main_window.on_header_clicked)
        self.refresh()

    def show_context_menu(self, pos):
        item = self.table.itemAt(pos)
        if not item: return
        
        # 排除汇总行 (Skip summary rows)
        if item.data(Qt.UserRole) == "SUMMARY" or item.text().startswith("汇 ("):
            return
        
        menu = QMenu(self)
        
        # 仅在未汇总模式下启用详细跳转
        if self.group_combo.currentIndex() == 0:
            original_index = item.data(Qt.UserRole)
            if original_index is not None:
                nav = self.main_window.create_navigation_menu(original_index, "query")
                menu.addActions(nav.actions())
                menu.addSeparator()
                
                # 新增查看 PDF
                action_pdf = menu.addAction("📄 查看源文件 PDF (Open)")
                action_pdf.triggered.connect(lambda: self.main_window.view_source_pdf(original_index))
        else:
            # 聚合模式下的钻取功能 (Drill down)
            action_drill = menu.addAction("🔍 钻取详细数据 (Drill Down)")
            action_drill.triggered.connect(lambda: self.drill_down(item))
            
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def drill_down(self, item):
        """聚合模式下钻取：提取名称 CNPJ 作为关键词搜"""
        row = item.row()
        
        # 排除汇总行
        first_item = self.table.item(row, 0)
        if first_item and first_item.data(Qt.UserRole) == "SUMMARY":
            return

        key_item = self.table.item(row, 0)
        if not key_item: return
        
        full_text = key_item.text()
        search_key = full_text
        
        # 如果 CNPJ 归一化后的名称，提取 Tax ID 部分
        if "[Tax ID:" in full_text:
            import re
            m = re.search(r'\[Tax ID:\s*([0-9\.\-/]+)\]', full_text)
            if m:
                # 提取 CNPJ
                search_key = m.group(1)
        elif "(" in search_key:
            search_key = search_key.split("(")[0].strip()
        
        self.search_input.setText(search_key)
        self.group_combo.setCurrentIndex(0) # 切换到不汇总模式展示明 
    def refresh(self):
        # Invalidate caches
        if hasattr(self, '_status_map_cache'):
            self._status_map_cache = None
        if hasattr(self, '_search_cache'):
            self._search_cache = {}
        self.apply_filter()

    def clear_filter(self):
        self.search_input.clear()
        self.use_date_check.setChecked(False)
        self.group_combo.setCurrentIndex(0)
        self.apply_filter()

    def apply_filter(self):
        term = self.search_input.text().strip().lower()
        status_filter = self.status_filter_combo.currentText()
        
        # Determine grouping mode
        group_col_index = -1
        group_text = self.group_combo.currentText()
        
        if "按月份" in group_text: group_mode = "MONTH"
        elif "按标准产品" in group_text: group_mode = "STD_PRODUCT"
        elif "按标准往来单位" in group_text: group_mode = "STD_PARTNER"
        else:
            group_mode = "COLUMN"
            try:
                group_col_index = self.main_window.combined_headers.index(group_text)
            except ValueError: group_mode = "STD_PARTNER" # Fallback
        
        use_date = self.use_date_check.isChecked()
        s_date = self.start_date.date().toPython()
        e_date = self.end_date.date().toPython()
        
        invoices = self.main_window.invoices or []
        
        if not hasattr(self, '_status_map_cache') or self._status_map_cache is None:
            self._status_map_cache = self._build_status_map()
        self._status_map = self._status_map_cache
        
        if not hasattr(self, '_search_cache'): self._search_cache = {}        

        # 1. 过滤与统计逻辑
        filtered_rows = [] 
        kpi = {'total': 0.0, 'paid': 0.0, 'bal': 0.0, 'tax': 0.0}
        
        for inv in invoices:
            # A. 日期过滤
            if use_date:
                d_obj = safe_parse_date_to_date(inv.data_emissao)
                if not d_obj or not (s_date <= d_obj <= e_date):
                    continue
            
            # B. 状态过滤
            inv_key = inv.numero or f"FILE-{inv.file_name}"
            p_data = self._status_map.get(inv_key, {})
            p_bal = p_data.get('balance', 0.0)
            
            if "已结清" in status_filter and (p_bal > 0.05 or p_data == {}): continue
            if "有欠款" in status_filter and p_bal <= 0.05: continue
            if "已逾期" in status_filter:
                # 简单逻辑：如果发票日期是 30 天前且还有余额，视为逾期
                if p_bal <= 0.05: continue
            
            # C. 关键字过滤
            if id(inv) not in self._search_cache:
                self._search_cache[id(inv)] = " ".join(str(x) for x in self.main_window.get_row_from_invoice_and_item(inv, Item())[:37]).lower()
            inv_match = term in self._search_cache[id(inv)]
            
            items_to_add = []
            invoice_added_to_kpi = False
            
            if inv.itens:
                for item in inv.itens:
                    if id(item) not in self._search_cache:
                        self._search_cache[id(item)] = " ".join(str(x) for x in self.main_window.get_row_from_invoice_and_item(inv, item)).lower()
                    
                    if inv_match or term in self._search_cache[id(item)]:
                        items_to_add.append(item)
                        # 累计明细金额 (增加安全校验)
                        v_total = item.valor_total or 0.0
                        v_tax = (item.valor_icms or 0.0) + (item.valor_ipi or 0.0)
                        
                        # 巴西单项金额极少超过 1 亿，过滤掉明显的解析错误
                        if abs(v_total) < 1e8:
                            kpi['total'] += v_total
                        if abs(v_tax) < 1e8:
                            kpi['tax'] += v_tax
                        
                        if not invoice_added_to_kpi:
                            p_paid = p_data.get('paid_amount', 0.0)
                            p_bal = p_data.get('balance', 0.0)
                            if abs(p_paid) < 1e9: kpi['paid'] += p_paid
                            if abs(p_bal) < 1e9: kpi['bal'] += p_bal
                            invoice_added_to_kpi = True
            elif inv_match:
                items_to_add.append(Item()) # Dummy for invoice with no items
                kpi['paid'] += p_data.get('paid_amount', 0.0)
                kpi['bal'] += p_data.get('balance', 0.0)

            for it in items_to_add:
                filtered_rows.append((inv, it))

            for it in items_to_add:
                filtered_rows.append((inv, it))

        # 2. 界面展示
        if self.group_combo.currentIndex() == 0:
            self._show_detailed_list(filtered_rows)
        else:
            self._show_aggregated_list(filtered_rows, group_mode, group_col_index)

        # 3. 重新计算看板数据 (基于表格显示的结果，确保跳过汇总行并应用了安全过滤)
        kpi = {'total': 0.0, 'paid': 0.0, 'bal': 0.0, 'tax': 0.0}
        # 寻找金额相关的列索引 (根据详细模式的表头)
        # 注意：此处逻辑需根据当前是聚合还是详细模式分别处理
        if self.group_combo.currentIndex() == 0:
            # 详细模式
            # 获取发票总额、税额、已收、待收对应的列
            # headers = self.main_window.invoice_headers + ["已收金额", "待收余额"] + self.main_window.item_headers
            inv_len = len(self.main_window.invoice_headers)
            total_col = 8 # total_nota
            tax_col = 9   # total_icms (此处仅为示例，实际应根据业务累加)
            # 详细模式下我们直接从 item 对象累加更准，但要跳过异常值
            for inv, it in filtered_rows:
                v_total = it.valor_total or 0.0
                if abs(v_total) < 1e8: kpi['total'] += v_total
                kpi['tax'] += (it.valor_icms or 0.0) + (it.valor_ipi or 0.0)
                
                inv_key = inv.numero or f"FILE-{inv.file_name}"
                p_data = self._status_map.get(inv_key, {})
                p_paid = p_data.get('paid_amount', 0.0)
                p_bal = p_data.get('balance', 0.0)
                if abs(p_paid) < 1e9: kpi['paid'] += p_paid
                if abs(p_bal) < 1e9: kpi['bal'] += p_bal
        else:
            # 聚合模式：从表格行中提取
            for r in range(self.table.rowCount()):
                # 跳过汇总行
                first_item = self.table.item(r, 0)
                if first_item and (first_item.data(Qt.UserRole) == "SUMMARY" or first_item.text().startswith("汇 (")):
                    continue
                
                # 获取列数据 (根据 _show_aggregated_list 的 headers)
                # headers = ["分组名称", "明细数", "发票总数", "已结清数", "产品总额", "总税额", "发票已收", "发票余额", "进度"]
                try:
                    v_total = br_to_float(self.table.item(r, 4).text()) or 0.0
                    v_tax = br_to_float(self.table.item(r, 5).text()) or 0.0
                    v_paid = br_to_float(self.table.item(r, 6).text()) or 0.0
                    v_bal = br_to_float(self.table.item(r, 7).text()) or 0.0
                    
                    if abs(v_total) < 1e12: kpi['total'] += v_total
                    kpi['tax'] += v_tax
                    kpi['paid'] += v_paid
                    kpi['bal'] += v_bal
                except: pass

        # 更新看板显示
        self.lbl_kpi_total.setText(f"总营收: {kpi['total']:,.2f}")
        self.lbl_kpi_paid.setText(f"总实收: {kpi['paid']:,.2f}")
        self.lbl_kpi_balance.setText(f"总待收: {kpi['bal']:,.2f}")
        self.lbl_kpi_tax.setText(f"总税额: {kpi['tax']:,.2f}")

        self.main_window.update_summary_row(self.table)
        
        # 4. Save for chart
        self.filtered_rows = filtered_rows

    def open_chart_dialog(self):
        """打开图表分析窗口"""
        if not self.filtered_rows:
            QMessageBox.warning(self, "无数据", "当前没有数据可分析，请先执行查询。")
            return
            
        dlg = ChartConfigDialog(self.main_window, self.filtered_rows, self)
        dlg.exec()

    def export_query_result(self):
        """导出当前查询结果"""
        rows_count = self.table.rowCount()
        cols_count = self.table.columnCount()
        if rows_count == 0:
            QMessageBox.warning(self, "无数据", "当前列表为空，无法导出")
            return
            
        path, _ = QFileDialog.getSaveFileName(self, "导出查询结果", "query_result.xlsx", "Excel Files (*.xlsx);;CSV Files (*.csv)")
        if not path: return
        
        # Collect data from table
        headers = []
        for c in range(cols_count):
            item = self.table.horizontalHeaderItem(c)
            headers.append(item.text() if item else f"Col {c}")
            
        data_rows = []
        for r in range(rows_count):
            if self.table.isRowHidden(r): continue
            
            # 排除汇总行 (Skip summary rows - Robust check)
            is_summary = False
            first_item = self.table.item(r, 0)
            if first_item:
                if first_item.data(Qt.UserRole) == "SUMMARY" or first_item.text().startswith("汇 ("):
                    is_summary = True
            
            # Double check all columns in the row for SUMMARY role
            if not is_summary:
                for c in range(cols_count):
                    it = self.table.item(r, c)
                    if it and it.data(Qt.UserRole) == "SUMMARY":
                        is_summary = True
                        break
            
            if is_summary: continue
                
            row_data = []
            for c in range(cols_count):
                item = self.table.item(r, c)
                txt = item.text() if item else ""
                row_data.append(txt)
            data_rows.append(row_data)
            
        # Delegate to MainWindow's helper
        self.main_window._export_data(path, headers, data_rows)

    def _show_detailed_list(self, rows):
        """显示详细列表 (包含发票与产品明细)"""
        # 组装完整表头
        headers = self.main_window.invoice_headers + ["已收金额", "待收余额"] + self.main_window.item_headers
        
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(rows))
        
        # 缓存原始索引
        inv_index_map = {id(inv): idx for idx, inv in enumerate(self.main_window.invoices)}
        
        self.table.blockSignals(True)
        for i, (inv, item) in enumerate(rows):
            # 获取发票级数据 + 收款统计 + 产品级数据
            raw_row = self.main_window.get_row_from_invoice_and_item(inv, item)
            inv_len = len(self.main_window.invoice_headers)
            
            inv_part = raw_row[:inv_len]
            item_part = raw_row[inv_len:]
            
            # 实时收款数据
            p_data = self._status_map.get(inv.numero or f"FILE-{inv.file_name}", {})
            stats = [p_data.get('paid_amount', 0.0), p_data.get('balance', 0.0)]
            
            full_data = inv_part + stats + item_part
            
            original_index = inv_index_map.get(id(inv), -1)
            for col, val in enumerate(full_data):
                txt = str(val) if val is not None else ""
                item_widget = QTableWidgetItem(txt)
                item_widget.setData(Qt.UserRole, original_index)
                self.table.setItem(i, col, item_widget)
        
        self.table.blockSignals(False)
        self.result_label.setText(f"匹配: {len(rows)} 条明细数据")

    def _show_aggregated_list(self, rows, mode, col_index=-1):
        """显示汇总统计，包含收款数据 (支持CNPJ归一化)"""
        # groups: key -> {count, total_amount, total_tax, paid_amount, balance, seen_invoices, display_name_counts}
        groups = {}
        
        # 识别是否需要通过 CNPJ 归一化
        # 获取表头名称
        target_header = ""
        if col_index >= 0 and col_index < len(self.main_window.combined_headers):
            target_header = self.main_window.combined_headers[col_index]
            
        use_cnpj_grouping = False
        is_issuer = False
        
        if target_header == "发货方名称" or target_header == "收货方名称":
            use_cnpj_grouping = True
            is_issuer = (target_header == "发货方名称")

        # --- 汇总判定：按往来单位汇总时，智能识别‘对方’单位 ---
        use_partner_logic = False
        my_cnpj = None
        if mode == "STD_PARTNER":
            use_partner_logic = True
            my_cnpj = self.main_window.identify_self_cnpj()

        for inv, item in rows:
            key = "未知"
            # 原始显示 (用于统计最常用的名 
            display_val = ""
            
            amount = item.valor_total if item.valor_total is not None else 0.0
            tax = (item.valor_icms or 0) + (item.valor_ipi or 0)
            
            # Get grouping key
            if mode == "MONTH": # Month
                date = inv.data_emissao
                if date:
                    parts = date.split('/')
                    key = f"{parts[2]}-{parts[1]}" if len(parts) >= 3 else date[:7]
                else: key = "无日期"
                display_val = key

            elif mode == "STD_PRODUCT":
                raw_code = item.codigo_produto or ""
                prod_std = self.main_window.mapping_mgr.get_product_std(raw_code)
                if prod_std and prod_std.get("std_name"):
                     key = prod_std.get("std_name")
                else:
                     key = raw_code or "未分类产 (Unmapped)"
                display_val = key

            elif mode == "STD_PARTNER":
                raw_cnpj, raw_name = self.main_window.get_other_party(inv, my_cnpj)
                
                clean_cnpj = re.sub(r'\D', '', raw_cnpj)
                
                std_name = None
                if clean_cnpj:
                    std_name = self.main_window.mapping_mgr.get_partner_std(clean_cnpj)
                
                if not std_name:
                    std_name = self.main_window.mapping_mgr.get_partner_std(raw_name)

                if std_name:
                    key = std_name
                else:
                    key = raw_name
                display_val = key
                
            elif mode == "COLUMN" and col_index >= 0:
                if use_cnpj_grouping:
                    # 使用 CNPJ 作为 Key
                    cnpj = inv.emitente_cnpj if is_issuer else inv.destinatario_cnpj
                    name = inv.emitente_nome if is_issuer else inv.destinatario_nome
                    
                    if cnpj:
                        # 简单清 CNPJ，只保留数字
                        clean_cnpj = re.sub(r'\D', '', cnpj)
                        if clean_cnpj:
                            key = clean_cnpj
                        else:
                            key = name or "未知 (无CNPJ)"
                    else:
                        key = name or "未知 (无CNPJ)"
                        
                    display_val = name or "(空名 "
                else:
                    # 常规提取
                    row_data = self.main_window.get_row_from_invoice_and_item(inv, item)
                    if col_index < len(row_data):
                        val = row_data[col_index]
                        key = str(val) if val is not None else ""
                        if not key: key = "(Empty)"
                    display_val = key
            
            if key not in groups:
                groups[key] = {
                    'count': 0, 'amount': 0.0, 'tax': 0.0, 
                    'paid': 0.0, 'balance': 0.0, 'seen_invs': set(),
                    'names': {}
                }
            g = groups[key]
            g['count'] += 1
            g['amount'] += amount
            g['tax'] += tax
            
            # 记录 Display Value 用于后续选出最常用
            if display_val:
                g['names'][display_val] = g['names'].get(display_val, 0) + 1
            
            # Aggregate payment info
            inv_key = inv.numero or f"FILE-{inv.file_name}"
            if inv_key not in g['seen_invs']:
                p_data = self._status_map.get(inv_key)
                if isinstance(p_data, dict):
                    g['paid'] += p_data.get('paid_amount', 0.0)
                    g['balance'] += p_data.get('balance', 0.0)
                    # 统计已结清数量
                    if p_data.get('balance', 0.0) <= 0.01 and p_data.get('paid_amount', 0.0) > 0:
                        g['fully_paid_invs'] = g.get('fully_paid_invs', 0) + 1
                g['seen_invs'].add(inv_key)
            
        # 显示
        headers = ["分组名称 (Group)", "明细数 (Items)", "发票总数", "已结清数", "产品总额", "总税额", "发票已收", "发票余额", "收款进度"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(groups))
        
        sorted_keys = sorted(groups.keys())
        
        self.table.blockSignals(True)
        for i, key in enumerate(sorted_keys):
            data = groups[key]
            
            # 决定最终显示名 (优先从映射库取标准名)
            final_name = key
            
            if not str(mode).startswith("STD_"):
                # 1. 尝试从映射库找标准名
                std_mapping = self.main_window.mapping_mgr.get_partner_std(key)
                if not std_mapping:
                    # 如果是产品编码，也查一下
                    prod_std = self.main_window.mapping_mgr.get_product_std(key)
                    if prod_std:
                        std_mapping = prod_std.get("std_name")

                if std_mapping:
                    final_name = f"{std_mapping}\n[Key: {key}]"
                elif data['names']:
                    # 找出出现次数最多的名字作为备选
                    most_common_name = max(data['names'], key=data['names'].get)
                    if use_cnpj_grouping and re.match(r'^\d+$', key): # key is clean CNPJ
                        fmt_cnpj = key
                        if len(key) == 14:
                            fmt_cnpj = f"{key[:2]}.{key[2:5]}.{key[5:8]}/{key[8:12]}-{key[12:]}"
                        final_name = f"{most_common_name}\n[Tax ID: {fmt_cnpj}]"
                        if len(data['names']) > 1:
                            final_name += f" (+{len(data['names'])-1} 别名)"
                    else:
                        final_name = most_common_name
            else:
                # Standard Mode: Key is already the name
                # append count of raw variations if interesting?
                pass
            
            self.table.setItem(i, 0, QTableWidgetItem(str(final_name)))
            self.table.setItem(i, 1, QTableWidgetItem(str(data['count'])))
            self.table.setItem(i, 2, QTableWidgetItem(str(len(data['seen_invs'])))) # Invoices Count
            
            # 对齐 9 列标题: 0:分组, 1:明细数, 2:发票数, 3:已结清数, 4:产品总额, 5:总税额, 6:发票已收, 7:发票余额, 8:进度
            self.table.setItem(i, 3, QTableWidgetItem(str(data.get('fully_paid_invs', 0))))
            self.table.setItem(i, 4, QTableWidgetItem(f"{data['amount']:.2f}"))
            self.table.setItem(i, 5, QTableWidgetItem(f"{data['tax']:.2f}"))
            self.table.setItem(i, 6, QTableWidgetItem(f"{data['paid']:.2f}"))
            self.table.setItem(i, 7, QTableWidgetItem(f"{data['balance']:.2f}"))
            
            # 汇总模式下进度显示百分比或比例
            prog_text = ""
            if len(data['seen_invs']) > 0:
                perc = (data.get('fully_paid_invs', 0) / len(data['seen_invs'])) * 100
                prog_text = f"{perc:.1f}%"
            self.table.setItem(i, 8, QTableWidgetItem(prog_text))
            
        self.table.blockSignals(False)
        self.table.setColumnWidth(0, 300) # 固定第一列宽度，提高性能
        self.result_label.setText(f"汇总组  {len(groups)}")
        return groups

    def _build_status_map(self) -> Dict[str, dict]:
        """构建包含金额的收款状态图"""
        status_map = {}
        try:
            db_invoices = self.main_window.payment_tab.db.get_invoices()
        except Exception:
            db_invoices = []

        for row in db_invoices:
            paid = row['total_paid_amount'] or 0.0
            total = row['total_amount'] or 0.0
            data = {
                'progress': f"{row['paid_terms_count']}/{row['total_terms_count']}",
                'paid_amount': paid,
                'balance': total - paid
            }
            if row['invoice_number']:
                status_map[row['invoice_number']] = data
            if row['file_name']:
                status_map[f"FILE-{row['file_name']}"] = data
        return status_map

    def _get_full_row_data(self, invoice: Invoice, item: Item) -> list:
        """构建完整的行数据 (详细列表模式)"""
        raw_row = self.main_window.get_row_from_invoice_and_item(invoice, item)
        inv_headers_count = len(self.main_window.invoice_headers)
        split_index = inv_headers_count - 1 # 排除最后的"收款进度"         
        inv_part = raw_row[:split_index]
        item_part = raw_row[split_index:]
        
        key = invoice.numero or f"FILE-{invoice.file_name}"
        p_data = self._status_map.get(key, {})
        
        if isinstance(p_data, dict):
            progress_val = p_data.get('progress', "N/A")
            # 在详细列表中也加入已收和待收数据
            payment_stats = [p_data.get('paid_amount', 0.0), p_data.get('balance', 0.0)]
        else:
            progress_val = "N/A"
            payment_stats = [0.0, 0.0]
        
        # 组装：发票信 + 收款进度 + 已收 + 待收 + 产品信息
        return inv_part + [progress_val] + payment_stats + item_part

class ColumnConfigDialog(QDialog):
    """列配置对话框：排序和显示/隐藏"""
    def __init__(self, table_widget: QTableWidget, parent=None):
        super().__init__(parent)
        self.table_widget = table_widget
        self.setWindowTitle("列设 (拖拽排序 / 勾选显 ")
        self.resize(400, 600)
        
        layout = QVBoxLayout(self)
        
        # 说明
        layout.addWidget(QLabel("提示：拖拽列表项可调整列顺序，取消勾选可隐藏列。"))
        
        # 列表控件
        from PySide6.QtWidgets import QListWidget, QListWidgetItem
        self.list_widget = QListWidget()
        self.list_widget.setDragDropMode(QAbstractItemView.InternalMove) # 允许内部拖拽排序
        layout.addWidget(self.list_widget)
        
        # 快捷选择按钮
        sel_btn_layout = QHBoxLayout()
        sel_all_btn = QPushButton("全")
        sel_none_btn = QPushButton("全不")
        sel_btn_layout.addWidget(sel_all_btn)
        sel_btn_layout.addWidget(sel_none_btn)
        sel_btn_layout.addStretch()
        layout.addLayout(sel_btn_layout)
        
        sel_all_btn.clicked.connect(lambda: self.set_all_checked(True))
        sel_none_btn.clicked.connect(lambda: self.set_all_checked(False))
        
        # 获取当前表头状态
        header = self.table_widget.horizontalHeader()
        count = self.table_widget.columnCount()
        
        # 构建 (visual_index, logical_index) 列表，以便按当前视觉顺序显示
        visual_order = []
        for logical_index in range(count):
            visual_index = header.visualIndex(logical_index)
            visual_order.append((visual_index, logical_index))
        
        # 按视觉顺序排列
        visual_order.sort(key=lambda x: x[0])
        
        for _, logical_index in visual_order:
            item_text = self.table_widget.horizontalHeaderItem(logical_index).text()
            item = QListWidgetItem(item_text)
            item.setData(Qt.UserRole, logical_index) # 存储逻辑索引
            
            # 设置勾选状态
            is_hidden = self.table_widget.isColumnHidden(logical_index)
            item.setCheckState(Qt.Unchecked if is_hidden else Qt.Checked)
            
            self.list_widget.addItem(item)
            
        # 按钮
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("应用")
        cancel_btn = QPushButton("取消")
        
        ok_btn.clicked.connect(self.accept)
        cancel_btn.clicked.connect(self.reject)
        
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)

    def set_all_checked(self, checked):
        state = Qt.Checked if checked else Qt.Unchecked
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            item.setCheckState(state)

    def apply_settings(self):
        """应用设置到表"""
        header = self.table_widget.horizontalHeader()
        
        # 1. 应用显隐
        # 2. 应用顺序
        # 注意：moveSection 是基于当前视觉位置移动         # 最简单的方法是：先隐藏，再移动         # 为了保证移动正确，我们根 list_widget 的顺序重         
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            logical_index = item.data(Qt.UserRole)
            is_checked = (item.checkState() == Qt.Checked)
            
            # 设置显隐
            self.table_widget.setColumnHidden(logical_index, not is_checked)
            
            # 设置顺序
            # 目标：将 logical_index 移动到视觉位 i
            current_visual_index = header.visualIndex(logical_index)
            if current_visual_index != i:
                header.moveSection(current_visual_index, i)

class UserDashboardTab(QWidget):
    """用户中心 / 仪表"""
    def __init__(self, mainwindow):
        super().__init__()
        self.mainwindow = mainwindow
        self.layout = QHBoxLayout(self)
        
        # --- Left: User Profile ---
        profile_frame = QFrame()
        profile_frame.setFrameShape(QFrame.StyledPanel)
        profile_frame.setFixedWidth(250)
        p_layout = QVBoxLayout(profile_frame)
        
        # Avatar (Placeholder)
        avatar_lbl = QLabel("👤")
        avatar_lbl.setAlignment(Qt.AlignCenter)
        font = QFont()
        font.setPointSize(48)
        avatar_lbl.setFont(font)
        p_layout.addWidget(avatar_lbl)
        
        user_lbl = QLabel("管理 (Admin)")
        user_lbl.setAlignment(Qt.AlignCenter)
        user_lbl.setStyleSheet("font-weight: bold; font-size: 14pt;")
        p_layout.addWidget(user_lbl)
        
        p_layout.addStretch()
        
        # Stats summary
        self.total_lbl = QLabel("总发票数: 0")
        self.untagged_lbl = QLabel("无标  0")
        p_layout.addWidget(self.total_lbl)
        p_layout.addWidget(self.untagged_lbl)
        
        p_layout.addStretch()
        self.layout.addWidget(profile_frame)
        
        # --- Right: Tag Stats ---
        stats_frame = QFrame()
        stats_layout = QVBoxLayout(stats_frame)
        stats_layout.addWidget(QLabel("<b>标签统计 (双击跳转)</b>"))
        
        self.tags_table = QTableWidget()
        self.tags_table.setColumnCount(3)
        self.tags_table.setHorizontalHeaderLabels(["标签名称", "颜色", "关联数量"])
        self.tags_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tags_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tags_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.tags_table.cellDoubleClicked.connect(self.on_tag_double_clicked)
        
        stats_layout.addWidget(self.tags_table)
        
        # Refresh Btn
        btn_layout = QHBoxLayout()
        self.btn_add_tag = QPushButton("➕ 新增标签")
        self.btn_edit_tag = QPushButton("✏️ 修改标签")
        self.btn_del_tag = QPushButton("🗑️ 删除标签")
        
        refresh_btn = QPushButton("🔄 刷新统计")
        refresh_btn.clicked.connect(self.refresh_stats)
        export_btn = QPushButton("📤 导出统计")
        export_btn.clicked.connect(lambda: export_qtable(self.tags_table, self, "tag_stats"))
        
        self.btn_add_tag.clicked.connect(self.add_tag)
        self.btn_edit_tag.clicked.connect(self.edit_selected_tag)
        self.btn_del_tag.clicked.connect(self.delete_selected_tag)

        btn_layout.addWidget(self.btn_add_tag)
        btn_layout.addWidget(self.btn_edit_tag)
        btn_layout.addWidget(self.btn_del_tag)
        btn_layout.addStretch()
        btn_layout.addWidget(refresh_btn)
        btn_layout.addWidget(export_btn)
        stats_layout.addLayout(btn_layout)
        
        self.layout.addWidget(stats_frame)

    def refresh_stats(self):
        invoices = self.mainwindow.invoices
        if not invoices: return
        
        self.total_lbl.setText(f"总发票数: {len(invoices)}")
        
        tag_counts = {}
        untagged_count = 0
        
        # Init with known tags
        for t in self.mainwindow.tag_mgr.get_tags():
            tag_counts[t['name']] = 0
            
        for inv in invoices:
            has_tag = False
            # Check invoice tags
            if inv.tags:
                has_tag = True
                for t in inv.tags:
                    tag_counts[t] = tag_counts.get(t, 0) + 1
            
            # Check item tags
            for item in inv.itens:
                if item.tags:
                    has_tag = True
                    for t in item.tags:
                        tag_counts[t] = tag_counts.get(t, 0) + 1
            
            if not has_tag:
                untagged_count += 1
                
        self.untagged_lbl.setText(f"无标  {untagged_count}")
        
        self.tags_table.setRowCount(len(tag_counts))
        # Sort by count desc
        sorted_tags = sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)
        
        # Get colors map
        color_map = {t['name']: t['color'] for t in self.mainwindow.tag_mgr.get_tags()}        
        for i, (name, count) in enumerate(sorted_tags):
            self.tags_table.setItem(i, 0, QTableWidgetItem(name))
            
            color = color_map.get(name, "#FFFFFF")
            color_item = QTableWidgetItem(color)
            color_item.setBackground(QColor(color))
            self.tags_table.setItem(i, 1, color_item)
            
            self.tags_table.setItem(i, 2, QTableWidgetItem(str(count)))

    def on_tag_double_clicked(self, row, col):
        tag_name = self.tags_table.item(row, 0).text()
        self.mainwindow.filter_by_tag(tag_name)

    def add_tag(self):
        name, ok = QInputDialog.getText(self, "新增标签", "请输入标签名称:")
        if ok and name.strip():
            # Let user pick a color or use random
            from PySide6.QtWidgets import QColorDialog
            color = QColorDialog.getColor(Qt.white, self, "选择标签颜色")
            if color.isValid():
                self.mainwindow.tag_mgr.add_tag(name.strip(), color.name())
                self.refresh_stats()

    def edit_selected_tag(self):
        row = self.tags_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一个要编辑的标签")
            return
        
        old_name = self.tags_table.item(row, 0).text()
        # Find current tag data
        tag_data = next((t for t in self.mainwindow.tag_mgr.get_tags() if t['name'] == old_name), None)
        if not tag_data: return

        new_name, ok = QInputDialog.getText(self, "修改标签", "标签名称:", text=old_name)
        if ok and new_name.strip():
            from PySide6.QtWidgets import QColorDialog
            color = QColorDialog.getColor(QColor(tag_data['color']), self, "选择新颜色")
            if color.isValid():
                # Remove old and add new (or update in mgr)
                self.mainwindow.tag_mgr.remove_tag(old_name)
                self.mainwindow.tag_mgr.add_tag(new_name.strip(), color.name())
                # Update any invoices using this tag if needed? 
                # For now, just update the library.
                self.refresh_stats()

    def delete_selected_tag(self):
        row = self.tags_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "提示", "请先选择一个要删除的标签")
            return
        
        tag_name = self.tags_table.item(row, 0).text()
        reply = QMessageBox.question(self, "确认删除", f"确定要彻底删除标签 '{tag_name}' 吗？\n这不会清除已标记发票上的文字，但该标签将从库中消失。", 
                                   QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.mainwindow.tag_mgr.remove_tag(tag_name)
            self.refresh_stats()

class AccountDialog(QDialog):
    """账户添加/编辑对话"""
    def __init__(self, account_data=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("账户信息")
        self.resize(400, 300)
        
        self.account_id = account_data['id'] if account_data else None
        
        layout = QVBoxLayout(self)
        from PySide6.QtWidgets import QFormLayout
        form = QFormLayout()
        
        self.name_edit = QLineEdit()
        if account_data: self.name_edit.setText(account_data['name'])
        form.addRow("账户名称:", self.name_edit)
        
        self.bank_edit = QLineEdit()
        self.bank_edit.setPlaceholderText("银行名称 / 账号信息")
        if account_data: self.bank_edit.setText(account_data['bank_info'])
        form.addRow("银行信息:", self.bank_edit)
        
        self.currency_edit = QComboBox()
        self.currency_edit.addItems(["BRL", "USD", "CNY", "EUR"])
        if account_data: self.currency_edit.setCurrentText(account_data['currency'])
        form.addRow("币种:", self.currency_edit)
        
        self.balance_spin = QDoubleSpinBox()
        self.balance_spin.setRange(-999999999.00, 999999999.00)
        self.balance_spin.setPrefix("R$ ")
        if account_data: 
            self.balance_spin.setValue(account_data['initial_balance'])
            self.balance_spin.setEnabled(False) # 初始余额通常创建后不可改? 或者允许改?
        form.addRow("初始余额:", self.balance_spin)
        
        self.note_edit = QTextEdit()
        self.note_edit.setMaximumHeight(80)
        if account_data: self.note_edit.setPlainText(account_data['note'])
        form.addRow("备注:", self.note_edit)
        
        layout.addLayout(form)
        
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)
        
    def get_data(self):
        return {
            "name": self.name_edit.text(),
            "bank_info": self.bank_edit.text(),
            "currency": self.currency_edit.currentText(),
            "initial_balance": self.balance_spin.value(),
            "note": self.note_edit.toPlainText()
        }
class PaymentRegistrationDialog(QDialog):
    """收款登记对话 (关联账户)"""
    def __init__(self, db_manager, default_amount=0.0, parent=None):
        super().__init__(parent)
        self.db = db_manager
        self.setWindowTitle("登记收款")
        self.resize(400, 250)
        
        layout = QVBoxLayout(self)
        from PySide6.QtWidgets import QFormLayout, QDoubleSpinBox, QDateEdit
        
        form = QFormLayout()
        
        self.account_combo = QComboBox()
        # Load accounts
        accounts = self.db.get_accounts()
        if not accounts:
            self.account_combo.addItem("无账 (请先创建账户)", None)
            self.account_combo.setEnabled(False)
        else:
            for acc in accounts:
                self.account_combo.addItem(f"{acc['name']} (余额: {acc['current_balance']:.2f})", acc['id'])
        
        form.addRow("收款账户:", self.account_combo)
        
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(datetime.now().date())
        self.date_edit.setDisplayFormat("yyyy/MM/dd")
        form.addRow("收款日期:", self.date_edit)
        
        self.amount_spin = QDoubleSpinBox()
        self.amount_spin.setRange(0.0, 999999999.0)
        self.amount_spin.setValue(default_amount)
        form.addRow("收款金额:", self.amount_spin)

        self.penalty_spin = QDoubleSpinBox()
        self.penalty_spin.setRange(0.0, 999999999.0)
        self.penalty_spin.setValue(0.0)
        self.penalty_spin.setPrefix("+")
        form.addRow("登记滞纳 ", self.penalty_spin)
        
        self.note_edit = QLineEdit()
        form.addRow("备注:", self.note_edit)
        
        layout.addLayout(form)
        
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)
        
    def get_data(self):
        return {
            "account_id": self.account_combo.currentData(),
            "date": self.date_edit.date().toString("yyyy-MM-dd"),
            "amount": self.amount_spin.value(),
            "penalty": self.penalty_spin.value(),
            "note": self.note_edit.text()
        }
class ReceivablesStatusTab(QWidget):
    """应收情况分析模块 (整合账龄与预收账款)"""
    def __init__(self, main_window):
        super().__init__(main_window)
        self.main_window = main_window
        self.db = self.main_window.payment_tab.db
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # 顶部控制栏
        controls = QHBoxLayout()
        controls.addWidget(QLabel("<b>应收情况分析 (Receivables & Advances)</b>"))
        
        controls.addSpacing(20)
        # 模式切换
        self.mode_group = QButtonGroup(self)
        self.rb_smart = QRadioButton("智能识别 (Smart)")
        self.rb_customer = QRadioButton("分析客户 (收货方)")
        self.rb_supplier = QRadioButton("分析供应商 (开票方)")
        self.rb_smart.setChecked(True)
        self.mode_group.addButton(self.rb_smart, 2)
        self.mode_group.addButton(self.rb_customer, 0)
        self.mode_group.addButton(self.rb_supplier, 1)
        
        controls.addWidget(self.rb_smart)
        controls.addWidget(self.rb_customer)
        controls.addWidget(self.rb_supplier)
        
        controls.addStretch()
        
        self.refresh_btn = QPushButton("🔄 刷新分析")
        self.refresh_btn.clicked.connect(self.refresh_analysis)
        controls.addWidget(self.refresh_btn)
        
        self.export_btn = QPushButton("📤 导出报表")
        self.export_btn.clicked.connect(lambda: export_qtable(self.table, self, "receivables_status_report"))
        controls.addWidget(self.export_btn)
        
        layout.addLayout(controls)

        # 统计表格
        self.table = QTableWidget()
        headers = [
            "往来单位名称", "纳税识别号 (CNPJ)", "发票未结 (Debt)", "预收余额 (Advances)", "净应收 (Net Receivable)",
            "未到期 (Current)", "1-30天", "31-60天", "61-90天", "90天+"
        ]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setSortingEnabled(True)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        
        layout.addWidget(self.table)
        
        # 启用右键菜单
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
        # 底部说明
        footer = QLabel("注：净应收 = 发票未结 - 预收余额。红色表示净欠款，绿色表示净预收(有结余)。")
        footer.setStyleSheet("color: gray; font-style: italic;")
        layout.addWidget(footer)
        
        # 注册表格统计
        if hasattr(self.main_window, 'register_table_for_stats'):
            self.main_window.register_table_for_stats(self.table)

    def refresh_analysis(self):
        # 1. 获取所有基础数据
        # Refactored for Network Support
        rows = self.db.get_all_installments_extended()
        
        # B. 预收账款数据
        advances = self.db.get_advances_by_customer()
        
        # conn.close() - Removed as db handles connection

        # 2. 确定分析对象
        analyze_mode = self.mode_group.checkedId()
        my_cnpj = None
        if analyze_mode == 2: # Smart
            my_cnpj = self.main_window.identify_self_cnpj()
        
        # 3. 聚合数据
        partners = {} # key -> {name, cnpj, debt, advances, buckets}
        now = datetime.now().date()
        
        # 处理发票欠款
        if rows:
            for r in rows:
                if analyze_mode == 2: # Smart
                    dest_clean = re.sub(r'\D', '', r['destinatario_cnpj'] or "")
                    emit_clean = re.sub(r'\D', '', r['issuer_cnpj'] or "")
                    if my_cnpj and dest_clean == my_cnpj:
                        name = r['issuer_name']
                        cnpj = r['issuer_cnpj']
                    elif my_cnpj and emit_clean == my_cnpj:
                        name = r['destinatario_name']
                        cnpj = r['destinatario_cnpj']
                    else:
                        # Fallback to customer mode
                        name = r['destinatario_name']
                        cnpj = r['destinatario_cnpj']
                elif analyze_mode == 0: # Customer
                    name = r['destinatario_name']
                    cnpj = r['destinatario_cnpj']
                else: # Supplier
                    name = r['issuer_name']
                    cnpj = r['issuer_cnpj']
                
                if not name: name = "Unknown"
                
                # 使用 Mapping 归一化名称
                clean_key = re.sub(r'\D', '', cnpj) if cnpj else name
                std_name = self.main_window.mapping_mgr.get_partner_std(clean_key)
                display_name = std_name if std_name else name
                
                group_key = display_name
                
                if group_key not in partners:
                    partners[group_key] = {
                        'name': display_name,
                        'cnpj': cnpj,
                        'debt': 0.0,
                        'advances': 0.0,
                        'buckets': [0.0] * 5
                    }
                
                p = partners[group_key]
                amount = r['amount'] or 0.0
                paid = r['paid_amount'] or 0.0
                balance = amount - paid
                
                p['debt'] += balance
                
                if balance > 0.01:
                    # 计算账龄
                    try:
                        due = datetime.strptime(r['due_date'], "%Y-%m-%d").date()
                        days_overdue = (now - due).days
                        
                        if days_overdue <= 0:
                            p['buckets'][0] += balance # 未到期
                        elif days_overdue <= 30:
                            p['buckets'][1] += balance # 1-30
                        elif days_overdue <= 60:
                            p['buckets'][2] += balance # 31-60
                        elif days_overdue <= 90:
                            p['buckets'][3] += balance # 61-90
                        else:
                            p['buckets'][4] += balance # 90+
                    except:
                        p['buckets'][0] += balance # 解析失败归入当前

        # 处理预收账款 (在客户模式或智能模式下处理)
        if (analyze_mode in [0, 2]) and advances:
            for adv in advances:
                name = adv['customer_name']
                cnpj = adv['customer_cnpj']
                rem = adv['remaining_amount']
                
                if not name: name = "Unknown"
                clean_key = re.sub(r'\D', '', cnpj) if cnpj else name
                std_name = self.main_window.mapping_mgr.get_partner_std(clean_key)
                display_name = std_name if std_name else name
                
                group_key = display_name
                
                if group_key not in partners:
                    partners[group_key] = {
                        'name': display_name,
                        'cnpj': cnpj,
                        'debt': 0.0,
                        'advances': 0.0,
                        'buckets': [0.0] * 5
                    }
                
                partners[group_key]['advances'] += rem

        # 4. 填充表格
        settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)
        overdue_color = QColor(settings.value("color_overdue", "#ffaa00"))
        
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(partners))
        for i, (key, p) in enumerate(partners.items()):
            net_receivable = p['debt'] - p['advances']
            
            items = [
                QTableWidgetItem(p['name']),
                QTableWidgetItem(p['cnpj']),
                QTableWidgetItem(f"{p['debt']:,.2f}"),
                QTableWidgetItem(f"{p['advances']:,.2f}"),
                QTableWidgetItem(f"{net_receivable:,.2f}"),
                QTableWidgetItem(f"{p['buckets'][0]:,.2f}"),
                QTableWidgetItem(f"{p['buckets'][1]:,.2f}"),
                QTableWidgetItem(f"{p['buckets'][2]:,.2f}"),
                QTableWidgetItem(f"{p['buckets'][3]:,.2f}"),
                QTableWidgetItem(f"{p['buckets'][4]:,.2f}")
            ]
            
            # 格式化与配色
            # 预收余额 (Col 3): 绿色
            if p['advances'] > 0.01:
                items[3].setForeground(QColor("green"))
                items[3].setFont(QFont("", -1, QFont.Bold))
            
            # 净应收 (Col 4): 红色(欠款) / 绿色(溢出)
            if net_receivable > 0.01:
                items[4].setForeground(QColor("#ff5555")) # Red
                items[4].setFont(QFont("", -1, QFont.Bold))
            elif net_receivable < -0.01:
                items[4].setForeground(QColor("green")) # Green
                items[4].setFont(QFont("", -1, QFont.Bold))
            
            # 逾期列 (Col 6-9)
            for col in range(6, 10):
                if p['buckets'][col-5] > 0:
                    items[col].setForeground(overdue_color)
            
            # 右对齐数值列
            for col in range(2, 10):
                items[col].setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

            for col, item in enumerate(items):
                self.table.setItem(i, col, item)
        
        self.table.setSortingEnabled(True)

    def show_context_menu(self, pos):
        item = self.table.itemAt(pos)
        if not item: return
        
        menu = QMenu(self)
        action_drill = menu.addAction("🔍 查看该单位明细数据 (Drill Down)")
        action_drill.triggered.connect(lambda: self.drill_down(item.row()))
        
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def drill_down(self, row):
        name_item = self.table.item(row, 0)
        cnpj_item = self.table.item(row, 1)
        if not name_item: return
        
        # 优先使用 CNPJ 搜索，否则使用名称
        search_term = cnpj_item.text() if cnpj_item and cnpj_item.text() else name_item.text()
        
        # 切换到数据查询 Tab
        self.main_window.tab_widget.setCurrentWidget(self.main_window.query_tab)
        self.main_window.query_tab.search_input.setText(search_term)
        self.main_window.query_tab.apply_filter()

class AccountManagerTab(QWidget):
    """账户管理模块"""
    def __init__(self, main_window):
        super().__init__(main_window)
        self.main_window = main_window
        # Assuming db is accessible via PaymentTab or shared
        self.db = self.main_window.payment_tab.db
        self._trans_load_seq = 0
        self._trans_view_state = {}
        self._trans_worker = None
        self._pending_invoice_filter_id = None
        self._pending_filter_account_id = None
        
        self.setup_ui()
        self.load_accounts()

    def setup_ui(self):
        layout = QHBoxLayout(self)
        
        # --- Left: Account List ---
        left_panel = QGroupBox("账户列表")
        left_layout = QVBoxLayout(left_panel)
        
        self.account_list = QListWidget()
        self.account_list.itemClicked.connect(self.on_account_selected)
        left_layout.addWidget(self.account_list)
        
        btn_layout = QHBoxLayout()
        add_btn = QPushButton(" 新建")
        edit_btn = QPushButton("✏️ 编辑")
        del_btn = QPushButton("🗑 删除")
        
        add_btn.clicked.connect(self.add_account)
        edit_btn.clicked.connect(self.edit_account)
        del_btn.clicked.connect(self.delete_account)
        
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(del_btn)
        left_layout.addLayout(btn_layout)
        
        # --- Right: Transactions ---
        right_panel = QGroupBox("账户流水 (Transactions)")
        right_layout = QVBoxLayout(right_panel)
        
        # Filter / Info
        self.balance_lbl = QLabel("当前余额: 0.00")
        self.balance_lbl.setStyleSheet("font-size: 12pt; font-weight: bold;")
        right_layout.addWidget(self.balance_lbl)
        
        self.trans_table = QTableWidget()
        self.trans_table.setColumnCount(8)
        self.trans_table.setHorizontalHeaderLabels(["ID", "日期", "摘要 (Description)", "发票号", "往来单位", "借方 (Income)", "贷方 (Expense)", "余额 (Balance)"])
        self.trans_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.trans_table.setColumnWidth(3, 100) # 发票号
        self.trans_table.setColumnWidth(4, 150) # 往来单位
        self.trans_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.trans_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.trans_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.trans_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.trans_table.customContextMenuRequested.connect(self.show_context_menu)
        # Double click to reset filter
        self.trans_table.itemDoubleClicked.connect(lambda: [self.trans_table.setRowHidden(r, False) for r in range(self.trans_table.rowCount())] or self.load_transactions(self.current_account_id))
        
        right_layout.addWidget(self.trans_table)
        
        # Transaction Actions
        t_btn_layout = QHBoxLayout()
        add_inc_btn = QPushButton(" 记收")
        add_exp_btn = QPushButton(" 记支")
        del_trans_btn = QPushButton("🗑 删除明细")
        export_btn = QPushButton("📤 导出流水")
        
        add_inc_btn.clicked.connect(lambda: self.add_manual_transaction('Income'))
        add_exp_btn.clicked.connect(lambda: self.add_manual_transaction('Expense'))
        del_trans_btn.clicked.connect(self.delete_transaction)
        export_btn.clicked.connect(self.export_transactions)
        
        t_btn_layout.addWidget(add_inc_btn)
        t_btn_layout.addWidget(add_exp_btn)
        t_btn_layout.addWidget(export_btn)
        t_btn_layout.addStretch()
        t_btn_layout.addWidget(del_trans_btn)
        right_layout.addLayout(t_btn_layout)
        
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)
        layout.addWidget(splitter)
        
        # 注册表格统计
        if hasattr(self.main_window, 'register_table_for_stats'):
            self.main_window.register_table_for_stats(self.trans_table)
        
        self.current_account_id = None

    def load_accounts(self):
        self.account_list.clear()
        accounts = self.db.get_accounts()
        for acc in accounts:
            item = QListWidgetItem(f"{acc['name']} ({acc['currency']})")
            # Store ID and full data
            item.setData(Qt.UserRole, acc['id'])
            item.setData(Qt.UserRole + 1, dict(acc))
            self.account_list.addItem(item)
            
        if self.current_account_id:
            # Try to reselect
            for i in range(self.account_list.count()):
                item = self.account_list.item(i)
                if item.data(Qt.UserRole) == self.current_account_id:
                    self.account_list.setCurrentItem(item)
                    self.load_transactions(self.current_account_id)
                    break

    def on_account_selected(self, item):
        self.current_account_id = item.data(Qt.UserRole)
        self._pending_invoice_filter_id = None
        self._pending_filter_account_id = None
        self.load_transactions(self.current_account_id)

    def filter_transactions_by_invoice(self, invoice_id):
        """Filters transaction table by invoice id (async safe)."""
        target_account_id = self.db.get_account_id_for_invoice(invoice_id)
        if not target_account_id:
            return False  # No transactions found for this invoice

        # 记录待执行筛选，待异步加载完成后应用
        self._pending_invoice_filter_id = invoice_id
        self._pending_filter_account_id = target_account_id

        # 切换左侧账户选中态
        for i in range(self.account_list.count()):
            item = self.account_list.item(i)
            if item.data(Qt.UserRole) == target_account_id:
                self.account_list.setCurrentItem(item)
                break

        self.current_account_id = target_account_id
        self.load_transactions(target_account_id, invoice_filter_id=invoice_id)
        return True

    def load_transactions(self, account_id, invoice_filter_id=None):
        if not account_id:
            self.trans_table.setRowCount(0)
            self.balance_lbl.setText("当前余额: 0.00")
            return

        if invoice_filter_id is None:
            self._pending_invoice_filter_id = None
            self._pending_filter_account_id = None
        else:
            self._pending_invoice_filter_id = invoice_filter_id
            self._pending_filter_account_id = account_id

        # --- 保存当前选中的交易ID和滚动条位置 ---
        selected_trans_id = None
        table = self.trans_table
        curr_row = table.currentRow()
        if curr_row >= 0:
            id_item = table.item(curr_row, 0)
            if id_item:
                stored_id = id_item.data(Qt.UserRole)
                if stored_id is not None:
                    selected_trans_id = stored_id
                else:
                    try:
                        selected_trans_id = int(id_item.text())
                    except Exception:
                        pass

        self._trans_load_seq += 1
        req_id = self._trans_load_seq
        self._trans_view_state[req_id] = {
            "account_id": account_id,
            "selected_trans_id": selected_trans_id,
            "v_scroll": table.verticalScrollBar().value(),
            "h_scroll": table.horizontalScrollBar().value(),
        }
        for key in list(self._trans_view_state.keys()):
            if key < req_id - 5:
                self._trans_view_state.pop(key, None)

        if self._trans_worker and self._trans_worker.isRunning():
            self._trans_worker.stop()

        worker = TransactionRowsLoadWorker(self.db, req_id, int(account_id), self)
        worker.finished.connect(self._on_transaction_rows_loaded)
        worker.error.connect(self._on_transaction_rows_error)
        worker.finished.connect(worker.deleteLater)
        worker.error.connect(worker.deleteLater)
        self._trans_worker = worker
        worker.start()

    def _on_transaction_rows_loaded(self, req_id, account_id, current_balance, rows):
        if req_id != self._trans_load_seq:
            return
        self._trans_worker = None

        self.balance_lbl.setText(f"当前余额: {current_balance:,.2f}")
        view_state = self._trans_view_state.pop(req_id, {})

        table = self.trans_table
        table.blockSignals(True)
        table.setUpdatesEnabled(False)
        table.clearContents()
        table.setRowCount(len(rows))

        # Calculate running balance for display (reverse calculation from current)
        running_balance = current_balance
        try:
            for i, row in enumerate(rows):
                trans_id = row.get('id')
                date_str = format_date_gui(str(row.get('date') or ""))
                desc = str(row.get('description') or "")
                inv_num = row.get('invoice_number') or "-"
                partner = row.get('destinatario_name') or row.get('issuer_name') or "-"

                try:
                    amount = float(row.get('amount') or 0.0)
                except Exception:
                    amount = 0.0

                table.setItem(i, 0, QTableWidgetItem(str(trans_id if trans_id is not None else "")))
                table.setItem(i, 1, QTableWidgetItem(date_str))
                table.setItem(i, 2, QTableWidgetItem(desc))
                table.setItem(i, 3, QTableWidgetItem(inv_num))
                table.setItem(i, 4, QTableWidgetItem(partner))

                income_val = ""
                expense_val = ""
                row_balance = running_balance

                if str(row.get('type')) == 'Income':
                    income_val = f"{amount:,.2f}"
                    running_balance -= amount
                else:
                    expense_val = f"{amount:,.2f}"
                    running_balance += amount

                inc_item = QTableWidgetItem(income_val)
                inc_item.setForeground(QBrush(QColor("green")))
                table.setItem(i, 5, inc_item)

                exp_item = QTableWidgetItem(expense_val)
                exp_item.setForeground(QBrush(QColor("red")))
                table.setItem(i, 6, exp_item)

                bal_item = QTableWidgetItem(f"{row_balance:,.2f}")
                bal_item.setFont(QFont(self.font().family(), -1, QFont.Bold))
                table.setItem(i, 7, bal_item)

                id_item = table.item(i, 0)
                if id_item:
                    id_item.setData(Qt.UserRole, trans_id)
                    id_item.setData(Qt.UserRole + 1, row.get('related_invoice_id'))
        finally:
            table.setUpdatesEnabled(True)
            table.blockSignals(False)

        # 恢复选中与滚动位置
        selected_trans_id = view_state.get("selected_trans_id")
        if selected_trans_id is not None:
            table.blockSignals(True)
            for r in range(table.rowCount()):
                item = table.item(r, 0)
                if not item:
                    continue
                curr_id = item.data(Qt.UserRole)
                if curr_id is None:
                    try:
                        curr_id = int(item.text())
                    except Exception:
                        curr_id = item.text()
                if str(curr_id) == str(selected_trans_id):
                    table.selectRow(r)
                    table.setCurrentCell(r, 0)
                    break
            table.blockSignals(False)

        table.verticalScrollBar().setValue(view_state.get("v_scroll", 0))
        table.horizontalScrollBar().setValue(view_state.get("h_scroll", 0))

        # 若存在待筛选请求，在加载完成后应用
        if (
            self._pending_invoice_filter_id is not None
            and str(self._pending_filter_account_id) == str(account_id)
        ):
            self._apply_transaction_invoice_filter(self._pending_invoice_filter_id)
            self._pending_invoice_filter_id = None
            self._pending_filter_account_id = None

    def _on_transaction_rows_error(self, req_id, account_id, error_msg):
        if req_id != self._trans_load_seq:
            return
        self._trans_worker = None
        if hasattr(self.main_window, "log_message"):
            self.main_window.log_message(f"加载账户流水失败 (account_id={account_id}): {error_msg}")
        else:
            logging.error("Failed to load transaction rows: %s", error_msg)

    def _apply_transaction_invoice_filter(self, invoice_id):
        match_count = 0
        for row in range(self.trans_table.rowCount()):
            id_cell = self.trans_table.item(row, 0)
            item_inv_id = id_cell.data(Qt.UserRole + 1) if id_cell else None
            is_match = str(item_inv_id) == str(invoice_id)
            self.trans_table.setRowHidden(row, not is_match)
            if is_match:
                match_count += 1
                self.trans_table.selectRow(row)

        if match_count > 0:
            self.balance_lbl.setText(f"筛选显  关联发票 #{invoice_id}  {match_count} 条记 (双击列表重置)")
            return True
        self.balance_lbl.setText(f"未找到关联发票 #{invoice_id} 的流水")
        return False

    def shutdown_workers(self):
        if self._trans_worker and self._trans_worker.isRunning():
            self._trans_worker.stop()
            self._trans_worker.wait(300)

    def export_transactions(self):
        if self.trans_table.rowCount() == 0:
            QMessageBox.warning(self, "无数据", "当前没有流水记录可导")
            return
        export_qtable(self.trans_table, self, "account_transactions")

    def add_account(self):
        dlg = AccountDialog(parent=self)
        if dlg.exec():
            data = dlg.get_data()
            new_id = self.db.add_account(**data)
            if not new_id:
                QMessageBox.warning(self, "失败", "账户创建失败（名称可能重复或连接异常）。")
                return
            self.load_accounts()

    def edit_account(self):
        item = self.account_list.currentItem()
        if not item: return
        data = item.data(Qt.UserRole + 1)
        
        dlg = AccountDialog(account_data=data, parent=self)
        if dlg.exec():
            new_data = dlg.get_data()
            ok = self.db.update_account(
                data['id'],
                new_data['name'],
                new_data['bank_info'],
                new_data['currency'],
                new_data['note'],
            )
            if ok is False:
                QMessageBox.warning(self, "失败", "账户更新失败（记录不存在或连接异常）。")
                return
            self.load_accounts()

    def delete_account(self):
        item = self.account_list.currentItem()
        if not item: return
        acc_id = item.data(Qt.UserRole)
        
        reply = QMessageBox.question(self, "确认", "确定要删 停用该账户吗？\n(如果存在交易记录，将仅标记为停用)", QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            ok = self.db.delete_account(acc_id)
            if ok is False:
                QMessageBox.warning(self, "失败", "账户删除/停用失败（记录不存在或连接异常）。")
                return
            self.load_accounts()
            self.trans_table.setRowCount(0)
            self.current_account_id = None

    def add_manual_transaction(self, t_type):
        if not self.current_account_id: return
        
        amount, ok = QInputDialog.getDouble(self, "金额", "请输入金 ", 0, 0, 999999999, 2)
        if not ok: return
        
        desc, ok = QInputDialog.getText(self, "描述", "请输入备 ")
        if not ok: return
        
        date = datetime.now().strftime("%Y-%m-%d")
        
        trans_id = self.db.add_transaction(self.current_account_id, date, t_type, amount, desc)
        if not trans_id:
            QMessageBox.warning(self, "失败", "流水新增失败，请检查输入或数据库连接。")
            return
        self.load_transactions(self.current_account_id)

    def delete_transaction(self):
        selected_rows = sorted(list(set(index.row() for index in self.trans_table.selectedIndexes())), reverse=True)
        if not selected_rows:
            return
        
        count = len(selected_rows)
        msg = f"确定要删除这 {count} 笔交易吗？余额将自动回滚。" if count > 1 else "确定要删除该交易吗？余额将回滚"
        
        reply = QMessageBox.question(self, "确认", msg, QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            failed_count = 0
            for row in selected_rows:
                trans_id = self.trans_table.item(row, 0).data(Qt.UserRole)
                ok = self.db.delete_transaction(trans_id)
                if ok is False:
                    failed_count += 1
                
            self.load_transactions(self.current_account_id)
            self.load_accounts() # 刷新账户余额列表显示

            # 自动同步刷新收款系统
            if self.main_window and hasattr(self.main_window, 'payment_tab'):
                self.main_window.payment_tab.load_invoices()
                # 如果当前收款系统正选中了某张发票，也尝试刷新其分期详情
                if self.main_window.payment_tab.current_invoice_id:
                    self.main_window.payment_tab.load_installments(self.main_window.payment_tab.current_invoice_id)

            if failed_count > 0:
                QMessageBox.warning(self, "部分失败", f"有 {failed_count} 笔流水删除失败，请检查连接后重试。")

    def show_context_menu(self, pos):
        selected_rows = sorted(list(set(index.row() for index in self.trans_table.selectedIndexes())))
        if not selected_rows: return
        
        menu = QMenu(self)
        
        if len(selected_rows) == 1:
            row = selected_rows[0]
            inv_id = self.trans_table.item(row, 0).data(Qt.UserRole + 1)
            if inv_id:
                action = menu.addAction("🔍 追踪关联发票 (Go to Invoice)")
                action.triggered.connect(lambda: self.trace_invoice(inv_id))
            
            del_action = menu.addAction("🗑 删除该交易")
            del_action.triggered.connect(self.delete_transaction)
        else:
            del_action = menu.addAction(f"🗑 批量删除 {len(selected_rows)} 笔交易")
            del_action.triggered.connect(self.delete_transaction)
            
        menu.exec(self.trans_table.viewport().mapToGlobal(pos))

    def trace_invoice(self, inv_id):
        # Find invoice in main window list
        # Need to map DB ID to Invoice Object
        # Actually navigate_to_payment uses index, but we can search by ID in payment tab
        
        self.main_window.tab_widget.setCurrentWidget(self.main_window.payment_tab)
        ptab = self.main_window.payment_tab
        
        for r in range(ptab.invoice_table.rowCount()):
            curr_id = ptab.invoice_table.item(r, 0).text()
            if str(curr_id) == str(inv_id):
                ptab.invoice_table.selectRow(r)
                ptab.on_invoice_selected(ptab.invoice_table.item(r, 0))
                break

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("DANFE 批量提取工具 v1.0")
        
        # 确保窗口可调整大小，并设置最小尺寸
        self.resize(1200, 800)
        self.setMinimumSize(800, 600)
        
        # 设置图标 (添加异常处理避免图标损坏影响启动)
        try:
            icon_data = QByteArray.fromBase64(APP_ICON_B64.encode())
            image = QImage.fromData(icon_data)
            if not image.isNull():
                icon = QIcon(QPixmap.fromImage(image))
                self.setWindowIcon(icon)
        except Exception:
            pass  # 图标加载失败不影响程序运行

        # 主窗口和布局
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # --- 顶部控制 ---
        top_layout = QHBoxLayout()
        self.select_files_btn = QPushButton("选择 PDF 文件")
        self.select_folder_btn = QPushButton("选择文件")
        self.start_btn = QPushButton("开始提")
        self.export_btn = QPushButton("导出 CSV")
        self.save_data_btn = QPushButton("💾 保存数据")
        self.load_data_btn = QPushButton("📂 加载数据")
        self.clear_data_btn = QPushButton("清空加载数据")
        self.mapping_btn = QPushButton("🔗 编码映射")
        self.mapping_conflict_btn = QPushButton("⚠️ 映射冲突")
        self.column_settings_btn = QPushButton("⚙️ 列设置") # 新增按钮
        self.clear_filters_btn = QPushButton("🧹 清除筛选") # 新增按钮
        self.settings_btn = QPushButton("设置")

        self.select_files_btn.clicked.connect(self.select_files)
        self.select_folder_btn.clicked.connect(self.select_folder)
        self.start_btn.clicked.connect(self.load_pdfs)
        self.export_btn.clicked.connect(self.export_to_csv)
        self.save_data_btn.clicked.connect(self.save_data)
        self.load_data_btn.clicked.connect(self.load_data)
        self.clear_data_btn.clicked.connect(self.clear_data)
        self.mapping_btn.clicked.connect(self.manage_mappings)
        self.mapping_conflict_btn.clicked.connect(self.show_product_mapping_conflicts)
        self.column_settings_btn.clicked.connect(self.open_column_settings)
        self.clear_filters_btn.clicked.connect(self.clear_current_tab_filters)
        self.settings_btn.clicked.connect(self.open_settings)

        top_layout.addWidget(self.select_files_btn)
        top_layout.addWidget(self.select_folder_btn)
        top_layout.addWidget(self.start_btn)
        top_layout.addWidget(self.export_btn)
        top_layout.addWidget(self.save_data_btn)
        top_layout.addWidget(self.load_data_btn)
        top_layout.addWidget(self.clear_data_btn)
        top_layout.addWidget(self.mapping_btn)
        top_layout.addWidget(self.mapping_conflict_btn)
        top_layout.addWidget(self.column_settings_btn)
        top_layout.addWidget(self.clear_filters_btn)
        top_layout.addStretch()
        top_layout.addWidget(self.settings_btn)
        main_layout.addLayout(top_layout)

        # --- 文件列表 ---
        self.file_list_label = QLabel("待处理文  0")
        main_layout.addWidget(self.file_list_label)

        # --- 结果视图 (Tab Widget) ---
        self.tab_widget = QTabWidget()

        self.invoices = []
        self.product_mapping_conflicts = []
        
        # Initialize Managers before creating tabs that depend on them
        self.mapping_mgr = MappingManager()
        self.tag_mgr = TagManager()
        
        # Tab 0: 用户中心 (User Dashboard)
        self.user_tab = UserDashboardTab(self)
        self.tab_widget.addTab(self.user_tab, "用户中心")

        # Tab 1: 明细数据
        self.tab1_container = QWidget()
        layout_t1 = QVBoxLayout(self.tab1_container)
        layout_t1.setContentsMargins(0, 0, 0, 0)
        layout_t1.setSpacing(0)

        self.table = QTableWidget()
        self.setup_table()
        layout_t1.addWidget(self.table)
        
        self.tab_widget.addTab(self.tab1_container, "明细数据")

        # Tab 2: 发票汇
        self.tab2_container = QWidget()
        layout_t2 = QVBoxLayout(self.tab2_container)
        layout_t2.setContentsMargins(0, 0, 0, 0)
        layout_t2.setSpacing(0)

        self.summary_table = QTableWidget()
        self.setup_summary_table()
        layout_t2.addWidget(self.summary_table)

        self.tab_widget.addTab(self.tab2_container, "发票汇")

        # Tab 3: 收款系统
        self.payment_tab = PaymentSystemTab(self)
        self.tab_widget.addTab(self.payment_tab, "收款系统")

        # Tab 3.5: 账户管理 (New)
        self.account_tab = AccountManagerTab(self)
        self.tab_widget.addTab(self.account_tab, "账户管理")
        
        self.aging_tab = ReceivablesStatusTab(self)
        self.tab_widget.addTab(self.aging_tab, "应收情况")
        
        self.recon_tab = AutoReconciliationTab(self)
        self.tab_widget.addTab(self.recon_tab, "自动核对")
        
        self.query_tab = DataQueryTab(self)
        self.tab_widget.addTab(self.query_tab, "数据查询汇")

        # --- 布局优化：使用 Splitter 包裹主内容和日志区 ---
        self.main_splitter = QSplitter(Qt.Vertical)
        self.main_splitter.addWidget(self.tab_widget)
        
        # 创建日志容器
        self.log_container = QWidget()
        log_layout = QVBoxLayout(self.log_container)
        log_layout.setContentsMargins(0, 5, 0, 0)
        log_layout.setSpacing(2)
        
        # 日志标题栏（含收起按钮）
        log_header = QHBoxLayout()
        log_title = QLabel("<b>系统日志 (Logs)</b>")
        self.toggle_log_btn = QPushButton("▼ 收起")
        self.toggle_log_btn.setCheckable(True)
        self.toggle_log_btn.setFixedWidth(80)
        self.toggle_log_btn.clicked.connect(self.toggle_log_panel)
        
        log_header.addWidget(log_title)
        log_header.addStretch()
        log_header.addWidget(self.toggle_log_btn)
        log_layout.addLayout(log_header)

        self.log_viewer = QTextEdit()
        self.log_viewer.setReadOnly(True)
        self_font = self.font()
        self_font.setPointSize(9)
        self.log_viewer.setFont(self_font)
        log_layout.addWidget(self.log_viewer)
        
        self.main_splitter.addWidget(self.log_container)
        # 设置初始分配比例：主视图 85%，日志 15%
        self.main_splitter.setStretchFactor(0, 85)
        self.main_splitter.setStretchFactor(1, 15)
        
        main_layout.addWidget(self.main_splitter)

        # --- 底部状态栏 ---
        bottom_layout = QHBoxLayout()
        self.progress_bar = QProgressBar()
        self.status_label = QLabel("准备就绪")
        
        # 类似 Excel 的状态栏统计
        self.stats_label = QLabel("")
        self.stats_label.setStyleSheet("color: #007bff; font-weight: bold; margin-right: 15px;")
        self.stats_label.setTextInteractionFlags(Qt.TextSelectableByMouse) # Enable selection
        
        self.stop_btn = QPushButton("停止")
        self.stop_btn.clicked.connect(self.stop_extraction)
        self.stop_btn.setEnabled(False)
        
        bottom_layout.addWidget(self.status_label, 1)
        bottom_layout.addWidget(self.stats_label) # Add stats label
        bottom_layout.addWidget(self.progress_bar, 2)
        bottom_layout.addWidget(self.stop_btn)
        main_layout.addLayout(bottom_layout)

        self.file_paths = []
        self.worker_thread = None
        self.data_modified = False
        self.auto_save_file = "danfe_data_autosave.json"
        
        # Legacy migration (one-off)
        self.mapping_file = "product_code_mapping.json"
        if os.path.exists(self.mapping_file) and not self.mapping_mgr.data["products"]:
            try:
                with open(self.mapping_file, 'r', encoding='utf-8') as f:
                    old_map = json.load(f)
                    for k, v in old_map.items():
                        self.mapping_mgr.set_product_std(k, v, "", source="legacy_migration", status="candidate")
                self.mapping_mgr.save()
                self.log_message("已迁移旧版产品映射到新库")
            except Exception as e:
                self.log_message(f"旧版产品映射迁移失败: {e}")
        # Ensure legacy mapping dict is initialized before any data load
        self.product_code_mapping = {}
        self.load_product_mapping()
        
        # 注册主表进行统计
        self.register_table_for_stats(self.table)
        self.register_table_for_stats(self.summary_table)

        # 自动保存定时 ( 0 
        self.auto_save_timer = QTimer()
        self.auto_save_timer.timeout.connect(self.auto_save)
        self.auto_save_timer.start(30000)  # 30 
        self.load_settings()
        
        # 恢复表格列状
        self.restore_column_state()

        # 尝试加载上次自动保存的数据 (延迟加载，避免阻塞窗口显示)
        if os.path.exists(self.auto_save_file):
            self.log_message(f"自动加载上次保存的数据 {self.auto_save_file}")
            # 使用 QTimer.singleShot 延迟加载，让窗口先显示
            QTimer.singleShot(100, self._deferred_auto_load)

    def _deferred_auto_load(self):
        """延迟加载自动保存的数据（使用后台线程，避免阻塞界面）"""
        self.status_label.setText("正在后台加载数据...")
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)

        # 创建后台加载线程
        self._data_load_worker = DataLoadWorker(self.auto_save_file)
        self._data_load_worker.progress.connect(self._on_load_progress)
        self._data_load_worker.finished.connect(self._on_load_finished)
        self._data_load_worker.error.connect(self._on_load_error)
        self._data_load_worker.log.connect(self.log_message)
        self._data_load_worker.start()

    def _on_load_progress(self, current: int, total: int):
        """后台加载进度回调"""
        if total > 0:
            percent = int(current * 100 / total)
            self.progress_bar.setValue(percent)
            self.status_label.setText(f"正在解析数据... {current}/{total}")

    def _on_load_finished(self, invoices: list, recon_data: dict):
        """后台加载完成回调"""
        self.status_label.setText("正在填充表格...")
        QApplication.processEvents()

        # 设置数据
        self.invoices = invoices

        # 恢复对账模块数据
        if recon_data and hasattr(self, 'recon_tab'):
            self.recon_tab.set_state(recon_data)

        # 填充表格
        self.populate_table()
        self.file_list_label.setText(f"已加载发票 {len(self.invoices)}")
        self.export_btn.setEnabled(True)
        self.data_modified = False

        # 刷新收款界面（不同步，只刷新显示）
        self.payment_tab.load_invoices()

        self.progress_bar.setValue(100)
        self.status_label.setText("准备就绪")
        self.log_message(f"自动加载完成 ({len(self.invoices)} 个发票)")

    def _on_load_error(self, error_msg: str):
        """后台加载错误回调"""
        self.log_message(f"自动加载失败: {error_msg}")
        self.status_label.setText("自动加载失败")
        self.progress_bar.setValue(0)

    def get_other_party(self, inv: Invoice, my_cnpj: Optional[str]) -> Tuple[str, str]:
        """获取交易对手方的 CNPJ 和 名称 (避开自己)"""
        dest_cnpj = inv.destinatario_cnpj or ""
        dest_name = inv.destinatario_nome or "未知收货人"
        emit_cnpj = inv.emitente_cnpj or ""
        emit_name = inv.emitente_nome or "未知发货人"
        
        dest_clean = re.sub(r'\D', '', dest_cnpj)
        emit_clean = re.sub(r'\D', '', emit_cnpj)
        
        # 如果我是收货人，那么对方是发货人
        if my_cnpj and dest_clean == my_cnpj:
            return emit_cnpj, emit_name
        # 如果我是发货人，那么对方是收货人
        if my_cnpj and emit_clean == my_cnpj:
            return dest_cnpj, dest_name
            
        # 如果没识别出自己，或者双方都不是自己，优先返回非空的那个 (默认客户模式)
        if dest_cnpj: return dest_cnpj, dest_name
        return emit_cnpj, emit_name

    def identify_self_cnpj(self) -> Optional[str]:
        """自动识别‘自身’企业的 CNPJ (通过统计出现频率最高的 CNPJ)"""
        if not self.invoices:
            return None
            
        counts = {}
        for inv in self.invoices:
            for c in [inv.emitente_cnpj, inv.destinatario_cnpj]:
                if not c: continue
                clean = re.sub(r'\D', '', c)
                if len(clean) == 14: # CNPJ 必须是 14 位
                    counts[clean] = counts.get(clean, 0) + 1
        
        if not counts:
            return None
            
        # 找出频率最高的 CNPJ
        top_cnpj = max(counts, key=counts.get)
        # 如果该 CNPJ 出现的次数超过发票总数的一半，基本可以确定是‘自己’
        if counts[top_cnpj] >= len(self.invoices) * 0.5:
            return top_cnpj
        return None

    def toggle_log_panel(self, checked):
        """收起或展开日志面板"""
        if checked:
            # 收起：将 Splitter 的第二个部件（日志区）尺寸设为 0
            self.main_splitter.setSizes([1000, 0])
            self.toggle_log_btn.setText("▲ 展开")
            # 隐藏日志内容只保留标题栏
            self.log_viewer.setVisible(False)
        else:
            # 展开：恢复尺寸
            self.main_splitter.setSizes([800, 200])
            self.toggle_log_btn.setText("▼ 收起")
            self.log_viewer.setVisible(True)

    def register_table_for_stats(self, table):
        """注册表格以启用底部状态栏统计功能"""
        if isinstance(table, QTableWidget):
            table.itemSelectionChanged.connect(self.update_selection_stats)
            
    def update_selection_stats(self):
        """计算当前选中的总和、平均值、计数 (增强版：使用 br_to_float 保证精度)"""
        table = self.sender()
        if not isinstance(table, QTableWidget):
            return
            
        selected_items = table.selectedItems()
        # 排除汇总行中的项
        valid_items = [i for i in selected_items if i.data(Qt.UserRole) != "SUMMARY"]
        
        if not valid_items:
            self.stats_label.setText("")
            return
            
        count = len(valid_items)
        values = []
        for item in valid_items:
            try:
                raw_txt = item.text().strip()
                if not raw_txt or raw_txt == "-": continue
                
                # 使用标准的 br_to_float 处理 (处理组合匹配 R$ 1.234,56 (组合) 等复杂情况)
                # 先清除括号内容以免干扰数值识别
                clean_txt = re.sub(r'\(.*?\)', '', raw_txt).strip()
                val = br_to_float(clean_txt)
                
                if val is not None:
                    values.append(val)
            except:
                pass
        
        txt_parts = [f"选中格数: {count}"]
        if values:
            total = sum(values)
            avg = total / len(values)
            # 汇总和合计保持 2 位小数，平均数使用 4 位小数以提高精度
            txt_parts.append(f"求和: {total:,.2f}")
            if len(values) > 1:
                txt_parts.append(f"平均: {avg:,.4f}")
            
        self.stats_label.setText("  |  ".join(txt_parts))

    def filter_by_tag(self, tag_name):
        """根据标签过滤明细"""
        # 1. Switch to Detail Tab
        self.tab_widget.setCurrentWidget(self.tab1_container)
        
        # 2. Find "User Tags" column index
        # Combined headers: invoice_headers + item_headers
        # "User Tags" is in invoice_headers, usually index 40
        try:
            col_idx = self.combined_headers.index("用户标签")
        except ValueError:
            QMessageBox.warning(self, "错误", "未找 用户标签'")
            return
            
        # 3. Apply Filter
        header = self.table.horizontalHeader()
        if isinstance(header, FilterHeader):
            header._filters[col_idx] = tag_name
            header.filterChanged.emit()
            self.log_message(f"已应用标签筛  {tag_name}。如需查看全部数据，请点击顶部'🧹 清除筛 按钮")
        else:
            QMessageBox.information(self, "提示", "当前表头不支持过")

    def open_column_settings(self):
        """打开当前激活表格的列设"""
        current_widget = self.tab_widget.currentWidget()
        target_table = None
        
        if current_widget == self.tab1_container:
            target_table = self.table
        elif current_widget == self.tab2_container:
            target_table = self.summary_table
        elif current_widget == self.payment_tab:
            # 优先配置获得焦点的表格，否则默认发票列表
            if self.payment_tab.installment_table.hasFocus():
                target_table = self.payment_tab.installment_table
            else:
                target_table = self.payment_tab.invoice_table
        elif hasattr(current_widget, "table"): # Handle DataQueryTab
             target_table = current_widget.table
        
        if not isinstance(target_table, QTableWidget):
             QMessageBox.information(self, "提示", "当前页面不支持列设置")
             return
             
        dialog = ColumnConfigDialog(target_table, self)
        if dialog.exec():
            dialog.apply_settings()
            self.save_column_state()

    def clear_current_tab_filters(self):
        """清除当前激活表格的所有筛"""
        current_widget = self.tab_widget.currentWidget()
        target_table = None
        
        if current_widget == self.tab1_container:
            target_table = self.table
        elif current_widget == self.tab2_container:
            target_table = self.summary_table
        elif current_widget == self.payment_tab:
            # 优先处理获得焦点的表
            if self.payment_tab.installment_table.hasFocus():
                target_table = self.payment_tab.installment_table
            else:
                target_table = self.payment_tab.invoice_table
        elif hasattr(current_widget, "table"): # Handle DataQueryTab
             target_table = current_widget.table
        
        if isinstance(target_table, QTableWidget):
             header = target_table.horizontalHeader()
             if isinstance(header, FilterHeader):
                 header.clear_filters()
                 self.log_message(f"已清除表 {target_table.objectName()} 的筛")
             else:
                 QMessageBox.information(self, "提示", "当前表格不支持高级筛选清")
        else:
             QMessageBox.information(self, "提示", "当前页面没有可清除筛选的表格")

    def save_column_state(self):
        """保存表格列状态"""
        settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)

        # 保存明细表状态
        settings.setValue(
            "table_header_state",
            self.table.horizontalHeader().saveState(),
        )

        # 保存汇总表状态
        settings.setValue(
            "summary_table_header_state",
            self.summary_table.horizontalHeader().saveState(),
        )

        # 保存收款系统表格状态
        settings.setValue(
            "pay_inv_header_state",
            self.payment_tab.invoice_table.horizontalHeader().saveState(),
        )
        settings.setValue(
            "pay_inst_header_state",
            self.payment_tab.installment_table.horizontalHeader().saveState(),
        )

    def restore_column_state(self):
        """恢复表格列状"""
        settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)
        
        state1 = settings.value("table_header_state")
        if state1:
            self.table.horizontalHeader().restoreState(state1)
            
        state2 = settings.value("summary_table_header_state")
        if state2:
            self.summary_table.horizontalHeader().restoreState(state2)
            
        state3 = settings.value("pay_inv_header_state")
        if state3:
            self.payment_tab.invoice_table.horizontalHeader().restoreState(state3)
            
        state4 = settings.value("pay_inst_header_state")
        if state4:
            self.payment_tab.installment_table.horizontalHeader().restoreState(state4)

    def create_navigation_menu(self, invoice_index, current_tab_name):
        """创建跨页面导航菜单"""
        from PySide6.QtWidgets import QMenu
        menu = QMenu()
        
        # 1. 跳转到明细 (Detail)
        if current_tab_name != "detail":
            action = menu.addAction("🔍 跳转到明细数据")
            action.triggered.connect(lambda: self.navigate_to_invoice(invoice_index, self.table))
            
        # 2. 跳转到汇总 (Summary)
        if current_tab_name != "summary":
            action = menu.addAction("📊 跳转到发票汇总")
            action.triggered.connect(lambda: self.navigate_to_invoice(invoice_index, self.summary_table))
            
        # 3. 跳转到收款 (Payment)
        if current_tab_name != "payment":
            action = menu.addAction("💰 跳转到收款系统")
            action.triggered.connect(lambda: self.navigate_to_payment(invoice_index))

        # 4. 跳转到账龄 (Aging)
        if current_tab_name != "aging":
            action = menu.addAction("⏳ 查看账龄分析")
            action.triggered.connect(lambda: self.navigate_to_aging(invoice_index))
            
        return menu

    def view_source_pdf(self, invoice_index):
        """打开当前发票的原始 PDF 文件"""
        if 0 <= invoice_index < len(self.invoices):
            invoice = self.invoices[invoice_index]
            path = invoice.file_path
            
            if not path or not os.path.exists(path):
                # 尝试通过文件名在当前选择的路径中寻找 (兜底逻辑)
                QMessageBox.warning(self, "文件未找到", f"无法定位原始文件：\n{path or invoice.file_name}\n\n请确保文件未被移动或删除。")
                return
                
            from PySide6.QtCore import QUrl
            from PySide6.QtGui import QDesktopServices
            QDesktopServices.openUrl(QUrl.fromLocalFile(path))

    def navigate_to_aging(self, invoice_index):
        """跳转到账龄分析并根据当前发票的往来单位进行过滤"""
        if 0 <= invoice_index < len(self.invoices):
            invoice = self.invoices[invoice_index]
            # 优先看客户
            partner_name = invoice.destinatario_nome or invoice.emitente_nome
            if partner_name:
                self.tab_widget.setCurrentWidget(self.aging_tab)
                # 可以在这里添加自动高亮/过滤逻辑
                self.aging_tab.refresh_analysis()
                # 寻找并选中
                for row in range(self.aging_tab.table.rowCount()):
                    item = self.aging_tab.table.item(row, 0)
                    if item and partner_name in item.text():
                        self.aging_tab.table.selectRow(row)
                        self.aging_tab.table.scrollToItem(item)
                        break

    def navigate_to_invoice(self, invoice_index, target_table):
        """通用跳转逻辑 (基于列表索引)"""
        # 切换到目标表格所在的 Tab
        if target_table == self.table:
            self.tab_widget.setCurrentWidget(self.tab1_container)
        elif target_table == self.summary_table:
            self.tab_widget.setCurrentWidget(self.tab2_container)
            
        # 查找：假设表格行顺序与 self.invoices 索引一致（对于未排序的情况）
        # 如果排过序，需要遍历 UserRole
        if 0 <= invoice_index < len(self.invoices):
            found = False
            for row in range(target_table.rowCount()):
                if target_table.isRowHidden(row): continue # Skip hidden rows
                
                item = target_table.item(row, 0)
                # 我们在填充表格时设置 UserRole  invoice_index (now tuple)
                if item:
                    data = item.data(Qt.UserRole)
                    # 排除汇总行
                    current_idx = data[0] if isinstance(data, tuple) else data
                    
                    if current_idx == invoice_index:
                        target_table.selectRow(row)
                        target_table.scrollToItem(item)
                        found = True
                        break
            if not found:
                self.log_message(f"在目标表格中未找到索引为 {invoice_index} 的发 (可能被过 ")

    def navigate_to_payment(self, invoice_index):
        """跳转到收款系统并选中发票"""
        if 0 <= invoice_index < len(self.invoices):
            invoice = self.invoices[invoice_index]
            invoice_number = invoice.numero or f"FILE-{invoice.file_name}"
            
            self.tab_widget.setCurrentWidget(self.payment_tab)
            
            # 在收款列表中查找
            table = self.payment_tab.invoice_table
            for row in range(table.rowCount()):
                # 发票号在  (index 1)
                item = table.item(row, 1) 
                if item and item.text() == invoice_number:
                    table.selectRow(row)
                    table.scrollToItem(item)
                    self.payment_tab.on_invoice_selected(item) # 触发加载分期
                    break

    def setup_table(self):
        # 定义完整表头
        self.invoice_headers = [
            "文件名", "发票号", "序列", "性质 (Natureza)", "访问密钥", "授权协议",
            "签发日期", "进出日期", "总金额", "ICMS总额", "Base ICMS", "Base ST", "Valor ST",
            "IPI总额", "运费", "保险", "总折扣", "其他费用",
            "发货方名称", "发货方CNPJ", "发货方IE",
            "收货方名称", "收货方CNPJ", "收货方IE", "收货方地址", "收货方街区",
            "收货方城市", "收货方州", "收货方CEP", "收货方电话",
            "运输公司", "运输公司CNPJ/CPF", "运费方式", "车辆车牌", "车辆UF",
            "毛重", "净重", "补充信息", "备注", "提取元数据", "用户标签", "收款进度"
        ]
        self.item_headers = [
            "商品编码", "国内编码", "商品描述", "NCM", "CST", "CFOP", "单位", "数量",
            "单价", "总价", "ICMS基数", "ICMS金额", "ICMS税率",
            "IPI金额", "IPI税率", "折扣"
        ]
        self.combined_headers = self.invoice_headers + self.item_headers
        
        self.table.setColumnCount(len(self.combined_headers))
        
        # Custom Filter Header
        header = FilterHeader(self.table)
        header.filterChanged.connect(lambda: self.apply_header_filters(self.table))
        # Disable native sorting to handle summary row
        self.table.setSortingEnabled(False)
        header.sectionClicked.connect(self.on_header_clicked)
        self.table.setHorizontalHeader(header)
        
        self.table.setHorizontalHeaderLabels(self.combined_headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        # 允许双击编辑
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.table.setSortingEnabled(True)

        # 连接单元格修改信号
        self.table.itemChanged.connect(self.on_table_item_changed)
        
        # 启用右键菜单
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_table_context_menu)

    def show_table_context_menu(self, pos):
        item = self.table.itemAt(pos)
        
        # 排除汇总行
        if item and item.data(Qt.UserRole) == "SUMMARY": return

        menu = QMenu(self)
        
        if item:
            data = item.data(Qt.UserRole)
            original_index = data[0] if isinstance(data, tuple) else data
            
            if original_index is not None and original_index >= 0:
                nav = self.create_navigation_menu(original_index, "detail")
                menu.addActions(nav.actions())
                menu.addSeparator()
                
                # 新增查看 PDF
                action_pdf = menu.addAction("📄 查看源文件 PDF (Open)")
                action_pdf.triggered.connect(lambda: self.view_source_pdf(original_index))
                menu.addSeparator()

        # Edit Actions
        menu.addAction("🏷 编辑标签 (Tags)", self.edit_tags)
        menu.addAction("🧹 清除所有标签 (Clear Tags)", self.clear_all_tags)
        menu.addAction("➕ 添加新行 (Add Manual Row)", self.add_manual_row)
        
        if item:
            menu.addAction("🗑 删除选中 (Delete Rows)", self.delete_selected_rows)
            
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def clear_all_tags(self):
        """清除选中行的所有标签"""
        rows = sorted(set(i.row() for i in self.table.selectedItems()))
        if not rows:
            QMessageBox.warning(self, "提示", "请选择至少一行")
            return

        reply = QMessageBox.question(self, "确认清除", f"确定要清除选中 {len(rows)} 行的所有标签吗？", 
                                   QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes: return

        for r in rows:
            it = self.table.item(r, 0)
            if not it: continue
            
            # 排除汇总行 (Skip summary rows)
            if it.data(Qt.UserRole) == "SUMMARY":
                continue
                
            d = it.data(Qt.UserRole)
            i_idx, it_idx = d if isinstance(d, tuple) else (d, -1)
            
            if 0 <= i_idx < len(self.invoices):
                target_inv = self.invoices[i_idx]
                if it_idx >= 0 and it_idx < len(target_inv.itens):
                    target_inv.itens[it_idx].tags = []
                else:
                    target_inv.tags = []
                count += 1
                    
        self.data_modified = True
        self.populate_table()
        self.log_message(f"已清除 {count} 行的标签")

    def edit_tags(self):
        """编辑选中行的标签"""
        rows = sorted(set(i.row() for i in self.table.selectedItems()))
        if not rows:
            QMessageBox.warning(self, "提示", "请选择至少一行")
            return

        # 获取第一个选中项的标签作为初始状态
        first_item = self.table.item(rows[0], 0)
        data = first_item.data(Qt.UserRole)
        inv_idx, item_idx = data if isinstance(data, tuple) else (data, -1)
        
        current_tags = []
        if 0 <= inv_idx < len(self.invoices):
            inv = self.invoices[inv_idx]
            if item_idx >= 0 and item_idx < len(inv.itens):
                current_tags = inv.itens[item_idx].tags
            else:
                current_tags = inv.tags
        
        dialog = TagEditDialog(self.tag_mgr, current_tags, self)
        if dialog.exec():
            new_tags = list(dialog.selected_tags)
            
            # Apply to all selected rows
            for r in rows:
                it = self.table.item(r, 0)
                if not it: continue
                
                # 排除汇总行 (Skip summary rows)
                if it.data(Qt.UserRole) == "SUMMARY":
                    continue
                    
                d = it.data(Qt.UserRole)
                i_idx, it_idx = d if isinstance(d, tuple) else (d, -1)
                
                if 0 <= i_idx < len(self.invoices):
                    target_inv = self.invoices[i_idx]
                    if it_idx >= 0 and it_idx < len(target_inv.itens):
                        target_inv.itens[it_idx].tags = new_tags[:]
                    else:
                        target_inv.tags = new_tags[:]
                        
            self.data_modified = True
            self.populate_table()

    def add_manual_row(self):
        """添加手动 (Dummy Invoice)"""
        new_inv = Invoice(file_name="Manual Entry", numero="MANUAL", data_emissao=datetime.now().strftime("%d/%m/%Y"))
        new_item = Item(descricao="New Item", quantidade=1, valor_unitario=0.0)
        new_inv.itens.append(new_item)
        
        self.invoices.append(new_inv)
        self.data_modified = True
        self.populate_table()
        self.table.scrollToBottom()

    def delete_selected_rows(self):
        """删除选中"""
        rows = sorted(set(i.row() for i in self.table.selectedItems()), reverse=True)
        if not rows: return
        
        reply = QMessageBox.question(self, "确认删除", f"确定要删除选中 {len(rows)} 行吗", QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes: return
        
        items_to_remove = [] 
        invs_to_check = set()
        
        for r in rows:
            it = self.table.item(r, 0)
            if not it: continue
            
            # 排除汇总行 (Skip summary rows)
            if it.data(Qt.UserRole) == "SUMMARY":
                continue
                
            d = it.data(Qt.UserRole)
            i_idx, it_idx = d if isinstance(d, tuple) else (d, -1)
            
            if 0 <= i_idx < len(self.invoices):
                inv = self.invoices[i_idx]
                if it_idx >= 0 and it_idx < len(inv.itens):
                    items_to_remove.append((inv, inv.itens[it_idx]))
                    invs_to_check.add(inv)
                elif it_idx == -1:
                    invs_to_check.add(inv)
        
        count = 0
        for inv, item in items_to_remove:
            if item in inv.itens:
                inv.itens.remove(item)
                count += 1
                
        for inv in list(invs_to_check): 
            if not inv.itens:
                if inv in self.invoices:
                    self.invoices.remove(inv)
                    # 同步删除数据库中的记录 (Sync deletion to DB)
                    if inv.numero:
                        self.payment_tab.db.delete_invoice_by_number(inv.numero)
                    else:
                        self.payment_tab.db.delete_invoice_by_number(f"FILE-{inv.file_name}")
            else:
                # 如果发票还在，但明细变了，同步更新到数据库 (可选：可能需要重新计算总额)
                # Recalculate total if all items have valor_total
                if all(it.valor_total is not None for it in inv.itens):
                    inv.total_nota = sum(it.valor_total for it in inv.itens)
                self.sync_single_invoice_to_db(inv)
                    
        # 刷新所有模块界面 (Refresh all relevant tabs)
        self.refresh_all_tabs()
            
        self.data_modified = True
        self.populate_table()
        self.log_message(f"已删 {count} 行，并同步到数据库")

    def setup_summary_table(self):
        """初始化汇总表"""
        # 使用发票层级的表
        self.summary_table.setColumnCount(len(self.invoice_headers))
        
        # Custom Filter Header
        header = FilterHeader(self.summary_table)
        header.filterChanged.connect(lambda: self.apply_header_filters(self.summary_table))
        # Disable native sorting
        self.summary_table.setSortingEnabled(False)
        header.sectionClicked.connect(self.on_header_clicked)
        self.summary_table.setHorizontalHeader(header)
        
        self.summary_table.setHorizontalHeaderLabels(self.invoice_headers)
        self.summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.summary_table.horizontalHeader().setStretchLastSection(True)
        self.summary_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.summary_table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        self.summary_table.setSortingEnabled(True)
        # 连接汇总表格的修改信号
        self.summary_table.itemChanged.connect(self.on_summary_item_changed)
        
        # 启用右键菜单
        self.summary_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.summary_table.customContextMenuRequested.connect(self.show_summary_context_menu)
        
        # Optimize: Set Delegate for Progress Column
        progress_col = len(self.invoice_headers) - 1
        self.summary_table.setItemDelegateForColumn(progress_col, ProgressBarDelegate(self.summary_table))
        self.summary_table.setColumnWidth(progress_col, 120)

    def show_summary_context_menu(self, pos):
        """汇总表右键菜单"""
        item = self.summary_table.itemAt(pos)
        if not item: return

        original_index = item.data(Qt.UserRole)
        # 排除汇总行
        if original_index == "SUMMARY": return

        if original_index is not None:
            menu = self.create_navigation_menu(original_index, "summary")
            
            menu.addSeparator()
            # 新增查看 PDF
            action_pdf = menu.addAction("📄 查看源文件 PDF (Open)")
            action_pdf.triggered.connect(lambda: self.view_source_pdf(original_index))
            
            menu.addSeparator()
            action_delete = menu.addAction("🗑 删除整单 (Delete Invoices)")
            action_delete.triggered.connect(self.delete_selected_summary_rows)
            
            # 保留原有的复制功能
            action_copy = menu.addAction("📋 复制当前行")
            action_copy.triggered.connect(lambda: self._copy_summary_row(item.row()))
            
            menu.exec(self.summary_table.viewport().mapToGlobal(pos))

    def delete_selected_summary_rows(self):
        """汇总表删除选中整单"""
        rows = sorted(set(i.row() for i in self.summary_table.selectedItems()), reverse=True)
        if not rows: return
        
        reply = QMessageBox.question(self, "确认删除", f"确定要删除选中 {len(rows)} 张发票及其所有明细吗？", 
                                   QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes: return
        
        indices_to_remove = []
        for r in rows:
            it = self.summary_table.item(r, 0)
            if not it: continue
            
            # 排除汇总行 (Skip summary rows)
            if it.data(Qt.UserRole) == "SUMMARY":
                continue
                
            idx = it.data(Qt.UserRole)
            if idx is not None and idx >= 0:
                indices_to_remove.append(idx)
        
        # Sort indices descending to avoid shifting issues
        indices_to_remove.sort(reverse=True)
        
        count = 0
        for idx in indices_to_remove:
            if 0 <= idx < len(self.invoices):
                inv = self.invoices.pop(idx)
                # 同步删除数据库中的记录 (Sync deletion to DB)
                if inv.numero:
                    self.payment_tab.db.delete_invoice_by_number(inv.numero)
                else:
                    # 如果没有发票号，尝试文件名标识
                    self.payment_tab.db.delete_invoice_by_number(f"FILE-{inv.file_name}")
                count += 1
        
        # 刷新所有模块界面 (Refresh all relevant tabs)
        self.refresh_all_tabs()
        
        self.data_modified = True
        self.populate_table()
        self.log_message(f"已从列表和数据库移除 {count} 张发票")

    def _copy_summary_row(self, row):
        values = []
        for col in range(self.summary_table.columnCount()):
            cell = self.summary_table.item(row, col)
            # cell might be None if it's a widget cell (progress bar)
            if cell:
                values.append(cell.text())
            else:
                # Handle widget cell (Progress Bar)
                widget = self.summary_table.cellWidget(row, col)
                if isinstance(widget, QProgressBar):
                    values.append(widget.format())
                else:
                    values.append("")
        QApplication.clipboard().setText("\t".join(values))

    def load_settings(self):
        self.settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)
        tesseract_cmd = self.settings.value("tesseract_cmd", "")
        
        # --- 自动寻找 Tesseract 路径 ---
        if not tesseract_cmd or not os.path.exists(tesseract_cmd):
            # 1. 尝试 PATH 环境变量寻找
            import shutil
            found_path = shutil.which("tesseract")
            if found_path:
                tesseract_cmd = found_path
            else:
                # 2. 尝试常见 Windows 安装路径
                common_paths = [
                    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                    os.path.expanduser(r"~\AppData\Local\Tesseract-OCR\tesseract.exe")
                ]
                for p in common_paths:
                    if os.path.exists(p):
                        tesseract_cmd = p
                        break
            
            # 如果找到了，保存到设置中
            if tesseract_cmd and os.path.exists(tesseract_cmd):
                self.settings.setValue("tesseract_cmd", tesseract_cmd)
                self.log_message(f"自动发现 Tesseract-OCR: {tesseract_cmd}")

        if tesseract_cmd and os.path.exists(tesseract_cmd):
            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
            self.log_message(f"Tesseract-OCR 已加  {tesseract_cmd}")
        else:
            self.log_message("警告: Tesseract-OCR 路径未设置或无效，OCR 功能将不可用。请在设置中手动指定")

    def get_current_config(self) -> dict:
        return {
            "prefer_pymupdf": self.settings.value("prefer_pymupdf", True, type=bool),
            "enable_ocr": self.settings.value("enable_ocr", False, type=bool),
            "ocr_lang": self.settings.value("ocr_lang", "por"),
            "poppler_path": self.settings.value("poppler_path"),
            "max_workers": os.cpu_count() or 1,
            "enable_llm": self.settings.value("enable_llm", False, type=bool),
            "llm_mode": self.settings.value("llm_mode", "local"),
            "llm_endpoint": self.settings.value("llm_endpoint", "http://localhost:1234/v1"),
            "llm_model": self.settings.value("llm_model", ""),
            "llm_api_key": self.settings.value("llm_api_key", ""),
            "llm_use_multimodal": self.settings.value("llm_use_multimodal", True, type=bool),
            "llm_timeout_sec": self.settings.value("llm_timeout_sec", "30"),
            "llm_max_chars": self.settings.value("llm_max_chars", "8000")
        }
    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "选择 PDF 文件", "", "PDF Files (*.pdf)")
        if files:
            self.file_paths = files
            self.file_list_label.setText(f"待处理文  {len(self.file_paths)}")
            self.log_message(f"已选择 {len(self.file_paths)} 个文件")

    def select_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "选择文件")
        if folder:
            self.file_paths = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith('.pdf')]
            self.file_list_label.setText(f"待处理文  {len(self.file_paths)}")
            self.log_message(f"已从文件 {folder} 加载 {len(self.file_paths)}  PDF 文件")

    def clear_data(self):
        """清空所有已加载的发票数据"""
        if not self.invoices and not self.file_paths:
            QMessageBox.information(self, "提示", "当前没有加载任何数据。")
            return

        reply = QMessageBox.question(self, "确认清空", 
                                   f"确定要清空当前已加载的数据吗？\n(包含 {len(self.invoices)} 个发票记录)",
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            self.invoices = []
            self.file_paths = []
            self.file_list_label.setText("待处理文件: 0")
            
            # Refresh tables (will be empty)
            self.populate_table()
            
            # Reset Status
            self.progress_bar.setValue(0)
            self.status_label.setText("数据已清空")
            self.export_btn.setEnabled(False)
            
            self.log_message("已清空所有加载的数据。")

    def load_pdfs(self):
        if not self.file_paths:
            QMessageBox.warning(self, "无文", "请先选择要处理的 PDF 文件或文件夹")
            return

        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.export_btn.setEnabled(False)
        self.table.setRowCount(0)
        self.invoices = []
        self.progress_bar.setValue(0)
        self.status_label.setText("正在提取...")

        config = self.get_current_config()
        self.worker_thread = WorkerThread(self.file_paths, config)
        self.worker_thread.progress.connect(self.update_progress)
        self.worker_thread.finished.connect(self.on_extraction_finished)
        self.worker_thread.log.connect(self.log_message)
        self.worker_thread.start()

    def stop_extraction(self):
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.stop()
            self.status_label.setText("正在停止...")
            self.stop_btn.setEnabled(False)
            self.log_message("用户请求停止提取")

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def on_extraction_finished(self, results):
        self.invoices = results
        self.populate_table()
        
        self.sync_invoices_to_payment(self.invoices)

        self.status_label.setText(f"提取完成！共处理 {len(self.invoices)} 个有效发票")
        self.progress_bar.setValue(100)
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.export_btn.setEnabled(True)
        self.worker_thread = None
        self.data_modified = True  # 标记有新数据需要保存
        QMessageBox.information(self, "完成", f"已成功处理 {len(self.invoices)} 个发票")

    def populate_table(self):
        self.populate_table_with_invoices(self.invoices)
        if hasattr(self, 'user_tab'):
            self.user_tab.refresh_stats()

    def populate_table_with_invoices(self, invoices: List[Invoice]):
        # 临时断开itemChanged信号，避免在填充表格时触发同步逻辑
        try:
            self.table.itemChanged.disconnect(self.on_table_item_changed)
        except RuntimeError:
            pass

        table = self.table
        table.blockSignals(True)
        table.setUpdatesEnabled(False)

        try:
            # 应用产品编码映射 (统一使用 MappingManager)
            for invoice in invoices:
                for item in invoice.itens:
                    if item.codigo_produto:
                        info = self.mapping_mgr.get_product_std(item.codigo_produto)
                        if info and info.get("std_code") and not item.codigo_domestico:
                            item.codigo_domestico = info.get("std_code")

            table.setRowCount(0) # 清空表格
            table.setSortingEnabled(False) # 排序时插入性能较差，先禁用

            # 预建索引映射，避免 O(N^2) 的 list.index 调用
            inv_index_map = {}
            if invoices is not self.invoices:
                inv_index_map = {id(inv): idx for idx, inv in enumerate(self.invoices)}

            row_data = []
            row_indices = [] # 存储每行对应 Invoice 索引

            for inv_list_idx, invoice in enumerate(invoices):
                if invoices is self.invoices:
                    original_index = inv_list_idx
                else:
                    original_index = inv_index_map.get(id(invoice), -1)

                if invoice.itens:
                    for idx, item in enumerate(invoice.itens):
                        row = self.get_row_from_invoice_and_item(invoice, item)
                        row_data.append(row)
                        row_indices.append((original_index, idx))
                else:
                    # 即使没有 item，也显示 invoice 信息
                    row = self.get_row_from_invoice_and_item(invoice, Item())
                    row_data.append(row)
                    row_indices.append((original_index, -1))

            table.setRowCount(len(row_data))

            # 准备标签颜色映射
            tag_colors = {t['name']: t['color'] for t in self.tag_mgr.get_tags()}
            for row_idx, data in enumerate(row_data):
                # 每 500 行处理一次 GUI 事件，保持界面响应
                if row_idx % 500 == 0 and row_idx > 0:
                    QApplication.processEvents()

                inv_idx, item_idx = row_indices[row_idx]
                bg_color = None

                if 0 <= inv_idx < len(self.invoices):
                    inv_obj = self.invoices[inv_idx]
                    # 收集当前行的所有标签
                    current_tags = inv_obj.tags[:]
                    if item_idx != -1 and 0 <= item_idx < len(inv_obj.itens):
                        current_tags.extend(inv_obj.itens[item_idx].tags)

                    # 查找第一个定义的颜色
                    for t in current_tags:
                        c = tag_colors.get(t)
                        if c and str(c).upper() != "#FFFFFF":
                            bg_color = QColor(c)
                            break

                for col_idx, cell_value in enumerate(data):
                    item = QTableWidgetItem(str(cell_value))
                    # 存储原始索引以便右键菜单使用 (tuple: inv_idx, item_idx)
                    item.setData(Qt.UserRole, row_indices[row_idx])

                    if bg_color:
                        item.setBackground(bg_color)

                    table.setItem(row_idx, col_idx, item)

            # self.table.resizeColumnsToContents()
            table.setSortingEnabled(False) # Keep false for custom sort
        finally:
            table.setUpdatesEnabled(True)
            table.blockSignals(False)
            # 重新连接itemChanged信号
            self.table.itemChanged.connect(self.on_table_item_changed)

        # 填充汇总表
        self.populate_summary_table(invoices)
        
        # Update Summary Row for main table
        self.update_summary_row(self.table)

    def populate_summary_table(self, invoices: Optional[List[Invoice]] = None):
        """填充汇总表 (优化版：使用 Delegate  缓存索引)"""
        if invoices is None:
            invoices = self.invoices

        # 临时断开信号
        try:
            self.summary_table.itemChanged.disconnect(self.on_summary_item_changed)
        except RuntimeError:
            pass 

        self.summary_table.blockSignals(True)
        self.summary_table.setUpdatesEnabled(False)
        self.summary_table.setSortingEnabled(False)
        self.summary_table.setRowCount(len(invoices)) # Pre-allocate rows

        # 优化：建立 Invoice -> Original Index 映射 (O(N))
        # 仅当 invoices 不是 self.invoices 时需要
        inv_index_map = {}
        if invoices is not self.invoices:
             for idx, inv in enumerate(self.invoices):
                 inv_index_map[id(inv)] = idx

        # 获取收款状态数
        try:
            db_invoices = self.payment_tab.db.get_invoices()
            status_map = {} 
            for row in db_invoices:
                data = {
                    'paid_terms': row['paid_terms_count'],
                    'total_terms': row['total_terms_count'],
                    'paid_amount': row['total_paid_amount'] or 0.0,
                    'total_amount': row['total_amount'] or 0.0
                }
                if row['invoice_number']:
                    status_map[row['invoice_number']] = data
                if row['file_name']:
                    status_map[f"FILE-{row['file_name']}"] = data
        except Exception as e:
            logging.error(f"Failed to load payment status: {e}")
            status_map = {}
        progress_col = len(self.invoice_headers) - 1
        data_col_count = len(self.invoice_headers) - 1

        total_invoices = len(invoices)
        for i, invoice in enumerate(invoices):
            # 每 500 条处理一次 GUI 事件
            if i % 500 == 0 and i > 0:
                QApplication.processEvents()

            # 获取原始索引
            if invoices is self.invoices:
                original_index = i
            else:
                original_index = inv_index_map.get(id(invoice), -1)

            # 获取行数 (仅发票部分数据)
            full_row = self.get_row_from_invoice_and_item(invoice, Item())
            
            # --- 核心修复：根据 invoice_headers 映射数据，确保列不偏移 ---
            # invoice_headers 的长度是 41, 最后一列是 "收款进度"
            # 我们取 full_row 的前 40 列 (对应发票属性)
            display_data = full_row[:data_col_count]
            
            for col, val in enumerate(display_data):
                item = QTableWidgetItem(str(val if val is not None else ""))
                item.setData(Qt.UserRole, original_index) # 存储原始索引
                self.summary_table.setItem(i, col, item)

            # 进度条数据准备
            key = invoice.numero or f"FILE-{invoice.file_name}"
            payment_data = status_map.get(key, {})
            paid_amt = payment_data.get('paid_amount', 0.0)
            
            percent = 0
            text = "0/0"
            
            if payment_data:
                paid_terms = payment_data.get('paid_terms', 0)
                total_terms = payment_data.get('total_terms', 0)
                total_amt = payment_data.get('total_amount', 0.0)
                
                if total_amt <= 0.01 and invoice.total_nota:
                    total_amt = invoice.total_nota

                if total_terms > 1:
                    # 多期：按期数显示
                    if total_terms > 0:
                        percent = int((paid_terms / total_terms) * 100)
                else:
                    # 单期：按金额显示进度条长度，但文本显 x/y
                    if total_amt > 0:
                        percent = int((paid_amt / total_amt) * 100)
                
                text = f"{paid_terms}/{total_terms}"
            
            if percent > 100: percent = 100
            
            # 使用 Delegate 的方式：设置数据
            progress_item = QTableWidgetItem()
            progress_item.setData(Qt.DisplayRole, percent) # 用于排序
            progress_item.setData(Qt.UserRole, original_index)
            progress_item.setData(Qt.UserRole + 1000, (percent, text)) # 用于 Delegate 绘制
            
            self.summary_table.setItem(i, progress_col, progress_item)

        # 移除自动调整列宽，极大提升性能
        # self.summary_table.resizeColumnsToContents()
        
        self.summary_table.setSortingEnabled(False)
        try:
            self.summary_table.itemChanged.connect(self.on_summary_item_changed)
        except RuntimeError:
            pass
        self.summary_table.setUpdatesEnabled(True)
        self.summary_table.blockSignals(False)
        
        self.update_summary_row(self.summary_table)

    def filter_invoices(self, term: str) -> List[Invoice]:
        invoices = self.invoices or []
        if not term:
            return invoices
        term = term.lower()
        filtered = []
        for inv in invoices:
            rows_to_check = [self.get_row_from_invoice_and_item(inv, Item())]
            if inv.itens:
                for item in inv.itens:
                    rows_to_check.append(self.get_row_from_invoice_and_item(inv, item))
            for row in rows_to_check:
                blob = " ".join("" if v is None else str(v) for v in row).lower()
                if term in blob:
                    filtered.append(inv)
                    break
        return filtered

    def apply_query_filter(self, term: str):
        filtered_invoices = self.filter_invoices(term)
        self.populate_table_with_invoices(filtered_invoices)
        self.populate_summary_table(filtered_invoices)

    def get_row_from_invoice_and_item(self, invoice: Invoice, item: Item) -> list:
        """将发票和项目对象转换为表格行数据"""
        return [
            invoice.file_name, invoice.numero, invoice.serie, invoice.natureza_operacao, invoice.chave_acesso,
            invoice.protocolo_autorizacao, format_date_gui(invoice.data_emissao), format_date_gui(invoice.data_saida_entrada),
            invoice.total_nota, invoice.total_icms, invoice.base_calculo_icms, invoice.base_calculo_st, invoice.valor_icms_st, # Updated fields
            invoice.total_ipi, invoice.frete, invoice.seguro,
            invoice.desconto_total, invoice.outras_despesas,
            invoice.emitente_nome, invoice.emitente_cnpj, invoice.emitente_ie,
            invoice.destinatario_nome, invoice.destinatario_cnpj, invoice.destinatario_ie,
            invoice.destinatario_endereco, invoice.destinatario_bairro, invoice.destinatario_municipio,
            invoice.destinatario_uf, invoice.destinatario_cep, invoice.destinatario_fone,
            invoice.transportador_nome, invoice.transportador_cnpjcpf, invoice.modalidade_frete_raw,
            invoice.placa_veiculo, invoice.uf_veiculo, invoice.peso_bruto, invoice.peso_liquido,
            invoice.info_compl_contribuinte, invoice.llm_table_note, str(invoice.extract_meta),
            ",".join(str(t) for t in (invoice.tags + item.tags)), # User Tags
            "-", # 收款进度 (Placeholder to avoid column shift)
            # Item data
            item.codigo_produto, item.codigo_domestico, item.descricao, item.ncm, item.cst, item.cfop, item.unidade,
            item.quantidade, item.valor_unitario, item.valor_total, item.bc_icms, item.valor_icms,
            item.aliquota_icms, item.valor_ipi, item.aliquota_ipi, item.desconto
        ]

    def export_to_csv(self):
        if not self.invoices:
            QMessageBox.warning(self, "无数据", "没有可导出的数据")
            return

        path, _ = QFileDialog.getSaveFileName(self, "导出文件", "", "Excel Files (*.xlsx);;CSV Files (*.csv)")
        if not path:
            return

        # Prepare full data
        header = self.combined_headers
        rows = []
        for invoice in self.invoices:
            if invoice.itens:
                for item in invoice.itens:
                    rows.append(self.get_row_from_invoice_and_item(invoice, item))
            else:
                rows.append(self.get_row_from_invoice_and_item(invoice, Item()))
        
        self._export_data(path, header, rows)

    def _export_data(self, path, headers, rows):
        """通用导出函数 (支持 .csv  .xlsx)"""
        try:
            ext = os.path.splitext(path)[1].lower()
            
            if ext == ".xlsx":
                if openpyxl is None:
                    raise ImportError("未安装 openpyxl 库，无法导出 Excel。请 pip install openpyxl")
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(headers)
                for r in rows:
                    # Clean data for Excel (remove illegal chars if any)
                    clean_row = []
                    for cell in r:
                        val = cell
                        if isinstance(val, str):
                            # Replace newlines with space for cleaner output
                            val = val.replace('\n', ' ')
                            # Remove illegal characters for Excel
                            val = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)
                        clean_row.append(val)
                    ws.append(clean_row)
                wb.save(path)
                
            else:
                # Default to CSV
                with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    # Clean newlines for CSV too
                    clean_rows = []
                    for r in rows:
                        clean_rows.append([str(c).replace('\n', ' ') if c is not None else "" for c in r])
                    writer.writerows(clean_rows)

            self.log_message(f"成功导出  {path}")
            QMessageBox.information(self, "导出成功", f"数据已成功保存到\n{path}")
            
            reply = QMessageBox.question(self, '打开文件', '要现在打开导出的文件吗？',
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                QDesktopServices.openUrl(PySide6.QtCore.QUrl.fromLocalFile(path))

        except Exception as e:
            self.log_message(f"导出失败: {e}")
            QMessageBox.critical(self, "导出错误", f"无法写入文件: {e}")

    def update_summary_row(self, table: QTableWidget):
        """更新表格第一行的汇总数 (优化版：确保唯一并处于顶端)"""
        if table is None:
            return

        # 临时禁用排序，防止操作中行位置变动
        was_sorting = table.isSortingEnabled()
        table.setSortingEnabled(False)

        try:
            # 1. 查找并清理所有现有的汇总行
            r = 0
            while r < table.rowCount():
                if is_summary_row(table, r):
                    table.removeRow(r)
                else:
                    r += 1

            # 2. 在第0行插入新的汇总行
            table.insertRow(0)
            for c in range(table.columnCount()):
                item = QTableWidgetItem()
                item.setData(Qt.UserRole, SUMMARY_ROW_ROLE)
                item.setBackground(QColor(255, 255, 224)) # Light Yellow
                font = QFont()
                font.setBold(True)
                item.setFont(font)
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                table.setItem(0, c, item)

            # 3. 统计数据
            cols = table.columnCount()
            totals = {}
            visible_count = 0
            rows = table.rowCount()

            # 识别数值列
            numeric_cols = set()
            rows_to_check = min(rows, 100)
            for r in range(rows_to_check):
                if table.isRowHidden(r) or is_summary_row(table, r):
                    continue
                for c in range(cols):
                    if c in numeric_cols:
                        continue
                    item = table.item(r, c)
                    if item:
                        raw_text = item.text().strip()
                        # 排除明显的 ID 类长数字 (如 CNPJ, 访问码等)
                        digits_only = re.sub(r'\D', '', raw_text)
                        if len(digits_only) >= 10:
                            if ',' not in raw_text and '.' not in raw_text:
                                continue
                            if len(digits_only) in [14, 44]:
                                continue

                        val = br_to_float(raw_text)
                        # 防止错把流水号当金额
                        if val is not None and abs(val) < 1e9:
                            numeric_cols.add(c)

            for r in range(rows):
                if table.isRowHidden(r) or is_summary_row(table, r):
                    continue
                visible_count += 1
                for c in numeric_cols:
                    item = table.item(r, c)
                    if not item:
                        continue
                    val = br_to_float(item.text())
                    # 再次校验单行金额合理性：单行金额不应超过 1 亿
                    if val is not None and abs(val) < 1e8:
                        totals[c] = totals.get(c, 0.0) + val

            # 4. 更新文本 (汇总和合计)
            for c in range(cols):
                item = table.item(0, c)
                if not item:
                    continue

                header_item = table.horizontalHeaderItem(c)
                header_text = header_item.text() if header_item else ""

                # Check if this column should show Average instead of Sum
                show_avg = False
                avg_keywords = ["单价", "PRICE", "UNIT", "PREÇO", "AVERAGE", "MÉDIA", "RATE", "TAXA", "进度", "PROGRESS", "%"]
                if any(k in header_text.upper() for k in avg_keywords):
                    show_avg = True

                if c == 0:
                    item.setText(f"汇总 ({visible_count})")
                    item.setToolTip(f"当前可见行数: {visible_count}")
                elif c in totals:
                    val = totals[c]
                    avg = val / visible_count if visible_count > 0 else 0

                    if abs(val) > 1e15:
                        item.setText("数据异常")
                    else:
                        sum_txt = f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        avg_txt = f"{avg:,.4f}".replace(",", "X").replace(".", ",").replace("X", ".")

                        if show_avg:
                            item.setText(avg_txt)
                            item.setToolTip(f"平均: {avg_txt}\n合计: {sum_txt}")
                        else:
                            item.setText(sum_txt)
                            item.setToolTip(f"合计: {sum_txt}\n平均: {avg_txt}")

                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                else:
                    item.setText("")
                    item.setToolTip("")
        finally:
            table.setSortingEnabled(was_sorting)

    def apply_header_filters(self, table: QTableWidget):
        """应用表头筛选 (支持多选列表及高级运算符)"""
        header = table.horizontalHeader()
        if not isinstance(header, FilterHeader):
            return

        filters = header._filters
        rows = table.rowCount()

        for r in range(rows):
            if is_summary_row(table, r):
                table.setRowHidden(r, False)
                continue

            visible = True
            for c, criterion in filters.items():
                item = table.item(r, c)
                val = item.text().strip() if item else ""

                if not check_filter_match(val, criterion):
                    visible = False
                    break

            table.setRowHidden(r, not visible)

        self.update_summary_row(table)

    def on_header_clicked(self, logicalIndex):
        """Handle header click for custom sorting with summary row"""
        sender_header = self.sender()
        if not sender_header: return
        table = sender_header.parent()
        if not isinstance(table, QTableWidget): return
        
        # Determine order
        current_order = sender_header.sortIndicatorOrder()
        # Toggle if clicking same column? No, QHeaderView handles visual toggle automatically usually.
        # But since we disabled sorting on table, we might need to manage it.
        # Actually, let's check current indicator section.
        if sender_header.sortIndicatorSection() == logicalIndex:
            new_order = Qt.DescendingOrder if current_order == Qt.AscendingOrder else Qt.AscendingOrder
        else:
            new_order = Qt.AscendingOrder
            
        sender_header.setSortIndicator(logicalIndex, new_order)
        self.sort_with_summary(table, logicalIndex, new_order)

    def sort_with_summary(self, table: QTableWidget, col: int, order: Qt.SortOrder):
        """Sorts table but keeps summary row at top"""
        if table is None:
            return

        r = 0
        while r < table.rowCount():
            if is_summary_row(table, r):
                table.removeRow(r)
            else:
                r += 1

        table.sortItems(col, order)
        self.update_summary_row(table)

    def open_settings(self):
        dialog = SettingsDialog(self)
        dialog.exec()
        self.load_settings() # 重新加载设置以应用
        self.log_message("设置已更新")

    def log_message(self, message):
        logging.info(message)
        self.log_viewer.append(f"[{time.strftime('%H:%M:%S')}] {message}")
        self.log_viewer.verticalScrollBar().setValue(self.log_viewer.verticalScrollBar().maximum())

    def closeEvent(self, event):
        # 保存列状
        self.save_column_state()

        # 在退出前自动保存
        if self.data_modified and self.invoices:
            self.auto_save()

        # 停止后台加载线程，避免退出时残留任务
        if hasattr(self, "payment_tab") and self.payment_tab:
            self.payment_tab.shutdown_workers()
        if hasattr(self, "account_tab") and self.account_tab:
            self.account_tab.shutdown_workers()
        if hasattr(self, "_data_load_worker") and self._data_load_worker and self._data_load_worker.isRunning():
            self._data_load_worker.stop()
            self._data_load_worker.wait(300)

        if self.worker_thread and self.worker_thread.isRunning():
            reply = QMessageBox.question(self, '确认退出', '提取任务仍在进行中，确定要退出吗？',
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                self.worker_thread.stop()
                self.worker_thread.wait() # 等待线程结束
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()

    # ==================== 数据持久化功 ====================

    def save_data(self):
        """手动保存数据到JSON文件"""
        if not self.invoices:
            QMessageBox.warning(self, "无数据", "没有可保存的数据")
            return

        default_filename = f"danfe_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        path, _ = QFileDialog.getSaveFileName(self, "保存数据", default_filename, "JSON Files (*.json)")
        if not path:
            return

        try:
            self.save_data_to_file(path)
            self.data_modified = False
            self.log_message(f"数据已保存到: {path}")
            QMessageBox.information(self, "保存成功", f"数据已成功保存到\n{path}")
        except Exception as e:
            self.log_message(f"保存失败: {e}")
            QMessageBox.critical(self, "保存错误", f"无法保存文件: {e}")

    def load_data(self):
        """从JSON文件加载数据"""
        path, _ = QFileDialog.getOpenFileName(self, "加载数据", "", "JSON Files (*.json)")
        if not path:
            return

        try:
            self.load_data_from_file(path)
            self.log_message(f"数据已从 {path} 加载")
            QMessageBox.information(self, "加载成功", f"已加 {len(self.invoices)} 个发票数")
        except Exception as e:
            self.log_message(f"加载失败: {e}")
            QMessageBox.critical(self, "加载错误", f"无法读取文件: {e}")

    def auto_save(self):
        """自动保存数据"""
        if self.invoices and self.data_modified:
            try:
                self.save_data_to_file(self.auto_save_file, include_payment_history=False, pretty=False)
                self.log_message(f"轻量自动保存完成 ({len(self.invoices)} 个发票)")
                self.data_modified = False
            except Exception as e:
                self.log_message(f"自动保存失败: {e}")

    def refresh_all_tabs(self):
        """刷新所有相关模块的界面 (Refresh all relevant UI tabs)"""
        # 1. 刷新收款模块
        if hasattr(self, 'payment_tab'):
            self.payment_tab.load_invoices()
        
        # 2. 刷新应收账龄模块
        if hasattr(self, 'aging_tab'):
            self.aging_tab.refresh_analysis()
            
        # 3. 刷新数据查询模块
        if hasattr(self, 'query_tab'):
            self.query_tab.refresh()
            
        # 4. 刷新账户管理模块
        if hasattr(self, 'account_tab'):
            self.account_tab.load_accounts()
            
        # 5. 刷新用户中心统计
        if hasattr(self, 'user_tab'):
            self.user_tab.refresh_stats()

    def save_data_to_file(self, filepath: str, include_payment_history: bool = True, pretty: bool = True):
        """保存发票数据到JSON文件

        Args:
            filepath: 输出文件路径
            include_payment_history: 是否包含付款分期明细（大数据量时可关闭以加快自动保存）
            pretty: 是否格式化输出（自动保存建议关闭）
        """
        invoice_list = []
        for inv in self.invoices:
            inv_dict = asdict(inv)
            if include_payment_history:
                # 获取对应的付款明细
                invoice_number = inv.numero or f"FILE-{inv.file_name}"
                inv_dict["payment_history"] = self.payment_tab.db.get_all_installments_for_export(invoice_number)
            else:
                # 轻量自动保存: 保持结构一致但跳过昂贵的分期查询
                inv_dict["payment_history"] = []
            invoice_list.append(inv_dict)

        data = {
            "version": "1.1",
            "saved_at": datetime.now().isoformat(),
            "invoices": invoice_list,
            "recon_data": self.recon_tab.get_state() if hasattr(self, 'recon_tab') else None
        }
        with open(filepath, 'w', encoding='utf-8') as f:
            if pretty:
                json.dump(data, f, ensure_ascii=False, indent=2, default=str)
            else:
                json.dump(data, f, ensure_ascii=False, default=str, separators=(",", ":"))

    def load_data_from_file(self, filepath: str, skip_db_sync: bool = False):
        """从JSON文件加载发票数据 (同时恢复付款明细)

        Args:
            filepath: JSON文件路径
            skip_db_sync: 是否跳过数据库同步（自动加载时使用，避免重复同步）
        """
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # 重建Invoice和Item对象
        self.invoices = []
        import_history = {} # {invoice_number: installments}
        total_invs = len(data.get("invoices", []))

        for idx, inv_dict in enumerate(data.get("invoices", [])):
            # 每 1000 条处理一次 GUI 事件
            if idx % 1000 == 0 and idx > 0:
                self.status_label.setText(f"正在解析数据... {idx}/{total_invs}")
                QApplication.processEvents()

            # 提取明细
            payment_history = inv_dict.pop("payment_history", [])

            # 提取items
            items_data = inv_dict.pop("itens", [])
            items = [Item(**item_dict) for item_dict in items_data]

            # 创建Invoice对象
            inv_dict["itens"] = items
            invoice = Invoice(**inv_dict)
            self.invoices.append(invoice)

            # 暂存付款记录
            inv_key = invoice.numero or f"FILE-{invoice.file_name}"
            import_history[inv_key] = payment_history

        # 恢复对账模块数据
        if "recon_data" in data and hasattr(self, 'recon_tab'):
            self.recon_tab.set_state(data["recon_data"])

        # 1. 更新表格显示
        self.status_label.setText("正在填充表格...")
        QApplication.processEvents()
        self.populate_table()
        self.file_list_label.setText(f"已加载发票 {len(self.invoices)}")
        self.export_btn.setEnabled(True)
        self.data_modified = False

        # 2. 同步基础发票信息到 DB（自动加载时跳过，因为数据已在数据库中）
        if not skip_db_sync:
            self.status_label.setText("正在同步到数据库...")
            QApplication.processEvents()
            self.sync_invoices_to_payment(self.invoices)

            # 3. 恢复详细的付款分期记录
            restore_failed = 0
            for inv_key, history in import_history.items():
                if history:
                    ok = self.payment_tab.db.restore_installments_from_import(inv_key, history)
                    if ok is False:
                        restore_failed += 1
            if restore_failed > 0:
                self.log_message(f"警告: {restore_failed} 张发票的分期明细恢复失败")

        # 4. 最后刷新一次收款界面
        self.payment_tab.load_invoices()

    def sync_invoices_to_payment(self, invoices: List[Invoice]):
        """同步发票到收款系统数据库（优化版：批量查询减少数据库操作）"""
        if not invoices:
            return

        sync_count = 0
        skipped_count = 0
        overwrite_all = False
        skip_all = False

        # --- 批量预取已存在的发票号（避免 N 次单独查询）---
        try:
            existing_numbers = self.payment_tab.db.get_all_existing_invoice_numbers()
        except Exception as e:
            self.log_message(f"获取已存在发票列表失败: {e}")
            existing_numbers = set()

        # --- PRE-FETCH MISSING PDF INVOICES ---
        try:
            pending_list = self.payment_tab.db.get_need_pdf_invoices()
            pending_map = {}
            for pid, pnum in pending_list:
                try:
                    digits = re.sub(r'\D', '', str(pnum))
                    if digits:
                        norm = int(digits)
                        pending_map[norm] = (pid, pnum)
                except:
                    pass
        except Exception as e:
            self.log_message(f"获取待补PDF列表失败: {e}")
            pending_map = {}

        total_count = len(invoices)
        for idx, inv in enumerate(invoices):
            # 每 500 条处理一次 GUI 事件，显示进度
            if idx % 500 == 0:
                self.status_label.setText(f"正在同步到数据库... {idx}/{total_count}")
                QApplication.processEvents()

            # 数据清洗与兜底
            invoice_number = inv.numero
            if not invoice_number:
                invoice_number = f"FILE-{inv.file_name}"
                self.log_message(f"警告: 发票 {inv.file_name} 未识别到发票号，将使用文件名作为标识")

            # --- MATCHING LOGIC ---
            try:
                digits = re.sub(r'\D', '', str(invoice_number))
                if digits:
                    curr_norm = int(digits)
                    if curr_norm in pending_map:
                        match_id, match_num = pending_map[curr_norm]
                        if match_num != invoice_number:
                            renamed = self.payment_tab.db.update_invoice_number(match_id, invoice_number)
                            if renamed is False:
                                self.log_message(f"⚠️ 自动关联失败: 无法将临时发票 {match_num} 更新为 {invoice_number}")
                            else:
                                self.log_message(f"⚡ 自动关联: 发现临时发票 {match_num} (ID: {match_id}) -> 匹配导入发票 {invoice_number}")
                        else:
                            self.log_message(f"⚡ 自动关联: 匹配到临时发票 {invoice_number}")
            except Exception:
                pass

            # 使用批量查询结果判断是否已存在（避免单独查询）
            is_existing = invoice_number in existing_numbers
            if is_existing:
                if skip_all:
                    skipped_count += 1
                    continue
                if not overwrite_all:
                    reply = QMessageBox.question(self, "重复录入",
                                               f"发票号 {invoice_number} ({inv.emitente_nome or '未知'}) 已存在于系统中。\n是否要进行覆盖操作？",
                                               QMessageBox.Yes | QMessageBox.No | QMessageBox.YesToAll | QMessageBox.NoToAll)
                    if reply == QMessageBox.YesToAll:
                        overwrite_all = True
                    elif reply == QMessageBox.NoToAll:
                        skip_all = True
                        skipped_count += 1
                        continue
                    elif reply == QMessageBox.No:
                        skipped_count += 1
                        continue

            try:
                db_data = {
                    "invoice_number": invoice_number,
                    "issuer_name": inv.emitente_nome or "Unknown",
                    "issuer_cnpj": inv.emitente_cnpj or "",
                    "issue_date": inv.data_emissao or "",
                    "total_amount": inv.total_nota or 0.0,
                    "file_name": inv.file_name,
                    "natureza_operacao": inv.natureza_operacao or "",
                    "destinatario_name": inv.destinatario_nome or "Unknown",
                    "destinatario_cnpj": inv.destinatario_cnpj or ""
                }
                inv_id = self.payment_tab.db.upsert_invoice(db_data)
                if not isinstance(inv_id, int) or inv_id <= 0:
                    inv_id = self.payment_tab.db.find_invoice_id_by_number(invoice_number)
                if not isinstance(inv_id, int) or inv_id <= 0:
                    raise RuntimeError("upsert_invoice 未返回有效 ID")

                # 默认生成1期计划 (如果不存在)
                installments = self.payment_tab.db.get_installments(inv_id) or []
                if installments == []:
                    idate = safe_parse_date(inv.data_emissao)
                    plan_ok = self.payment_tab.db.generate_payment_plan(inv_id, 1, start_date=idate)
                    if plan_ok is False:
                        raise RuntimeError("生成默认分期计划失败")
                sync_count += 1
                # 更新缓存
                existing_numbers.add(invoice_number)
            except Exception as e:
                self.log_message(f"同步发票 {invoice_number} 到收款系统失败: {e}")
                skipped_count += 1

        if sync_count > 0:
            self.payment_tab.load_invoices()
            self.log_message(f"已自动同步 {sync_count} 个发票到收款系统")

        if skipped_count > 0:
            self.log_message(f"警告: {skipped_count} 个发票同步失败或跳过")

    def load_product_mapping(self):
        """从JSON文件加载产品编码映射"""
        if os.path.exists(self.mapping_file):
            try:
                with open(self.mapping_file, 'r', encoding='utf-8') as f:
                    loaded_mapping = json.load(f)
                cleaned_mapping, rejected_mapping = MappingManager.sanitize_legacy_product_mapping(loaded_mapping)
                self.product_code_mapping = cleaned_mapping
                self.log_message(f"已加 {len(self.product_code_mapping)} 个编码映")
                if rejected_mapping:
                    self.log_message(f"已忽略 {len(rejected_mapping)} 条旧版疑似金额/脏数据映射")
            except Exception as e:
                self.log_message(f"加载映射文件失败: {e}")
                self.product_code_mapping = {}
        else:
            self.product_code_mapping = {}
    def sync_single_invoice_to_db(self, invoice: Invoice):
        """将单个发票的修改同步到收款系统数据库"""
        try:
            # 只有当发票已存在 PaymentManager 时才更新
            # 我们通过 upsert_invoice 更新基本信息
            
            # 构造数据字典，类似 sync_invoices_to_payment 中的逻辑
            invoice_number = invoice.numero
            if not invoice_number:
                invoice_number = f"FILE-{invoice.file_name}"
            
            total_amount = invoice.total_nota if invoice.total_nota is not None else 0.0
            issuer_name = invoice.emitente_nome if invoice.emitente_nome else "未知开票人"
            issue_date = invoice.data_emissao if invoice.data_emissao else datetime.now().strftime("%d/%m/%Y")
            
            data = {
                'invoice_number': invoice_number,
                'issuer_name': issuer_name,
                'issuer_cnpj': invoice.emitente_cnpj or "",
                'issue_date': issue_date,
                'total_amount': total_amount,
                'file_name': invoice.file_name,
                'natureza_operacao': invoice.natureza_operacao,
                'destinatario_name': invoice.destinatario_nome or "",
                'destinatario_cnpj': invoice.destinatario_cnpj or ""
            }            
            # 更新数据
            inv_id = self.payment_tab.db.upsert_invoice(data)
            if not isinstance(inv_id, int) or inv_id <= 0:
                inv_id = self.payment_tab.db.find_invoice_id_by_number(invoice_number)
            if not isinstance(inv_id, int) or inv_id <= 0:
                raise RuntimeError("同步发票更新失败：未获取有效数据库 ID")
            
            # 刷新收款界面的列表（如果它当前显示的是发票列表）
            # 只有在收款页面可见时才刷新，避免太频
            if self.tab_widget.currentWidget() == self.payment_tab:
                self.payment_tab.load_invoices()
                
        except Exception as e:
            self.log_message(f"同步发票更新失败: {e}")

    def save_product_mapping(self):
        """保存产品编码映射到JSON文件"""
        try:
            cleaned_mapping, rejected_mapping = MappingManager.sanitize_legacy_product_mapping(self.product_code_mapping)
            self.product_code_mapping = cleaned_mapping
            with open(self.mapping_file, 'w', encoding='utf-8') as f:
                json.dump(self.product_code_mapping, f, ensure_ascii=False, indent=2)
            self.log_message(f"已保 {len(self.product_code_mapping)} 个编码映")
            if rejected_mapping:
                self.log_message(f"保存时忽略了 {len(rejected_mapping)} 条疑似金额/脏数据映射")
            return True
        except Exception as e:
            self.log_message(f"保存映射文件失败: {e}")
            return False

    def apply_mapping_to_items(self):
        """将映射应用到所有Item对象（冲突感知模式）"""
        stats = {
            "scanned": 0,
            "auto_pass": 0,
            "code_only": 0,
            "name_only": 0,
            "conflict": 0,
            "unmapped": 0,
            "applied": 0
        }
        self.product_mapping_conflicts = []

        for invoice in self.invoices:
            for idx, item in enumerate(invoice.itens or []):
                raw_code = str(item.codigo_produto or "").strip()
                raw_desc = str(item.descricao or "").strip()
                if not raw_code and not raw_desc:
                    continue

                stats["scanned"] += 1
                evaluated = self.mapping_mgr.evaluate_product_match(raw_code, raw_desc)
                status = evaluated.get("status", "UNMAPPED")
                key = status.lower()
                if key in stats:
                    stats[key] += 1

                code_candidate = evaluated.get("code_candidate")
                name_candidate = evaluated.get("name_candidate")
                apply_candidate = None
                if status == "AUTO_PASS":
                    apply_candidate = code_candidate or name_candidate

                if apply_candidate and apply_candidate.get("std_code"):
                    item.codigo_domestico = apply_candidate.get("std_code")
                    stats["applied"] += 1
                    continue

                if not isinstance(item.tags, list):
                    item.tags = []
                if status in ("CODE_ONLY", "NAME_ONLY", "CONFLICT"):
                    if "待确认" not in item.tags:
                        item.tags.append("待确认")
                    if status == "CONFLICT" and "映射冲突" not in item.tags:
                        item.tags.append("映射冲突")

                    self.product_mapping_conflicts.append({
                        "resolution": None,
                        "status": status,
                        "invoice_number": invoice.numero or f"FILE-{invoice.file_name}",
                        "file_name": invoice.file_name or "",
                        "item_index": idx,
                        "raw_code": raw_code,
                        "raw_name": raw_desc,
                        "code_candidate": code_candidate,
                        "name_candidate": name_candidate,
                        "code_score": code_candidate.get("score", 0.0) if code_candidate else 0.0,
                        "name_score": name_candidate.get("score", 0.0) if name_candidate else 0.0,
                        "item_ref": item
                    })

        return stats

    def _resolve_product_conflict(self, conflict: dict, strategy: str) -> bool:
        if strategy == "code":
            candidate = conflict.get("code_candidate")
        elif strategy == "name":
            candidate = conflict.get("name_candidate")
        else:
            return False

        if not candidate:
            return False

        item = conflict.get("item_ref")
        std_code = str(candidate.get("std_code", "")).strip()
        std_name = str(candidate.get("std_name", "")).strip()
        raw_code = str(conflict.get("raw_code", "")).strip()

        if item:
            if std_code:
                item.codigo_domestico = std_code
            if not isinstance(item.tags, list):
                item.tags = []
            item.tags = [t for t in item.tags if t not in ("待确认", "映射冲突")]
            if "已核实" not in item.tags:
                item.tags.append("已核实")

        if raw_code and (std_code or std_name):
            current = self.mapping_mgr.get_product_std(raw_code) or {}
            final_code = std_code or str(current.get("std_code", "")).strip()
            final_name = std_name or str(current.get("std_name", "")).strip()
            if final_code or final_name:
                self.mapping_mgr.set_product_std(raw_code, final_code, final_name, source="conflict_resolution", status="reviewed")
        return True

    def show_product_mapping_conflicts(self):
        unresolved = [c for c in self.product_mapping_conflicts if not c.get("resolution")]
        if not unresolved:
            QMessageBox.information(self, "映射冲突", "当前没有待处理的产品映射冲突/待确认项。")
            return

        dlg = ProductConflictBatchDialog(unresolved, self)
        if not dlg.exec():
            return

        changed = False
        processed = 0
        for c, strategy in dlg.decisions:
            if strategy == "ignored":
                c["resolution"] = "ignored"
                processed += 1
            elif strategy == "code":
                if self._resolve_product_conflict(c, "code"):
                    c["resolution"] = "code"
                    changed = True
                    processed += 1
                else:
                    c["resolution"] = "failed"
            elif strategy == "name":
                if self._resolve_product_conflict(c, "name"):
                    c["resolution"] = "name"
                    changed = True
                    processed += 1
                else:
                    c["resolution"] = "failed"

        if changed:
            self.mapping_mgr.save()
            self.populate_table()
            self.data_modified = True

        remaining = len([c for c in self.product_mapping_conflicts if not c.get("resolution")])
        self.log_message(f"映射冲突处理完成: 已处理 {processed} 条，剩余 {remaining} 条。")

    def manage_mappings(self):
        """打开编码映射管理对话框"""
        dialog = MappingLibraryDialog(self.mapping_mgr, self.invoices, self)
        if dialog.exec():
            stats = self.apply_mapping_to_items()
            self.populate_table()
            self.data_modified = True
            pending = stats["code_only"] + stats["name_only"] + stats["conflict"]
            self.log_message(
                f"映射扫描完成: 自动通过 {stats['auto_pass']}，待确认 {pending}，未命中 {stats['unmapped']}，已自动应用 {stats['applied']}"
            )
            if pending > 0:
                reply = QMessageBox.question(
                    self,
                    "发现待确认映射",
                    f"发现 {pending} 条待确认项（其中冲突 {stats['conflict']} 条）。\n是否现在进入处理队列？",
                    QMessageBox.Yes | QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    self.show_product_mapping_conflicts()

    def on_table_item_changed(self, item: QTableWidgetItem):
        """表格单元格被修改时调用"""
        if not self.invoices:
            return

        # 排除汇总行 (如果是汇总行被修改，直接忽略)
        if item.data(Qt.UserRole) == "SUMMARY" or item.text().startswith("汇 ("):
            return

        self.data_modified = True
        row = item.row()
        col = item.column()
        new_value = item.text()

        # 调整 Row Index (如果存在汇总行，数据行索引需要减1)
        first_item = self.table.item(0, 0)
        if first_item and first_item.data(Qt.UserRole) == "SUMMARY":
            row -= 1
            if row < 0: return

        try:
            # 同步修改到invoices数据
            self.sync_table_to_invoices(row, col, new_value)
            self.log_message(f"已修改行{row+1}，列{col+1}: {new_value}")
            self.update_summary_row(self.table)
        except Exception as e:
            self.log_message(f"数据同步失败: {e}")

    def sync_table_to_invoices(self, row: int, col: int, value: str):
        """将表格修改同步到invoices数据结构"""
        # 找到对应的invoice和item
        current_row = 0
        target_invoice = None
        target_item = None

        for invoice in self.invoices:
            items_count = len(invoice.itens) if invoice.itens else 1
            if current_row <= row < current_row + items_count:
                target_invoice = invoice
                if invoice.itens:
                    item_index = row - current_row
                    target_item = invoice.itens[item_index]
                break
            current_row += items_count

        if not target_invoice:
            return

        # 根据列索引更新对应字段
        col_name = self.combined_headers[col]

        # Invoice字段
        if col == 0: target_invoice.file_name = value
        elif col == 1: target_invoice.numero = value
        elif col == 2: target_invoice.serie = value
        elif col == 3: target_invoice.natureza_operacao = value
        elif col == 4: target_invoice.chave_acesso = value
        elif col == 5: target_invoice.protocolo_autorizacao = value
        elif col == 6: target_invoice.data_emissao = value
        elif col == 7: target_invoice.data_saida_entrada = value
        elif col == 8: target_invoice.total_nota = br_to_float(value)
        elif col == 9: target_invoice.total_icms = br_to_float(value)
        elif col == 10: target_invoice.base_calculo_icms = br_to_float(value) # Added
        elif col == 11: target_invoice.base_calculo_st = br_to_float(value)   # Added
        elif col == 12: target_invoice.valor_icms_st = br_to_float(value)     # Added
        elif col == 13: target_invoice.total_ipi = br_to_float(value)         # Shifted +3
        elif col == 14: target_invoice.frete = br_to_float(value)             # Shifted +3
        elif col == 15: target_invoice.seguro = br_to_float(value)            # Shifted +3
        elif col == 16: target_invoice.desconto_total = br_to_float(value)    # Shifted +3
        elif col == 17: target_invoice.outras_despesas = br_to_float(value)   # Shifted +3
        elif col == 18: target_invoice.emitente_nome = value                  # Shifted +3
        elif col == 19: target_invoice.emitente_cnpj = value                  # Shifted +3
        elif col == 20: target_invoice.emitente_ie = value                    # Shifted +3
        elif col == 21: target_invoice.destinatario_nome = value              # Shifted +3
        elif col == 22: target_invoice.destinatario_cnpj = value              # Shifted +3
        elif col == 23: target_invoice.destinatario_ie = value                # Shifted +3
        elif col == 24: target_invoice.destinatario_endereco = value          # Shifted +3
        elif col == 25: target_invoice.destinatario_bairro = value            # Shifted +3
        elif col == 26: target_invoice.destinatario_municipio = value         # Shifted +3
        elif col == 27: target_invoice.destinatario_uf = value                # Shifted +3
        elif col == 28: target_invoice.destinatario_cep = value               # Shifted +3
        elif col == 29: target_invoice.destinatario_fone = value              # Shifted +3
        elif col == 30: target_invoice.transportador_nome = value             # Shifted +3
        elif col == 31: target_invoice.transportador_cnpjcpf = value          # Shifted +3
        elif col == 32: target_invoice.modalidade_frete_raw = value           # Shifted +3
        elif col == 33: target_invoice.placa_veiculo = value                  # Shifted +3
        elif col == 34: target_invoice.uf_veiculo = value                     # Shifted +3
        elif col == 35: target_invoice.peso_bruto = br_to_float(value)        # Shifted +3
        elif col == 36: target_invoice.peso_liquido = br_to_float(value)      # Shifted +3
        elif col == 37: target_invoice.info_compl_contribuinte = value        # Shifted +3
        elif col == 38: target_invoice.llm_table_note = value
        elif col == 39:
            try:
                # Meta is stored as string representation of dict
                import ast
                target_invoice.extract_meta = ast.literal_eval(value)
            except: pass
        elif col == 40:
            target_invoice.tags = [t.strip() for t in value.split(",") if t.strip()]
        elif col == 41:
            pass # 收款进度 is read-only or calculated

        # Item字段 (如果有target_item)
        if target_item:
            item_col_start = len(self.invoice_headers)
            if col == item_col_start + 0: target_item.codigo_produto = value
            elif col == item_col_start + 1:
                target_item.codigo_domestico = value
                # 如果用户编辑了国内编码，同时更新映射
                if target_item.codigo_produto and value:
                    # Keep existing std_name if any, or use current desc
                    current = self.mapping_mgr.get_product_std(target_item.codigo_produto) or {}
                    std_name = current.get("std_name") or target_item.descricao or ""
                    self.mapping_mgr.set_product_std(target_item.codigo_produto, value, std_name, source="table_edit", status="reviewed", sample_text=target_item.descricao or "")
                    self.mapping_mgr.save()
            elif col == item_col_start + 2: target_item.descricao = value
            elif col == item_col_start + 3: target_item.ncm = value
            elif col == item_col_start + 4: target_item.cst = value
            elif col == item_col_start + 5: target_item.cfop = value
            elif col == item_col_start + 6: target_item.unidade = value
            elif col == item_col_start + 7: target_item.quantidade = br_to_float(value)
            elif col == item_col_start + 8: target_item.valor_unitario = br_to_float(value)
            elif col == item_col_start + 9: target_item.valor_total = br_to_float(value)
            elif col == item_col_start + 10: target_item.bc_icms = br_to_float(value)
            elif col == item_col_start + 11: target_item.valor_icms = br_to_float(value)
            elif col == item_col_start + 12: target_item.aliquota_icms = br_to_float(value)
            elif col == item_col_start + 13: target_item.valor_ipi = br_to_float(value)
            elif col == item_col_start + 14: target_item.aliquota_ipi = br_to_float(value)
            elif col == item_col_start + 15: target_item.desconto = br_to_float(value)

        # 同步更新到收款数据库
        if target_invoice:
            self.sync_single_invoice_to_db(target_invoice)

    def on_summary_item_changed(self, item: QTableWidgetItem):
        """汇总表格修改处"""
        if not self.invoices:
            return
            
        self.data_modified = True
        col = item.column()
        new_value = item.text()
        
        # 获取原始发票索引 (支持排序)
        original_index = item.data(Qt.UserRole)
        if isinstance(original_index, str):
            if not original_index.isdigit():
                return
            original_index = int(original_index)
        elif isinstance(original_index, float):
            if not original_index.is_integer():
                return
            original_index = int(original_index)
        elif not isinstance(original_index, int):
            return

        if 0 <= original_index < len(self.invoices):
            invoice = self.invoices[original_index]
            self.update_invoice_field(invoice, col, new_value)
            self.update_summary_row(self.summary_table)

    def update_invoice_field(self, invoice: Invoice, col: int, value: str):
        """根据列索引更新发票字"""
        # 注意：这里的 col 是基 invoice_headers 的索
        if col == 0: invoice.file_name = value
        elif col == 1: invoice.numero = value
        elif col == 2: invoice.serie = value
        elif col == 3: invoice.natureza_operacao = value
        elif col == 4: invoice.chave_acesso = value
        elif col == 5: invoice.protocolo_autorizacao = value
        elif col == 6: invoice.data_emissao = value
        elif col == 7: invoice.data_saida_entrada = value
        elif col == 8: invoice.total_nota = br_to_float(value)
        elif col == 9: invoice.total_icms = br_to_float(value)
        elif col == 10: invoice.base_calculo_icms = br_to_float(value)
        elif col == 11: invoice.base_calculo_st = br_to_float(value)
        elif col == 12: invoice.valor_icms_st = br_to_float(value)
        elif col == 13: invoice.total_ipi = br_to_float(value)
        elif col == 14: invoice.frete = br_to_float(value)
        elif col == 15: invoice.seguro = br_to_float(value)
        elif col == 16: invoice.desconto_total = br_to_float(value)
        elif col == 17: invoice.outras_despesas = br_to_float(value)
        elif col == 18: invoice.emitente_nome = value
        elif col == 19: invoice.emitente_cnpj = value
        elif col == 20: invoice.emitente_ie = value
        elif col == 21: invoice.destinatario_nome = value
        elif col == 22: invoice.destinatario_cnpj = value
        elif col == 23: invoice.destinatario_ie = value
        elif col == 24: invoice.destinatario_endereco = value
        elif col == 25: invoice.destinatario_bairro = value
        elif col == 26: invoice.destinatario_municipio = value
        elif col == 27: invoice.destinatario_uf = value
        elif col == 28: invoice.destinatario_cep = value
        elif col == 29: invoice.destinatario_fone = value
        elif col == 30: invoice.transportador_nome = value
        elif col == 31: invoice.transportador_cnpjcpf = value
        elif col == 32: invoice.modalidade_frete_raw = value
        elif col == 33: invoice.placa_veiculo = value
        elif col == 34: invoice.uf_veiculo = value
        elif col == 35: invoice.peso_bruto = br_to_float(value)
        elif col == 36: invoice.peso_liquido = br_to_float(value)
        elif col == 37: invoice.info_compl_contribuinte = value
        elif col == 38: invoice.llm_table_note = value
        elif col == 39:
            try:
                import ast
                invoice.extract_meta = ast.literal_eval(value)
            except: pass
        elif col == 40:
            invoice.tags = [t.strip() for t in value.split(",") if t.strip()]
        
        # 同步更新到收款数据库
        self.sync_single_invoice_to_db(invoice)

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    
    # 加载设置并应用
    settings = QSettings(SETTINGS_FILE, QSettings.IniFormat)
    theme = settings.value("app_theme", "System")
    ThemeManager.apply_theme(app, theme)
    
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
